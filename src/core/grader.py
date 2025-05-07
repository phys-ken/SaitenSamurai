"""
解答用紙の採点結果管理を行うモジュール
"""
import os
import csv
import glob
import shutil
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple, Union
import openpyxl
from PIL import Image
import tempfile

from ..models.grade_data import GradeData
from ..utils.file_utils import SETTING_DIR, get_sorted_image_files


class Grader:
    """採点結果の収集・分析クラス"""
    
    def __init__(self, output_dir: str = None, excel_path: str = None):
        """
        初期化処理
        
        Args:
            output_dir: 分割された解答用紙の保存ディレクトリ
            excel_path: 採点結果を出力するExcelファイルのパス
        """
        self.output_dir = output_dir if output_dir else str(SETTING_DIR / "output")
        self.excel_path = excel_path if excel_path else str(SETTING_DIR / "saiten.xlsx")
        print(f"出力ディレクトリ: {self.output_dir}")
        print(f"Excel出力先: {self.excel_path}")
    
    def get_question_directories(self) -> List[str]:
        """
        出力ディレクトリ内の問題ごとのディレクトリ名を取得します
        
        Returns:
            List[str]: 問題ディレクトリ名のリスト
        """
        try:
            dirs = []
            output_dir_path = Path(self.output_dir)
            if not output_dir_path.exists():
                print(f"ディレクトリが存在しません: {self.output_dir}")
                return []
            
            for item in output_dir_path.iterdir():
                if item.is_dir():
                    dirs.append(item.name)
            
            print(f"問題ディレクトリ一覧: {dirs}")
            return sorted(dirs)
        except Exception as e:
            print(f"問題ディレクトリの取得中にエラー: {e}")
            return []
    
    def get_student_files_for_question(self, question_id: str) -> List[str]:
        """
        指定した問題の学生解答ファイル一覧を取得します
        
        Args:
            question_id: 問題ID
            
        Returns:
            List[str]: 学生ファイル名のリスト
        """
        question_dir = os.path.join(self.output_dir, question_id)
        if not os.path.exists(question_dir):
            return []
        
        # 未採点ファイル（最上位ディレクトリにあるファイル）を取得
        try:
            files = [os.path.basename(f) for f in get_sorted_image_files(os.path.join(question_dir, "*"))]
        except Exception as e:
            print(f"ファイル検索エラー: {e}")
            # 直接glob関数を使用
            files = [os.path.basename(f) for f in glob.glob(os.path.join(question_dir, "*.*")) 
                   if os.path.isfile(f) and f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))]
        
        return files
    
    def get_graded_student_files_for_question(self, question_id: str) -> Dict[str, str]:
        """
        採点済みの学生ファイル（スコア別にディレクトリ分けされたもの）を取得します
        
        Args:
            question_id: 問題ID
            
        Returns:
            Dict[str, str]: {学生ファイル名: スコア} の辞書
        """
        question_dir = os.path.join(self.output_dir, question_id)
        result = {}
        
        if not os.path.exists(question_dir):
            print(f"問題ディレクトリが存在しません: {question_dir}")
            return result
        
        try:
            # サブディレクトリ（スコア）を取得
            for score_dir in [d for d in os.listdir(question_dir) if os.path.isdir(os.path.join(question_dir, d))]:
                score_path = os.path.join(question_dir, score_dir)
                # 各スコアディレクトリ内のファイルを取得
                for file_path in glob.glob(os.path.join(score_path, "*.*")):
                    if os.path.isfile(file_path) and file_path.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                        file_name = os.path.basename(file_path)
                        result[file_name] = score_dir
        except Exception as e:
            print(f"採点済みファイルの取得中にエラー: {e}")
        
        return result
    
    def grade_answer(self, question_id: str, student_file: str, score: Union[int, str]) -> bool:
        """
        解答を採点し、対応するディレクトリに移動します
        
        Args:
            question_id: 問題ID
            student_file: 学生ファイル名
            score: 採点結果（数値またはskipなどの文字列）
            
        Returns:
            bool: 処理成功時True、失敗時False
        """
        try:
            # 元のファイルパスと移動先パス
            question_dir = os.path.join(self.output_dir, question_id)
            original_path = os.path.join(question_dir, student_file)
            
            # ファイルの存在確認
            if not os.path.exists(original_path):
                print(f"採点対象ファイルが存在しません: {original_path}")
                return False
            
            # スコアディレクトリを作成
            score_dir = os.path.join(question_dir, str(score))
            os.makedirs(score_dir, exist_ok=True)
            
            # ファイルを移動
            target_path = os.path.join(score_dir, student_file)
            shutil.move(original_path, target_path)
            print(f"ファイル移動: {original_path} → {target_path}")
            
            return True
        except Exception as e:
            print(f"採点処理中にエラーが発生しました: {e}")
            return False
    
    def create_excel_report(self) -> bool:
        """
        採点結果をExcelファイルに出力します
        
        Returns:
            bool: 成功時True、失敗時False
        """
        try:
            print("Excel出力を開始します")
            
            # 採点データを収集
            grade_data = self._collect_grade_data()
            if not grade_data:
                print("採点データがありません")
                return False
            
            print(f"採点データ取得完了: {list(grade_data.keys())}")
            
            # Excelブックを新規作成またはロード
            try:
                wb = openpyxl.load_workbook(self.excel_path)
                print(f"既存のExcelファイルをロードしました: {self.excel_path}")
                # ワークシートがなければ作成、あれば初期化
                if "採点シート" in wb.sheetnames:
                    ws = wb["採点シート"]
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                        for cell in row:
                            cell.value = None
                else:
                    ws = wb.create_sheet("採点シート")
            except Exception as e:
                print(f"新規Excelファイルを作成します (理由: {e})")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "採点シート"
            
            # ヘッダ行の作成 - オリジナルのsaitenGiri2021と同じ形式
            headers = ["ファイル名", "画像", "生徒番号", "名前"]
            
            # 問題列のヘッダを追加
            for q_dir in sorted(grade_data.keys()):
                if q_dir.startswith("Q_"):
                    headers.append(q_dir)
            
            # 合計列
            headers.append("合計")
            
            # ヘッダー行を設定
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx).value = header
            
            # 解答用紙ファイル名を収集
            answer_sheets = set()
            for question_dir, files_dict in grade_data.items():
                answer_sheets.update(files_dict.keys())
            
            # 一時ディレクトリを作成（openpyxlの画像処理用）
            temp_dir = os.path.join(os.path.dirname(self.excel_path), "temp_excel_images")
            os.makedirs(temp_dir, exist_ok=True)
            print(f"一時ディレクトリを作成しました: {temp_dir}")
            
            # 使用された一時ファイルのパスを記録するリスト
            temp_files_created = []
            
            # 行の最大高さを格納する配列
            max_height = []
            max_width = 0
            
            try:
                # 各解答用紙の行を作成
                for row_idx, sheet_name in enumerate(sorted(answer_sheets), 2):
                    print(f"処理中: {sheet_name}")
                    
                    # A列：ファイル名
                    ws.cell(row=row_idx, column=1).value = sheet_name
                    
                    # B列：画像を追加（nameフォルダの画像）
                    if "name" in grade_data:
                        name_dir_path = os.path.join(self.output_dir, "name")
                        name_img_path = None
                        
                        # nameディレクトリ内のファイルを検索
                        if os.path.exists(name_dir_path):
                            # まず未採点ファイルを確認
                            direct_file = os.path.join(name_dir_path, sheet_name)
                            if os.path.exists(direct_file) and os.path.isfile(direct_file):
                                name_img_path = direct_file
                                print(f"未採点の名前画像を発見: {direct_file}")
                            else:
                                # 採点済みファイルを検索
                                for score_dir in os.listdir(name_dir_path):
                                    score_path = os.path.join(name_dir_path, score_dir)
                                    if os.path.isdir(score_path):
                                        scored_file = os.path.join(score_path, sheet_name)
                                        if os.path.exists(scored_file) and os.path.isfile(scored_file):
                                            name_img_path = scored_file
                                            print(f"採点済みの名前画像を発見: {scored_file}")
                                            break
                        
                        # 画像が見つかった場合は挿入
                        if name_img_path and os.path.exists(name_img_path):
                            print(f"名前画像を挿入します: {name_img_path}")
                            try:
                                # 画像サイズを取得
                                pil_img = Image.open(name_img_path)
                                width, height = pil_img.size
                                
                                # 配列「max_height」において、「row_idx-2」番目の要素が存在しなければ、挿入
                                while len(max_height) < row_idx - 1:
                                    max_height.append(0)
                                
                                # 最大の高さと幅を更新
                                if max_height[row_idx-2] < height:
                                    max_height[row_idx-2] = height
                                if max_width < width:
                                    max_width = width
                                
                                # 行高さをピクセルに基づいて調整（元のsaitengiri2021.pyと同じ係数を使用）
                                ws.row_dimensions[row_idx].height = max_height[row_idx-2] * 0.75
                                
                                # B列の幅を調整（元のsaitengiri2021.pyと同じ係数を使用）
                                ws.column_dimensions['B'].width = max_width * 0.13
                                
                                # 一時ファイルを作成して画像を挿入
                                temp_img_path = os.path.join(temp_dir, f"temp_{sheet_name}")
                                # 画像をコピーして使用
                                shutil.copy2(name_img_path, temp_img_path)
                                temp_files_created.append(temp_img_path)
                                
                                # 画像オブジェクトを作成
                                img = openpyxl.drawing.image.Image(temp_img_path)
                                
                                # セル座標を取得して画像をアンカー
                                cell_address = ws.cell(row=row_idx, column=2).coordinate
                                img.anchor = cell_address
                                
                                # 画像をシートに追加
                                ws.add_image(img)
                                print(f"画像を追加しました: {temp_img_path}")
                                
                            except Exception as e:
                                print(f"画像追加エラー: {e}")
                                # 画像追加に失敗しても続行
                    
                    # 各問題の得点を入力
                    total_score = 0
                    for col_idx, header in enumerate(headers[4:], 5):  # 問題列は5列目（E列）から
                        if header == "合計":
                            # 合計点を表示
                            ws.cell(row=row_idx, column=col_idx).value = total_score
                        else:
                            # 各問題の点数を表示
                            question_id = header
                            if question_id in grade_data and sheet_name in grade_data[question_id]:
                                score = grade_data[question_id][sheet_name]
                                try:
                                    # 数値に変換できる場合は数値として表示
                                    score_value = int(score)
                                    ws.cell(row=row_idx, column=col_idx).value = score_value
                                    total_score += score_value  # 合計点に加算
                                except ValueError:
                                    # "skip"など、数値以外の場合
                                    ws.cell(row=row_idx, column=col_idx).value = score
                            else:
                                # データがない場合は空欄
                                ws.cell(row=row_idx, column=col_idx).value = ""
                
                # Excelファイルを保存
                wb.save(self.excel_path)
                print(f"Excelファイルを保存しました: {self.excel_path}")
                
                return True
                
            finally:
                # 一時ファイルの削除処理
                print(f"一時ファイルの削除を開始します...")
                for temp_file in temp_files_created:
                    if os.path.exists(temp_file):
                        try:
                            os.remove(temp_file)
                            print(f"一時ファイルを削除しました: {temp_file}")
                        except Exception as e:
                            print(f"一時ファイルの削除中にエラー: {e}")
                
                # 一時ディレクトリの削除
                try:
                    if os.path.exists(temp_dir) and len(os.listdir(temp_dir)) == 0:
                        os.rmdir(temp_dir)
                        print(f"空の一時ディレクトリを削除しました: {temp_dir}")
                    elif os.path.exists(temp_dir):
                        print(f"一時ディレクトリにまだファイルが残っているため削除しません: {temp_dir}")
                except Exception as e:
                    print(f"一時ディレクトリの削除中にエラー: {e}")
                
        except Exception as e:
            print(f"Excelレポート作成中にエラーが発生しました: {e}")
            return False
    
    def _collect_grade_data(self) -> Dict[str, Dict[str, str]]:
        """
        採点結果データを収集します
        
        Returns:
            Dict[str, Dict[str, str]]: {問題ID: {ファイル名: スコア}} の辞書
        """
        grade_data = {}
        
        # 出力ディレクトリの存在確認
        if not os.path.exists(self.output_dir):
            print(f"出力ディレクトリが存在しません: {self.output_dir}")
            return grade_data
        
        # 各問題ディレクトリを処理
        for question_dir in self.get_question_directories():
            question_path = os.path.join(self.output_dir, question_dir)
            grade_data[question_dir] = {}
            
            # 問題ディレクトリ内の未採点ファイル一覧を取得
            for item in os.listdir(question_path):
                item_path = os.path.join(question_path, item)
                
                if os.path.isdir(item_path):
                    # スコアディレクトリ内のファイル一覧を取得
                    score = os.path.basename(item_path)
                    for filename in os.listdir(item_path):
                        file_path = os.path.join(item_path, filename)
                        if os.path.isfile(file_path) and file_path.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                            grade_data[question_dir][filename] = score
                elif os.path.isfile(item_path) and item_path.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                    # 直接ファイルがある場合は採点前の状態
                    grade_data[question_dir][os.path.basename(item_path)] = "未"
        
        return grade_data