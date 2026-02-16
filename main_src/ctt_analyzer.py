"""
CTT (Classical Test Theory) 分析モジュール

古典テスト理論に基づくテスト分析機能を提供する。
Mark2の採点結果（配点ベース）を0/1バイナリに変換し、
CTT統計量（α係数, P値, D値, I-T相関, 選択肢分析）を算出。
Excel（openpyxl）とPDF（reportlab+matplotlib）のレポートを生成する。

採点ルール: score_answers() と完全に同一のロジック
  - 0⇔10等価判定（10番目のマーク位置 = 選択肢"0"）
  - 複数正答 (';' 区切り) は集合の一致判定
"""

import io
import logging
import sys
import re
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    Paragraph as ChartParagraph,
    ParagraphProperties,
    CharacterProperties,
)
from openpyxl.drawing.text import Font as ChartFont

# Optional: matplotlib for CTT plots
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    plt.rcParams['font.family'] = 'MS Gothic'
    HAS_MATPLOTLIB = True
except ImportError:
    matplotlib = None
    plt = None
    HAS_MATPLOTLIB = False

# Optional: reportlab for PDF reports
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph as RLParagraph, Spacer,
        Table as RLTable, TableStyle, Image as RLImage,
        PageBreak, KeepTogether,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from constants import safe_print
from scoring_engine import (
    number_to_circled,
    normalize_value,
    load_template,
    load_mark2_results,
    score_answers,
)


def convert_mark2_to_ctt_data(template_path, mark2_result_path, skip_questions=0,
                              descriptive_config=None, descriptive_scores=None,
                              template_dict=None, mark2_results=None):
    """
    Mark2のデータをCTT分析用のDataFrameに変換する。
    
    Mark2: 配点ベース・選択肢文字列
    CTT:   0/1バイナリ・設問ID文字列
    
    注意: マークシートは最大10択択肢(1,2,...,9,0)。
    10番目のマーク位置 = 選択肢"0"。"選択肢10"は存在しない。
    テンプレートの正答に"10"が記載されている場合は"0"に正規化する。
    
    記述問題の統合:
        descriptive_config と descriptive_scores が指定されている場合、
        記述問題をバイナリ(0/1)項目として追加する。
        - 満点 → 正答(1), 解答="1"
        - 満点未満 → 誤答(0), 解答="0"
        CTTの選択肢分析では正答="1" の2値問題として扱われる。
    
    Args:
        template_path: テンプレートExcelのパス（記述のみモードではNone可）
        mark2_result_path: Mark2読取結果Excelのパス（記述のみモードではNone可）
        skip_questions: スキップする列数（学籍番号等）
        descriptive_config: 記述問題設定dict（オプション）
        descriptive_scores: {ファイル名: {問題ID: 得点}} の辞書（オプション）
        template_dict: 事前読込済みテンプレート（省略時は内部でExcelパース）
        mark2_results: 事前読込済みOMR結果（省略時は内部でExcelパース）
    
    Returns:
        tuple: (ans_df, key_df)
            ans_df: DataFrame (行=学生, 列=設問ID文字列, 値=選択肢文字列)
            key_df: DataFrame (QuestionID, Key)
    """
    # --- 記述のみモード: マーク問題なし ---
    has_mark = (template_path is not None or template_dict is not None)
    
    if has_mark:
        if template_dict is None:
            template_dict = load_template(template_path)
        if mark2_results is None:
            mark2_results = load_mark2_results(mark2_result_path, skip_questions)
        
        question_numbers = sorted(template_dict.keys())
        questions = [str(q) for q in question_numbers]
        
        # 正答キーの正規化: "10" → "0" (マークシートの10番目の位置 = 選択肢"0")
        keys = []
        for q in question_numbers:
            raw_key = normalize_value(template_dict[q]['正答'])
            # 単一正答の "10" → "0" 変換
            if raw_key == '10':
                raw_key = '0'
            # 複数正答の各要素も正規化
            elif ';' in raw_key or '|' in raw_key:
                parts = raw_key.replace('|', ';').split(';')
                parts = ['0' if p.strip() == '10' else p.strip() for p in parts]
                raw_key = ';'.join(parts)
            keys.append(raw_key)
        
        # key_df: 設問IDと正答のペア
        key_df = pd.DataFrame({'QuestionID': questions, 'Key': keys})
        
        # ans_df: 各学生の解答マトリクス
        rows = []
        for result_data in mark2_results:
            student_answers = result_data['answers']
            row = {}
            row['StudentID'] = result_data['image']  # ファイル名をIDに
            for q_no in question_numbers:
                ans = student_answers.get(q_no, '')
                row[str(q_no)] = normalize_value(ans)
            rows.append(row)
        
        ans_df = pd.DataFrame(rows)
    else:
        # 記述のみモード: マーク解答なし、空の DataFrame から開始
        key_df = pd.DataFrame(columns=['QuestionID', 'Key'])
        if descriptive_scores:
            student_ids = sorted(descriptive_scores.keys())
            ans_df = pd.DataFrame({'StudentID': student_ids})
        else:
            ans_df = pd.DataFrame(columns=['StudentID'])
    
    # --- 記述問題の統合 ---
    if descriptive_config and descriptive_scores and descriptive_config.get('questions'):
        desc_questions = descriptive_config['questions']
        
        for dq in desc_questions:
            q_id = dq['id']       # "D1", "D2", ...
            max_score = dq['max_score']
            
            # key_df に記述問題を追加（正答キー = "1"）
            new_key = pd.DataFrame({'QuestionID': [q_id], 'Key': ['1']})
            key_df = pd.concat([key_df, new_key], ignore_index=True)
            
            # ans_df に記述問題のバイナリ解答を追加
            # 満点 → "1"（正答）, 満点未満 → "0"（誤答）
            desc_answers = []
            for _, row_data in ans_df.iterrows():
                student_id = row_data.get('StudentID', '')
                student_scores = descriptive_scores.get(student_id, {})
                score = student_scores.get(q_id)
                if score is not None and score >= max_score:
                    desc_answers.append('1')
                else:
                    desc_answers.append('0')
            
            ans_df[q_id] = desc_answers
        
        logger.info("  ✓ 記述問題 %d問をバイナリ(0/1)として統合", len(desc_questions))
    
    return ans_df, key_df


def _is_invalid_response(value_str: str) -> bool:
    """値が「無効回答」とみなすべきか判定する。

    無効回答となる値:
    - '' (空文字)、'nan' — 無回答
    - '-1'、'-1.0' — エラーチェックで無回答・ダブルマークを示す値
    - ';' を含む文字列 (ダブルマーク: '3;4' 等)
    - '無効回答' (集約後のラベル)
    - '無答' (旧ラベル、後方互換)

    これらは正規の選択肢(1,2,3,...,0)ではないが、
    分析対象として「無効回答」カテゴリに集約し集計に含める。
    """
    if value_str in ('', 'nan', '無答', '無効回答', '-1', '-1.0'):
        return True
    if ';' in value_str:
        return True
    return False


# 後方互換: 旧名でも呼び出せるようにする
_is_no_answer = _is_invalid_response


def _sort_choices(choices):
    """
    選択肢を数値順にソートするユーティリティ。
    
    マークシートは最大10択: 1,2,3,...,9,0 (10番目のマーク位置 = 選択肢"0")。
    "選択肢10" は原理的に存在しない。
    
    ソート順:
      1) 正規の選択肢を数値順 (1, 2, 3, ..., 9)
      2) "0" がある場合は正規選択肢の後に置く (10番目のマーク位置)
      3) "無効回答" は常に末尾
    
    -1、ダブルマーク ("3;4" 等)、空文字、nan は
    個別の選択肢としては表示せず「無効回答」カテゴリに集約する。
    ただし分析対象として集計には含める（呼び出し側で制御）。
    
    「無効回答」が入力に含まれていなくても追加はしない（呼び出し側で制御）。
    """
    regular = []     # 1,2,3,...,9 (1以上の数値)
    zero = []        # "0" (10番目のマーク位置)
    invalid = []     # "無効回答"

    for c in choices:
        cs = str(c)
        if cs == '無効回答':
            invalid.append(cs)
            continue
        if _is_invalid_response(cs):
            # -1, ダブルマーク, 空文字, nan → 「無効回答」に集約
            continue
        try:
            v = int(cs)
            if v == 0 or v == 10:
                # "0" = 10番目のマーク位置。"10" は後方互換で "0" と同義
                zero.append('0')
            else:
                regular.append((v, cs))
        except (ValueError, TypeError):
            # その他の非数値は無視
            continue

    regular.sort(key=lambda x: x[0])
    return [s for _, s in regular] + zero + invalid


class CTTAnalyzer:
    """
    古典テスト理論 (CTT) 分析エンジン
    
    0/1バイナリデータに基づき各種統計量を計算する。
    - テスト全体: 平均, SD, Cronbach α, 歪度, 尖度
    - 項目統計: P値, D値, I-T相関(修正/非修正), 削除α
    - 選択肢分析: 群別(上位/中位/下位 27%)の選択率
    """

    def __init__(self, ans_df, key_df):
        self.ans_df = ans_df.copy()
        self.key_df = key_df
        self.questions = [str(q) for q in key_df['QuestionID'].tolist()]
        self.keys = [str(k) for k in key_df['Key'].tolist()]
        self.n_students = len(ans_df)
        self.n_questions = len(self.questions)
        
        self.score_matrix = self._calculate_score_matrix()
        self._validate_binary_matrix()
        self.total_scores = self.score_matrix.sum(axis=1)
        self.ans_df['TotalScore'] = self.total_scores
        
        self.ans_df['Rank'] = self.ans_df['TotalScore'].rank(ascending=False, method='min')
        
        sorted_scores = self.total_scores.sort_values(ascending=False)
        score_freq = sorted_scores.value_counts().sort_index(ascending=False)
        self.score_cum_freq = score_freq.cumsum() / self.n_students
        self.d_margin = 0.0  # S-P分析のマージン（デフォルト0）

    def get_cum_freq(self, score):
        if score in self.score_cum_freq:
            return self.score_cum_freq[score]
        else:
            higher_scores = self.score_cum_freq.index[self.score_cum_freq.index > score]
            if len(higher_scores) > 0:
                return self.score_cum_freq[min(higher_scores)]
            else:
                return 0

    def _calculate_score_matrix(self):
        """
        解答マトリクスを0/1バイナリスコアに変換。
        
        採点ルール (score_answers と同一ロジック):
        - 単一正答: 文字列完全一致 + 0⇔10等価判定
        - 複数正答 (';' 区切り): 集合の一致判定
        - マークシートは最大10択択肢: 1,2,...,9,0
          (10番目のマーク位置 = 選択肢"0")
        """
        matrix = pd.DataFrame(index=self.ans_df.index, columns=self.questions)
        for i, q in enumerate(self.questions):
            correct_key = str(self.keys[i]).strip()
            if q not in self.ans_df.columns:
                matrix[q] = 0
                continue
            student_ans = self.ans_df[q].astype(str).str.strip()
            
            if ';' in correct_key or '|' in correct_key:
                # 複数正答: 集合比較 (score_answers と同一)
                correct_set = set(correct_key.replace('|', ';').split(';'))
                # 0⇔10正規化を集合内でも行う
                correct_set = {'0' if v == '10' else v for v in correct_set}
                def check_multi(ans):
                    if not ans or ans == 'nan' or ans == '':
                        return 0
                    ans_set = set(ans.replace('|', ';').split(';'))
                    ans_set = {'0' if v == '10' else v for v in ans_set}
                    return 1 if correct_set == ans_set else 0
                matrix[q] = student_ans.apply(check_multi)
            else:
                # 単一正答: 0⇔10等価判定付き
                def check_single(ans, key=correct_key):
                    ans = str(ans).strip()
                    if ans == key:
                        return 1
                    # 0⇔10等価判定 (後方互換性)
                    if (key == '0' and ans == '10') or (key == '10' and ans == '0'):
                        return 1
                    return 0
                matrix[q] = student_ans.apply(check_single)
        return matrix

    def _validate_binary_matrix(self):
        for q in self.questions:
            if q not in self.score_matrix.columns:
                continue
            col = self.score_matrix[q]
            unique_vals = set(col.dropna().unique())
            invalid_vals = unique_vals - {0, 1}
            if invalid_vals:
                raise ValueError(
                    f"⚠ バイナリデータ検証エラー: 設問 '{q}' に 0/1 以外の値 {invalid_vals}")
        logger.info("   ✅ バイナリデータ検証OK: 全設問が0/1データ")

    def calculate_test_stats(self):
        return {
            '平均点 (Mean)': self.total_scores.mean(),
            '中央値 (Median)': self.total_scores.median(),
            '標準偏差 (SD)': self.total_scores.std(),
            '最低点 (Min)': self.total_scores.min(),
            '最高点 (Max)': self.total_scores.max(),
            '分散 (Variance)': self.total_scores.var(ddof=1),
            '信頼性係数 (α)': self._calculate_cronbach_alpha(),
            '受験者数 (N)': self.n_students,
            '項目数 (K)': self.n_questions
        }

    def _calculate_cronbach_alpha(self):
        item_vars = self.score_matrix.var(axis=0, ddof=1).sum()
        total_var = self.total_scores.var(ddof=1)
        k = self.n_questions
        if total_var == 0 or k <= 1:
            return 0
        return (k / (k - 1)) * (1 - (item_vars / total_var))

    def calculate_item_stats(self):
        stats_list = []
        n_grp = max(1, int(self.n_students * 0.27))
        sorted_df = self.ans_df.sort_values(by='TotalScore', ascending=False)
        upper_idx = sorted_df.index[:n_grp]
        lower_idx = sorted_df.index[-n_grp:]
        upper_matrix = self.score_matrix.loc[upper_idx]
        lower_matrix = self.score_matrix.loc[lower_idx]

        for i, q in enumerate(self.questions):
            p_val = self.score_matrix[q].mean()
            
            rest_scores = self.total_scores - self.score_matrix[q]
            if self.score_matrix[q].std() == 0 or rest_scores.std() == 0:
                it_cor = 0
            else:
                it_cor = self.score_matrix[q].corr(rest_scores)
            
            if self.score_matrix[q].std() == 0 or self.total_scores.std() == 0:
                it_cor_incl = 0
            else:
                it_cor_incl = self.score_matrix[q].corr(self.total_scores)
            
            p_upper = upper_matrix[q].mean()
            p_lower = lower_matrix[q].mean()
            d_val = p_upper - p_lower
            
            idx_dropped = self.score_matrix.drop(columns=[q])
            total_var_dropped = idx_dropped.sum(axis=1).var(ddof=1)
            item_vars_dropped = idx_dropped.var(axis=0, ddof=1).sum()
            k_dropped = self.n_questions - 1
            if total_var_dropped == 0 or k_dropped <= 1:
                alpha_dropped = 0
            else:
                alpha_dropped = (k_dropped / (k_dropped - 1)) * (1 - (item_vars_dropped / total_var_dropped))

            stats_list.append({
                'QuestionID': q,
                'Key': self.keys[i],
                '正答率 (P)': p_val,
                '識別指数 (D)': d_val,
                'I-T相関': it_cor,
                'I-T相関(含)': it_cor_incl,
                '削除α': alpha_dropped,
                'P_Upper': p_upper,
                'P_Lower': p_lower
            })
        return pd.DataFrame(stats_list)

    def calculate_distractor_analysis(self):
        distractor_data = []
        n_grp = max(1, int(self.n_students * 0.27))
        sorted_df = self.ans_df.sort_values(by='TotalScore', ascending=False)
        upper_idx = sorted_df.index[:n_grp]
        lower_idx = sorted_df.index[-n_grp:]
        middle_idx = sorted_df.index[n_grp:-n_grp]
        
        groups = {
            '全体': self.ans_df.index,
            '高群': upper_idx,
            '中群': middle_idx,
            '低群': lower_idx
        }

        for i, q in enumerate(self.questions):
            found_choices = self.ans_df[q].astype(str).unique()
            raw_set = set(found_choices) | {str(self.keys[i])}
            # 無効回答相当の値を除去（「無効回答」カテゴリに集約して集計）
            raw_set = {v for v in raw_set if not _is_invalid_response(v)}
            all_choices = _sort_choices(list(raw_set) + ["無効回答"])
            
            for choice in all_choices:
                is_key = (str(choice) == str(self.keys[i]))
                row = {'QuestionID': q, 'Choice': choice, 'IsKey': is_key}
                
                for grp_name, grp_idx in groups.items():
                    subset = self.ans_df.loc[grp_idx, q].astype(str)
                    if choice == "無効回答":
                        count = sum(1 for v in subset if _is_invalid_response(v))
                    else:
                        count = (subset == choice).sum()
                    ratio = count / len(subset) if len(subset) > 0 else 0
                    row[f'Count_{grp_name}'] = count
                    row[f'Ratio_{grp_name}'] = ratio
                
                distractor_data.append(row)
        return pd.DataFrame(distractor_data)


class CTTPlotGenerator:
    """CTT分析用グラフ生成 (matplotlib)"""

    def __init__(self):
        if not HAS_MATPLOTLIB:
            raise RuntimeError("CTTグラフ生成にはmatplotlibが必要です。\npip install matplotlib でインストールしてください。")

    def generate_score_histogram(self, total_scores):
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.hist(total_scores, bins=15, color='skyblue', edgecolor='black')
        ax.set_title('得点分布 (Distribution of Total Scores)')
        ax.set_xlabel('得点 (Total Score)')
        ax.set_ylabel('人数 (Frequency)')
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return buf

    def generate_item_curve(self, item_stats_row, distractor_df):
        qid = item_stats_row['QuestionID']
        item_dist = distractor_df[distractor_df['QuestionID'] == qid]
        
        fig, ax = plt.subplots(figsize=(6, 4))
        groups = ['低群', '中群', '高群']
        group_cols = ['Ratio_低群', 'Ratio_中群', 'Ratio_高群']
        
        for _, row in item_dist.iterrows():
            choice = row['Choice']
            is_key = row['IsKey']
            try:
                y_vals = [row[col] for col in group_cols]
            except KeyError:
                continue
            
            label = f"{choice} {'(正答)' if is_key else ''}"
            color = 'blue' if is_key else None
            linewidth = 2.5 if is_key else 1
            linestyle = '-' if is_key else '--'
            
            if color:
                ax.plot(groups, y_vals, marker='o', label=label, color=color,
                        linewidth=linewidth, linestyle=linestyle)
            else:
                ax.plot(groups, y_vals, marker='o', label=label,
                        linewidth=linewidth, linestyle=linestyle)
        
        ax.set_title(f'項目分析: {qid}')
        ax.set_ylabel('選択率')
        ax.set_xlabel('学力群')
        ax.set_ylim(-0.05, 1.05)
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1))
        fig.tight_layout()
        
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return buf

    def generate_mini_trace_grid(self, distractor_stats, questions, cols=10, rows_per_page=6):
        per_page = cols * rows_per_page
        n_pages = (len(questions) + per_page - 1) // per_page
        group_cols = ['Ratio_低群', 'Ratio_中群', 'Ratio_高群']
        x_pos = [0, 1, 2]
        cell_w, cell_h = 1.5, 1.3
        pages = []

        for page_idx in range(n_pages):
            start = page_idx * per_page
            end = min(start + per_page, len(questions))
            page_questions = questions[start:end]
            n_q = len(page_questions)
            actual_rows = (n_q + cols - 1) // cols
            fig_w, fig_h = cols * cell_w, actual_rows * cell_h

            fig, axes = plt.subplots(actual_rows, cols, figsize=(fig_w, fig_h), squeeze=False)
            fig.subplots_adjust(hspace=0.4, wspace=0.25)

            for idx, qid in enumerate(page_questions):
                r, c = idx // cols, idx % cols
                ax = axes[r][c]
                item_dist = distractor_stats[distractor_stats['QuestionID'] == qid]

                for _, row in item_dist.iterrows():
                    is_key = row['IsKey']
                    try:
                        y_vals = [row[col] for col in group_cols]
                    except KeyError:
                        continue
                    if is_key:
                        ax.plot(x_pos, y_vals, color='#2060C0', linewidth=1.8,
                                linestyle='-', marker='o', markersize=2)
                    else:
                        ax.plot(x_pos, y_vals, color='#AAAAAA', linewidth=0.5,
                                linestyle=':', marker='', markersize=0)

                ax.set_ylim(-0.05, 1.05)
                ax.set_xlim(-0.3, 2.3)
                ax.set_xticks([])
                ax.set_yticks([])
                ax.set_title(str(qid), fontsize=7, pad=2)
                ax.set_aspect('auto')

                key_rows = item_dist[item_dist['IsKey'] == True]
                if len(key_rows) > 0:
                    key_row = key_rows.iloc[0]
                    try:
                        if key_row['Ratio_高群'] < key_row['Ratio_低群']:
                            ax.set_facecolor('#FFE0E0')
                    except KeyError:
                        pass

            for idx in range(n_q, actual_rows * cols):
                r, c = idx // cols, idx % cols
                axes[r][c].set_visible(False)

            fig.suptitle(f'トレースライン一覧 ({start+1}〜{end}問目)', fontsize=11, y=1.02)
            buf = io.BytesIO()
            fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
            plt.close(fig)
            buf.seek(0)
            pages.append((buf, fig_w / fig_h))

        return pages

    def generate_correlation_chart(self, item_stats, corr_col, title):
        qids = item_stats['QuestionID'].astype(str).tolist()
        values = item_stats[corr_col].values
        n = len(qids)
        fig_width = max(8, n * 0.4)
        fig, ax = plt.subplots(figsize=(fig_width, 4.5))

        bar_colors = ['#CC3333' if v < 0 else '#4472C4' for v in values]
        x = np.arange(n)
        ax.bar(x, values, color=bar_colors, width=0.7, edgecolor='white', linewidth=0.5)
        ax.axhline(y=0.2, color='#228B22', linestyle='--', linewidth=0.8, label='基準 (0.20)')
        ax.axhline(y=0.0, color='black', linestyle='-', linewidth=0.5)
        ax.set_xticks(x)
        ax.set_xticklabels(qids, fontsize=8,
                           rotation=45 if n > 15 else 0,
                           ha='right' if n > 15 else 'center')
        ax.set_ylabel('相関係数', fontsize=10)
        ax.set_title(title, fontsize=12, pad=10)
        ax.set_ylim(min(-0.3, min(values) - 0.1), max(1.0, max(values) + 0.1))
        ax.legend(fontsize=8, loc='upper right')

        for i, v in enumerate(values):
            ax.text(i, v + (0.02 if v >= 0 else -0.05), f'{v:.2f}',
                    ha='center', va='bottom' if v >= 0 else 'top', fontsize=6)

        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=120, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return buf

    def generate_correlation_heatmap(self, score_matrix, questions):
        corr_matrix = score_matrix[questions].corr()
        n = len(questions)
        fig_size = max(6, n * 0.35)
        fig, ax = plt.subplots(figsize=(fig_size, fig_size))
        cmap = plt.cm.RdBu
        im = ax.imshow(corr_matrix.values, cmap=cmap, vmin=-1, vmax=1,
                       interpolation='nearest', aspect='equal')
        cbar = fig.colorbar(im, ax=ax, shrink=0.8)
        cbar.ax.tick_params(labelsize=8)
        q_labels = [str(q) for q in questions]
        ax.set_xticks(np.arange(n))
        ax.set_yticks(np.arange(n))
        ax.set_xticklabels(q_labels, fontsize=7, rotation=45, ha='left', color='#CC0000')
        ax.set_yticklabels(q_labels, fontsize=7, color='#CC0000')
        ax.xaxis.set_ticks_position('top')
        ax.xaxis.set_label_position('top')

        for i in range(n):
            ax.add_patch(plt.Rectangle((i - 0.5, i - 0.5), 1, 1,
                                       fill=False, edgecolor='#333333', linewidth=0.8))
        if n <= 30:
            for i in range(n):
                for j in range(n):
                    val = corr_matrix.values[i, j]
                    color = 'white' if abs(val) > 0.6 else 'black'
                    fontsize = 6 if n > 20 else 7
                    if i != j:
                        ax.text(j, i, f'{val:.2f}', ha='center', va='center',
                                fontsize=fontsize, color=color)

        ax.set_title('設問間相関行列（Phi係数）', fontsize=12, pad=15)
        fig.tight_layout()
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=130, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return buf


class CTTExcelExporter:
    """CTT分析Excelレポート生成 (openpyxl)"""

    def __init__(self, output_path, analyzer):
        self.output_path = output_path
        self.az = analyzer
        self.wb = Workbook()
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        FN = 'Yu Gothic'
        self.f_title   = Font(name=FN, bold=True, size=16)
        self.f_h1      = Font(name=FN, bold=True, size=12)
        self.f_h2w     = Font(name=FN, bold=True, size=11, color='FFFFFF')
        self.f_h3      = Font(name=FN, bold=True, size=10)
        self.f_n       = Font(name=FN, size=10)
        self.f_s       = Font(name=FN, size=9, color='555555')
        self.f_red     = Font(name=FN, size=10, color='CC0000')
        self.f_link    = Font(name=FN, size=11, color='0563C1', underline='single')
        self.f_wh      = Font(name=FN, size=1, color='FFFFFF')

        self.bg_hdr    = PatternFill('solid', fgColor='4472C4')
        self.bg_hdr2   = PatternFill('solid', fgColor='5B9BD5')
        self.bg_sub    = PatternFill('solid', fgColor='D6E4F0')
        self.bg_lt     = PatternFill('solid', fgColor='F2F2F2')
        self.bg_grn    = PatternFill('solid', fgColor='C6EFCE')
        self.bg_red    = PatternFill('solid', fgColor='FFC7CE')
        self.bg_org    = PatternFill('solid', fgColor='FFE0B2')
        self.bg_key    = PatternFill('solid', fgColor='BDD7EE')
        self.bg_sp_y   = PatternFill('solid', fgColor='FFF2CC')
        self.bg_sp_b   = PatternFill('solid', fgColor='DAEEF3')
        self.bg_toc    = PatternFill('solid', fgColor='1F4E79')

        _t = Side(style='thin', color='AAAAAA')
        _m = Side(style='medium', color='444444')
        self.b_thin = Border(left=_t, right=_t, top=_t, bottom=_t)
        self.b_med  = Border(left=_m, right=_m, top=_m, bottom=_m)
        self.b_none = Border()

        self.a_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.a_l = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.a_w = Alignment(wrap_text=True, vertical='top')

    def _c(self, ws, r, c, val, font=None, fill=None, fmt=None, align=None, border=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = font or self.f_n
        cell.border = border or self.b_thin
        cell.alignment = align or self.a_c
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        return cell

    def _merge_h(self, ws, r, c1, c2, val, font=None, fill=None):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r, column=c1, value=val)
        cell.font = font or self.f_h2w
        cell.fill = fill or self.bg_hdr
        cell.alignment = self.a_c
        cell.border = self.b_med
        for col in range(c1, c2 + 1):
            ws.cell(row=r, column=col).border = self.b_med

    def _note(self, ws, r, c, text, end_c=None):
        ec = end_c or c + 8
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=ec)
        cell = ws.cell(row=r, column=c, value=text)
        cell.font = self.f_s
        cell.alignment = self.a_w
        cell.border = self.b_none
        ws.row_dimensions[r].height = 14
        return r + 1

    def _widths(self, ws, d):
        for k, v in d.items():
            ws.column_dimensions[k].width = v

    def export(self, test_stats, student_df, item_stats, distractor_stats):
        self._create_toc(item_stats)
        self._create_score_sheet(test_stats)
        self._create_overview(item_stats, distractor_stats, test_stats)
        self._create_examinee(student_df)
        self._create_items(item_stats, distractor_stats, test_stats)
        self.wb.save(self.output_path)
        logger.info("   ✅ CTT Excel: %s", self.output_path)

    def _create_toc(self, item_stats):
        ws = self.wb.create_sheet("目次", 0)
        ws.sheet_properties.tabColor = '1F4E79'
        ws.merge_cells('A1:F1')
        self._c(ws, 1, 1, 'テスト分析レポート',
                font=Font(name='Yu Gothic', bold=True, size=18, color='FFFFFF'),
                fill=self.bg_toc, border=self.b_med)
        ws.row_dimensions[1].height = 45
        ws.merge_cells('A2:F2')
        self._c(ws, 2, 1, 'リンクをクリックすると各シートに移動します',
                font=self.f_s, align=self.a_l, border=self.b_none)

        self._c(ws, 4, 1, 'No.', font=self.f_h2w, fill=self.bg_hdr, border=self.b_med)
        self._c(ws, 4, 2, 'シート名', font=self.f_h2w, fill=self.bg_hdr, border=self.b_med)
        ws.merge_cells('C4:F4')
        self._c(ws, 4, 3, '内容', font=self.f_h2w, fill=self.bg_hdr, border=self.b_med)

        links = [
            ('テスト得点', '基本統計量・得点分布ヒストグラム・要約統計量'),
            ('項目全体',   '全項目の識別指数・選択率一覧（ワイド形式）'),
            ('受験者分析', '受験者別 正誤パターンとS-P分析'),
        ]
        for _, row in item_stats.iterrows():
            links.append((str(row['QuestionID']),
                          f"項目 {row['QuestionID']} 詳細分析・トレースライン"))

        for i, (name, desc) in enumerate(links, 1):
            r = 4 + i
            self._c(ws, r, 1, i)
            cell = self._c(ws, r, 2, name, font=self.f_link, align=self.a_l)
            cell.hyperlink = f"#'{name}'!A1"
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            self._c(ws, r, 3, desc, align=self.a_l)
            if i > 3:
                for col in range(1, 7):
                    ws.cell(row=r, column=col).fill = self.bg_lt

        self._widths(ws, {'A': 6, 'B': 16, 'C': 15, 'D': 15, 'E': 12, 'F': 12})
        ws.freeze_panes = 'A5'

        warn_row = 4 + len(links) + 2
        ws.merge_cells(start_row=warn_row, start_column=1,
                       end_row=warn_row + 3, end_column=6)
        warn_cell = ws.cell(row=warn_row, column=1,
            value=(
                '【データ処理に関する注意事項】\n'
                '本レポートは、各設問を「正解=1 / 不正解=0」の二値（バイナリ）データとして処理しています。\n'
                '・中間点（部分点）のあるデータ→不正解(0点)として処理  ・配点(重みづけ)のあるデータ→正解は一律1点として処理\n'
                '・得点・満点・平均点等は「各問1点換算」で算出されています'))
        warn_cell.font = Font(name='Yu Gothic', size=9, color='8B4513')
        warn_cell.fill = PatternFill('solid', fgColor='FFF8E1')
        warn_cell.alignment = Alignment(wrap_text=True, vertical='top')
        warn_cell.border = Border(
            left=Side(style='thin', color='D4A017'),
            right=Side(style='thin', color='D4A017'),
            top=Side(style='thin', color='D4A017'),
            bottom=Side(style='thin', color='D4A017'))
        ws.row_dimensions[warn_row].height = 60

    def _create_score_sheet(self, test_stats):
        ws = self.wb.create_sheet("テスト得点")
        ws.sheet_properties.tabColor = '2F5496'

        n     = int(test_stats.get('受験者数 (N)', 0))
        k     = int(test_stats.get('項目数 (K)', 0))
        mean  = test_stats.get('平均点 (Mean)', 0)
        sd    = test_stats.get('標準偏差 (SD)', 0)
        mn    = test_stats.get('最低点 (Min)', 0)
        mx    = test_stats.get('最高点 (Max)', 0)
        med   = test_stats.get('中央値 (Median)', 0)
        var_  = test_stats.get('分散 (Variance)', sd**2)
        alpha = test_stats.get('信頼性係数 (α)', 0)

        scores = np.array(self.az.total_scores)
        q1  = np.percentile(scores, 25)
        q3  = np.percentile(scores, 75)
        iqr = q3 - q1
        sem = sd * np.sqrt(1 - alpha) if alpha < 1 else 0
        if n > 2 and sd > 0:
            skew = (n / ((n-1)*(n-2))) * np.sum(((scores - mean) / sd)**3)
        else:
            skew = 0
        if n > 3 and sd > 0:
            kurt = ((n*(n+1)) / ((n-1)*(n-2)*(n-3))) * np.sum(((scores - mean)/sd)**4) \
                   - (3*(n-1)**2)/((n-2)*(n-3))
        else:
            kurt = 0

        ws.merge_cells('A1:E1')
        self._c(ws, 1, 1, 'テスト得点  ─  要約統計量',
                font=Font(name='Yu Gothic', bold=True, size=14, color='FFFFFF'),
                fill=self.bg_toc, border=self.b_med)
        ws.row_dimensions[1].height = 32

        stats_data = [
            ('受験者数',           n,                  ''),
            ('項目数',             k,                  ''),
            ('平均点',             round(mean, 2),     '0.00'),
            ('平均得点率',         f"{round(mean/k*100) if k else 0}%", ''),
            ('標準偏差 (SD)',      round(sd, 2),       '0.00'),
            ('分散',               round(var_, 2),     '0.00'),
            ('最小値',             int(mn),            ''),
            ('第1四分位 (Q1)',     round(q1, 1),       '0.0'),
            ('中央値',             round(med, 1),      '0.0'),
            ('第3四分位 (Q3)',     round(q3, 1),       '0.0'),
            ('最大値',             int(mx),            ''),
            ('四分位範囲 (IQR)',   round(iqr, 1),      '0.0'),
            ('歪度',               round(skew, 2),     '0.00'),
            ('尖度',               round(kurt, 2),     '0.00'),
            ('α係数',             round(alpha, 3),    '0.000'),
            ('測定標準誤差 (SEM)', round(sem, 2),      '0.00'),
        ]

        self._c(ws, 3, 1, '統計量', font=self.f_h2w, fill=self.bg_hdr, border=self.b_med)
        self._c(ws, 3, 2, '値',     font=self.f_h2w, fill=self.bg_hdr, border=self.b_med)
        self._c(ws, 3, 3, '解説',   font=self.f_h2w, fill=self.bg_hdr, border=self.b_med)
        ws.merge_cells('C3:E3')

        explanations = {
            '受験者数': '分析対象の総受験者数',
            '項目数':   'テストの問題数',
            '平均点':   '全受験者の平均得点（各問1点換算）',
            '平均得点率': '平均点÷項目数×100',
            '標準偏差 (SD)': '得点のばらつき（大きいほど差が大きい）',
            '分散': '標準偏差の2乗',
            '最小値': '最低得点',
            '第1四分位 (Q1)': '下位25%の境界点',
            '中央値': '受験者の半数が超える得点',
            '第3四分位 (Q3)': '上位25%の境界点',
            '最大値': '最高得点',
            '四分位範囲 (IQR)': 'Q3−Q1。中央50%の得点幅',
            '歪度': '正=低得点寄り / 負=高得点寄り / 0=対称',
            '尖度': '正=鋭い分布 / 負=平坦な分布',
            'α係数': '≧0.90=非常に高い / ≧0.80=十分 / ≧0.70=許容',
            '測定標準誤差 (SEM)': 'SD×√(1−α)。小さいほど測定精度が高い',
        }

        for i, (label, val, fmt) in enumerate(stats_data, 4):
            bg = self.bg_lt if i % 2 == 0 else None
            self._c(ws, i, 1, label, font=self.f_h3, fill=bg)
            c = self._c(ws, i, 2, val, fill=bg)
            if fmt:
                c.number_format = fmt
            if label == 'α係数':
                c.fill = self.bg_grn if alpha >= 0.8 else (self.bg_org if alpha >= 0.7 else self.bg_red)
            ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=5)
            self._c(ws, i, 3, explanations.get(label, ''), font=self.f_s, fill=bg, align=self.a_l)

        last_stat_row = 3 + len(stats_data)

        data_col_score = 7
        data_col_count = 8
        score_counts = pd.Series(scores).value_counts().sort_index()
        all_scores_range = range(0, k + 1)

        ws.cell(row=1, column=data_col_score, value='得点').font = self.f_wh
        ws.cell(row=1, column=data_col_count, value='人数').font = self.f_wh
        for idx, s in enumerate(all_scores_range, 2):
            ws.cell(row=idx, column=data_col_score, value=s).font = self.f_wh
            ws.cell(row=idx, column=data_col_count, value=int(score_counts.get(s, 0))).font = self.f_wh
        n_scores = len(list(all_scores_range))

        chart = BarChart()
        chart.type = 'col'
        chart.title = '得点分布'
        chart.y_axis.title = '人数'
        chart.x_axis.title = '得点'
        chart.y_axis.numFmt = '0'
        chart.y_axis.delete = False
        chart.x_axis.delete = False
        chart.x_axis.tickLblPos = 'low'
        chart.style = 2
        chart.legend = None
        chart.width = 14
        chart.height = 10
        chart.gapWidth = 80

        data_ref = Reference(ws, min_col=data_col_count, min_row=1, max_row=1 + n_scores)
        cats_ref = Reference(ws, min_col=data_col_score, min_row=2, max_row=1 + n_scores)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        s = chart.series[0]
        s.graphicalProperties.solidFill = '5B9BD5'
        s.graphicalProperties.line.solidFill = '2F5496'
        s.graphicalProperties.line.width = 6000
        ws.add_chart(chart, f'A{last_stat_row + 2}')

        self._widths(ws, {'A': 18, 'B': 12, 'C': 12, 'D': 12, 'E': 12})
        ws.freeze_panes = 'A4'

    def _create_overview(self, item_stats, distractor_stats, test_stats):
        ws = self.wb.create_sheet("項目全体")
        ws.sheet_properties.tabColor = '4472C4'

        uq = _sort_choices([c for c in distractor_stats['Choice'].unique() if c != '無効回答'])
        alpha = self.az._calculate_cronbach_alpha()
        n_base = 8
        ch_list = list(uq) + ['無効回答']
        n_ch = len(ch_list)
        total_cols = n_base + n_ch

        self._merge_h(ws, 1, 1, n_base, '項目統計')
        self._merge_h(ws, 1, n_base + 1, total_cols, '全体選択率', fill=self.bg_hdr2)

        for c, h in enumerate(['項目', '受験者数', '正答率', 'D値', 'I-T相関',
                                'α係数', '削除α', '正答'], 1):
            self._c(ws, 2, c, h, font=self.f_h3, fill=self.bg_sub, border=self.b_med)
        for i, ch in enumerate(ch_list):
            self._c(ws, 2, n_base + 1 + i, ch, font=self.f_h3, fill=self.bg_sub, border=self.b_med)

        for ri, (_, row) in enumerate(item_stats.iterrows(), 3):
            qid = row['QuestionID']
            kv = str(row['Key'])
            pv = row['正答率 (P)']
            dv = row['識別指数 (D)']
            iv = row['I-T相関']
            da = row['削除α']

            cell = self._c(ws, ri, 1, qid, font=self.f_link)
            cell.hyperlink = f"#'{qid}'!A1"
            self._c(ws, ri, 2, self.az.n_students)
            self._c(ws, ri, 3, round(pv * 100), fmt='0"%"',
                    fill=self.bg_grn if pv >= 0.8 else (self.bg_red if pv <= 0.2 else None))
            self._c(ws, ri, 4, round(dv * 100), fmt='0"%"',
                    font=self.f_red if dv < 0 else self.f_n,
                    fill=self.bg_red if dv < 0 else (self.bg_org if dv < 0.2 else None))
            self._c(ws, ri, 5, round(iv, 2), fmt='0.00',
                    font=self.f_red if iv < 0 else self.f_n,
                    fill=self.bg_red if iv < 0 else None)
            self._c(ws, ri, 6, round(alpha, 2), fmt='0.00')
            self._c(ws, ri, 7, round(da, 2), fmt='0.00',
                    fill=self.bg_org if da > alpha else None)
            self._c(ws, ri, 8, kv, font=self.f_h3)

            d_sub = distractor_stats[distractor_stats['QuestionID'] == qid]
            cm = {r2['Choice']: r2['Ratio_全体'] for _, r2 in d_sub.iterrows()}
            is_desc = str(qid).startswith("D")
            for i, ch in enumerate(ch_list):
                if is_desc and str(ch) not in ('0', '1', '無効回答'):
                    self._c(ws, ri, n_base + 1 + i, '–',
                            fill=self.bg_lt, font=self.f_n)
                else:
                    ratio = cm.get(ch, 0)
                    self._c(ws, ri, n_base + 1 + i, round(ratio * 100), fmt='0"%"',
                            fill=self.bg_key if str(ch) == kv else None)

        lr = 2 + len(item_stats) + 1
        r = lr + 1
        r = self._note(ws, r, 1,
            '【凡例】 緑=正答率高(≧80%) / 赤=正答率低(≦20%)・D値負・I-T相関負 / '
            '橙=D値低(<20%)・削除αがαより大 / 水色=正答選択肢', total_cols)

        widths = {'A': 8}
        for i in range(2, total_cols + 1):
            widths[get_column_letter(i)] = 9
        self._widths(ws, widths)
        ws.freeze_panes = 'B3'
        ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{2 + len(item_stats)}"

    def _create_examinee(self, student_df):
        ws = self.wb.create_sheet("受験者分析")
        ws.sheet_properties.tabColor = '70AD47'

        qs = self.az.questions
        sp = ['TotalScore', 'Rank']
        attrs = [c for c in student_df.columns if c not in qs and c not in sp]
        k = self.az.n_questions

        df = student_df.copy()
        df['得点率'] = df['TotalScore'].apply(lambda x: round(x / k * 100) if k > 0 else 0)
        dcols = attrs + ['TotalScore', '得点率'] + qs
        dcols = [c for c in dcols if c in df.columns]

        na = len(attrs)
        ns = 2
        ae = na
        se = ae + ns
        qe = se + len(qs)

        if na > 0:
            self._merge_h(ws, 1, 1, ae, '受験者属性')
        self._merge_h(ws, 1, ae + 1, se, '得点', fill=self.bg_hdr2)
        self._merge_h(ws, 1, se + 1, qe, '各項目  (1=正答 / 0=誤答)')

        jp = {'TotalScore': '得点', '得点率': '得点率(%)'}
        for c, cn in enumerate(dcols, 1):
            self._c(ws, 2, c, jp.get(cn, cn), font=self.f_h3, fill=self.bg_sub, border=self.b_med)

        sm = self.az.score_matrix
        ip = {}
        idf = self.az.calculate_item_stats()
        for _, r2 in idf.iterrows():
            ip[r2['QuestionID']] = r2['正答率 (P)']
        dm = getattr(self.az, 'd_margin', 0.0)
        dfs = df.sort_values(by='TotalScore', ascending=False)

        for ri, (oi, row) in enumerate(dfs.iterrows(), 3):
            sc = row['TotalScore']
            cb = self.az.get_cum_freq(sc)
            rbg = self.bg_lt if ri % 2 == 1 else None

            for ci, cn in enumerate(dcols, 1):
                rv = row[cn]
                if cn in qs:
                    try:
                        cv = int(sm.loc[oi, cn])
                    except Exception:
                        cv = rv
                    cell = self._c(ws, ri, ci, cv, fill=rbg)
                    try:
                        pv = ip.get(cn, 0)
                        if cv == 1 and (cb - dm) > pv:
                            cell.fill = self.bg_sp_b
                        elif cv == 0 and (cb + dm) < pv:
                            cell.fill = self.bg_sp_y
                    except Exception:
                        pass
                elif cn == '得点率':
                    self._c(ws, ri, ci, int(rv), fmt='0"%"', fill=rbg)
                else:
                    self._c(ws, ri, ci, rv, fill=rbg)

        _d = Side(style='double', color='444444')
        for r in range(1, 3 + len(dfs)):
            if ae > 0 and ae < qe:
                c = ws.cell(row=r, column=ae)
                c.border = Border(left=c.border.left, right=_d, top=c.border.top, bottom=c.border.bottom)
            c2 = ws.cell(row=r, column=se)
            c2.border = Border(left=c2.border.left, right=_d, top=c2.border.top, bottom=c2.border.bottom)

        lr = 2 + len(dfs) + 1
        r = lr + 1
        r = self._note(ws, r, 1,
            '【S-P分析】 黄=易しい項目に誤答（意外な不正解） / '
            '水色=難しい項目に正答（意外な正解）', qe)

        widths = {}
        for i, cn in enumerate(dcols, 1):
            lt = get_column_letter(i)
            widths[lt] = 5 if cn in qs else (8 if cn in ['TotalScore', '得点率'] else 10)
        self._widths(ws, widths)
        ws.freeze_panes = f'{get_column_letter(se + 1)}3'

    def _create_items(self, item_stats, distractor_stats, test_stats):
        uq = _sort_choices([c for c in distractor_stats['Choice'].unique() if c != '無効回答'])
        alpha = self.az._calculate_cronbach_alpha()
        all_ch_list = list(uq) + ['無効回答']
        palette = ['4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '70AD47',
                   '9B57A0', '264478', 'C55A11']

        for idx, (_, row) in enumerate(item_stats.iterrows()):
            qid = str(row['QuestionID'])
            # 記述問題は選択肢を 0, 1, 無効回答 のみに制限
            if qid.startswith("D"):
                ch_list = ['0', '1', '無効回答']
            else:
                ch_list = all_ch_list
            n_ch = len(ch_list)
            kv = str(row['Key'])
            pv = row['正答率 (P)']
            dv = row['識別指数 (D)']
            iv = row['I-T相関']
            da = row['削除α']

            ws = self.wb.create_sheet(qid)
            if dv < 0 or iv < 0:
                ws.sheet_properties.tabColor = 'CC0000'
            elif dv < 0.2:
                ws.sheet_properties.tabColor = 'ED7D31'
            elif dv >= 0.3 and iv >= 0.2:
                ws.sheet_properties.tabColor = '70AD47'
            else:
                ws.sheet_properties.tabColor = '5B9BD5'

            ws.merge_cells('A1:G1')
            self._c(ws, 1, 1, f'項目 {qid}',
                    font=Font(name='Yu Gothic', bold=True, size=14, color='FFFFFF'),
                    fill=self.bg_toc, border=self.b_med)
            ws.row_dimensions[1].height = 32

            cell = self._c(ws, 2, 1, '← 目次に戻る', font=self.f_link, align=self.a_l, border=self.b_none)
            cell.hyperlink = "#'目次'!A1"

            for c, h in enumerate(['受験者数', '正答率', 'D値', 'I-T相関', 'α係数', '削除α', '正答キー'], 1):
                self._c(ws, 3, c, h, font=self.f_h3, fill=self.bg_sub, border=self.b_med)

            self._c(ws, 4, 1, self.az.n_students)
            self._c(ws, 4, 2, f"{round(pv * 100)}%",
                    fill=self.bg_grn if pv >= 0.8 else (self.bg_red if pv <= 0.2 else None))
            self._c(ws, 4, 3, f"{round(dv * 100)}%",
                    font=self.f_red if dv < 0 else self.f_n,
                    fill=self.bg_red if dv < 0 else (self.bg_org if dv < 0.2 else None))
            self._c(ws, 4, 4, round(iv, 2), fmt='0.00',
                    font=self.f_red if iv < 0 else self.f_n,
                    fill=self.bg_red if iv < 0 else None)
            self._c(ws, 4, 5, round(alpha, 2), fmt='0.00')
            self._c(ws, 4, 6, round(da, 2), fmt='0.00',
                    fill=self.bg_org if da > alpha else None)
            self._c(ws, 4, 7, kv, font=self.f_h3, fill=self.bg_grn)

            # 選択率テーブル
            self._merge_h(ws, 6, 1, 1 + n_ch, '選択率（群別）')
            self._c(ws, 7, 1, '', font=self.f_h3, fill=self.bg_sub, border=self.b_med)
            for i, ch in enumerate(ch_list, 2):
                self._c(ws, 7, i, ch, font=self.f_h3,
                        fill=self.bg_key if str(ch) == kv else self.bg_sub, border=self.b_med)

            d_sub = distractor_stats[distractor_stats['QuestionID'] == qid]
            cd = {r2['Choice']: r2 for _, r2 in d_sub.iterrows()}

            for gi, (gl, gk) in enumerate([('全体', 'Ratio_全体'), ('高群', 'Ratio_高群'),
                                             ('中群', 'Ratio_中群'), ('低群', 'Ratio_低群')]):
                r = 8 + gi
                self._c(ws, r, 1, gl, font=self.f_h3 if gi == 0 else self.f_n,
                        fill=self.bg_sub if gi == 0 else self.bg_lt)
                for i, ch in enumerate(ch_list, 2):
                    d = cd.get(ch)
                    val = d[gk] if d is not None and gk in d else 0
                    self._c(ws, r, i, f"{round(val * 100)}%",
                            fill=self.bg_key if str(ch) == kv else None)

            # 識別指標
            self._merge_h(ws, 13, 1, 1 + n_ch, '識別指標（各選択肢）')
            self._c(ws, 14, 1, '', font=self.f_h3, fill=self.bg_sub, border=self.b_med)
            for i, ch in enumerate(ch_list, 2):
                self._c(ws, 14, i, ch, font=self.f_h3, fill=self.bg_sub, border=self.b_med)

            self._c(ws, 15, 1, 'D値', font=self.f_h3, fill=self.bg_lt)
            self._c(ws, 16, 1, 'I-T相関', font=self.f_h3, fill=self.bg_lt)

            for i, ch in enumerate(ch_list, 2):
                d = cd.get(ch)
                ik = str(ch) == kv
                if d is not None:
                    hr = d.get('Ratio_高群', 0)
                    lr_ = d.get('Ratio_低群', 0)
                    chd = hr - lr_
                    bad_d = (ik and chd < 0) or (not ik and chd > 0.1)
                    self._c(ws, 15, i, f"{round(chd * 100)}%",
                            font=self.f_red if bad_d else self.f_n,
                            fill=self.bg_red if bad_d else None)
                    try:
                        chose = (self.az.ans_df[qid].astype(str).str.strip() == str(ch).strip()).astype(int)
                        rest = self.az.total_scores - self.az.score_matrix[qid]
                        chit = chose.corr(rest) if chose.std() > 0 and rest.std() > 0 else 0
                    except Exception:
                        chit = 0
                    bad_it = (ik and chit < 0) or (not ik and chit > 0.15)
                    self._c(ws, 16, i, round(chit, 2), fmt='0.00',
                            font=self.f_red if bad_it else self.f_n,
                            fill=self.bg_red if bad_it else None)

            # フィードバック
            r = 18
            r = self._note(ws, r, 1, '── フィードバック ──', 1 + n_ch)
            fb = self._feedback(pv, dv, iv, da, alpha)
            for f in fb:
                r = self._note(ws, r, 1, f, 1 + n_ch)

            # トレースライン（折れ線グラフ）
            dc = max(n_ch + 4, 8)  # A1:G1 のマージ範囲(列7)を超える
            ws.cell(row=1, column=dc, value='群').font = self.f_wh
            ws.cell(row=2, column=dc, value='低群').font = self.f_wh
            ws.cell(row=3, column=dc, value='中群').font = self.f_wh
            ws.cell(row=4, column=dc, value='高群').font = self.f_wh

            for ci, ch in enumerate(ch_list):
                col = dc + 1 + ci
                label = f"★{ch}(正答)" if str(ch) == kv else str(ch)
                ws.cell(row=1, column=col, value=label).font = self.f_wh
                d = cd.get(ch)
                if d is not None:
                    ws.cell(row=2, column=col, value=round(d.get('Ratio_低群', 0), 3)).font = self.f_wh
                    ws.cell(row=3, column=col, value=round(d.get('Ratio_中群', 0), 3)).font = self.f_wh
                    ws.cell(row=4, column=col, value=round(d.get('Ratio_高群', 0), 3)).font = self.f_wh
                else:
                    ws.cell(row=2, column=col, value=0).font = self.f_wh
                    ws.cell(row=3, column=col, value=0).font = self.f_wh
                    ws.cell(row=4, column=col, value=0).font = self.f_wh

            chart = LineChart()
            chart.title = f'トレースライン: {qid}'
            chart.y_axis.title = '選択率'
            chart.x_axis.title = '学力群'
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1
            chart.y_axis.numFmt = '0%'
            chart.y_axis.majorUnit = 0.2
            chart.y_axis.delete = False
            chart.x_axis.delete = False
            chart.x_axis.tickLblPos = 'low'
            chart.style = 2
            chart.width = 14
            chart.height = 10
            chart.legend.position = 'r'

            cats = Reference(ws, min_col=dc, min_row=2, max_row=4)
            chart.set_categories(cats)

            for ci, ch in enumerate(ch_list):
                col = dc + 1 + ci
                ref = Reference(ws, min_col=col, min_row=1, max_row=4)
                chart.add_data(ref, titles_from_data=True)
                s = chart.series[-1]
                s.smooth = False
                ik = str(ch) == kv
                color = palette[ci % len(palette)]
                s.graphicalProperties.line.solidFill = color
                s.graphicalProperties.line.width = 28000 if ik else 12000
                if not ik:
                    s.graphicalProperties.line.dashStyle = 'dash'
                s.marker.symbol = 'circle' if ik else 'diamond'
                s.marker.size = 8 if ik else 5
                s.marker.graphicalProperties.solidFill = color
                s.marker.graphicalProperties.line.solidFill = color

            ws.add_chart(chart, f'{get_column_letter(n_ch + 3)}1')

            widths = {'A': 10}
            for i in range(2, n_ch + 3):
                widths[get_column_letter(i)] = 8
            self._widths(ws, widths)

    def _feedback(self, p, d, it, da, a):
        fb = []
        if p >= 0.9:
            fb.append('📝 正答率90%以上: 易しすぎる項目。識別力が出にくい（天井効果）。')
        elif p >= 0.3:
            fb.append('✅ 正答率が適切範囲(30-70%)。難易度バランス良好。' if p <= 0.7
                      else '✅ 正答率がやや高め。')
        elif p >= 0.2:
            fb.append('⚠ 正答率がやや低い。テスト目的に照らして確認を。')
        else:
            fb.append('❌ 正答率20%未満: 非常に難しい（床効果・当て推量に近い）。')

        if d < 0:
            fb.append('❌ D値が負: 低得点者のほうが正答しやすい異常項目。正答キー・問題文を確認。')
        elif d < 0.2:
            fb.append('⚠ D値が低い: 識別力不足。選択肢の改善を検討。')
        else:
            fb.append(f'✅ D値{round(d*100)}%: 十分な識別力。')

        if it < 0:
            fb.append('❌ I-T相関が負: テスト全体との整合性に問題。')
        elif it < 0.2:
            fb.append('⚠ I-T相関が低い: テスト全体との一貫性が弱い。')
        else:
            fb.append(f'✅ I-T相関{it:.2f}: テストとの一貫性良好。')

        if da > a:
            fb.append(f'⚠ 削除α({da:.2f})>α({a:.2f}): この項目を除くとテスト信頼性が向上。')
        return fb


class CTTPDFReporter:
    """CTT分析PDFレポート生成 (ReportLab)"""

    def __init__(self, output_path):
        if not HAS_REPORTLAB:
            raise RuntimeError("CTT PDFレポート生成にはreportlabが必要です。\npip install reportlab でインストールしてください。")
        self.output_path = output_path
        self.plot_gen = CTTPlotGenerator()

        try:
            pdfmetrics.registerFont(TTFont('Gothic', 'C:\\Windows\\Fonts\\msgothic.ttc'))
            self.fn = 'Gothic'
        except Exception:
            try:
                pdfmetrics.registerFont(TTFont('Gothic', 'C:\\Windows\\Fonts\\msmincho.ttc'))
                self.fn = 'Gothic'
            except Exception:
                self.fn = 'Helvetica'

        self.styles = getSampleStyleSheet()
        self._create_styles()
        self.elements = []

    def _create_styles(self):
        fn = self.fn
        self.styles.add(ParagraphStyle('JP_Title',   fontName=fn, fontSize=22, leading=28, spaceAfter=8, alignment=1))
        self.styles.add(ParagraphStyle('JP_Sub',     fontName=fn, fontSize=12, leading=16, spaceAfter=12, alignment=1,
                                       textColor=rl_colors.HexColor('#555555')))
        self.styles.add(ParagraphStyle('JP_H1',      fontName=fn, fontSize=16, leading=22, spaceBefore=16, spaceAfter=8))
        self.styles.add(ParagraphStyle('JP_H2',      fontName=fn, fontSize=13, leading=18, spaceBefore=10, spaceAfter=6))
        self.styles.add(ParagraphStyle('JP_Body',    fontName=fn, fontSize=10, leading=14))
        self.styles.add(ParagraphStyle('JP_Small',   fontName=fn, fontSize=9,  leading=12,
                                       textColor=rl_colors.HexColor('#444444')))
        self.styles.add(ParagraphStyle('JP_Feedback', fontName=fn, fontSize=10, leading=14, leftIndent=16,
                                       textColor=rl_colors.HexColor('#333333')))
        self.styles.add(ParagraphStyle('JP_Good',    fontName=fn, fontSize=10, leading=14, leftIndent=16,
                                       textColor=rl_colors.HexColor('#006600')))
        self.styles.add(ParagraphStyle('JP_Warn',    fontName=fn, fontSize=10, leading=14, leftIndent=16,
                                       textColor=rl_colors.HexColor('#CC6600')))
        self.styles.add(ParagraphStyle('JP_Bad',     fontName=fn, fontSize=10, leading=14, leftIndent=16,
                                       textColor=rl_colors.HexColor('#CC0000')))
        self.styles.add(ParagraphStyle('TOC_Item',   fontName=fn, fontSize=11, leading=16, leftIndent=20,
                                       textColor=rl_colors.HexColor('#0563C1')))

    CLR_HDR  = rl_colors.HexColor('#4472C4')
    CLR_HDR2 = rl_colors.HexColor('#5B9BD5')
    CLR_SUB  = rl_colors.HexColor('#D6E4F0')
    CLR_GRN  = rl_colors.HexColor('#C6EFCE')
    CLR_RED  = rl_colors.HexColor('#FFC7CE')
    CLR_ORG  = rl_colors.HexColor('#FFE0B2')
    CLR_KEY  = rl_colors.HexColor('#BDD7EE')
    CLR_GRAY = rl_colors.HexColor('#F2F2F2')

    def generate_report(self, test_stats, item_stats, distractor_stats,
                        total_scores, questions=None, score_matrix=None):
        doc = SimpleDocTemplate(
            str(self.output_path), pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=15*mm, bottomMargin=15*mm)

        self._add_title_page(test_stats)
        self._add_toc(item_stats)
        self.elements.append(PageBreak())

        if questions is not None:
            self._add_trace_overview(distractor_stats, questions)

        self._add_summary(test_stats, total_scores, item_stats)
        self.elements.append(PageBreak())

        self._section('■ 得点分布')
        try:
            buf = self.plot_gen.generate_score_histogram(total_scores)
            self.elements.append(RLImage(buf, width=420, height=260))
        except Exception as e:
            self.elements.append(RLParagraph(f"ヒストグラム生成エラー: {e}", self.styles['JP_Body']))

        self.elements.append(Spacer(1, 8))
        self._add_distribution_analysis(total_scores)
        self.elements.append(PageBreak())

        self._add_correlation_charts(item_stats)

        if score_matrix is not None and questions is not None:
            self._add_correlation_heatmap(score_matrix, questions)
        self.elements.append(PageBreak())

        self._add_overall_evaluation(test_stats, item_stats)
        self.elements.append(PageBreak())

        self._section('■ 項目統計一覧')
        self._add_item_table(item_stats)
        self.elements.append(PageBreak())

        self._section('■ 項目別詳細分析')
        for count, (_, row) in enumerate(item_stats.iterrows()):
            if count > 0 and count % 2 == 0:
                self.elements.append(PageBreak())
            self._add_item_detail(row, distractor_stats)

        doc.build(self.elements)
        logger.info("   ✅ CTT PDF: %s", self.output_path)

    def _add_title_page(self, stats):
        self.elements.append(Spacer(1, 1.5*inch))
        self.elements.append(RLParagraph('テスト分析レポート', self.styles['JP_Title']))
        self.elements.append(Spacer(1, 0.2*inch))
        self.elements.append(RLParagraph(
            f"作成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}", self.styles['JP_Sub']))

        n = int(stats.get('受験者数 (N)', 0))
        k = int(stats.get('項目数 (K)', 0))
        alpha = stats.get('信頼性係数 (α)', 0)
        self.elements.append(Spacer(1, 0.5*inch))
        summary_data = [['受験者数', str(n)], ['項目数', str(k)], ['α係数', f"{alpha:.2f}"]]
        t = RLTable(summary_data, colWidths=[1.5*inch, 1.5*inch])
        t.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), self.fn),
            ('FONTSIZE', (0,0), (-1,-1), 12),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, rl_colors.grey),
            ('BACKGROUND', (0,0), (0,-1), self.CLR_HDR),
            ('TEXTCOLOR', (0,0), (0,-1), rl_colors.white),
        ]))
        self.elements.append(t)

        self.elements.append(Spacer(1, 0.4*inch))
        warning_text = (
            '<b>【データ処理に関する注意事項】</b><br/>'
            '本レポートは、各設問を <b>正解=1 / 不正解=0</b> の二値（バイナリ）データとして処理しています。<br/>'
            '• 中間点（部分点）のあるデータ → 不正解（0点）として処理されます<br/>'
            '• 配点（重みづけ）のあるデータ → 配点に関わらず正解は一律1点として処理されます<br/>'
            '• 得点・満点・平均点等は <b>各問1点換算</b> で算出されています')
        warning_style = ParagraphStyle(
            'Warning', parent=self.styles['JP_Small'],
            backColor=rl_colors.Color(1.0, 0.95, 0.85),
            borderColor=rl_colors.Color(0.9, 0.7, 0.3),
            borderWidth=1, borderPadding=8,
            spaceBefore=6, spaceAfter=6, fontSize=8, leading=12)
        self.elements.append(RLParagraph(warning_text, warning_style))

    def _add_toc(self, item_stats):
        self.elements.append(Spacer(1, 0.6*inch))
        self.elements.append(RLParagraph('── 目次 ──', self.styles['JP_H2']))
        toc_items = [
            '1. トレースライン一覧（逆転項目スクリーニング）',
            '2. テスト要約統計量', '3. 得点分布と分布特性',
            '4. 設問別相関係数', '5. 設問間相関行列（Phi係数ヒートマップ）',
            '6. テスト全体の評価', '7. 項目統計一覧', '8. 項目別詳細分析',
        ]
        for qid_idx, (_, row) in enumerate(item_stats.iterrows()):
            toc_items.append(f"   8-{qid_idx+1}. 項目 {row['QuestionID']}")
        for item in toc_items:
            self.elements.append(RLParagraph(item, self.styles['TOC_Item']))

    def _add_summary(self, stats, total_scores, item_stats):
        self._section('■ テスト要約統計量')
        n = int(stats.get('受験者数 (N)', 0))
        k = int(stats.get('項目数 (K)', 0))
        mean = stats.get('平均点 (Mean)', 0)
        sd   = stats.get('標準偏差 (SD)', 0)
        mn   = stats.get('最低点 (Min)', 0)
        mx   = stats.get('最高点 (Max)', 0)
        med  = stats.get('中央値 (Median)', 0)
        alpha = stats.get('信頼性係数 (α)', 0)

        scores_arr = np.array(total_scores)
        q1 = np.percentile(scores_arr, 25)
        q3 = np.percentile(scores_arr, 75)
        iqr = q3 - q1
        variance = stats.get('分散 (Variance)', np.var(scores_arr, ddof=1))
        n_s = len(scores_arr)
        if n_s > 2 and sd > 0:
            skewness = (n_s / ((n_s - 1) * (n_s - 2))) * np.sum(((scores_arr - mean) / sd) ** 3)
        else:
            skewness = 0
        if n_s > 3 and sd > 0:
            kurtosis = ((n_s * (n_s + 1)) / ((n_s - 1) * (n_s - 2) * (n_s - 3))) * \
                       np.sum(((scores_arr - mean) / sd) ** 4) - \
                       (3 * (n_s - 1) ** 2) / ((n_s - 2) * (n_s - 3))
        else:
            kurtosis = 0
        sem = sd * np.sqrt(1 - alpha) if alpha < 1 else 0

        data = [
            ['統計量', '値', '解説'],
            ['受験者数',         str(n),          '分析対象の総受験者数'],
            ['項目数',           str(k),          'テストの問題数'],
            ['平均点',           f'{mean:.2f}',   '全受験者の平均得点（各問1点換算）'],
            ['平均得点率',       f'{round(mean/k*100) if k>0 else 0}%', '平均点÷項目数×100'],
            ['標準偏差 (SD)',    f'{sd:.2f}',     '得点のばらつきの大きさ'],
            ['分散',             f'{variance:.2f}', 'SDの2乗'],
            ['最小値',           str(int(mn)),    '最低得点'],
            ['第1四分位 (Q1)',   f'{q1:.1f}',     '下位25%の境界点'],
            ['中央値',           f'{med:.1f}',    '受験者の半数が超える得点'],
            ['第3四分位 (Q3)',   f'{q3:.1f}',     '上位25%の境界点'],
            ['最大値',           str(int(mx)),    '最高得点'],
            ['四分位範囲 (IQR)', f'{iqr:.1f}',    'Q3−Q1: 中央50%の得点幅'],
            ['歪度',             f'{skewness:.2f}', '正=右に裾が長い / 負=左に裾が長い'],
            ['尖度',             f'{kurtosis:.2f}', '正=鋭い分布 / 負=平坦な分布'],
            ['α係数',           f'{alpha:.2f}',  'テストの内的整合性（信頼性）'],
            ['測定標準誤差 (SEM)', f'{sem:.2f}',  'SD×√(1−α): 得点の測定精度'],
        ]
        t = RLTable(data, colWidths=[2*inch, 1*inch, 3.2*inch])
        t.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), self.fn),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('BACKGROUND', (0,0), (-1,0), self.CLR_HDR),
            ('TEXTCOLOR', (0,0), (-1,0), rl_colors.white),
            ('ALIGN', (1,1), (1,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-1), 0.5, rl_colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [rl_colors.white, self.CLR_GRAY]),
        ]))
        self.elements.append(t)
        self.elements.append(Spacer(1, 0.2*inch))

        self.elements.append(RLParagraph('【指標の読み方ガイド】', self.styles['JP_H2']))
        guide = (
            '<b>正答率 (P値):</b> 0.30〜0.70が識別に最適な範囲。0.50で最大識別力。<br/>'
            '<b>識別指数 (D値):</b> 上位27%群の正答率−下位27%群の正答率。≧0.30=優良 / 0.20〜0.29=改善余地 / ＜0.20=要注意。<br/>'
            '<b>I-T相関:</b> 項目得点と残余合計得点のpoint-biserial相関。0.20以上が目安。負は不適切項目。<br/>'
            '<b>α係数 (Cronbach):</b> ≧0.90=非常に高い / ≧0.80=十分 / ≧0.70=許容 / ＜0.70=改善必要。<br/>'
            '<b>削除α:</b> その項目を除いたα係数。α係数より大きい場合、その項目がテストの信頼性を下げている。<br/>'
            '<b>SEM:</b> 生徒の真の得点は「観測得点±1.96×SEM」の範囲に95%の確率で含まれる。<br/>')
        self.elements.append(RLParagraph(guide, self.styles['JP_Small']))

    def _add_distribution_analysis(self, total_scores):
        scores = np.array(total_scores)
        mean = np.mean(scores)
        sd = np.std(scores, ddof=1)
        n = len(scores)
        if n > 2 and sd > 0:
            skew = (n / ((n-1)*(n-2))) * np.sum(((scores - mean) / sd) ** 3)
        else:
            skew = 0

        self.elements.append(RLParagraph('【分布の特徴】', self.styles['JP_H2']))
        if abs(skew) < 0.5:
            self.elements.append(RLParagraph(
                f"✅ 歪度 = {skew:.2f}: 得点分布はほぼ対称。テストの難易度は適切と考えられます。",
                self.styles['JP_Good']))
        elif skew > 0:
            self.elements.append(RLParagraph(
                f"📝 歪度 = {skew:.2f}: 得点分布が右に偏っています（低得点者が多い）。",
                self.styles['JP_Feedback']))
        else:
            self.elements.append(RLParagraph(
                f"📝 歪度 = {skew:.2f}: 得点分布が左に偏っています（高得点者が多い）。",
                self.styles['JP_Feedback']))

        k = len(scores)
        pct_low  = np.sum(scores < mean - sd) / k * 100
        pct_mid  = np.sum((scores >= mean - sd) & (scores <= mean + sd)) / k * 100
        pct_high = np.sum(scores > mean + sd) / k * 100
        self.elements.append(RLParagraph(
            f"得点帯: 低得点(平均−1SD以下) {pct_low:.0f}% ／ 中間 {pct_mid:.0f}% ／ 高得点(平均+1SD以上) {pct_high:.0f}%",
            self.styles['JP_Feedback']))

    def _add_overall_evaluation(self, stats, item_stats):
        self._section('■ テスト全体の評価と改善提案')
        alpha = stats.get('信頼性係数 (α)', 0)
        mean = stats.get('平均点 (Mean)', 0)
        k = int(stats.get('項目数 (K)', 0))

        self.elements.append(RLParagraph('<b>1. 信頼性の評価</b>', self.styles['JP_H2']))
        if alpha >= 0.9:
            self.elements.append(RLParagraph(
                f"✅ α係数 = {alpha:.2f}: 非常に高い信頼性。個人の得点に基づく判断にも十分使用できます。",
                self.styles['JP_Good']))
        elif alpha >= 0.8:
            self.elements.append(RLParagraph(
                f"✅ α係数 = {alpha:.2f}: 十分な信頼性。グループ間の比較や指導改善に適しています。",
                self.styles['JP_Good']))
        elif alpha >= 0.7:
            self.elements.append(RLParagraph(
                f"⚠ α係数 = {alpha:.2f}: やや低い信頼性。項目の改善により信頼性向上が期待できます。",
                self.styles['JP_Warn']))
        else:
            self.elements.append(RLParagraph(
                f"❌ α係数 = {alpha:.2f}: 信頼性が低い状態。テスト得点を重要な判断に使用する際は注意が必要です。",
                self.styles['JP_Bad']))

        self.elements.append(RLParagraph('<b>2. 難易度の評価</b>', self.styles['JP_H2']))
        if k > 0:
            mean_ratio = mean / k
            p_values = item_stats['正答率 (P)'].values
            easy_count = int(np.sum(p_values >= 0.8))
            hard_count = int(np.sum(p_values <= 0.2))
            optimal_count = int(np.sum((p_values >= 0.3) & (p_values <= 0.7)))
            self.elements.append(RLParagraph(
                f"平均得点率: {round(mean_ratio*100)}%　（理想的な範囲: 50〜70%）",
                self.styles['JP_Body']))
            self.elements.append(RLParagraph(
                f"項目の難易度分布: 易しい(≧80%) {easy_count}問 / 適切(30-70%) {optimal_count}問 / 難しい(≦20%) {hard_count}問",
                self.styles['JP_Feedback']))

        self.elements.append(RLParagraph('<b>3. 識別力の評価</b>', self.styles['JP_H2']))
        d_values = item_stats['識別指数 (D)'].values
        it_values = item_stats['I-T相関'].values
        neg_d = item_stats[item_stats['識別指数 (D)'] < 0]['QuestionID'].tolist()
        low_d = item_stats[(item_stats['識別指数 (D)'] >= 0) & (item_stats['識別指数 (D)'] < 0.2)]['QuestionID'].tolist()
        neg_it = item_stats[item_stats['I-T相関'] < 0]['QuestionID'].tolist()

        if neg_d:
            self.elements.append(RLParagraph(
                f"❌ D値が負の項目: {', '.join(str(q) for q in neg_d)}　→ 正答キーの確認が必要です。",
                self.styles['JP_Bad']))
        if low_d:
            self.elements.append(RLParagraph(
                f"⚠ D値が低い項目(0〜0.20未満): {', '.join(str(q) for q in low_d)}　→ 選択肢の改善を検討してください。",
                self.styles['JP_Warn']))
        if neg_it:
            self.elements.append(RLParagraph(
                f"❌ I-T相関が負の項目: {', '.join(str(q) for q in neg_it)}　→ テスト全体との整合性に問題があります。",
                self.styles['JP_Bad']))

        good_items = item_stats[(item_stats['識別指数 (D)'] >= 0.3) & (item_stats['I-T相関'] >= 0.2)]
        self.elements.append(RLParagraph(
            f"✅ 良好な項目(D≧0.30 かつ I-T≧0.20): {len(good_items)}問 / {k}問 ({round(len(good_items)/k*100) if k else 0}%)",
            self.styles['JP_Good']))

        self.elements.append(RLParagraph('<b>4. 改善のための提案</b>', self.styles['JP_H2']))
        improve_alpha = item_stats[item_stats['削除α'] > alpha]['QuestionID'].tolist()
        if improve_alpha:
            self.elements.append(RLParagraph(
                f"📋 以下の項目を削除するとα係数が改善: {', '.join(str(q) for q in improve_alpha)}",
                self.styles['JP_Feedback']))

        mean_d = np.mean(d_values)
        mean_it = np.mean(it_values)
        self.elements.append(RLParagraph(
            f"📊 D値の平均: {mean_d:.2f} / I-T相関の平均: {mean_it:.2f}",
            self.styles['JP_Feedback']))

    def _add_item_table(self, df):
        headers = ['項目', '正答', '正答率', 'D値', 'I-T相関', '削除α', '判定']
        data = [headers]
        row_flags = []
        for _, row in df.iterrows():
            d_val = row['識別指数 (D)']
            it_val = row['I-T相関']
            if d_val < 0 or it_val < 0:
                judge = '❌ 要修正'
            elif d_val < 0.2 or it_val < 0.2:
                judge = '⚠ 要改善'
            elif d_val >= 0.3 and it_val >= 0.3:
                judge = '✅ 優良'
            else:
                judge = '○ 良好'
            r = [str(row['QuestionID']), str(row['Key']),
                 f"{round(row['正答率 (P)'] * 100)}%", f"{round(d_val * 100)}%",
                 f"{it_val:.2f}", f"{row['削除α']:.2f}", judge]
            row_flags.append(d_val < 0 or it_val < 0)
            data.append(r)

        t = RLTable(data, colWidths=[0.7*inch, 0.5*inch, 0.8*inch, 0.7*inch, 0.8*inch, 0.7*inch, 1*inch],
                    repeatRows=1)
        cmds = [
            ('FONTNAME', (0,0), (-1,-1), self.fn),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('BACKGROUND', (0,0), (-1,0), self.CLR_HDR),
            ('TEXTCOLOR', (0,0), (-1,0), rl_colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, rl_colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [rl_colors.white, self.CLR_GRAY]),
        ]
        for i, bad in enumerate(row_flags):
            if bad:
                cmds.append(('BACKGROUND', (0, i+1), (-1, i+1), self.CLR_RED))
                cmds.append(('TEXTCOLOR', (0, i+1), (-1, i+1), rl_colors.HexColor('#CC0000')))
        t.setStyle(TableStyle(cmds))
        self.elements.append(t)

    def _add_item_detail(self, row, distractor_stats):
        qid = row['QuestionID']
        key_val = str(row['Key'])
        p_val = row['正答率 (P)']
        d_val = row['識別指数 (D)']
        it_val = row['I-T相関']
        del_alpha = row['削除α']

        if d_val < 0 or it_val < 0:
            judge = '❌ 要修正'
        elif d_val < 0.2 or it_val < 0.2:
            judge = '⚠ 要改善'
        elif d_val >= 0.3 and it_val >= 0.3:
            judge = '✅ 優良'
        else:
            judge = '○ 良好'

        self.elements.append(RLParagraph(
            f"<b>● 項目 {qid}</b>　正答: {key_val}　正答率: {round(p_val*100)}%　"
            f"D値: {round(d_val*100)}%　I-T: {it_val:.2f}　<b>{judge}</b>",
            self.styles['JP_H2']))

        if d_val < 0:
            self.elements.append(RLParagraph(
                "❌ D値が負: 低得点者が正答しやすい異常項目。正答キー・問題文の確認が必要。",
                self.styles['JP_Bad']))
        elif d_val < 0.2:
            self.elements.append(RLParagraph(
                "⚠ D値が低い: 識別力不足。選択肢の魅力度や問題の明確さを改善してください。",
                self.styles['JP_Warn']))
        else:
            self.elements.append(RLParagraph(
                f"✅ D値が{round(d_val*100)}%で十分な識別力。", self.styles['JP_Good']))

        if it_val < 0:
            self.elements.append(RLParagraph(
                "❌ I-T相関が負: テスト全体と整合しない可能性。", self.styles['JP_Bad']))

        self.elements.append(Spacer(1, 4))

        try:
            img_buf = self.plot_gen.generate_item_curve(row, distractor_stats)
            img = RLImage(img_buf, width=280, height=200)
        except Exception as e:
            img = RLParagraph(f"チャートエラー: {e}", self.styles['JP_Body'])

        d_subset = distractor_stats[distractor_stats['QuestionID'] == qid]
        t_data = [['選択肢', '全体', '高群', '中群', '低群']]
        for _, d_row in d_subset.iterrows():
            choice = d_row['Choice']
            is_key = d_row['IsKey']
            label = f"★{choice}" if is_key else str(choice)
            t_data.append([
                label,
                f"{d_row['Ratio_全体']:.0%}",
                f"{d_row['Ratio_高群']:.0%}",
                f"{d_row.get('Ratio_中群', 0):.0%}",
                f"{d_row['Ratio_低群']:.0%}"])

        stats_table = RLTable(t_data, colWidths=[0.6*inch, 0.55*inch, 0.55*inch, 0.55*inch, 0.55*inch])
        t_cmds = [
            ('FONTNAME', (0,0), (-1,-1), self.fn),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, rl_colors.grey),
            ('BACKGROUND', (0,0), (-1,0), self.CLR_HDR),
            ('TEXTCOLOR', (0,0), (-1,0), rl_colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ]
        for ri, (_, d_row) in enumerate(d_subset.iterrows(), 1):
            if d_row['IsKey']:
                t_cmds.append(('BACKGROUND', (0, ri), (-1, ri), self.CLR_GRN))
        stats_table.setStyle(TableStyle(t_cmds))

        layout = RLTable([[img, stats_table]], colWidths=[290, 220])
        layout.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ]))
        self.elements.append(layout)
        self.elements.append(Spacer(1, 0.25*inch))

    def _add_trace_overview(self, distractor_stats, questions):
        self._section('■ トレースライン一覧（逆転項目スクリーニング）')
        self.elements.append(RLParagraph(
            '正答トレースライン: 青実線 / 誤答: 灰色点線 / 背景ピンク: 逆転の疑い',
            self.styles['JP_Small']))
        self.elements.append(Spacer(1, 4))
        try:
            pages = self.plot_gen.generate_mini_trace_grid(distractor_stats, questions)
            for i, (buf, aspect_ratio) in enumerate(pages):
                if i > 0:
                    self.elements.append(PageBreak())
                img_width = 520
                img_height = img_width / aspect_ratio
                self.elements.append(RLImage(buf, width=img_width, height=img_height))
        except Exception as e:
            self.elements.append(RLParagraph(f"トレースライン一覧生成エラー: {e}", self.styles['JP_Body']))
        self.elements.append(PageBreak())

    def _add_correlation_charts(self, item_stats):
        self._section('■ 設問別 I-T 相関係数')
        self.elements.append(RLParagraph(
            'I-T相関: 各設問の得点(0/1)と合計得点の点双列相関係数。0.20以上が望ましい。',
            self.styles['JP_Small']))
        self.elements.append(Spacer(1, 6))

        self.elements.append(RLParagraph(
            '<b>【修正I-T相関】合計得点からその設問を除いたバージョン（推奨）</b>',
            self.styles['JP_H2']))
        try:
            buf1 = self.plot_gen.generate_correlation_chart(
                item_stats, 'I-T相関', '修正I-T相関（合計得点からその設問を除外）')
            self.elements.append(RLImage(buf1, width=500, height=230))
        except Exception as e:
            self.elements.append(RLParagraph(f"グラフ生成エラー: {e}", self.styles['JP_Body']))

        self.elements.append(Spacer(1, 12))
        if 'I-T相関(含)' in item_stats.columns:
            self.elements.append(RLParagraph(
                '<b>【非修正I-T相関】合計得点にその設問を含むバージョン</b>',
                self.styles['JP_H2']))
            try:
                buf2 = self.plot_gen.generate_correlation_chart(
                    item_stats, 'I-T相関(含)', '非修正I-T相関（合計得点にその設問を含む）')
                self.elements.append(RLImage(buf2, width=500, height=230))
            except Exception as e:
                self.elements.append(RLParagraph(f"グラフ生成エラー: {e}", self.styles['JP_Body']))

    def _add_correlation_heatmap(self, score_matrix, questions):
        self.elements.append(PageBreak())
        self._section('■ 設問間相関行列（Phi係数ヒートマップ）')
        self.elements.append(RLParagraph(
            'Phi係数: 0/1バイナリデータ間のピアソン相関係数。青=正の相関 / 赤=負の相関。',
            self.styles['JP_Small']))
        self.elements.append(Spacer(1, 6))
        try:
            buf = self.plot_gen.generate_correlation_heatmap(score_matrix, questions)
            map_size = min(480, 480)
            self.elements.append(RLImage(buf, width=map_size, height=map_size))
        except Exception as e:
            self.elements.append(RLParagraph(f"ヒートマップ生成エラー: {e}", self.styles['JP_Body']))

    def _section(self, text):
        self.elements.append(RLParagraph(text, self.styles['JP_H1']))
        self.elements.append(Spacer(1, 0.05*inch))


def generate_ctt_analysis(template_path, mark2_result_path, excel_output_path,
                          pdf_output_path, skip_questions=0,
                          descriptive_config=None, descriptive_scores=None,
                          template_dict=None, mark2_results=None):
    """
    古典テスト理論(CTT)分析のExcel+PDFレポートを生成する統括関数。
    
    Args:
        template_path: テンプレートExcelのパス
        mark2_result_path: Mark2読取結果Excelのパス
        excel_output_path: CTT分析Excel出力パス
        pdf_output_path: CTT分析PDF出力パス
        skip_questions: スキップする列数
        descriptive_config: 記述問題設定dict（オプション）
        descriptive_scores: {ファイル名: {問題ID: 得点}} の辞書（オプション）
        template_dict: 事前読込済みのテンプレートdict（Noneなら内部で読込）
        mark2_results: 事前読込済みのMark2結果list（Noneなら内部で読込）
    
    Returns:
        dict: 成功フラグと統計情報
    """
    logger.info("=" * 60)
    logger.info("古典テスト理論(CTT)分析レポート生成")
    logger.info("=" * 60)
    
    try:
        # 1. データ変換（記述問題を含む場合は統合）
        logger.info("  データ変換中...")
        ans_df, key_df = convert_mark2_to_ctt_data(
            template_path, mark2_result_path, skip_questions,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
            template_dict=template_dict,
            mark2_results=mark2_results,
        )
        logger.info("  ✓ 受験者数: %d, 設問数: %d", len(ans_df), len(key_df))
        
        # 2. CTT分析実行
        logger.info("  CTT分析実行中...")
        analyzer = CTTAnalyzer(ans_df, key_df)
        test_stats = analyzer.calculate_test_stats()
        item_stats = analyzer.calculate_item_stats()
        distractor_stats = analyzer.calculate_distractor_analysis()
        student_stats = analyzer.ans_df
        
        logger.info("  ✓ 平均点(各問1点換算): %.2f", test_stats['平均点 (Mean)'])
        logger.info("  ✓ α係数: %.2f", test_stats['信頼性係数 (α)'])
        
        # 3. Excelレポート
        logger.info("  Excelレポート生成中...")
        exporter = CTTExcelExporter(str(excel_output_path), analyzer)
        exporter.export(test_stats, student_stats, item_stats, distractor_stats)
    
        # 4. PDFレポート
        pdf_success = False
        if HAS_REPORTLAB and HAS_MATPLOTLIB:
            logger.info("  PDFレポート生成中...")
            try:
                reporter = CTTPDFReporter(str(pdf_output_path))
                reporter.generate_report(
                    test_stats, item_stats, distractor_stats,
                    analyzer.total_scores, analyzer.questions, analyzer.score_matrix)
                pdf_success = True
            except Exception as e:
                logger.warning("  ⚠ PDF生成エラー: %s", e)
                import traceback
                traceback.print_exc()
        else:
            missing = []
            if not HAS_REPORTLAB:
                missing.append('reportlab')
            if not HAS_MATPLOTLIB:
                missing.append('matplotlib')
            logger.warning("  ⚠ PDFレポートスキップ（%s 未インストール）", ', '.join(missing))
        
        logger.info("=" * 60)
        logger.info("CTT分析レポート生成完了")
        logger.info("=" * 60)
        
        return {
            'success': True,
            'pdf_success': pdf_success,
            'test_stats': test_stats,
            'excel_path': str(excel_output_path),
            'pdf_path': str(pdf_output_path) if pdf_success else None
        }

    except Exception as e:
        import traceback
        logger.error("  ✗ CTT分析エラー: %s", e)
        traceback.print_exc()
        return {
            'success': False,
            'pdf_success': False,
            'test_stats': None,
            'excel_path': None,
            'pdf_path': None,
            'error': str(e)
        }
