"""
test_special_all_correct.py — 特例（全員正解）機能のテスト

不適切問題の救済措置として answer_key.xlsx の「特例」列に「全員正解」を
指定した場合の挙動を検証する。

- load_template: 特例列の読み取り・正答空欄の許容・後方互換
- score_answers: 無回答含め全員満点、resultsへのspecial伝播
- CTT: score_matrix全員正答化、項目分析(item_stats/選択肢分析/α)からの除外
- image_renderer: ★マーク描画（正答位置／正答未登録なら左端、トグルOFFで非表示）
- omr_engine: テンプレート生成に特例列とドロップダウンが含まれる
"""
import sys
from pathlib import Path

import pandas as pd
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "main_src"))

from scoring_engine import load_template, score_answers, SPECIAL_ALL_CORRECT


# ── テンプレート生成ヘルパー ──────────────────────────


def _create_template_xlsx(path, questions, include_special_col=True):
    """テスト用 answer_key.xlsx を生成する。

    Args:
        path: 保存先パス
        questions: [{問題番号, 正答, 配点, 観点, 特例}, ...] のリスト
        include_special_col: Falseなら特例列なし（旧形式テンプレートを模倣）
    """
    rows = []
    for q in questions:
        row = {
            '問題番号': q['問題番号'],
            '正答': q.get('正答', ''),
            '配点': q.get('配点', ''),
            '観点': q.get('観点', ''),
        }
        if include_special_col:
            row['特例'] = q.get('特例', '')
        rows.append(row)
    df = pd.DataFrame(rows)
    df.to_excel(path, index=False)


# ── load_template ──────────────────────────────────────


class TestLoadTemplateSpecial:
    """load_template() の特例列読み取りテスト"""

    def test_all_correct_with_answer(self, tmp_path):
        """特例=全員正解（正答登録あり）→ 採点対象、特例が設定される"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '4', '配点': 3, '観点': 1, '特例': '全員正解'},
        ])
        td = load_template(path)
        assert set(td.keys()) == {1, 2}
        assert td[1]['特例'] == ''
        assert td[2]['特例'] == SPECIAL_ALL_CORRECT

    def test_all_correct_without_answer(self, tmp_path):
        """特例=全員正解なら正答空欄でもスキップされない"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '', '配点': 3, '観点': 1, '特例': '全員正解'},
        ])
        td = load_template(path)
        assert set(td.keys()) == {1, 2}
        assert td[2]['正答'] == ''
        assert td[2]['特例'] == SPECIAL_ALL_CORRECT
        assert td[2]['配点'] == 3

    def test_all_correct_requires_points(self, tmp_path):
        """特例=全員正解でも配点空欄ならスキップ（配点は必須のまま）"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '', '配点': '', '観点': '', '特例': '全員正解'},
        ])
        td = load_template(path)
        assert set(td.keys()) == {1}

    def test_unknown_special_value_falls_back(self, tmp_path):
        """未対応の特例値は通常扱い（正答登録済みなら採点対象・特例なし）"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1, '特例': '配点ゼロ'},
        ])
        td = load_template(path)
        assert td[1]['特例'] == ''

    def test_unknown_special_with_empty_answer_skipped(self, tmp_path):
        """未対応の特例値+正答空欄 → 通常ルール通りスキップ"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '', '配点': 2, '観点': 1, '特例': '謎の値'},
        ])
        td = load_template(path)
        assert td == {}

    def test_no_special_column_backward_compat(self, tmp_path):
        """特例列なし（旧形式）→ 全問特例なしで読み込める"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
        ], include_special_col=False)
        td = load_template(path)
        assert td[1]['特例'] == ''


# ── score_answers ──────────────────────────────────────


def _template_dict(entries):
    """{q_no: (正答, 配点, 観点, 特例)} から template_dict を組み立てる"""
    return {
        q: {'正答': ans, '配点': pts, '観点': asp, '問題概要': '', '特例': sp}
        for q, (ans, pts, asp, sp) in entries.items()
    }


class TestScoreAnswersAllCorrect:
    """score_answers() の全員正解特例テスト"""

    def test_wrong_answer_gets_full_points(self):
        """誤答でも正解扱いで満点"""
        td = _template_dict({
            1: ('3', 2, 1, ''),
            2: ('4', 3, 1, SPECIAL_ALL_CORRECT),
        })
        result = score_answers({1: '3', 2: '1'}, td)  # 問2は誤答
        assert result['results'][2]['correct'] is True
        assert result['results'][2]['points'] == 3
        assert result['total_score'] == 5
        assert result['max_score'] == 5

    def test_no_answer_gets_full_points(self):
        """無回答でも正解扱いで満点"""
        td = _template_dict({1: ('4', 3, 1, SPECIAL_ALL_CORRECT)})
        result = score_answers({}, td)  # 無回答
        assert result['results'][1]['correct'] is True
        assert result['results'][1]['points'] == 3
        assert result['total_score'] == 3

    def test_empty_correct_answer(self):
        """正答未登録（空欄）の特例問題も全員満点"""
        td = _template_dict({1: ('', 2, 1, SPECIAL_ALL_CORRECT)})
        result = score_answers({1: '5'}, td)
        assert result['results'][1]['correct'] is True
        assert result['results'][1]['points'] == 2

    def test_special_propagated_to_results(self):
        """resultsに特例区分が伝播する（描画・レポートで参照）"""
        td = _template_dict({
            1: ('3', 2, 1, ''),
            2: ('', 3, 1, SPECIAL_ALL_CORRECT),
        })
        result = score_answers({1: '3'}, td)
        assert result['results'][1]['special'] == ''
        assert result['results'][2]['special'] == SPECIAL_ALL_CORRECT

    def test_aspect_scores_include_special(self):
        """観点別得点にも特例分が加算される"""
        td = _template_dict({
            1: ('3', 2, 1, ''),
            2: ('', 3, 2, SPECIAL_ALL_CORRECT),
        })
        result = score_answers({1: '3'}, td)
        assert result['aspect_scores'][2] == 3
        assert result['aspect_max_scores'][2] == 3

    def test_zero_points_question_regression(self):
        """配点0の問題: ○×は正誤通り、得点・満点に影響しない（既存挙動の回帰確認）"""
        td = _template_dict({
            1: ('3', 0, 1, ''),
            2: ('4', 5, 1, ''),
        })
        result = score_answers({1: '3', 2: '4'}, td)
        assert result['results'][1]['correct'] is True
        assert result['results'][1]['points'] == 0
        assert result['total_score'] == 5
        assert result['max_score'] == 5


# ── CTT分析 ────────────────────────────────────────────


def _make_ctt_analyzer():
    """3問×6人のCTTAnalyzerを作る。設問3が特例(全員正解・正答未登録)"""
    from ctt_analyzer import CTTAnalyzer
    ans_df = pd.DataFrame({
        'StudentID': [f's{i}.png' for i in range(1, 7)],
        '1': ['1', '1', '1', '2', '2', '3'],
        '2': ['2', '2', '3', '2', '1', '1'],
        '3': ['5', '', '1', '2', '4', ''],  # 実際のマークはバラバラ/無回答
    })
    key_df = pd.DataFrame({
        'QuestionID': ['1', '2', '3'],
        'Key': ['1', '2', ''],
        'Summary': ['', '', ''],
        'AllCorrect': [False, False, True],
    })
    return CTTAnalyzer(ans_df, key_df)


class TestCTTAllCorrect:
    """CTTAnalyzer の特例(全員正解)除外テスト"""

    def test_score_matrix_all_ones(self):
        """特例設問は無回答含め全員正答(1)"""
        az = _make_ctt_analyzer()
        assert (az.score_matrix['3'] == 1).all()

    def test_total_scores_include_special(self):
        """合計得点に特例設問の1点が含まれる"""
        az = _make_ctt_analyzer()
        # 生徒1: 問1正答+問2正答+問3特例 = 3
        assert az.total_scores.iloc[0] == 3

    def test_analysis_questions_exclude_special(self):
        """項目分析対象から特例設問が除外される"""
        az = _make_ctt_analyzer()
        assert az.analysis_questions == ['1', '2']

    def test_item_stats_exclude_special(self):
        """項目統計(P/D/I-T相関)に特例設問が現れない"""
        az = _make_ctt_analyzer()
        item_stats = az.calculate_item_stats()
        assert set(item_stats['QuestionID']) == {'1', '2'}

    def test_distractor_exclude_special(self):
        """選択肢分析に特例設問が現れない"""
        az = _make_ctt_analyzer()
        dist = az.calculate_distractor_analysis()
        assert set(dist['QuestionID']) == {'1', '2'}

    def test_alpha_excludes_special(self):
        """α係数は特例設問を除いて計算される（定数項目でkが増えない）"""
        from ctt_analyzer import CTTAnalyzer
        az = _make_ctt_analyzer()
        # 特例設問を除いた2問だけのanalyzerと同じαになるはず
        ans_df2 = az.ans_df[['StudentID', '1', '2']].copy()
        key_df2 = pd.DataFrame({
            'QuestionID': ['1', '2'], 'Key': ['1', '2'],
            'Summary': ['', ''], 'AllCorrect': [False, False],
        })
        az2 = CTTAnalyzer(ans_df2, key_df2)
        assert az._calculate_cronbach_alpha() == pytest.approx(az2._calculate_cronbach_alpha())

    def test_backward_compat_no_allcorrect_column(self):
        """AllCorrect列なしのkey_df（旧形式）でも動作する"""
        from ctt_analyzer import CTTAnalyzer
        ans_df = pd.DataFrame({
            'StudentID': ['s1.png', 's2.png'],
            '1': ['1', '2'],
        })
        key_df = pd.DataFrame({'QuestionID': ['1'], 'Key': ['1']})
        az = CTTAnalyzer(ans_df, key_df)
        assert az.analysis_questions == ['1']


class TestConvertMark2ToCttData:
    """convert_mark2_to_ctt_data が特例フラグをkey_dfに伝播するテスト"""

    def test_allcorrect_flag_in_key_df(self):
        from ctt_analyzer import convert_mark2_to_ctt_data
        template_dict = {
            1: {'正答': '3', '配点': 2, '観点': 1, '問題概要': '', '特例': ''},
            2: {'正答': '', '配点': 3, '観点': 1, '問題概要': '', '特例': SPECIAL_ALL_CORRECT},
        }
        mark2_results = [
            {'image': 's1.png', 'answers': {1: '3', 2: '1'}},
            {'image': 's2.png', 'answers': {1: '2', 2: ''}},
        ]
        _, key_df = convert_mark2_to_ctt_data(
            None, None, 0, template_dict=template_dict, mark2_results=mark2_results)
        flags = dict(zip(key_df['QuestionID'], key_df['AllCorrect']))
        assert flags == {'1': False, '2': True}

    def test_descriptive_questions_not_excluded(self):
        """記述問題(concat追加)がNaN→True誤判定で除外されないこと(回帰)"""
        from ctt_analyzer import convert_mark2_to_ctt_data, CTTAnalyzer
        template_dict = {
            1: {'正答': '3', '配点': 2, '観点': 1, '問題概要': '', '特例': ''},
        }
        mark2_results = [
            {'image': 's1.png', 'answers': {1: '3'}},
            {'image': 's2.png', 'answers': {1: '2'}},
        ]
        descriptive_config = {'questions': [
            {'id': 'D1', 'max_score': 5, 'name': '記述1', 'aspect': 1},
        ]}
        descriptive_scores = {'s1.png': {'D1': 5}, 's2.png': {'D1': 3}}
        ans_df, key_df = convert_mark2_to_ctt_data(
            None, None, 0,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
            template_dict=template_dict, mark2_results=mark2_results)
        az = CTTAnalyzer(ans_df, key_df)
        # 記述問題D1は項目分析対象に残る
        assert 'D1' in az.analysis_questions


# ── R連携（分散ゼロ自動除外の確認） ────────────────────


class TestRExportExcludesSpecial:
    """特例設問(全員正答=分散ゼロ)がR連携キットから自動除外されることの確認"""

    def test_constant_column_dropped(self, tmp_path):
        from r_export import export_r_analysis_kit
        key_path = tmp_path / "key.xlsx"
        _create_template_xlsx(key_path, [
            {'問題番号': 1, '正答': '1', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '2', '配点': 2, '観点': 1},
            {'問題番号': 3, '正答': '', '配点': 3, '観点': 1, '特例': '全員正解'},
        ])
        # Mark2結果xlsx（3行ヘッダー構造）を作成
        import openpyxl
        mark2_path = tmp_path / "mark2.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['No', 'File', '1', '2', '3'])
        ws.append([None, None, '1', '2', '3'])
        answers = [('1', '2', '5'), ('2', '2', ''), ('1', '1', '3'), ('2', '1', '1')]
        for i, (a1, a2, a3) in enumerate(answers, 1):
            ws.append([i, f's{i}.png', a1, a2, a3])
        wb.save(mark2_path)
        wb.close()

        result = export_r_analysis_kit(
            str(key_path), str(mark2_path), str(tmp_path), skip_questions=0)
        assert result['success'], result.get('error')
        scored = pd.read_csv(Path(result['output_dir']) / 'scored_data.csv', index_col=0)
        # 特例設問(Q003)は分散ゼロで自動除外される
        assert 'Q003' not in scored.columns
        assert {'Q001', 'Q002'}.issubset(set(scored.columns))


# ── 画像描画（★マーク） ────────────────────────────────


def _render_scoring(results_dict, num_choices=5, rendering_settings=None):
    """_draw_scoring_on_pil を白画像上で実行し、PIL Imageを返す"""
    from PIL import Image, ImageDraw
    from image_renderer import _draw_scoring_on_pil

    cell = 40
    coordinates = []
    for q_no, _ in results_dict.items():
        for ci in range(num_choices):
            coordinates.append({
                'question_no': q_no,
                'x': 10 + ci * cell, 'y': 10 + (q_no - 1) * cell,
                'width': cell - 8, 'height': cell - 8,
            })
    img = Image.new('RGB', (10 + num_choices * cell + 50, 10 + len(results_dict) * cell + 50),
                    (255, 255, 255))
    draw = ImageDraw.Draw(img)
    scoring_result = {'results': results_dict}
    _draw_scoring_on_pil(draw, coordinates, scoring_result,
                         rendering_settings=rendering_settings)
    return img


def _box_has_ink(img, q_no, choice_index):
    """指定選択肢ボックス内に非白ピクセルがあるか"""
    cell = 40
    x0 = 10 + choice_index * cell
    y0 = 10 + (q_no - 1) * cell
    region = img.crop((x0, y0, x0 + cell - 8, y0 + cell - 8))
    return any(px != (255, 255, 255) for px in region.getdata())


def _result(correct=True, points=2, student='3', key='3', special=''):
    return {
        'correct': correct, 'points': points, 'max_points': points,
        'student_answer': student, 'correct_answer': key,
        'aspect': 1, 'special': special,
    }


# ○×・得点等を消して★だけを検証するための設定
_STAR_ONLY = {
    'show_ox_mark': False, 'show_score': False,
    'show_aspect': False, 'show_correct_answer': False,
}


class TestStarRendering:
    """解答用紙への★マーク描画テスト"""

    def test_star_at_correct_answer_position(self):
        """正答登録済みの特例設問: 正答位置(選択肢4=index3)に★"""
        img = _render_scoring(
            {1: _result(key='4', special=SPECIAL_ALL_CORRECT)},
            rendering_settings=_STAR_ONLY)
        assert _box_has_ink(img, 1, 3)
        assert not _box_has_ink(img, 1, 0)

    def test_star_at_leftmost_when_no_answer(self):
        """正答未登録の特例設問: 左端(index0)に★"""
        img = _render_scoring(
            {1: _result(key='', special=SPECIAL_ALL_CORRECT)},
            rendering_settings=_STAR_ONLY)
        assert _box_has_ink(img, 1, 0)

    def test_no_star_when_toggle_off(self):
        """トグルOFFなら★は描画されない"""
        settings = dict(_STAR_ONLY, show_all_correct_star=False)
        img = _render_scoring(
            {1: _result(key='4', special=SPECIAL_ALL_CORRECT)},
            rendering_settings=settings)
        assert not _box_has_ink(img, 1, 3)
        assert not _box_has_ink(img, 1, 0)

    def test_no_star_for_normal_question(self):
        """通常設問には★は描画されない"""
        img = _render_scoring(
            {1: _result(key='4', special='')},
            rendering_settings=_STAR_ONLY)
        assert not _box_has_ink(img, 1, 3)

    def test_legacy_results_without_special_key(self):
        """special キーのない旧形式 results でもクラッシュしない"""
        r = _result(key='4')
        del r['special']
        img = _render_scoring({1: r}, rendering_settings=_STAR_ONLY)
        assert not _box_has_ink(img, 1, 3)


# ── テンプレート生成 ────────────────────────────────────


class TestTemplateGeneration:
    """generate_template が特例列を含むテスト"""

    def test_template_has_special_column(self, tmp_path):
        from omr_engine import generate_template
        # 座標定義ファイル(3行ヘッダー+問題行)を模倣
        import openpyxl
        coord_path = tmp_path / "coord.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(3):
            ws.append([None])
        for q in range(1, 4):
            ws.append([q, 10, 10, 30, 30])
        wb.save(coord_path)
        wb.close()

        template_path = generate_template(str(coord_path), str(tmp_path))
        df = pd.read_excel(template_path)
        assert '特例' in df.columns

        # ドロップダウン(データ入力規則)が設定されている
        wb2 = openpyxl.load_workbook(template_path)
        ws2 = wb2.active
        dvs = list(ws2.data_validations.dataValidation)
        assert len(dvs) == 1
        assert '全員正解' in str(dvs[0].formula1)
        wb2.close()
