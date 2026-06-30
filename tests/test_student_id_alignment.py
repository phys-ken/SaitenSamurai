"""
test_student_id_alignment.py — 学生別サマリーの学籍番号セルの整合性テスト

回帰テスト対象のバグ:
    OMR結果xlsxの行順はマルチスレッド読込の完了順（as_completed）で、ファイル名の
    自然順とは一致しない。generate_student_summary は mark2_results をファイル名で
    再ソートする一方、学籍番号だけを original_df の「行の位置インデックス」で読んで
    いたため、学籍番号セルが別の生徒の行とズレていた（得点は answers 由来なので正常）。

このテストは、入力xlsxの行順をわざとシャッフルしても、出力サマリーの各行で
学籍番号がそのファイルの値と一致することを検証する。学籍番号の桁数（skip列数）が
可変であることも parametrize でカバーする。
"""

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from summary_generator import generate_student_summary


# 自然順: page1 < page2 < page3 < page10
FILES_NATURAL = ["page1.jpg", "page2.jpg", "page3.jpg", "page10.jpg"]
# わざと自然順と異なる（完了順を模した）入力行順
FILES_SHUFFLED = ["page10.jpg", "page1.jpg", "page3.jpg", "page2.jpg"]

# 各ファイルの設問解答（Q1正答=1, Q2正答=2, Q3正答=3, 各10点）→ 合計が一意になるよう設定
ANSWERS = {
    "page1.jpg": ["1", "2", "3"],   # 全問正解 → 30
    "page2.jpg": ["1", "2", "9"],   # 2問正解 → 20
    "page3.jpg": ["1", "9", "9"],   # 1問正解 → 10
    "page10.jpg": ["9", "9", "9"],  # 全問不正解 → 0
}
EXPECTED_TOTAL = {"page1.jpg": 30, "page2.jpg": 20, "page3.jpg": 10, "page10.jpg": 0}


def _student_ids(fname, skip):
    """ファイル固有の学籍番号桁を skip 個生成（一意・文字列）。"""
    stem = fname.split(".")[0]
    return [f"{stem}-id{k + 1}" for k in range(skip)]


def _build_template(path):
    """採点テンプレート（問題番号,正答,配点,観点）を作成。"""
    df = pd.DataFrame({
        "問題番号": [1, 2, 3],
        "正答": [1, 2, 3],
        "配点": [10, 10, 10],
        "観点": [1, 1, 1],
    })
    df.to_excel(path, index=False)


def _build_mark2_result(path, file_order, skip):
    """Mark2結果xlsxを save_recognition_results 互換の構造で作成。

    列構成: No, File, [学籍番号 x skip], [設問1..3]
    Row0: ヘッダー（No, File, 元インデックス番号...）
    Row1: 設問名行（学籍番号列は空、設問列に Scored Index 1,2,3）
    Row2+: データ（file_order の順に並べる＝シャッフル可能）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    n_q = 3
    # Row0 ヘッダー: No, File, 1..(skip+n_q) を元インデックスとして付与
    header = ["No", "File"] + [str(i) for i in range(1, skip + n_q + 1)]
    ws.append(header)

    # Row1 設問名行: 学籍番号列は空、設問列に Scored Index(1,2,3)
    name_row = ["", ""] + [""] * skip + [str(q) for q in range(1, n_q + 1)]
    ws.append(name_row)

    # データ行
    for i, fname in enumerate(file_order, 1):
        ids = _student_ids(fname, skip)
        row = [i, fname] + ids + ANSWERS[fname]
        ws.append(row)

    wb.save(path)


@pytest.mark.parametrize("skip", [1, 2, 3])
def test_student_id_follows_filename_when_rows_shuffled(tmp_path, skip):
    """入力行順をシャッフルしても、各行の学籍番号がファイル名と一致すること。"""
    template_path = tmp_path / "template.xlsx"
    mark2_path = tmp_path / "Mark2-Result.xlsx"
    out_path = tmp_path / "student_summary.xlsx"

    _build_template(template_path)
    _build_mark2_result(mark2_path, FILES_SHUFFLED, skip)

    generate_student_summary(
        str(template_path), str(mark2_path), str(out_path), skip_questions=skip
    )

    wb = load_workbook(out_path)
    ws = wb.active

    # ヘッダー行2（列名）から各列のインデックスを引く
    header2 = [c.value for c in ws[2]]
    file_col = header2.index("File")
    id_cols = [header2.index(f"学籍番号{k + 1}") for k in range(skip)]

    # データ行（3行目以降）を走査
    seen = set()
    for r in range(3, ws.max_row + 1):
        fname = ws.cell(row=r, column=file_col + 1).value
        if not fname:
            continue
        seen.add(fname)
        expected_ids = _student_ids(fname, skip)
        actual_ids = [ws.cell(row=r, column=ci + 1).value for ci in id_cols]
        assert actual_ids == expected_ids, (
            f"{fname}: 学籍番号が不一致 expected={expected_ids} actual={actual_ids}"
        )

    # 全ファイルが出力されていること
    assert seen == set(FILES_NATURAL)


def test_scores_correct_and_independent_of_row_order(tmp_path):
    """得点は元々ファイル名に紐付くので、行順に関わらず正しいこと（回帰の確認）。"""
    skip = 2
    template_path = tmp_path / "template.xlsx"
    mark2_path = tmp_path / "Mark2-Result.xlsx"
    out_path = tmp_path / "student_summary.xlsx"

    _build_template(template_path)
    _build_mark2_result(mark2_path, FILES_SHUFFLED, skip)

    generate_student_summary(
        str(template_path), str(mark2_path), str(out_path), skip_questions=skip
    )

    wb = load_workbook(out_path)
    ws = wb.active
    header2 = [c.value for c in ws[2]]
    file_col = header2.index("File")
    total_col = header2.index("合計得点")

    for r in range(3, ws.max_row + 1):
        fname = ws.cell(row=r, column=file_col + 1).value
        if not fname:
            continue
        total = ws.cell(row=r, column=total_col + 1).value
        assert total == EXPECTED_TOTAL[fname], (
            f"{fname}: 合計得点が不一致 expected={EXPECTED_TOTAL[fname]} actual={total}"
        )
