#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
v3機能のユニットテスト
"""

import sys
import os
import unittest
import tempfile
import shutil
from pathlib import Path

from saitensamurai import (
    generate_template,
    extract_pdf_to_images,
    combine_images_to_pdf,
    HAS_PYMUPDF,
    ANSWER_KEY_FILE,
    RESULTS_FOLDER,
)


class TestTemplateOverwriteProtection(unittest.TestCase):
    """v3-1: テンプレート上書き防止のテスト"""
    
    def setUp(self):
        """テスト用の一時ディレクトリとダミー座標Excelを作成"""
        self.test_dir = tempfile.mkdtemp()
        self.output_dir = Path(self.test_dir) / "output"
        self.output_dir.mkdir()
        
        # ダミー座標Excelを作成
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # ヘッダー行（3行目以降が設問データ）
        ws.append(['', '', '', ''])  # row 1
        ws.append(['', '', '', ''])  # row 2
        ws.append(['', '', '', ''])  # row 3
        # 設問データ: question_no, name, skip, skip, x, y, w, h
        ws.append([1, 'Q1', '', '', 100, 100, 20, 20])
        ws.append([2, 'Q2', '', '', 100, 200, 20, 20])
        ws.append([3, 'Q3', '', '', 100, 300, 20, 20])
        self.coord_path = Path(self.test_dir) / "coord.xlsx"
        wb.save(str(self.coord_path))
    
    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)
    
    def test_01_creates_new_template(self):
        """新規テンプレートが作成されること"""
        result = generate_template(self.coord_path, self.output_dir, skip_questions=0)
        self.assertTrue(result.exists())
        self.assertEqual(result.name, ANSWER_KEY_FILE)
    
    def test_02_does_not_overwrite_existing(self):
        """既存テンプレートが上書きされないこと"""
        template_path = self.output_dir / ANSWER_KEY_FILE
        
        # 最初に作成
        generate_template(self.coord_path, self.output_dir, skip_questions=0)
        self.assertTrue(template_path.exists())
        
        # ユーザーがデータを入力した状態をシミュレート
        import openpyxl
        wb = openpyxl.load_workbook(str(template_path))
        ws = wb.active
        ws.cell(row=2, column=2, value="3")  # 正答を入力
        ws.cell(row=2, column=3, value="5")  # 配点を入力
        wb.save(str(template_path))
        
        # ファイルの更新時刻を記録
        mtime_before = template_path.stat().st_mtime
        
        # 再度generate_template を呼び出し
        import time
        time.sleep(0.1)  # mtimeが変わることを保証するための待機
        generate_template(self.coord_path, self.output_dir, skip_questions=0)
        
        # ファイルが上書きされていないことを確認
        mtime_after = template_path.stat().st_mtime
        self.assertEqual(mtime_before, mtime_after, "既存テンプレートが上書きされました")
        
        # ユーザーが入力したデータが保持されていることを確認
        wb2 = openpyxl.load_workbook(str(template_path))
        ws2 = wb2.active
        self.assertEqual(ws2.cell(row=2, column=2).value, "3")
        self.assertEqual(ws2.cell(row=2, column=3).value, "5")


@unittest.skipUnless(HAS_PYMUPDF, "PyMuPDFがインストールされていません")
class TestPdfExtraction(unittest.TestCase):
    """v3-4: PDF入力サポートのテスト"""
    
    def setUp(self):
        self.test_dir = tempfile.mkdtemp()
    
    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)
    
    def _create_dummy_pdf(self, page_count=3):
        """テスト用のダミーPDFを作成"""
        import fitz
        pdf_path = Path(self.test_dir) / "test_input.pdf"
        doc = fitz.open()
        for i in range(page_count):
            page = doc.new_page(width=595, height=842)  # A4
            page.insert_text((100, 100), f"Page {i+1}", fontsize=20)
        doc.save(str(pdf_path))
        doc.close()
        return pdf_path
    
    def test_01_extract_creates_images(self):
        """PDFから正しい枚数の画像が生成されること"""
        pdf_path = self._create_dummy_pdf(page_count=3)
        output_folder = extract_pdf_to_images(pdf_path)
        
        self.assertTrue(output_folder.exists())
        png_files = sorted(output_folder.glob("*.png"))
        self.assertEqual(len(png_files), 3)
        # ファイル名が {PDF名}_pNNN 形式でソート順を保つこと
        self.assertEqual(png_files[0].name, "test_input_p001.png")
        self.assertEqual(png_files[2].name, "test_input_p003.png")
    
    def test_02_extract_to_custom_folder(self):
        """カスタム出力先フォルダにも展開できること"""
        pdf_path = self._create_dummy_pdf(page_count=2)
        custom_output = Path(self.test_dir) / "custom_output"
        output_folder = extract_pdf_to_images(pdf_path, output_folder=custom_output)
        
        self.assertEqual(output_folder, custom_output)
        self.assertEqual(len(list(output_folder.glob("*.png"))), 2)
    
    def test_03_default_output_folder_name(self):
        """デフォルト出力先が {PDF名}_images であること"""
        pdf_path = self._create_dummy_pdf()
        output_folder = extract_pdf_to_images(pdf_path)
        
        self.assertEqual(output_folder.name, "test_input_images")
    
    def test_04_file_not_found(self):
        """存在しないPDFでFileNotFoundErrorが発生すること"""
        fake_path = Path(self.test_dir) / "nonexistent.pdf"
        with self.assertRaises(FileNotFoundError):
            extract_pdf_to_images(fake_path)
    
    def test_05_reextract_overwrites(self):
        """既にある展開画像を再展開で上書きできること"""
        pdf_path = self._create_dummy_pdf(page_count=2)
        output_folder = extract_pdf_to_images(pdf_path)
        first_mtime = (output_folder / "test_input_p001.png").stat().st_mtime

        import time
        time.sleep(0.1)

        extract_pdf_to_images(pdf_path, output_folder=output_folder)
        second_mtime = (output_folder / "test_input_p001.png").stat().st_mtime
        self.assertGreater(second_mtime, first_mtime)

    def test_06_multiple_pdfs_no_collision(self):
        """複数PDFを同一フォルダへ展開してもファイル名が衝突しないこと"""
        import fitz
        paths = []
        for name in ("classA", "classB"):
            pdf_path = Path(self.test_dir) / f"{name}.pdf"
            doc = fitz.open()
            for i in range(2):
                page = doc.new_page(width=595, height=842)
                page.insert_text((100, 100), f"{name} Page {i+1}", fontsize=20)
            doc.save(str(pdf_path))
            doc.close()
            paths.append(pdf_path)

        common = Path(self.test_dir) / "pdf_import_images"
        for p in paths:
            extract_pdf_to_images(p, output_folder=common)

        png_names = sorted(f.name for f in common.glob("*.png"))
        self.assertEqual(png_names, [
            "classA_p001.png", "classA_p002.png",
            "classB_p001.png", "classB_p002.png",
        ])


class TestCombineImagesToPdf(unittest.TestCase):
    """v3-5: 統合PDF出力のテスト"""
    
    def setUp(self):
        self.test_dir = tempfile.mkdtemp()
        self.image_folder = Path(self.test_dir) / "scored"
        self.image_folder.mkdir()
    
    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)
    
    def _create_dummy_images(self, count=3):
        """テスト用のダミー画像を作成"""
        from PIL import Image, ImageDraw
        for i in range(count):
            img = Image.new('RGB', (200, 300), color=(255, 255, 255))
            draw = ImageDraw.Draw(img)
            draw.text((50, 100), f"Page {i+1}", fill=(0, 0, 0))
            img.save(str(self.image_folder / f"student_{i+1:03d}.jpg"))
    
    def test_01_creates_pdf(self):
        """画像からPDFが生成されること"""
        self._create_dummy_images(3)
        output_pdf = Path(self.test_dir) / "output" / "combined.pdf"
        result = combine_images_to_pdf(self.image_folder, output_pdf)
        
        self.assertIsNotNone(result)
        self.assertTrue(output_pdf.exists())
        self.assertGreater(output_pdf.stat().st_size, 0)
    
    def test_02_empty_folder_returns_none(self):
        """画像がない場合はNoneを返すこと"""
        result = combine_images_to_pdf(self.image_folder, Path(self.test_dir) / "empty.pdf")
        self.assertIsNone(result)
    
    @unittest.skipUnless(HAS_PYMUPDF, "PyMuPDFがインストールされていません")
    def test_03_pdf_page_count(self):
        """生成PDFのページ数が正しいこと"""
        import fitz
        self._create_dummy_images(5)
        output_pdf = Path(self.test_dir) / "combined.pdf"
        combine_images_to_pdf(self.image_folder, output_pdf)
        
        doc = fitz.open(str(output_pdf))
        self.assertEqual(len(doc), 5)
        doc.close()


if __name__ == '__main__':
    unittest.main()
