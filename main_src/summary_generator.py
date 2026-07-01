"""
summary_generator.py - サマリー生成モジュール

学生別得点サマリーと試験統計サマリーをExcelファイルとして生成する。
マークシート採点結果＋記述問題得点を統合し、観点別・設問別の統計を出力。
"""

from pathlib import Path
import logging
import re
import tempfile
import pandas as pd
import numpy as np
from PIL import Image

logger = logging.getLogger(__name__)
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib import rcParams
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

from constants import (
    combine_images_to_pdf,
    get_app_temp_dir,
    RESULTS_FOLDER,
    SCORED_FOLDER,
    FINAL_REPORT_FOLDER,
    STUDENT_SUMMARY_FILE,
    EXAM_SUMMARY_FILE,
    CTT_ANALYSIS_EXCEL_FILE,
    CTT_ANALYSIS_PDF_FILE,
    SCORED_PDF_FILE,
    R_EXPORT_FOLDER,
    escape_excel_formula,
)

from scoring_engine import (
    number_to_circled,
    load_template,
    load_mark2_results,
    score_answers,
)


def _natural_sort_key(text: str):
    """Windows Explorer 互換の自然順ソートキー

    文字列内の数値部分を int に変換し、非数値部分は小文字化して比較する。
    例: 'page2.jpg' < 'page10.jpg' (辞書順だと逆になる)
    """
    return [int(c) if c.isdigit() else c.lower()
            for c in re.split(r'(\d+)', text)]


def generate_student_summary(template_path, mark2_result_path, output_path, skip_questions=0, name_images=None,
                             descriptive_config=None, descriptive_scores=None,
                             template_dict=None, mark2_results=None):
    """学生別サマリーを生成
    
    Args:
        template_path: 正答データExcelファイルのパス
        mark2_result_path: OMR読取結果Excelファイルのパス
        output_path: 出力先パス
        skip_questions: スキップする問題数（学籍番号用）
        name_images: {ファイル名: トリミング画像パス} の辞書。Noneの場合は氏名欄列なし。
        descriptive_config: 記述問題設定dict（オプション）
        descriptive_scores: {ファイル名: {問題ID: 得点}} の辞書（オプション）
        template_dict: 事前読込済みのテンプレートdict（Noneなら内部で読込）
        mark2_results: 事前読込済みのMark2結果list（Noneなら内部で読込）
    """
    logger.info("=" * 60)
    logger.info("学生別サマリー生成")
    logger.info("=" * 60)
    
    if template_dict is None:
        template_dict = load_template(template_path)
    logger.info("✓ テンプレート読込: %d問", len(template_dict))
    
    if mark2_results is None:
        mark2_results = load_mark2_results(mark2_result_path, skip_questions)
    logger.info("✓ Mark2結果読込: %d件", len(mark2_results))
    
    original_df = pd.read_excel(mark2_result_path, header=None)

    # 学籍番号を「ファイル名キー」で引けるdictを構築する。
    # OMR結果xlsxの行順はマルチスレッド読込の完了順（as_completed）で、ファイル名の
    # 自然順とは一致しない。一方この後 mark2_results はファイル名で再ソートするため、
    # 行の位置インデックスで original_df を参照すると学籍番号だけが別の生徒の行とズレる
    # （得点は result_data['answers'] 由来なので正しい）。File列を動的検出して
    # {File値: [学籍番号セル...]} を作り、image_name で引くことで順序非依存にする。
    # 学籍番号の桁数（列数）は skip_questions で可変。
    student_id_by_file = {}
    if skip_questions > 0 and len(original_df) > 2:
        header_row = original_df.iloc[0]
        file_col_idx = next(
            (i for i, v in enumerate(header_row) if str(v).strip().lower() == 'file'),
            1,
        )
        id_cols = [file_col_idx + 1 + k for k in range(skip_questions)]
        n_cols = original_df.shape[1]
        for r in range(2, len(original_df)):
            file_val = original_df.iloc[r, file_col_idx]
            if pd.isna(file_val):
                continue
            key = str(file_val).strip()
            if not key:
                continue
            student_id_by_file[key] = [
                ('' if c >= n_cols or pd.isna(original_df.iloc[r, c])
                 else original_df.iloc[r, c])
                for c in id_cols
            ]

    aspects = sorted(set(data['観点'] for data in template_dict.values()))
    
    # 記述問題の観点をマージ
    has_descriptive = descriptive_config and descriptive_config.get('questions') and descriptive_scores
    if has_descriptive:
        desc_aspects = set(q['aspect'] for q in descriptive_config['questions'])
        all_aspects = sorted(set(aspects) | desc_aspects)
        logger.info("✓ 観点(マーク+記述): %s", all_aspects)
        logger.info("✓ 記述問題: %d問", len(descriptive_config['questions']))
    else:
        all_aspects = list(aspects)
        logger.info("✓ 観点: %s", aspects)
    
    # 氏名欄画像の有無を判定
    has_name_images = name_images is not None and len(name_images) > 0
    if has_name_images:
        logger.info("✓ 氏名欄画像: %d枚", len(name_images))
    
    question_numbers = sorted(template_dict.keys())
    rows = []

    # 自然順ソート: page1, page2, ..., page10 の順になる (Windows Explorer互換)
    mark2_results = sorted(mark2_results, key=lambda r: _natural_sort_key(r['image']))
    
    for idx, result_data in enumerate(mark2_results, 1):
        image_name = result_data['image']
        student_answers = result_data['answers']
        scoring_result = score_answers(student_answers, template_dict)
        
        row = {'No': idx, 'File': image_name}

        # 学籍番号はファイル名キーで引く（行順に依存しない）
        student_ids = student_id_by_file.get(image_name, [])
        for skip_idx in range(skip_questions):
            col_name = f'学籍番号{skip_idx + 1}'
            row[col_name] = student_ids[skip_idx] if skip_idx < len(student_ids) else ''
        
        row['合計得点'] = scoring_result['total_score']
        
        for aspect in all_aspects:
            aspect_name = f'観点{number_to_circled(aspect)}'
            row[aspect_name] = scoring_result['aspect_scores'].get(aspect, 0)
        
        for q_no in question_numbers:
            result = scoring_result['results'].get(q_no, {})
            row[f'問{q_no}'] = result.get('points', 0)
        
        # 記述問題の得点を追加
        if has_descriptive:
            desc_scores_for_student = descriptive_scores.get(image_name, {})
            desc_total = 0
            for q in descriptive_config['questions']:
                q_score = desc_scores_for_student.get(q['id'], 0)
                desc_total += q_score
                row[q['id']] = q_score
                # 対応する観点に加算
                asp_name = f'観点{number_to_circled(q["aspect"])}'
                if asp_name in row:
                    row[asp_name] += q_score
            row['合計得点'] += desc_total
        
        rows.append(row)
        total_display = row['合計得点']
        logger.info("  [%d/%d] %s: %s点", idx, len(mark2_results), image_name, total_display)
    
    df = pd.DataFrame(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "学生別得点"
    
    # --- ヘッダー行1: 問題番号 ---
    # 氏名欄がある場合は1列追加
    name_col_extra = [''] if has_name_images else []
    header_row1 = ['', ''] + name_col_extra + [''] * skip_questions + [''] + [''] * len(all_aspects) + list(question_numbers)
    if has_descriptive:
        header_row1 += [q['name'] for q in descriptive_config['questions']]
    ws.append(header_row1)
    
    # --- ヘッダー行2: 列名 ---
    header_row2 = ['No', 'File'] + (['氏名欄'] if has_name_images else []) + [f'学籍番号{i+1}' for i in range(skip_questions)] + ['合計得点']
    for aspect in all_aspects:
        header_row2.append(f'観点{number_to_circled(aspect)}')
    for q_no in question_numbers:
        header_row2.append(f'問{q_no}')
    if has_descriptive:
        for q in descriptive_config['questions']:
            header_row2.append(q['id'])
    ws.append(header_row2)
    
    # --- データ行 ---
    for row in rows:
        data_row = [row['No'], escape_excel_formula(row['File'])]
        if has_name_images:
            data_row.append('')  # 氏名欄画像用のプレースホルダ
        for skip_idx in range(skip_questions):
            data_row.append(escape_excel_formula(row.get(f'学籍番号{skip_idx + 1}', '')))
        data_row.append(row['合計得点'])
        for aspect in all_aspects:
            aspect_name = f'観点{number_to_circled(aspect)}'
            data_row.append(row[aspect_name])
        for q_no in question_numbers:
            data_row.append(row[f'問{q_no}'])
        if has_descriptive:
            for q in descriptive_config['questions']:
                data_row.append(row.get(q['id'], 0))
        ws.append(data_row)
    
    # --- 氏名欄画像の埋め込み ---
    if has_name_images:
        name_col_letter = 'C'  # 氏名欄は常にC列
        max_img_width = 0
        embedded_count = 0
        
        for idx, result_data in enumerate(mark2_results):
            image_name = result_data['image']
            row_num = idx + 3  # ヘッダー2行 + 1始まり
            
            if image_name in name_images:
                img_path = name_images[image_name]
                if Path(img_path).exists():
                    try:
                        xl_img = XlImage(img_path)
                        pil_img = Image.open(img_path)
                        img_w, img_h = pil_img.size
                        pil_img.close()
                        
                        if img_w > max_img_width:
                            max_img_width = img_w
                        
                        # 行の高さを画像に合わせる（ピクセル → ポイント変換: 1pt ≈ 0.75px）
                        ws.row_dimensions[row_num].height = img_h * 0.75
                        
                        # セルにアンカー設定
                        cell_address = f'{name_col_letter}{row_num}'
                        xl_img.anchor = cell_address
                        ws.add_image(xl_img)
                        embedded_count += 1
                    except Exception as e:
                        logger.warning("  氏名画像埋め込みエラー (%s): %s", image_name, e)
        
        # 氏名欄の列幅を画像に合わせる（ピクセル → 文字幅: 1文字幅 ≈ 7.5px）
        ws.column_dimensions[name_col_letter].width = max(max_img_width * 0.13, 10)
        logger.info("✓ 氏名欄画像埋め込み: %d枚", embedded_count)
    
    # --- スタイル設定 ---
    header_fill_dark = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=10)
    header_fill_light = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    header_font_dark = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # Row 1: 問題番号ヘッダー (ダークブルー + 白文字)
    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill_dark
        cell.alignment = center_align
        cell.border = thin_border

    # Row 2: 列名ヘッダー (ライトブルー)
    for cell in ws[2]:
        cell.font = header_font_dark
        cell.fill = header_fill_light
        cell.alignment = center_align
        cell.border = thin_border

    # データ行: 罫線 + 中央揃え
    # 氏名欄がある場合、数値列の開始は1列ずれる
    data_start_col = 4 if has_name_images else 3
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.column >= data_start_col:
                cell.alignment = center_align

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    # 氏名欄以外の数値列の幅設定
    width_start_col = 4 if has_name_images else 3
    for col_idx in range(width_start_col, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # ウィンドウ枠固定 (ヘッダー2行 + No/File列 [+ 氏名欄列])
    if has_name_images:
        ws.freeze_panes = 'D3'
    else:
        ws.freeze_panes = 'C3'

    wb.save(output_path)
    logger.info("✓ 保存: %s", output_path)
    logger.info("")
    return df


# ============================================================
# matplotlib グラフ生成ヘルパー
# ============================================================

# 日本語フォント設定
_FONT_CANDIDATES = ['Yu Gothic', 'Yu Gothic UI', 'Meiryo', 'MS Gothic',
                     'Hiragino Kaku Gothic Pro', 'IPAexGothic', 'Noto Sans CJK JP']

def _setup_japanese_font():
    """日本語表示可能なフォントを検出して設定する"""
    import matplotlib.font_manager as fm
    available = {f.name for f in fm.fontManager.ttflist}
    for candidate in _FONT_CANDIDATES:
        if candidate in available:
            rcParams['font.family'] = candidate
            return candidate
    # 見つからない場合は sans-serif にフォールバック
    rcParams['font.family'] = 'sans-serif'
    return 'sans-serif'

_setup_japanese_font()


def _apply_clean_style(ax):
    """R/matplotlib風のクリーンなスタイルをaxに適用"""
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_linewidth(0.8)
    ax.spines['bottom'].set_linewidth(0.8)
    ax.tick_params(direction='out', length=4, width=0.8)
    ax.grid(axis='y', linestyle='--', alpha=0.3, linewidth=0.6)


def _save_fig_to_temp(fig, dpi=150, temp_dir=None):
    """figをPNGに保存して一時ファイルパスを返す"""
    tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False, dir=temp_dir)
    fig.savefig(tmp.name, dpi=dpi, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    return tmp.name


def _create_bar_chart(labels, values, title, xlabel, ylabel,
                      ylim=None, color='#5B9BD5', edgecolor='#2E75B6',
                      figsize=(10, 4.5)):
    """シンプルな棒グラフを生成"""
    fig, ax = plt.subplots(figsize=figsize)
    _apply_clean_style(ax)
    x = np.arange(len(labels))
    ax.bar(x, values, color=color, edgecolor=edgecolor, linewidth=0.6, width=0.7)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
    ax.set_title(title, fontsize=13, fontweight='bold', pad=10)
    ax.set_xlabel(xlabel, fontsize=11)
    ax.set_ylabel(ylabel, fontsize=11)
    if ylim:
        ax.set_ylim(ylim)
    ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    fig.tight_layout()
    return fig, ax


def _create_scatter_matrix(all_aspects, aspect_scores_list, corr_matrix, aspect_labels):
    """観点間の散布図行列を生成する

    対角: ヒストグラム
    上三角: 相関係数テキスト
    下三角: 散布図
    """
    n = len(all_aspects)
    fig_size = max(4, n * 2.2)
    fig, axes = plt.subplots(n, n, figsize=(fig_size, fig_size))
    if n == 1:
        axes = np.array([[axes]])
    elif n > 1 and axes.ndim == 1:
        axes = axes.reshape(n, n)

    for i in range(n):
        for j in range(n):
            ax = axes[i, j]
            si = aspect_scores_list[all_aspects[i]]
            sj = aspect_scores_list[all_aspects[j]]

            if i == j:
                # 対角: ヒストグラム
                ax.hist(si, bins='auto', color='#AAAAAA', edgecolor='#333333', linewidth=0.6)
                ax.set_title(aspect_labels[i], fontsize=9, fontweight='bold')
            elif i > j:
                # 下三角: 散布図
                ax.scatter(sj, si, s=12, alpha=0.6, color='#5B9BD5', edgecolor='#2E75B6', linewidth=0.3)
            else:
                # 上三角: 相関係数
                ax.set_xlim(0, 1)
                ax.set_ylim(0, 1)
                r = corr_matrix[i][j]
                fontsize = max(12, 20 - n * 2)
                ax.text(0.5, 0.5, f'r = {r:.3f}', ha='center', va='center',
                        fontsize=fontsize, fontweight='bold',
                        color='#1F4E79' if abs(r) >= 0.5 else '#666666')
                ax.set_xticks([])
                ax.set_yticks([])

            # 軸ラベル
            if i < n - 1:
                ax.set_xticklabels([])
            else:
                ax.tick_params(axis='x', labelsize=7)
                if j == n // 2:
                    ax.set_xlabel('得点', fontsize=9)
            if j > 0:
                ax.set_yticklabels([])
            else:
                ax.tick_params(axis='y', labelsize=7)
                if i == n // 2:
                    ax.set_ylabel('得点', fontsize=9)

            # 枠線
            for spine in ax.spines.values():
                spine.set_linewidth(0.5)
                spine.set_color('#CCCCCC')

    fig.suptitle('観点間の散布図行列', fontsize=13, fontweight='bold', y=1.02)
    fig.tight_layout()
    return fig


def generate_exam_summary(template_path, mark2_result_path, output_path, skip_questions=0,
                          descriptive_config=None, descriptive_scores=None,
                          template_dict=None, mark2_results=None):
    """試験サマリーを生成

    4シート構成の見やすい Excel レポートを出力する。
    シグネチャ・戻り値は従来互換（stats dict）。

    シート構成:
        1. 試験概要 — 基本情報 + 要約統計量
        2. 設問分析 — 設問ごとの正答率（記述含む）+ 棒グラフ
        3. 得点分布 — 10点刻み度数分布 + ヒストグラム
        4. 観点別統計 — 観点ごとの集計

    Args:
        template_dict: 事前読込済みのテンプレートdict（Noneなら内部で読込）
        mark2_results: 事前読込済みのMark2結果list（Noneなら内部で読込）
    """
    logger.info("=" * 60)
    logger.info("試験サマリー生成")
    logger.info("=" * 60)

    # 一時ファイル用ディレクトリ（出力パスから画像フォルダを逆算）
    _chart_temp_dir = get_app_temp_dir(str(Path(output_path).parent.parent.parent))

    if template_dict is None:
        template_dict = load_template(template_path)
    logger.info("✓ テンプレート読込: %d問", len(template_dict))

    if mark2_results is None:
        mark2_results = load_mark2_results(mark2_result_path, skip_questions)
    logger.info("✓ Mark2結果読込: %d件", len(mark2_results))

    aspects = sorted(set(data['観点'] for data in template_dict.values()))
    question_numbers = sorted(template_dict.keys())

    # 記述問題判定
    has_descriptive = descriptive_config and descriptive_config.get('questions') and descriptive_scores
    desc_questions = descriptive_config['questions'] if has_descriptive else []
    if has_descriptive:
        desc_aspects = set(q['aspect'] for q in desc_questions)
        all_aspects = sorted(set(aspects) | desc_aspects)
    else:
        all_aspects = list(aspects)

    # ── 全受験者の得点計算 ──
    _tmp_chart_files = []  # matplotlib一時ファイル管理
    all_scores = []
    aspect_scores_list = {aspect: [] for aspect in all_aspects}
    question_correct_count = {q_no: 0 for q_no in question_numbers}
    # 記述問題ごとの得点合計（正答率用）
    desc_score_totals = {q['id']: 0 for q in desc_questions}
    desc_full_mark_count = {q['id']: 0 for q in desc_questions}

    for result_data in mark2_results:
        student_answers = result_data['answers']
        scoring_result = score_answers(student_answers, template_dict)
        student_total = scoring_result['total_score']

        # 観点別得点（マーク）
        student_aspect = {aspect: scoring_result['aspect_scores'].get(aspect, 0) for aspect in all_aspects}

        # 記述問題の得点を加算
        if has_descriptive:
            image_name = result_data['image']
            desc_scores_for_student = descriptive_scores.get(image_name, {})
            for q in desc_questions:
                q_score = desc_scores_for_student.get(q['id'], 0)
                student_total += q_score
                student_aspect[q['aspect']] = student_aspect.get(q['aspect'], 0) + q_score
                desc_score_totals[q['id']] += q_score
                if q_score >= q['max_score']:
                    desc_full_mark_count[q['id']] += 1

        all_scores.append(student_total)
        for aspect in all_aspects:
            aspect_scores_list[aspect].append(student_aspect.get(aspect, 0))

        for q_no, result in scoring_result['results'].items():
            if result['correct']:
                question_correct_count[q_no] += 1

    total_students = len(mark2_results)
    mark_max_score = sum(data['配点'] for data in template_dict.values())
    desc_max_score = sum(q['max_score'] for q in desc_questions) if has_descriptive else 0
    total_max_score = mark_max_score + desc_max_score
    total_questions_mark = len(question_numbers)
    total_questions_desc = len(desc_questions)
    total_questions = total_questions_mark + total_questions_desc

    # 要約統計量
    stats = {
        '受験者数': total_students,
        '満点': total_max_score,
        '平均点': round(float(np.mean(all_scores)), 2) if all_scores else 0,
        '中央値': round(float(np.median(all_scores)), 2) if all_scores else 0,
        '最高点': int(np.max(all_scores)) if all_scores else 0,
        '最低点': int(np.min(all_scores)) if all_scores else 0,
        '標準偏差': round(float(np.std(all_scores, ddof=1)), 2) if len(all_scores) > 1 else 0,
        '分散': round(float(np.var(all_scores, ddof=1)), 2) if len(all_scores) > 1 else 0,
    }

    # 得点率
    score_rate = round(stats['平均点'] / total_max_score * 100, 1) if total_max_score > 0 else 0

    logger.info("✓ 統計計算完了")

    # ══════ 共通スタイル定義 ══════
    HEADER_FONT = Font(name='Yu Gothic UI', bold=True, size=10, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    TITLE_FONT = Font(name='Yu Gothic UI', bold=True, size=14, color='1F4E79')
    SUBTITLE_FONT = Font(name='Yu Gothic UI', bold=True, size=11, color='2E75B6')
    DATA_FONT = Font(name='Yu Gothic UI', size=10)
    NUM_FONT = Font(name='Yu Gothic UI', size=10)
    LABEL_FONT = Font(name='Yu Gothic UI', bold=True, size=10, color='333333')
    THIN_BORDER = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF'),
    )
    LIGHT_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    ALT_FILL = PatternFill(start_color='E9EFF5', end_color='E9EFF5', fill_type='solid')

    def _style_header_row(ws, row, max_col):
        """ヘッダー行にスタイルを適用"""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

    def _style_data_rows(ws, start_row, end_row, max_col, alternate=True):
        """データ行にスタイルと罫線を適用"""
        for r in range(start_row, end_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical='center')
                if alternate and (r - start_row) % 2 == 1:
                    cell.fill = LIGHT_FILL

    # ══════ シート1: 試験概要 ══════
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "試験概要"
    ws1.sheet_properties.tabColor = '1F4E79'

    # タイトル
    ws1.merge_cells('A1:D1')
    ws1['A1'] = '試験サマリーレポート'
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = Alignment(vertical='center')
    ws1.row_dimensions[1].height = 30

    # 基本情報セクション
    ws1.merge_cells('A3:D3')
    ws1['A3'] = '■ 基本情報'
    ws1['A3'].font = SUBTITLE_FONT

    info_data = [
        ('受験者数', f'{total_students} 名'),
        ('設問数（マーク式）', f'{total_questions_mark} 問'),
    ]
    if has_descriptive:
        info_data.append(('設問数（記述式）', f'{total_questions_desc} 問'))
    info_data.extend([
        ('設問数（合計）', f'{total_questions} 問'),
        ('満点', f'{total_max_score} 点'),
    ])
    if has_descriptive:
        info_data.append(('　（マーク式）', f'{mark_max_score} 点'))
        info_data.append(('　（記述式）', f'{desc_max_score} 点'))

    row = 4
    for label, value in info_data:
        ws1.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws1.cell(row=row, column=2, value=value).font = DATA_FONT
        ws1.cell(row=row, column=1).border = THIN_BORDER
        ws1.cell(row=row, column=2).border = THIN_BORDER
        ws1.cell(row=row, column=1).fill = LIGHT_FILL
        row += 1

    row += 1  # 空行

    # 要約統計量セクション
    ws1.merge_cells(f'A{row}:D{row}')
    ws1.cell(row=row, column=1, value='■ 要約統計量').font = SUBTITLE_FONT
    row += 1

    stat_header_row = row
    ws1.cell(row=row, column=1, value='統計量')
    ws1.cell(row=row, column=2, value='値')
    _style_header_row(ws1, row, 2)
    row += 1

    stat_items = [
        ('平均点', f"{stats['平均点']:.2f} 点"),
        ('得点率', f'{score_rate}%'),
        ('中央値', f"{stats['中央値']:.2f} 点"),
        ('最高点', f"{stats['最高点']} 点"),
        ('最低点', f"{stats['最低点']} 点"),
        ('標準偏差', f"{stats['標準偏差']:.2f}"),
        ('分散', f"{stats['分散']:.2f}"),
    ]
    stat_start_row = row
    for label, value in stat_items:
        ws1.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws1.cell(row=row, column=2, value=value).font = DATA_FONT
        ws1.cell(row=row, column=1).fill = LIGHT_FILL
        ws1.cell(row=row, column=1).border = THIN_BORDER
        ws1.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # 列幅
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 12

    # ══════ シート2: 設問分析 ══════
    ws2 = wb.create_sheet("設問分析")
    ws2.sheet_properties.tabColor = '2E75B6'

    ws2.merge_cells('A1:G1')
    ws2['A1'] = '設問分析（正答率）'
    ws2['A1'].font = TITLE_FONT
    ws2['A1'].alignment = Alignment(vertical='center')
    ws2.row_dimensions[1].height = 30

    # ヘッダー
    q_headers = ['No.', '種別', '配点', '観点', '正答者数', '正答率(%)', '評価']
    for c, h in enumerate(q_headers, 1):
        ws2.cell(row=3, column=c, value=h)
    _style_header_row(ws2, 3, len(q_headers))

    # マーク式設問
    q_row = 4
    for q_no in question_numbers:
        correct_count = question_correct_count[q_no]
        correct_rate = round(correct_count / total_students * 100, 1) if total_students > 0 else 0
        tpl = template_dict[q_no]
        evaluation = _evaluate_correct_rate(correct_rate)

        ws2.cell(row=q_row, column=1, value=q_no)
        ws2.cell(row=q_row, column=2, value='マーク')
        ws2.cell(row=q_row, column=3, value=tpl['配点'])
        ws2.cell(row=q_row, column=4, value=f'観点{number_to_circled(tpl["観点"])}')
        ws2.cell(row=q_row, column=5, value=correct_count)
        ws2.cell(row=q_row, column=6, value=correct_rate)
        ws2.cell(row=q_row, column=7, value=evaluation)
        # 正答率セルの数値書式
        ws2.cell(row=q_row, column=6).number_format = '0.0'
        q_row += 1

    # 記述式設問
    if has_descriptive:
        for q in desc_questions:
            full_count = desc_full_mark_count[q['id']]
            full_rate = round(full_count / total_students * 100, 1) if total_students > 0 else 0
            avg_score = round(desc_score_totals[q['id']] / total_students, 2) if total_students > 0 else 0
            evaluation = _evaluate_correct_rate(full_rate)

            ws2.cell(row=q_row, column=1, value=q['id'])
            ws2.cell(row=q_row, column=2, value='記述')
            ws2.cell(row=q_row, column=3, value=q['max_score'])
            ws2.cell(row=q_row, column=4, value=f'観点{number_to_circled(q["aspect"])}')
            ws2.cell(row=q_row, column=5, value=full_count)
            ws2.cell(row=q_row, column=6, value=full_rate)
            ws2.cell(row=q_row, column=7, value=evaluation)
            ws2.cell(row=q_row, column=6).number_format = '0.0'
            q_row += 1

    q_data_end = q_row - 1
    _style_data_rows(ws2, 4, q_data_end, len(q_headers))

    # 正答率の数値セルを右寄せ
    for r in range(4, q_row):
        for c in [3, 5, 6]:
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal='right', vertical='center')

    # 列幅
    col_widths = [8, 8, 8, 10, 10, 12, 10]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # 正答率 水平棒グラフ（matplotlib 生成 → 画像をテーブル右側に配置）
    if q_data_end >= 4:
        q_labels = []
        q_rates = []
        for r in range(4, q_data_end + 1):
            q_labels.append(str(ws2.cell(row=r, column=1).value))
            q_rates.append(float(ws2.cell(row=r, column=6).value))

        n_q = len(q_labels)
        fig_h = max(4, n_q * 0.35 + 1.5)
        fig, ax = plt.subplots(figsize=(7, fig_h))
        _apply_clean_style(ax)
        ax.spines['left'].set_visible(False)
        ax.tick_params(axis='y', length=0)

        y = np.arange(n_q)
        bars = ax.barh(y, q_rates, color='#AAAAAA', edgecolor='#555555',
                       linewidth=0.6, height=0.65)
        ax.set_yticks(y)
        ax.set_yticklabels(q_labels, fontsize=8)
        ax.invert_yaxis()  # 上から設問1
        ax.set_xlim(0, 105)
        ax.set_title('設問別正答率', fontsize=13, fontweight='bold', pad=10)
        ax.set_xlabel('正答率 (%)', fontsize=11)
        # 目安ライン
        ax.axvline(x=80, color='#999999', linewidth=0.8, linestyle='--', alpha=0.5)
        ax.axvline(x=60, color='#999999', linewidth=0.8, linestyle='--', alpha=0.5)
        # 数値ラベル
        for bar_obj, rate in zip(bars, q_rates):
            ax.text(bar_obj.get_width() + 0.8, bar_obj.get_y() + bar_obj.get_height() / 2,
                    f'{rate:.1f}%', va='center', ha='left', fontsize=8, color='#333333')
        fig.tight_layout()

        chart_path = _save_fig_to_temp(fig, temp_dir=_chart_temp_dir)
        img = XlImage(chart_path)
        # テーブル右側に配置 (I列 = col9, row3)
        ws2.add_image(img, 'I3')
        _tmp_chart_files.append(chart_path)

    # ══════ シート3: 得点分布 ══════
    ws3 = wb.create_sheet("得点分布")
    ws3.sheet_properties.tabColor = '548235'

    ws3.merge_cells('A1:E1')
    ws3['A1'] = '得点分布'
    ws3['A1'].font = TITLE_FONT
    ws3['A1'].alignment = Alignment(vertical='center')
    ws3.row_dimensions[1].height = 30

    # 度数分布表
    bins = list(range(0, total_max_score + 11, 10))
    hist, bin_edges = np.histogram(all_scores, bins=bins)

    dist_headers = ['得点範囲', '人数', '割合(%)', '累積人数', '累積割合(%)']
    for c, h in enumerate(dist_headers, 1):
        ws3.cell(row=3, column=c, value=h)
    _style_header_row(ws3, 3, len(dist_headers))

    cumulative = 0
    dist_start = 4
    for i in range(len(hist)):
        count = int(hist[i])
        percentage = round(count / total_students * 100, 1) if total_students > 0 else 0
        cumulative += count
        cum_pct = round(cumulative / total_students * 100, 1) if total_students > 0 else 0

        r = dist_start + i
        low = int(bin_edges[i])
        high = int(bin_edges[i + 1]) - 1
        ws3.cell(row=r, column=1, value=f'{low}〜{high}点')
        ws3.cell(row=r, column=2, value=count)
        ws3.cell(row=r, column=3, value=percentage)
        ws3.cell(row=r, column=4, value=cumulative)
        ws3.cell(row=r, column=5, value=cum_pct)
        ws3.cell(row=r, column=3).number_format = '0.0'
        ws3.cell(row=r, column=5).number_format = '0.0'
        for c in [2, 3, 4, 5]:
            ws3.cell(row=r, column=c).alignment = Alignment(horizontal='right', vertical='center')

    dist_end = dist_start + len(hist) - 1
    _style_data_rows(ws3, dist_start, dist_end, len(dist_headers))

    # 列幅
    for i, w in enumerate([14, 10, 12, 12, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ヒストグラム（matplotlib 生成 → 画像埋め込み）
    if dist_end >= dist_start and len(all_scores) > 0:
        fig, ax = plt.subplots(figsize=(8, 4.5))
        _apply_clean_style(ax)
        n_bins = len(hist)
        bin_centers = [(int(bin_edges[i]) + int(bin_edges[i + 1]) - 1) / 2 for i in range(n_bins)]
        bar_width = (bin_edges[1] - bin_edges[0]) * 0.85
        bars = ax.bar(bin_centers, hist, width=bar_width,
                      color='#AAAAAA', edgecolor='#333333', linewidth=0.8)
        # データラベル
        for bar_obj, count in zip(bars, hist):
            if count > 0:
                ax.text(bar_obj.get_x() + bar_obj.get_width() / 2, bar_obj.get_height() + 0.15,
                        str(int(count)), ha='center', va='bottom', fontsize=9)
        ax.set_title('得点分布ヒストグラム', fontsize=13, fontweight='bold', pad=10)
        ax.set_xlabel('得点', fontsize=11)
        ax.set_ylabel('人数', fontsize=11)
        ax.set_ylim(bottom=0)
        ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
        # X軸: ビン範囲ラベル
        tick_labels = [f'{int(bin_edges[i])}–{int(bin_edges[i+1])-1}' for i in range(n_bins)]
        ax.set_xticks(bin_centers)
        ax.set_xticklabels(tick_labels, rotation=45, ha='right', fontsize=8)
        fig.tight_layout()

        chart_path = _save_fig_to_temp(fig, temp_dir=_chart_temp_dir)
        img = XlImage(chart_path)
        ws3.add_image(img, f'A{dist_end + 3}')
        _tmp_chart_files.append(chart_path)

    # ══════ シート4: 観点別統計 ══════
    ws4 = wb.create_sheet("観点別統計")
    ws4.sheet_properties.tabColor = 'BF8F00'

    ws4.merge_cells('A1:G1')
    ws4['A1'] = '観点別統計'
    ws4['A1'].font = TITLE_FONT
    ws4['A1'].alignment = Alignment(vertical='center')
    ws4.row_dimensions[1].height = 30

    # 観点別の満点集計
    aspect_max_scores = {}
    for data in template_dict.values():
        asp = data['観点']
        aspect_max_scores[asp] = aspect_max_scores.get(asp, 0) + data['配点']
    if has_descriptive:
        for q in desc_questions:
            asp = q['aspect']
            aspect_max_scores[asp] = aspect_max_scores.get(asp, 0) + q['max_score']

    asp_headers = ['観点', '満点', '平均点', '得点率(%)', '標準偏差', '最高点', '最低点']
    for c, h in enumerate(asp_headers, 1):
        ws4.cell(row=3, column=c, value=h)
    _style_header_row(ws4, 3, len(asp_headers))

    asp_row = 4
    for aspect in all_aspects:
        scores = aspect_scores_list[aspect]
        max_sc = aspect_max_scores.get(aspect, 0)
        mean_sc = round(float(np.mean(scores)), 2) if scores else 0
        rate = round(mean_sc / max_sc * 100, 1) if max_sc > 0 else 0

        ws4.cell(row=asp_row, column=1, value=f'観点{number_to_circled(aspect)}')
        ws4.cell(row=asp_row, column=2, value=max_sc)
        ws4.cell(row=asp_row, column=3, value=mean_sc)
        ws4.cell(row=asp_row, column=4, value=rate)
        ws4.cell(row=asp_row, column=5, value=round(float(np.std(scores, ddof=1)), 2) if len(scores) > 1 else 0)
        ws4.cell(row=asp_row, column=6, value=int(np.max(scores)) if scores else 0)
        ws4.cell(row=asp_row, column=7, value=int(np.min(scores)) if scores else 0)
        # 数値書式
        ws4.cell(row=asp_row, column=3).number_format = '0.00'
        ws4.cell(row=asp_row, column=4).number_format = '0.0'
        ws4.cell(row=asp_row, column=5).number_format = '0.00'
        for c in range(2, 8):
            ws4.cell(row=asp_row, column=c).alignment = Alignment(horizontal='right', vertical='center')
        asp_row += 1

    asp_end = asp_row - 1
    _style_data_rows(ws4, 4, asp_end, len(asp_headers))

    for i, w in enumerate([12, 8, 10, 12, 10, 10, 10], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ── 相関係数テーブル ──
    corr_start_row = asp_end + 3
    ws4.merge_cells(f'A{corr_start_row}:G{corr_start_row}')
    ws4.cell(row=corr_start_row, column=1, value='■ 観点間の相関係数').font = SUBTITLE_FONT
    corr_start_row += 1

    # 相関行列を計算
    aspect_labels = [f'観点{number_to_circled(a)}' for a in all_aspects]
    n_asp = len(all_aspects)
    corr_matrix = np.zeros((n_asp, n_asp))
    for i in range(n_asp):
        for j in range(n_asp):
            si = aspect_scores_list[all_aspects[i]]
            sj = aspect_scores_list[all_aspects[j]]
            if len(si) > 1 and np.std(si) > 0 and np.std(sj) > 0:
                corr_matrix[i][j] = round(float(np.corrcoef(si, sj)[0, 1]), 3)
            elif i == j:
                corr_matrix[i][j] = 1.0
            else:
                corr_matrix[i][j] = 0.0

    # ヘッダー行
    ws4.cell(row=corr_start_row, column=1, value='')
    for j, label in enumerate(aspect_labels):
        ws4.cell(row=corr_start_row, column=2 + j, value=label)
    _style_header_row(ws4, corr_start_row, 1 + n_asp)
    corr_start_row += 1

    # データ行
    for i, label in enumerate(aspect_labels):
        ws4.cell(row=corr_start_row + i, column=1, value=label).font = LABEL_FONT
        ws4.cell(row=corr_start_row + i, column=1).fill = LIGHT_FILL
        ws4.cell(row=corr_start_row + i, column=1).border = THIN_BORDER
        for j in range(n_asp):
            cell = ws4.cell(row=corr_start_row + i, column=2 + j, value=corr_matrix[i][j])
            cell.number_format = '0.000'
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right', vertical='center')

    corr_end_row = corr_start_row + n_asp - 1

    # ── 散布図行列（matplotlib 生成 → 画像埋め込み）──
    if n_asp >= 2 and total_students >= 2:
        fig = _create_scatter_matrix(all_aspects, aspect_scores_list, corr_matrix, aspect_labels)
        chart_path = _save_fig_to_temp(fig, temp_dir=_chart_temp_dir)
        img = XlImage(chart_path)
        scatter_row = corr_end_row + 3
        ws4.add_image(img, f'A{scatter_row}')
        _tmp_chart_files.append(chart_path)
        logger.info("✓ 散布図行列を生成")

    # ══════ 保存 ══════
    wb.save(output_path)
    logger.info("✓ 保存: %s", output_path)
    logger.info("")

    # 一時ファイル削除
    for f in _tmp_chart_files:
        try:
            Path(f).unlink(missing_ok=True)
        except Exception:
            pass

    return stats


def _evaluate_correct_rate(rate):
    """正答率に基づく簡易評価ラベルを返す"""
    if rate >= 80:
        return '◎ 易'
    elif rate >= 60:
        return '○ 適正'
    elif rate >= 40:
        return '△ やや難'
    elif rate >= 20:
        return '▽ 難'
    else:
        return '× 極難'



# ============================================================
#   記述のみモードのサマリー生成
# ============================================================

def process_descriptive_only_summary(
    image_folder,
    descriptive_config,
    descriptive_scores,
    name_images=None,
    output_base_folder=None,
):
    """記述のみモードのサマリー生成。

    マーク採点結果なしで、記述採点データだけから
    学生別サマリーと試験統計を生成する。

    Args:
        image_folder: 画像フォルダパス
        descriptive_config: descriptive_config dict
        descriptive_scores: {filename: {question_id: score}}
        name_images: {filename: trimmed_image_path}
        output_base_folder: 出力先 (None→image_folder)

    Returns:
        {'success': bool, 'stats': dict, ...} or {'success': False, 'error': str}
    """
    image_folder = Path(image_folder)
    if output_base_folder is None:
        output_base_folder = image_folder
    else:
        output_base_folder = Path(output_base_folder)

    results_folder = output_base_folder / RESULTS_FOLDER
    final_report = results_folder / FINAL_REPORT_FOLDER
    final_report.mkdir(parents=True, exist_ok=True)

    student_summary_path = final_report / STUDENT_SUMMARY_FILE
    exam_summary_path = final_report / EXAM_SUMMARY_FILE

    questions = descriptive_config.get("questions", [])
    if not questions:
        return {"success": False, "error": "記述問題が設定されていません"}

    logger.info("=" * 60)
    logger.info("サマリー生成（記述のみモード）")
    logger.info("=" * 60)
    logger.info("✓ 記述問題: %d問", len(questions))
    logger.info("✓ 対象画像: %d件", len(descriptive_scores))
    if name_images:
        logger.info("✓ 氏名欄画像: %d枚", len(name_images))
    logger.info("")

    try:
        # --- 学生別サマリー ---
        wb_student = Workbook()
        ws = wb_student.active
        ws.title = "学生別サマリー"

        # ヘッダースタイル
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # ヘッダー構築
        headers = ["No.", "ファイル名"]
        if name_images:
            headers.append("氏名欄")
        for q in questions:
            headers.append(f"{q['name']} ({q['max_score']})")
        headers.append("合計")
        full_score = sum(q["max_score"] for q in questions)
        headers.append(f"配点計 ({full_score})")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        # データ行
        sorted_files = sorted(descriptive_scores.keys())
        totals = []
        for row_idx, fname in enumerate(sorted_files, 2):
            scores_for_file = descriptive_scores.get(fname, {})
            col = 1
            ws.cell(row=row_idx, column=col, value=row_idx - 1).border = thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=escape_excel_formula(fname)).border = thin_border
            col += 1

            if name_images:
                name_path = name_images.get(fname)
                if name_path and Path(name_path).exists():
                    try:
                        img = XlImage(str(name_path))
                        img.width = 120
                        img.height = 30
                        ws.add_image(img, get_column_letter(col) + str(row_idx))
                    except Exception:
                        pass
                ws.cell(row=row_idx, column=col).border = thin_border
                col += 1

            student_total = 0
            for q in questions:
                sc = scores_for_file.get(q["id"], 0)
                if sc is None:
                    sc = 0
                ws.cell(row=row_idx, column=col, value=sc).border = thin_border
                ws.cell(row=row_idx, column=col).alignment = center
                student_total += sc
                col += 1

            ws.cell(row=row_idx, column=col, value=student_total).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = center
            ws.cell(row=row_idx, column=col).font = Font(bold=True)
            col += 1
            ws.cell(row=row_idx, column=col, value=full_score).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = center
            totals.append(student_total)

        # 列幅調整
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 30
        if name_images:
            ws.column_dimensions["C"].width = 18
            ws.row_dimensions[1].height = 20
            for r in range(2, len(sorted_files) + 2):
                ws.row_dimensions[r].height = 25

        wb_student.save(str(student_summary_path))
        logger.info("✓ 学生別サマリー: %s", student_summary_path.name)

        # --- 試験統計 ---
        totals_arr = np.array(totals) if totals else np.array([0])
        exam_stats = {
            "受験者数": len(totals),
            "満点": full_score,
            "平均点": float(np.mean(totals_arr)) if totals else 0.0,
            "標準偏差": float(np.std(totals_arr, ddof=1)) if len(totals) > 1 else 0.0,
            "最高点": int(np.max(totals_arr)) if totals else 0,
            "最低点": int(np.min(totals_arr)) if totals else 0,
        }

        wb_exam = Workbook()
        ws_exam = wb_exam.active
        ws_exam.title = "試験統計"
        stat_items = [
            ("受験者数", exam_stats["受験者数"]),
            ("満点", exam_stats["満点"]),
            ("平均点", f"{exam_stats['平均点']:.2f}"),
            ("標準偏差", f"{exam_stats['標準偏差']:.2f}"),
            ("最高点", exam_stats["最高点"]),
            ("最低点", exam_stats["最低点"]),
        ]
        ws_exam.cell(row=1, column=1, value="項目").font = header_font
        ws_exam.cell(row=1, column=2, value="値").font = header_font
        for r, (k, v) in enumerate(stat_items, 2):
            ws_exam.cell(row=r, column=1, value=k)
            ws_exam.cell(row=r, column=2, value=v)
        ws_exam.column_dimensions["A"].width = 20
        ws_exam.column_dimensions["B"].width = 20

        # 設問別統計シート
        ws_q = wb_exam.create_sheet("設問別統計")
        q_headers = ["設問", "配点", "平均", "標準偏差", "最高", "最低", "正答率(%)"]
        for ci, h in enumerate(q_headers, 1):
            cell = ws_q.cell(row=1, column=ci, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center

        for qi, q in enumerate(questions, 2):
            qid = q["id"]
            q_scores = [
                descriptive_scores.get(f, {}).get(qid, 0) or 0
                for f in sorted_files
            ]
            q_arr = np.array(q_scores) if q_scores else np.array([0])
            ws_q.cell(row=qi, column=1, value=q["name"])
            ws_q.cell(row=qi, column=2, value=q["max_score"]).alignment = center
            ws_q.cell(row=qi, column=3, value=f"{np.mean(q_arr):.2f}").alignment = center
            sd = float(np.std(q_arr, ddof=1)) if len(q_scores) > 1 else 0.0
            ws_q.cell(row=qi, column=4, value=f"{sd:.2f}").alignment = center
            ws_q.cell(row=qi, column=5, value=int(np.max(q_arr))).alignment = center
            ws_q.cell(row=qi, column=6, value=int(np.min(q_arr))).alignment = center
            rate = float(np.mean(q_arr)) / q["max_score"] * 100 if q["max_score"] > 0 else 0.0
            ws_q.cell(row=qi, column=7, value=f"{rate:.1f}").alignment = center

        wb_exam.save(str(exam_summary_path))
        logger.info("✓ 試験統計: %s", exam_summary_path.name)

        # 統合PDF
        scored_folder = results_folder / SCORED_FOLDER
        scored_pdf_path = final_report / SCORED_PDF_FILE
        if scored_folder.exists():
            try:
                combine_images_to_pdf(scored_folder, scored_pdf_path)
                logger.info("✓ 統合PDF: %s", scored_pdf_path.name)
            except Exception as pdf_e:
                logger.warning("統合PDF生成エラー: %s", pdf_e)

        # CTT分析レポート生成（記述のみモード: マーク問題なし）
        ctt_excel_path = final_report / CTT_ANALYSIS_EXCEL_FILE
        ctt_pdf_path = final_report / CTT_ANALYSIS_PDF_FILE
        ctt_result = None
        try:
            from ctt_analyzer import generate_ctt_analysis
            ctt_result = generate_ctt_analysis(
                template_path=None,
                mark2_result_path=None,
                excel_output_path=ctt_excel_path,
                pdf_output_path=ctt_pdf_path,
                skip_questions=0,
                descriptive_config=descriptive_config,
                descriptive_scores=descriptive_scores,
            )
        except Exception as ctt_e:
            logger.warning("CTT分析レポート生成エラー: %s", ctt_e, exc_info=True)

        # R連携エクスポート（記述のみモード）
        r_export_result = None
        try:
            from r_export import export_r_analysis_kit
            r_export_result = export_r_analysis_kit(
                template_path=None,
                mark2_result_path=None,
                output_folder=final_report,
                skip_questions=0,
                descriptive_config=descriptive_config,
                descriptive_scores=descriptive_scores,
            )
        except Exception as r_e:
            logger.warning("R連携エクスポートエラー: %s", r_e, exc_info=True)

        logger.info("")
        logger.info("=" * 60)
        logger.info("サマリー生成完了（記述のみモード）")
        logger.info("=" * 60)
        logger.info("✓ 学生別サマリー: %s", student_summary_path.name)
        logger.info("✓ 試験統計: %s", exam_summary_path.name)
        if ctt_result and ctt_result.get('success'):
            logger.info("✓ CTT分析Excel: %s", ctt_excel_path.name)
            if ctt_result.get('pdf_success'):
                logger.info("✓ CTT分析PDF: %s", ctt_pdf_path.name)
        if r_export_result and r_export_result.get('success'):
            logger.info("✓ R分析キット: %s/", R_EXPORT_FOLDER)

        result = {
            "success": True,
            "student_summary_path": str(student_summary_path),
            "exam_summary_path": str(exam_summary_path),
            "stats": exam_stats,
        }
        if ctt_result and ctt_result.get('success'):
            result['ctt_excel_path'] = str(ctt_excel_path)
            if ctt_result.get('pdf_success'):
                result['ctt_pdf_path'] = str(ctt_pdf_path)
        if r_export_result and r_export_result.get('success'):
            result['r_export_dir'] = r_export_result['output_dir']
        return result

    except Exception as e:
        logger.error("エラー: %s", e, exc_info=True)
        return {"success": False, "error": str(e)}


def process_summary_generation(image_folder, coord_excel_path, template_path, 
                               mark2_result_path, skip_questions=0, output_base_folder=None,
                               name_images=None, descriptive_config=None, descriptive_scores=None,
                               include_descriptive_in_analysis=False, progress_callback=None,
                               cancel_event=None):
    """サマリー生成処理を実行
    
    Args:
        image_folder: 画像フォルダのパス
        coord_excel_path: 座標ファイルのパス
        template_path: 正答データExcelファイルのパス
        mark2_result_path: OMR読取結果Excelファイルのパス
        skip_questions: スキップする問題数
        output_base_folder: 出力ベースフォルダ（Noneの場合はimage_folder）
        name_images: {ファイル名: トリミング画像パス} の辞書（オプション）
        descriptive_config: 記述問題設定dict（オプション）
        descriptive_scores: {ファイル名: {問題ID: 得点}} の辞書（オプション）
        include_descriptive_in_analysis: 記述採点結果をCTT/R分析に含めるか（デフォルトFalse）
        progress_callback: 進捗コールバック(current, total)（オプション、GUIプログレスバー用）
        cancel_event: threading.Event — set()されると処理を中断
    """
    # 遅延インポート: ctt_analyzerとsummary_generatorの循環参照を避ける
    from ctt_analyzer import generate_ctt_analysis

    image_folder = Path(image_folder)
    
    if output_base_folder is None:
        output_base_folder = image_folder
    else:
        output_base_folder = Path(output_base_folder)
    
    results_folder = output_base_folder / RESULTS_FOLDER
    results_folder.mkdir(exist_ok=True)
    
    final_report_folder = results_folder / FINAL_REPORT_FOLDER
    final_report_folder.mkdir(exist_ok=True)
    
    student_summary_path = final_report_folder / STUDENT_SUMMARY_FILE
    exam_summary_path = final_report_folder / EXAM_SUMMARY_FILE
    
    logger.info("=" * 60)
    logger.info("サマリー生成処理")
    logger.info("=" * 60)
    logger.info("✓ 入力フォルダ: %s", image_folder)
    logger.info("✓ 出力フォルダ: %s", final_report_folder)
    logger.info("✓ テンプレート: %s", Path(template_path).name)
    logger.info("✓ Mark2結果: %s", Path(mark2_result_path).name)
    if name_images:
        logger.info("✓ 氏名欄画像: %d枚", len(name_images))
    if descriptive_config and descriptive_config.get('questions'):
        logger.info("✓ 記述問題: %d問", len(descriptive_config['questions']))
    if include_descriptive_in_analysis and descriptive_config and descriptive_scores:
        logger.info("✓ 記述を分析に含む: ON")
    logger.info("")
    
    try:
        # ★ Excel読込を1回だけ行い、全サブ関数に渡す（6→2回に削減）
        template_dict = load_template(template_path)
        mark2_results = load_mark2_results(mark2_result_path, skip_questions)

        def _progress(step):
            """Step 3 は5サブタスク。各完了時にプログレス更新。"""
            if progress_callback:
                try:
                    progress_callback(step, 5)
                except Exception:
                    pass

        def _cancelled():
            """中断が要求されているか確認。"""
            return cancel_event and cancel_event.is_set()

        _progress(0)

        student_df = generate_student_summary(
            template_path, 
            mark2_result_path, 
            student_summary_path, 
            skip_questions,
            name_images=name_images,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
            template_dict=template_dict,
            mark2_results=mark2_results
        )
        _progress(1)

        if _cancelled():
            logger.info("⏹ 中断されました")
            return {'success': False, 'cancelled': True}
        
        exam_stats = generate_exam_summary(
            template_path, 
            mark2_result_path, 
            exam_summary_path, 
            skip_questions,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
            template_dict=template_dict,
            mark2_results=mark2_results
        )
        _progress(2)

        if _cancelled():
            logger.info("⏹ 中断されました")
            return {'success': False, 'cancelled': True}
        
        # 統合PDF生成（採点済み画像を1つのPDFにまとめる）
        scored_folder = results_folder / SCORED_FOLDER
        scored_pdf_path = final_report_folder / SCORED_PDF_FILE
        pdf_result = None
        if scored_folder.exists():
            try:
                pdf_result = combine_images_to_pdf(scored_folder, scored_pdf_path)
            except Exception as pdf_e:
                logger.warning("統合PDF生成エラー: %s", pdf_e)
        _progress(3)

        if _cancelled():
            logger.info("⏹ 中断されました")
            return {'success': False, 'cancelled': True}
        
        # CTT分析レポート生成（古典テスト理論）
        ctt_excel_path = final_report_folder / CTT_ANALYSIS_EXCEL_FILE
        ctt_pdf_path = final_report_folder / CTT_ANALYSIS_PDF_FILE
        ctt_result = None
        # 記述問題を分析に含めるかの判定
        _desc_config_for_analysis = None
        _desc_scores_for_analysis = None
        if include_descriptive_in_analysis and descriptive_config and descriptive_scores:
            _desc_config_for_analysis = descriptive_config
            _desc_scores_for_analysis = descriptive_scores
        try:
            ctt_result = generate_ctt_analysis(
                template_path,
                mark2_result_path,
                ctt_excel_path,
                ctt_pdf_path,
                skip_questions,
                descriptive_config=_desc_config_for_analysis,
                descriptive_scores=_desc_scores_for_analysis,
                template_dict=template_dict,
                mark2_results=mark2_results,
            )
        except Exception as ctt_e:
            logger.warning("CTT分析レポート生成エラー: %s", ctt_e, exc_info=True)
        
        _progress(4)

        # R連携エクスポート (exametrika分析キット)
        r_export_result = None
        try:
            from r_export import export_r_analysis_kit
            r_export_result = export_r_analysis_kit(
                template_path,
                mark2_result_path,
                final_report_folder,
                skip_questions,
                descriptive_config=_desc_config_for_analysis,
                descriptive_scores=_desc_scores_for_analysis,
            )
        except Exception as r_e:
            logger.warning("R連携エクスポートエラー: %s", r_e, exc_info=True)
        
        _progress(5)

        logger.info("=" * 60)
        logger.info("サマリー生成完了")
        logger.info("=" * 60)
        logger.info("✓ 学生別サマリー: %s", student_summary_path.name)
        logger.info("✓ 試験サマリー: %s", exam_summary_path.name)
        if ctt_result and ctt_result.get('success'):
            logger.info("✓ CTT分析Excel: %s", ctt_excel_path.name)
            if ctt_result.get('pdf_success'):
                logger.info("✓ CTT分析PDF: %s", ctt_pdf_path.name)
        if pdf_result:
            logger.info("✓ 統合PDF: %s", scored_pdf_path.name)
        if r_export_result and r_export_result.get('success'):
            logger.info("✓ R分析キット: %s/", R_EXPORT_FOLDER)
        logger.info("")
        
        result = {
            'success': True,
            'student_summary_path': str(student_summary_path),
            'exam_summary_path': str(exam_summary_path),
            'stats': exam_stats
        }
        if ctt_result and ctt_result.get('success'):
            result['ctt_excel_path'] = str(ctt_excel_path)
            if ctt_result.get('pdf_success'):
                result['ctt_pdf_path'] = str(ctt_pdf_path)
        if pdf_result:
            result['scored_pdf_path'] = str(scored_pdf_path)
        if r_export_result and r_export_result.get('success'):
            result['r_export_dir'] = r_export_result['output_dir']
        return result
        
    except Exception as e:
        logger.error("エラー: %s", e, exc_info=True)
        return {
            'success': False,
            'error': str(e)
        }
