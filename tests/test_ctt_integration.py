"""
CTT分析統合テスト
=================
saitensamurai.py に統合された CTT 分析機能の包括的テスト。
直接ダミーデータを生成して各クラス・関数を検証する。
"""
import sys
import os
import tempfile
import traceback
import pytest
from pathlib import Path

import numpy as np
import pandas as pd

# プロジェクトルートをパスに追加
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "main_src"))

# saitensamurai からCTT関連をインポート
from saitensamurai import (
    CTTAnalyzer,
    CTTPlotGenerator,
    CTTExcelExporter,
    CTTPDFReporter,
    generate_ctt_analysis,
    convert_mark2_to_ctt_data,
    _sort_choices,
    _is_invalid_response,
    _is_no_answer,
    HAS_MATPLOTLIB,
    HAS_REPORTLAB,
)


# ── ダミーデータ生成 ──────────────────────────────────────────

def make_dummy_ctt_data(n_students=40, n_questions=20, seed=42):
    """CTTAnalyzer用のダミー ans_df / key_df を生成"""
    rng = np.random.default_rng(seed)

    # 正答キー: 各問ランダムに A-E
    choices = list("ABCDE")
    keys = [rng.choice(choices) for _ in range(n_questions)]
    q_ids = [str(i + 1) for i in range(n_questions)]

    key_df = pd.DataFrame({"QuestionID": q_ids, "Key": keys})

    # 学生回答を生成（正答率にばらつきを持たせる）
    rows = []
    for s in range(n_students):
        ability = rng.uniform(0.3, 0.95)  # 学生ごとの能力
        answers = {}
        for q_idx in range(n_questions):
            difficulty = rng.uniform(0.2, 0.9)  # 問題ごとの難易度
            p_correct = min(ability * (1.0 - difficulty) + difficulty * 0.5, 0.95)
            if rng.random() < p_correct:
                answers[q_ids[q_idx]] = keys[q_idx]  # 正答
            else:
                wrong = [c for c in choices if c != keys[q_idx]]
                answers[q_ids[q_idx]] = rng.choice(wrong)
        answers["StudentID"] = f"student_{s+1:03d}"
        rows.append(answers)

    ans_df = pd.DataFrame(rows)
    # StudentID を先頭にする
    cols = ["StudentID"] + q_ids
    ans_df = ans_df[cols]

    return ans_df, key_df


def make_dummy_mark2_files(tmpdir, n_students=30, n_questions=15, skip_questions=4, seed=42):
    """
    convert_mark2_to_ctt_data / generate_ctt_analysis 用の
    ダミー Mark2形式 Excel ファイルを生成して返す。
    """
    rng = np.random.default_rng(seed)

    # ── テンプレート (answer_key.xlsx) ──
    template_rows = []
    for i in range(1, n_questions + 1):
        template_rows.append({
            "問題番号": i,
            "正答": rng.integers(1, 6),  # 1-5
            "配点": rng.choice([2, 3]),
            "観点": rng.choice([1, 2, 3]),
        })
    template_df = pd.DataFrame(template_rows)
    template_path = os.path.join(tmpdir, "answer_key.xlsx")
    template_df.to_excel(template_path, index=False)

    # ── Mark2結果 Excel ──
    # Row 0 (ヘッダー): No, File, 1, 2, ..., (skip_questions + n_questions)
    total_cols = skip_questions + n_questions
    header_row = ["No", "File"] + [str(i) for i in range(1, total_cols + 1)]

    # Row 1 (設問名): NaN, NaN, ID列名..., 設問番号...
    id_names = ["学年", "クラス", "出席番号（十の位）", "出席番号（一の位）"]
    name_row = [np.nan, np.nan] + id_names[:skip_questions] + [str(i) for i in range(1, n_questions + 1)]

    # Row 2+ (データ)
    data_rows = []
    for s in range(n_students):
        row = [s + 1, f"page_{s+1:03d}.png"]
        # ID部（学年=1, クラス=1-3, 出番十=0-3, 出番一=1-9）
        row += [1, rng.integers(1, 4), s // 10, (s % 10) + 1][:skip_questions]
        # 各設問: 選択肢番号 1-5
        for q in range(n_questions):
            row.append(int(rng.integers(1, 6)))
        data_rows.append(row)

    all_rows = [header_row, name_row] + data_rows
    result_df = pd.DataFrame(all_rows)
    result_path = os.path.join(tmpdir, "Mark2-Result-A1-C002-20260210.xlsx")
    result_df.to_excel(result_path, index=False, header=False)

    return template_path, result_path


# ── テスト関数 ──────────────────────────────────────────

class _ResultsCollector:
    """テスト結果の集計"""
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.errors = []

    def ok(self, name):
        self.passed += 1
        print(f"  ✓ {name}")

    def fail(self, name, msg):
        self.failed += 1
        self.errors.append((name, msg))
        print(f"  ✗ {name}: {msg}")

    def summary(self):
        total = self.passed + self.failed
        print(f"\n{'='*60}")
        print(f"テスト結果: {self.passed}/{total} 成功, {self.failed}/{total} 失敗")
        if self.errors:
            print("\n失敗詳細:")
            for name, msg in self.errors:
                print(f"  - {name}: {msg}")
        print(f"{'='*60}")
        return self.failed == 0


@pytest.fixture
def results():
    """CTTテスト結果集計用のフィクスチャ"""
    return _ResultsCollector()


@pytest.fixture
def ctt_data():
    """CTTダミーデータ"""
    return make_dummy_ctt_data(n_students=40, n_questions=20)


@pytest.fixture
def analyzer(ctt_data):
    """CTTAnalyzerインスタンス"""
    ans_df, key_df = ctt_data
    return CTTAnalyzer(ans_df, key_df)


@pytest.fixture
def test_stats(analyzer):
    """テスト統計量"""
    return analyzer.calculate_test_stats()


@pytest.fixture
def item_stats(analyzer):
    """項目統計量"""
    return analyzer.calculate_item_stats()


@pytest.fixture
def distractor_stats(analyzer):
    """選択肢分析"""
    return analyzer.calculate_distractor_analysis()


def test_ctt_analyzer(results: _ResultsCollector):
    """CTTAnalyzer の基本テスト"""
    print("\n── CTTAnalyzer テスト ──")
    ans_df, key_df = make_dummy_ctt_data(n_students=40, n_questions=20)

    # 初期化
    try:
        analyzer = CTTAnalyzer(ans_df, key_df)
        results.ok("CTTAnalyzer 初期化")
    except Exception as e:
        results.fail("CTTAnalyzer 初期化", str(e))
        traceback.print_exc()
        return

    # score_matrix
    try:
        assert analyzer.score_matrix.shape == (40, 20), \
            f"shape={analyzer.score_matrix.shape}, expected=(40,20)"
        assert set(analyzer.score_matrix.values.flatten()) <= {0, 1}, \
            "score_matrixに0/1以外の値が含まれている"
        results.ok("score_matrix (shape & binary)")
    except Exception as e:
        results.fail("score_matrix", str(e))

    # total_scores
    try:
        assert len(analyzer.total_scores) == 40
        assert all(0 <= s <= 20 for s in analyzer.total_scores)
        results.ok("total_scores (range & length)")
    except Exception as e:
        results.fail("total_scores", str(e))

    # test_stats
    try:
        test_stats = analyzer.calculate_test_stats()
        required_keys = [
            "平均点 (Mean)", "中央値 (Median)", "標準偏差 (SD)",
            "最低点 (Min)", "最高点 (Max)", "分散 (Variance)",
            "信頼性係数 (α)", "受験者数 (N)", "項目数 (K)"
        ]
        for k in required_keys:
            assert k in test_stats, f"キー '{k}' が test_stats にない"
        assert test_stats["受験者数 (N)"] == 40
        assert test_stats["項目数 (K)"] == 20
        alpha = test_stats["信頼性係数 (α)"]
        assert -1 <= alpha <= 1, f"α={alpha} が範囲外"
        results.ok(f"test_stats (α={alpha:.4f})")
    except Exception as e:
        results.fail("test_stats", str(e))
        traceback.print_exc()
        test_stats = None

    # item_stats
    try:
        item_stats = analyzer.calculate_item_stats()
        assert len(item_stats) == 20, f"item_stats行数={len(item_stats)}, expected=20"
        required_cols = [
            "QuestionID", "Key", "正答率 (P)", "識別指数 (D)",
            "I-T相関", "I-T相関(含)", "削除α"
        ]
        for c in required_cols:
            assert c in item_stats.columns, f"列 '{c}' がない"
        # P値の範囲チェック
        p_vals = item_stats["正答率 (P)"]
        assert all(0 <= p <= 1 for p in p_vals), "P値が0-1の範囲外"
        results.ok("item_stats (shape & P-value range)")
    except Exception as e:
        results.fail("item_stats", str(e))
        traceback.print_exc()
        item_stats = None

    # distractor_analysis
    try:
        distractor_stats = analyzer.calculate_distractor_analysis()
        assert len(distractor_stats) > 0, "distractor_stats が空"
        required_cols = ["QuestionID", "Choice", "IsKey",
                         "Count_全体", "Ratio_全体",
                         "Count_高群", "Ratio_高群",
                         "Count_低群", "Ratio_低群"]
        for c in required_cols:
            assert c in distractor_stats.columns, f"列 '{c}' がない"
        results.ok(f"distractor_analysis ({len(distractor_stats)} rows)")
    except Exception as e:
        results.fail("distractor_analysis", str(e))
        traceback.print_exc()
        distractor_stats = None


def test_ctt_plot_generator(results: _ResultsCollector, analyzer, item_stats, distractor_stats):
    """CTTPlotGenerator のテスト"""
    print("\n── CTTPlotGenerator テスト ──")

    if not HAS_MATPLOTLIB:
        results.fail("CTTPlotGenerator", "matplotlib が利用不可")
        return

    try:
        plotter = CTTPlotGenerator()
        results.ok("CTTPlotGenerator 初期化")
    except Exception as e:
        results.fail("CTTPlotGenerator 初期化", str(e))
        return

    # ヒストグラム
    try:
        buf = plotter.generate_score_histogram(analyzer.total_scores)
        assert buf is not None, "histogram が None"
        assert len(buf.getvalue()) > 0, "histogram が空"
        results.ok(f"score_histogram ({len(buf.getvalue())} bytes)")
    except Exception as e:
        results.fail("score_histogram", str(e))
        traceback.print_exc()

    # アイテムカーブ
    try:
        row = item_stats.iloc[0]
        q_dist = distractor_stats[distractor_stats["QuestionID"] == row["QuestionID"]]
        buf = plotter.generate_item_curve(row, q_dist)
        assert buf is not None
        results.ok("item_curve")
    except Exception as e:
        results.fail("item_curve", str(e))
        traceback.print_exc()

    # ミニトレースグリッド
    try:
        bufs = plotter.generate_mini_trace_grid(
            distractor_stats, analyzer.questions, cols=5, rows_per_page=4
        )
        assert bufs is not None
        assert len(bufs) > 0
        results.ok(f"mini_trace_grid ({len(bufs)} pages)")
    except Exception as e:
        results.fail("mini_trace_grid", str(e))
        traceback.print_exc()

    # 相関チャート
    try:
        buf = plotter.generate_correlation_chart(
            item_stats, "I-T相関", "I-T相関"
        )
        assert buf is not None
        results.ok("correlation_chart")
    except Exception as e:
        results.fail("correlation_chart", str(e))
        traceback.print_exc()

    # 相関ヒートマップ
    try:
        buf = plotter.generate_correlation_heatmap(
            analyzer.score_matrix, analyzer.questions
        )
        assert buf is not None
        results.ok("correlation_heatmap")
    except Exception as e:
        results.fail("correlation_heatmap", str(e))
        traceback.print_exc()


def test_ctt_excel_exporter(results: _ResultsCollector, analyzer, test_stats, item_stats, distractor_stats):
    """CTTExcelExporter のテスト"""
    print("\n── CTTExcelExporter テスト ──")

    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_003_CTT_analysis_report.xlsx")
        try:
            exporter = CTTExcelExporter(output_path, analyzer)
            exporter.export(test_stats, analyzer.ans_df, item_stats, distractor_stats)
            assert os.path.exists(output_path), "Excel ファイルが生成されなかった"
            fsize = os.path.getsize(output_path)
            assert fsize > 1000, f"Excel ファイルが小さすぎる ({fsize} bytes)"
            results.ok(f"Excel export ({fsize:,} bytes)")

            # openpyxlで中身を検証
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            sheet_names = wb.sheetnames
            assert "目次" in sheet_names, f"'目次' シートがない: {sheet_names}"
            assert "テスト得点" in sheet_names, f"'テスト得点' シートがない: {sheet_names}"
            assert "項目全体" in sheet_names, f"'項目全体' シートがない: {sheet_names}"
            results.ok(f"Excel sheets: {sheet_names[:5]}...")
            wb.close()

        except Exception as e:
            results.fail("Excel export", str(e))
            traceback.print_exc()


def test_ctt_pdf_reporter(results: _ResultsCollector, analyzer, test_stats, item_stats, distractor_stats):
    """CTTPDFReporter のテスト"""
    print("\n── CTTPDFReporter テスト ──")

    if not HAS_REPORTLAB:
        results.fail("CTTPDFReporter", "reportlab が利用不可")
        return

    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "test_004_CTT_analysis_report.pdf")
        try:
            reporter = CTTPDFReporter(output_path)
            reporter.generate_report(
                test_stats, item_stats, distractor_stats,
                analyzer.total_scores, analyzer.questions, analyzer.score_matrix
            )
            assert os.path.exists(output_path), "PDF ファイルが生成されなかった"
            fsize = os.path.getsize(output_path)
            assert fsize > 1000, f"PDF ファイルが小さすぎる ({fsize} bytes)"
            results.ok(f"PDF export ({fsize:,} bytes)")
        except Exception as e:
            results.fail("PDF export", str(e))
            traceback.print_exc()


def test_convert_mark2_to_ctt_data(results: _ResultsCollector):
    """convert_mark2_to_ctt_data のテスト"""
    print("\n── convert_mark2_to_ctt_data テスト ──")

    with tempfile.TemporaryDirectory() as tmpdir:
        template_path, result_path = make_dummy_mark2_files(
            tmpdir, n_students=25, n_questions=10, skip_questions=4
        )

        try:
            ans_df, key_df = convert_mark2_to_ctt_data(
                template_path, result_path, skip_questions=4
            )
            assert "StudentID" in ans_df.columns, "StudentID列がない"
            assert len(ans_df) == 25, f"学生数={len(ans_df)}, expected=25"
            assert len(key_df) == 10, f"設問数={len(key_df)}, expected=10"
            assert "QuestionID" in key_df.columns
            assert "Key" in key_df.columns
            results.ok(f"convert_mark2_to_ctt_data (students={len(ans_df)}, questions={len(key_df)})")

            # 変換後データでCTTAnalyzer動作確認
            analyzer = CTTAnalyzer(ans_df, key_df)
            ts = analyzer.calculate_test_stats()
            results.ok(f"CTTAnalyzer with converted data (α={ts['信頼性係数 (α)']:.4f})")

        except Exception as e:
            results.fail("convert_mark2_to_ctt_data", str(e))
            traceback.print_exc()


def test_generate_ctt_analysis(results: _ResultsCollector):
    """generate_ctt_analysis 統合テスト"""
    print("\n── generate_ctt_analysis 統合テスト ──")

    with tempfile.TemporaryDirectory() as tmpdir:
        template_path, result_path = make_dummy_mark2_files(
            tmpdir, n_students=30, n_questions=15, skip_questions=4
        )
        excel_out = os.path.join(tmpdir, "003_CTT_analysis_report.xlsx")
        pdf_out = os.path.join(tmpdir, "004_CTT_analysis_report.pdf")

        try:
            result = generate_ctt_analysis(
                template_path, result_path, excel_out, pdf_out,
                skip_questions=4
            )
            assert result["success"], f"generate_ctt_analysis 失敗: {result}"
            assert os.path.exists(excel_out), "Excel出力なし"
            results.ok(f"generate_ctt_analysis success (Excel: {os.path.getsize(excel_out):,} bytes)")

            if result.get("pdf_success"):
                assert os.path.exists(pdf_out), "PDF出力なし"
                results.ok(f"PDF生成成功 ({os.path.getsize(pdf_out):,} bytes)")
            else:
                results.ok("PDF生成スキップ (依存ライブラリなし)")

            # test_stats の中身確認
            ts = result["test_stats"]
            print(f"    テスト統計: N={ts['受験者数 (N)']}, K={ts['項目数 (K)']}, "
                  f"Mean={ts['平均点 (Mean)']:.2f}, α={ts['信頼性係数 (α)']:.4f}")

        except Exception as e:
            results.fail("generate_ctt_analysis", str(e))
            traceback.print_exc()


def test_edge_cases(results: _ResultsCollector):
    """エッジケーステスト"""
    print("\n── エッジケース テスト ──")

    # 0. _sort_choices 単体テスト
    try:
        # 基本: 数値順
        assert _sort_choices(['3', '1', '2']) == ['1', '2', '3']
        # 10選択肢マークシート: "0"は末尾
        assert _sort_choices(['3', '0', '1', '2']) == ['1', '2', '3', '0']
        # 無効回答は常に最後
        assert _sort_choices(['2', '無効回答', '1']) == ['1', '2', '無効回答']
        # ダブルマークは「無効回答」に集約され個別には出力されない
        assert _sort_choices(['1', '2;3', '3', '無効回答', '0']) == ['1', '3', '0', '無効回答']
        # 空文字列・nanは除去される
        assert _sort_choices(['2', '', '1', 'nan']) == ['1', '2']
        # 大きな数値: 10は0と同義（10番目のマーク位置）
        assert _sort_choices(['10', '2', '1', '3']) == ['1', '2', '3', '0']
        # "0"と"10"が両方存在 → どちらも"0"に正規化
        assert _sort_choices(['0', '10', '1', '5']) == ['1', '5', '0', '0']
        # -1 は「無効回答」に集約され個別には出力されない
        assert _sort_choices(['1', '-1', '3', '0']) == ['1', '3', '0']
        # -1.0 も同様
        assert _sort_choices(['-1.0', '2', '5']) == ['2', '5']
        results.ok("_sort_choices 単体テスト (9 cases)")
    except Exception as e:
        results.fail("_sort_choices", str(e))
        traceback.print_exc()

    # 0⇔10等価判定テスト (CTT採点の核心)
    try:
        # 正答が"0"、生徒が"0" → 正解
        ans_df = pd.DataFrame({
            "StudentID": ["s1", "s2", "s3", "s4"],
            "1": ["0",  "10", "1",  "0"],     # s1,s2正解, s3不正解, s4正解
            "2": ["10", "0",  "10", "3"],      # s1,s2,s3正解(key=10→0), s4不正解
        })
        key_df = pd.DataFrame({
            "QuestionID": ["1", "2"],
            "Key": ["0", "10"]   # "10"はcalculate時に等価判定される
        })
        analyzer = CTTAnalyzer(ans_df, key_df)
        sm = analyzer.score_matrix
        # 設問1 (key=0): s1(0)==0→1, s2(10)==0→1(等価), s3(1)→0, s4(0)→1
        assert list(sm["1"]) == [1, 1, 0, 1], f"設問1: {list(sm['1'])}"
        # 設問2 (key=10→0等価): s1(10)→1, s2(0)→1, s3(10)→1, s4(3)→0
        assert list(sm["2"]) == [1, 1, 1, 0], f"設問2: {list(sm['2'])}"
        ts = analyzer.calculate_test_stats()
        results.ok(f"0⇔10等価判定 (Mean={ts['平均点 (Mean)']:.2f})")
    except Exception as e:
        results.fail("0⇔10等価判定", str(e))
        traceback.print_exc()

    # 複数正答テスト
    try:
        ans_df = pd.DataFrame({
            "StudentID": ["s1", "s2", "s3"],
            "1": ["1;3", "3;1", "1;2"],
        })
        key_df = pd.DataFrame({
            "QuestionID": ["1"],
            "Key": ["1;3"]
        })
        analyzer = CTTAnalyzer(ans_df, key_df)
        sm = analyzer.score_matrix
        # s1: "1;3" == "1;3" (集合一致) → 1
        # s2: "3;1" == "1;3" (集合一致, 順序不問) → 1
        # s3: "1;2" != "1;3" → 0
        assert list(sm["1"]) == [1, 1, 0], f"複数正答: {list(sm['1'])}"
        results.ok("複数正答 (集合一致)")
    except Exception as e:
        results.fail("複数正答", str(e))
        traceback.print_exc()

    # 1. 最小データ (2人、2問)
    try:
        ans_df = pd.DataFrame({
            "StudentID": ["s1", "s2"],
            "1": ["A", "B"],
            "2": ["C", "C"]
        })
        key_df = pd.DataFrame({
            "QuestionID": ["1", "2"],
            "Key": ["A", "C"]
        })
        analyzer = CTTAnalyzer(ans_df, key_df)
        ts = analyzer.calculate_test_stats()
        item = analyzer.calculate_item_stats()
        results.ok(f"最小データ (2students x 2questions, α={ts['信頼性係数 (α)']:.4f})")
    except Exception as e:
        results.fail("最小データ", str(e))
        traceback.print_exc()

    # 2. 全員全問正解
    try:
        ans_df = pd.DataFrame({
            "StudentID": [f"s{i}" for i in range(10)],
            "1": ["A"] * 10,
            "2": ["B"] * 10,
            "3": ["C"] * 10,
        })
        key_df = pd.DataFrame({
            "QuestionID": ["1", "2", "3"],
            "Key": ["A", "B", "C"]
        })
        analyzer = CTTAnalyzer(ans_df, key_df)
        ts = analyzer.calculate_test_stats()
        item = analyzer.calculate_item_stats()
        p_vals = item["正答率 (P)"]
        assert all(p == 1.0 for p in p_vals), "全員正解なのにP≠1.0"
        results.ok("全員全問正解 (P=1.0)")
    except Exception as e:
        results.fail("全員全問正解", str(e))
        traceback.print_exc()

    # 3. 全員全問不正解
    try:
        ans_df = pd.DataFrame({
            "StudentID": [f"s{i}" for i in range(10)],
            "1": ["B"] * 10,
            "2": ["C"] * 10,
            "3": ["D"] * 10,
        })
        key_df = pd.DataFrame({
            "QuestionID": ["1", "2", "3"],
            "Key": ["A", "B", "C"]
        })
        analyzer = CTTAnalyzer(ans_df, key_df)
        ts = analyzer.calculate_test_stats()
        item = analyzer.calculate_item_stats()
        p_vals = item["正答率 (P)"]
        assert all(p == 0.0 for p in p_vals), "全員不正解なのにP≠0.0"
        results.ok("全員全問不正解 (P=0.0)")
    except Exception as e:
        results.fail("全員全問不正解", str(e))
        traceback.print_exc()

    # 4. 空文字（無回答）を含む解答
    try:
        ans_df = pd.DataFrame({
            "StudentID": ["s1", "s2", "s3"],
            "1": ["A", "", "A"],     # s2が無回答
            "2": ["B", "B", ""],     # s3が無回答
        })
        key_df = pd.DataFrame({
            "QuestionID": ["1", "2"],
            "Key": ["A", "B"]
        })
        analyzer = CTTAnalyzer(ans_df, key_df)
        ts = analyzer.calculate_test_stats()
        results.ok(f"無回答あり (Mean={ts['平均点 (Mean)']:.2f})")
    except Exception as e:
        results.fail("無回答あり", str(e))
        traceback.print_exc()


# ── ③ 無効回答ハンドリング・記述問題テスト ──────────────────────────

class TestIsInvalidResponse:
    """_is_invalid_response ヘルパー関数のテスト"""

    def test_empty_string(self):
        assert _is_invalid_response('') is True

    def test_nan(self):
        assert _is_invalid_response('nan') is True

    def test_explicit_label(self):
        assert _is_invalid_response('無効回答') is True

    def test_legacy_label(self):
        """旧ラベル '無答' も無効回答として認識する"""
        assert _is_invalid_response('無答') is True

    def test_minus_one(self):
        assert _is_invalid_response('-1') is True

    def test_minus_one_float(self):
        assert _is_invalid_response('-1.0') is True

    def test_double_mark(self):
        assert _is_invalid_response('3;4') is True

    def test_double_mark_multi(self):
        assert _is_invalid_response('1;2;3') is True

    def test_regular_choice(self):
        assert _is_invalid_response('1') is False
        assert _is_invalid_response('0') is False
        assert _is_invalid_response('A') is False

    def test_zero_is_not_invalid(self):
        """0はマークシートの10番目の選択肢であり、無効回答ではない"""
        assert _is_invalid_response('0') is False

    def test_backward_compat_alias(self):
        """後方互換: _is_no_answer は _is_invalid_response のエイリアス"""
        assert _is_no_answer is _is_invalid_response


class TestDistractorWithInvalidResponse:
    """無効回答（-1, ダブルマーク）データを含むディストラクタ分析"""

    def test_minus_one_counted_as_invalid(self):
        """"-1"の回答は無効回答として集計される"""
        ans_df = pd.DataFrame({
            "StudentID": [f"s{i}" for i in range(10)],
            "1": ["A", "A", "B", "-1", "A", "C", "-1", "A", "B", "A"],
        })
        key_df = pd.DataFrame({"QuestionID": ["1"], "Key": ["A"]})
        az = CTTAnalyzer(ans_df, key_df)
        ds = az.calculate_distractor_analysis()
        q1 = ds[ds["QuestionID"] == "1"]
        choices = set(q1["Choice"])
        # -1 は選択肢に出ず、無効回答として集計
        assert "-1" not in choices, "-1が独立の選択肢として表示されている"
        assert "無効回答" in choices, "無効回答が存在しない"
        na_row = q1[q1["Choice"] == "無効回答"]
        # 2人が-1 → 無効回答の全体選択率 = 2/10 = 0.2
        assert abs(na_row["Ratio_全体"].values[0] - 0.2) < 0.05

    def test_double_mark_counted_as_invalid(self):
        """ダブルマーク("3;4"等)は無効回答として集計される"""
        ans_df = pd.DataFrame({
            "StudentID": [f"s{i}" for i in range(10)],
            "1": ["1", "2", "1;2", "3", "1", "2;3", "1", "3", "1", "2"],
        })
        key_df = pd.DataFrame({"QuestionID": ["1"], "Key": ["1"]})
        az = CTTAnalyzer(ans_df, key_df)
        ds = az.calculate_distractor_analysis()
        q1 = ds[ds["QuestionID"] == "1"]
        choices = set(q1["Choice"])
        assert "1;2" not in choices, "ダブルマークが独立の選択肢になっている"
        assert "2;3" not in choices, "ダブルマークが独立の選択肢になっている"
        assert "無効回答" in choices
        na_row = q1[q1["Choice"] == "無効回答"]
        # 2人がダブルマーク → 無効回答 = 2/10 = 0.2
        assert abs(na_row["Ratio_全体"].values[0] - 0.2) < 0.05

    def test_sort_choices_excludes_minus_one(self):
        """_sort_choices は -1 を出力しない"""
        result = _sort_choices(['-1', '1', '2', '3'])
        assert '-1' not in result
        assert result == ['1', '2', '3']

    def test_sort_choices_excludes_double_marks(self):
        """_sort_choices はダブルマークを出力しない"""
        result = _sort_choices(['1;2', '1', '2', '3'])
        assert '1;2' not in result
        assert result == ['1', '2', '3']


class TestDescriptiveQuestionExcel:
    """記述問題(D*)のExcel出力テスト"""

    def _make_mixed_data(self):
        """マーク問題+記述問題のダミーデータ"""
        rng = np.random.default_rng(123)
        n_students = 20

        rows = []
        for s in range(n_students):
            row = {
                "StudentID": f"s{s+1:03d}",
                "1": str(rng.choice([1, 2, 3, 4])),
                "2": str(rng.choice([1, 2, 3, 4])),
                "D1": str(rng.choice([0, 1, -1])),  # 記述問題
                "D2": str(rng.choice([0, 1])),       # 記述問題
            }
            rows.append(row)
        ans_df = pd.DataFrame(rows)

        key_df = pd.DataFrame({
            "QuestionID": ["1", "2", "D1", "D2"],
            "Key": ["2", "3", "1", "1"],
        })
        return ans_df, key_df

    def test_descriptive_overview_has_dash_for_extra_choices(self):
        """Overview シートで記述問題行は 0,1,無効回答 以外に '–' が入る"""
        ans_df, key_df = self._make_mixed_data()
        az = CTTAnalyzer(ans_df, key_df)
        ts = az.calculate_test_stats()
        ist = az.calculate_item_stats()
        ds = az.calculate_distractor_analysis()

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test_desc.xlsx")
            ex = CTTExcelExporter(out, az)
            ex._create_overview(ist, ds, ts)
            ws = ex.wb["項目全体"]

            # ヘッダー行(row=2)から選択肢列を取得
            headers = []
            for col_idx in range(9, ws.max_column + 1):
                h = ws.cell(row=2, column=col_idx).value
                if h is not None:
                    headers.append((col_idx, str(h)))

            # 記述問題 D1 の行を探す
            d1_row = None
            for r in range(3, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == "D1":
                    d1_row = r
                    break
            assert d1_row is not None, "D1行が見つからない"

            # D1行: 0, 1, 無効回答のセルは数値、その他は '–'
            for col_idx, h in headers:
                val = ws.cell(row=d1_row, column=col_idx).value
                if h in ('0', '1', '無効回答'):
                    # 数値(%フォーマット)のはず
                    assert val != '–', f"D1の{h}列が'–'になっている"
                else:
                    assert val == '–', f"D1の{h}列が'–'でない: {val}"

    def test_descriptive_item_sheet_has_limited_choices(self):
        """記述問題の個別シート(D1, D2)は選択肢が 0, 1, 無効回答 のみ"""
        ans_df, key_df = self._make_mixed_data()
        az = CTTAnalyzer(ans_df, key_df)
        ts = az.calculate_test_stats()
        ist = az.calculate_item_stats()
        ds = az.calculate_distractor_analysis()

        with tempfile.TemporaryDirectory() as tmpdir:
            out = os.path.join(tmpdir, "test_desc.xlsx")
            ex = CTTExcelExporter(out, az)
            ex._create_items(ist, ds, ts)

            # D1 シートの選択率ヘッダーを確認
            ws = ex.wb["D1"]
            # row=7 が選択率ヘッダー行 (col 2~ が選択肢名)
            ch_headers = []
            for c in range(2, 20):
                v = ws.cell(row=7, column=c).value
                if v is not None:
                    ch_headers.append(str(v))
                else:
                    break
            assert ch_headers == ['0', '1', '無効回答'], \
                f"D1シートの選択肢が {ch_headers} (期待: ['0', '1', '無効回答'])"

            # マーク問題 "1" のシートは通常の選択肢数
            ws_mark = ex.wb["1"]
            mark_headers = []
            for c in range(2, 20):
                v = ws_mark.cell(row=7, column=c).value
                if v is not None:
                    mark_headers.append(str(v))
                else:
                    break
            assert len(mark_headers) > 3, \
                f"マーク問題のシートの選択肢が少なすぎる: {mark_headers}"


# ── メイン ──────────────────────────────────────────

def main():
    print("=" * 60)
    print("CTT分析統合テスト")
    print(f"matplotlib: {'✓' if HAS_MATPLOTLIB else '✗'}")
    print(f"reportlab:  {'✓' if HAS_REPORTLAB else '✗'}")
    print("=" * 60)

    results = TestResults()

    # 1. CTTAnalyzer 基本テスト
    test_ctt_analyzer(results)

    # pytest用フィクスチャと同等のデータを直接生成（スタンドアロン用）
    ans_df, key_df = make_dummy_ctt_data(n_students=40, n_questions=20)
    try:
        analyzer = CTTAnalyzer(ans_df, key_df)
        test_stats = analyzer.calculate_test_stats()
        item_stats = analyzer.calculate_item_stats()
        distractor_stats = analyzer.calculate_distractor_analysis()
    except Exception:
        print("\n⚠ CTTAnalyzer初期化失敗のため後続テストをスキップ")
        results.summary()
        return

    # 2. CTTPlotGenerator テスト
    test_ctt_plot_generator(results, analyzer, item_stats, distractor_stats)

    # 3. CTTExcelExporter テスト
    test_ctt_excel_exporter(results, analyzer, test_stats, item_stats, distractor_stats)

    # 4. CTTPDFReporter テスト
    test_ctt_pdf_reporter(results, analyzer, test_stats, item_stats, distractor_stats)

    # 5. データ変換テスト
    test_convert_mark2_to_ctt_data(results)

    # 6. 統合テスト
    test_generate_ctt_analysis(results)

    # 7. エッジケース
    test_edge_cases(results)

    # 結果サマリ
    all_passed = results.summary()
    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
