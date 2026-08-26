"""webui Bridge の認識実行ジョブ（L1: pywebview なし・実ドライバ使用）。

合成答案画像（tests/test_multidigit_image_e2e のヘルパー再利用）で
process_box_drawer を実際に回し、進捗push・完了イベント・OMR結果の
自動選択までを固定する。
"""
import json
import sys
import time
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "webui"))
sys.path.insert(0, str(PROJECT_ROOT / "main_src"))
sys.path.insert(0, str(PROJECT_ROOT / "tests"))

from api.bridge import Bridge  # noqa: E402
from test_multidigit_image_e2e import (  # noqa: E402
    make_coord_xlsx, make_sheet, sym_positions,
)


class RecordingAdapter:
    """ダイアログはキュー応答、eval_js は push イベントを記録"""

    def __init__(self):
        self.folder_returns = []
        self.file_returns = []
        self.files_returns = []   # open_files_dialog（複数選択）用
        self.events = []

    def open_folder_dialog(self, directory=""):
        return self.folder_returns.pop(0) if self.folder_returns else None

    def open_file_dialog(self, file_types=None, directory=""):
        return self.file_returns.pop(0) if self.file_returns else None

    def open_files_dialog(self, file_types=None, directory=""):
        return self.files_returns.pop(0) if self.files_returns else None

    def eval_js(self, code):
        # window.saitenEvents({...}) から JSON部分を取り出す
        start = code.index("(") + 1
        self.events.append(json.loads(code[start:-1]))


def _wait_job_done(bridge, timeout=60):
    deadline = time.time() + timeout
    while bridge.state["job"]["running"]:
        if time.time() > deadline:
            pytest.fail("ジョブがタイムアウトしました")
        time.sleep(0.1)


@pytest.fixture
def ready_bridge(tmp_path):
    """合成答案2枚＋座標ファイルを選択済みの bridge（複数桁モード）"""
    import cv2
    adapter = RecordingAdapter()
    bridge = Bridge(window_adapter=adapter)
    bridge.set_mode("mark_only", "multi_digit")
    bridge.set_skip_questions(0)

    coord = make_coord_xlsx(tmp_path / "coord.xlsx", 3)
    scans = tmp_path / "scans"
    scans.mkdir()
    s1 = {q: p for q, p in zip([1, 2, 3], sym_positions('-24'))}
    s2 = {1: sym_positions('9')[0], 2: sym_positions('9')[0],
          3: sym_positions('9')[0]}
    for name, filled in [("s1.png", s1), ("s2.png", s2)]:
        cv2.imwrite(str(scans / name), make_sheet(filled, with_markers=True))

    adapter.folder_returns = [str(scans)]
    adapter.file_returns = [str(coord)]
    assert bridge.select_image_folder()["ok"]
    assert bridge.select_coord_file()["ok"]
    return bridge, adapter, tmp_path


class TestRecognitionJob:
    def test_full_run_pushes_progress_and_autoselects_result(self, ready_bridge):
        bridge, adapter, tmp_path = ready_bridge
        res = bridge.run_recognition()
        assert res["ok"] and res["state"]["job"]["running"] is True
        _wait_job_done(bridge)

        progress = [e for e in adapter.events if e["type"] == "progress"]
        done = [e for e in adapter.events if e["type"] == "job_done"]
        assert progress and progress[-1]["current"] == 2 and progress[-1]["total"] == 2
        assert len(done) == 1 and done[0]["ok"] is True
        assert done[0]["success_count"] == 2 and done[0]["error_count"] == 0

        # OMR結果が自動選択され、実在し、記号表記であること
        omr = bridge.state["omr_result"]
        assert omr and Path(omr).exists() and "Mark2-Result-" in omr
        import openpyxl
        ws = openpyxl.load_workbook(omr)["Sheet1"]
        rows = {ws.cell(r, 2).value: [ws.cell(r, c).value for c in (3, 4, 5)]
                for r in (3, 4)}
        assert rows["s1.png"] == ['-', '2', '4']

    def test_prerequisite_guards(self, tmp_path):
        bridge = Bridge(window_adapter=RecordingAdapter())
        assert "画像フォルダ" in bridge.run_recognition()["error"]

    def test_busy_guard(self, ready_bridge):
        bridge, adapter, _ = ready_bridge
        bridge.state["job"]["running"] = True  # 実行中を模擬
        assert "実行中" in bridge.run_recognition()["error"]
        bridge.state["job"]["running"] = False

    def test_cancel_sets_event(self, ready_bridge):
        bridge, adapter, _ = ready_bridge
        bridge.state["job"]["running"] = True
        assert bridge.cancel_job()["ok"]
        assert bridge._cancel_event.is_set()
        bridge.state["job"]["running"] = False

    def test_manual_omr_result_selection(self, ready_bridge):
        bridge, adapter, tmp_path = ready_bridge
        fake = tmp_path / "old-result.xlsx"
        fake.write_bytes(b"x")
        adapter.file_returns = [str(fake)]
        res = bridge.select_omr_result()
        assert res["state"]["omr_result"] == str(fake)
