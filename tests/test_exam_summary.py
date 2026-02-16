#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_exam_summary() の改善版テスト

4シート構成 (試験概要, 設問分析, 得点分布, 観点別統計)、
グラフ挿入、記述問題対応、評価ラベル等を検証する。
"""

import sys
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "main_src"))

from summary_generator import generate_exam_summary, _evaluate_correct_rate


# ============================================================
# フィクスチャ
# ============================================================

@pytest.fixture
def template_path(tmp_path):
    """3問・2観点のテンプレート"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["問題番号", "正答", "配点", "観点"])
    ws.append([1, "1", 2, 1])
    ws.append([2, "3", 3, 2])
    ws.append([3, "2", 5, 1])
    path = tmp_path / "template.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def mark2_result_path(tmp_path):
    """4名分のダミー結果"""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["No", "File", 1, 2, 3])
    ws.append([None, None, 1, 2, 3])
    # 全問正解
    ws.append([1, "img001.jpg", 1, 3, 2])
    # 問1のみ正解
    ws.append([2, "img002.jpg", 1, 1, 1])
    # 問2のみ正解
    ws.append([3, "img003.jpg", 2, 3, 1])
    # 全問不正解
    ws.append([4, "img004.jpg", 4, 4, 4])
    path = tmp_path / "mark2_result.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def descriptive_config():
    """記述問題2問"""
    return {
        "questions": [
            {"id": "D1", "name": "記述1", "region": [0, 0, 100, 100],
             "max_score": 5, "aspect": 2},
            {"id": "D2", "name": "記述2", "region": [0, 100, 100, 200],
             "max_score": 10, "aspect": 3},
        ],
        "total_display_region": [0, 0, 100, 50],
    }


@pytest.fixture
def descriptive_scores():
    """記述問題の得点"""
    return {
        "img001.jpg": {"D1": 5, "D2": 10},
        "img002.jpg": {"D1": 3, "D2": 6},
        "img003.jpg": {"D1": 5, "D2": 8},
        "img004.jpg": {"D1": 0, "D2": 2},
    }


# ============================================================
# _evaluate_correct_rate テスト
# ============================================================

class TestEvaluateCorrectRate:
    """正答率評価ラベルの境界値テスト"""

    def test_very_easy(self):
        assert _evaluate_correct_rate(100) == '◎ 易'
        assert _evaluate_correct_rate(80) == '◎ 易'

    def test_appropriate(self):
        assert _evaluate_correct_rate(79.9) == '○ 適正'
        assert _evaluate_correct_rate(60) == '○ 適正'

    def test_slightly_hard(self):
        assert _evaluate_correct_rate(59.9) == '△ やや難'
        assert _evaluate_correct_rate(40) == '△ やや難'

    def test_hard(self):
        assert _evaluate_correct_rate(39.9) == '▽ 難'
        assert _evaluate_correct_rate(20) == '▽ 難'

    def test_very_hard(self):
        assert _evaluate_correct_rate(19.9) == '× 極難'
        assert _evaluate_correct_rate(0) == '× 極難'


# ============================================================
# シート構成テスト
# ============================================================

class TestExamSummarySheets:
    """4シート構成の検証"""

    def test_sheet_names(self, template_path, mark2_result_path, tmp_path):
        """4シート名が正しい"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        assert wb.sheetnames == ["試験概要", "設問分析", "得点分布", "観点別統計"]

    def test_returns_stats_dict(self, template_path, mark2_result_path, tmp_path):
        """stats辞書の後方互換性"""
        out = str(tmp_path / "out.xlsx")
        stats = generate_exam_summary(template_path, mark2_result_path, out)
        assert isinstance(stats, dict)
        for key in ['受験者数', '満点', '平均点', '中央値', '最高点', '最低点', '標準偏差', '分散']:
            assert key in stats

    def test_stats_values(self, template_path, mark2_result_path, tmp_path):
        """基本統計量の値が正しい"""
        out = str(tmp_path / "out.xlsx")
        stats = generate_exam_summary(template_path, mark2_result_path, out)
        assert stats['受験者数'] == 4
        assert stats['満点'] == 10  # 2+3+5
        # img001=10, img002=2, img003=3, img004=0 → 平均=3.75
        assert stats['最高点'] == 10
        assert stats['最低点'] == 0


# ============================================================
# シート1: 試験概要
# ============================================================

class TestSheet1Overview:
    """試験概要シートの内容検証"""

    def test_title(self, template_path, mark2_result_path, tmp_path):
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["試験概要"]
        assert ws['A1'].value == '試験サマリーレポート'

    def test_basic_info_section(self, template_path, mark2_result_path, tmp_path):
        """基本情報セクションが存在し値が正しい"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["試験概要"]
        # A3 に ■ 基本情報
        assert '基本情報' in str(ws['A3'].value)
        # 受験者数を探す
        values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                  for r in range(4, 15)}
        assert '4 名' in str(values.get('受験者数', ''))

    def test_summary_stats_section(self, template_path, mark2_result_path, tmp_path):
        """要約統計量セクションが存在する"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["試験概要"]
        all_text = [str(ws.cell(row=r, column=1).value) for r in range(1, 25)]
        assert any('要約統計量' in t for t in all_text)
        assert any('平均点' in t for t in all_text)
        assert any('得点率' in t for t in all_text)


# ============================================================
# シート2: 設問分析
# ============================================================

class TestSheet2QuestionAnalysis:
    """設問分析シートの検証"""

    def test_title(self, template_path, mark2_result_path, tmp_path):
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["設問分析"]
        assert '設問分析' in str(ws['A1'].value)

    def test_headers(self, template_path, mark2_result_path, tmp_path):
        """ヘッダー行が正しい"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["設問分析"]
        expected = ['No.', '種別', '配点', '観点', '正答者数', '正答率(%)', '評価']
        actual = [ws.cell(row=3, column=c).value for c in range(1, 8)]
        assert actual == expected

    def test_mark_questions_present(self, template_path, mark2_result_path, tmp_path):
        """マーク式設問のデータが正しい"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["設問分析"]
        # 問1: 正答=1, img001・img002が正解 → 正答者2名, 正答率50%
        assert ws.cell(row=4, column=1).value == 1
        assert ws.cell(row=4, column=2).value == 'マーク'
        assert ws.cell(row=4, column=5).value == 2  # 正答者数
        assert ws.cell(row=4, column=6).value == 50.0  # 正答率

    def test_evaluation_column(self, template_path, mark2_result_path, tmp_path):
        """評価列に正しいラベルが入る"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["設問分析"]
        eval_val = ws.cell(row=4, column=7).value
        assert eval_val is not None
        # 50%→△ やや難
        assert '△' in eval_val

    def test_chart_image_exists(self, template_path, mark2_result_path, tmp_path):
        """正答率棒グラフ画像が挿入されている"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["設問分析"]
        assert len(ws._images) >= 1

    def test_descriptive_questions_in_analysis(
        self, template_path, mark2_result_path, tmp_path,
        descriptive_config, descriptive_scores
    ):
        """記述問題が設問分析に含まれる"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(
            template_path, mark2_result_path, out,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
        )
        wb = load_workbook(out)
        ws = wb["設問分析"]
        # マーク3問 + 記述2問 = 5行のデータ (row4-8)
        types = [ws.cell(row=r, column=2).value for r in range(4, 9)]
        assert types.count('マーク') == 3
        assert types.count('記述') == 2

    def test_descriptive_question_ids(
        self, template_path, mark2_result_path, tmp_path,
        descriptive_config, descriptive_scores
    ):
        """記述問題のIDが正しく表示される"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(
            template_path, mark2_result_path, out,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
        )
        wb = load_workbook(out)
        ws = wb["設問分析"]
        ids = [ws.cell(row=r, column=1).value for r in range(7, 9)]
        assert 'D1' in ids
        assert 'D2' in ids


# ============================================================
# シート3: 得点分布
# ============================================================

class TestSheet3Distribution:
    """得点分布シートの検証"""

    def test_title(self, template_path, mark2_result_path, tmp_path):
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["得点分布"]
        assert '得点分布' in str(ws['A1'].value)

    def test_headers(self, template_path, mark2_result_path, tmp_path):
        """ヘッダーに累積列が含まれる"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["得点分布"]
        headers = [ws.cell(row=3, column=c).value for c in range(1, 6)]
        assert '累積人数' in headers
        assert '累積割合(%)' in headers

    def test_cumulative_total(self, template_path, mark2_result_path, tmp_path):
        """累積人数の最終行が受験者数に一致"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["得点分布"]
        # 最終データ行を探す
        last_cum = None
        for r in range(4, 50):
            val = ws.cell(row=r, column=4).value
            if val is None:
                break
            last_cum = val
        assert last_cum == 4  # 受験者4名

    def test_histogram_image_exists(self, template_path, mark2_result_path, tmp_path):
        """ヒストグラム画像が挿入されている"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["得点分布"]
        assert len(ws._images) >= 1


# ============================================================
# シート4: 観点別統計
# ============================================================

class TestSheet4AspectStats:
    """観点別統計シートの検証"""

    def test_title(self, template_path, mark2_result_path, tmp_path):
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["観点別統計"]
        assert '観点別統計' in str(ws['A1'].value)

    def test_score_rate_column(self, template_path, mark2_result_path, tmp_path):
        """得点率(%)列が含まれる"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["観点別統計"]
        headers = [ws.cell(row=3, column=c).value for c in range(1, 8)]
        assert '得点率(%)' in headers

    def test_aspect_count(self, template_path, mark2_result_path, tmp_path):
        """観点数が正しい"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["観点別統計"]
        # データ行はrow4から始まり、空行までが観点データ
        rows = 0
        for r in range(4, 20):
            if ws.cell(row=r, column=2).value is None:  # 満点列が空ならデータ終了
                break
            rows += 1
        assert rows == 2  # 観点1, 観点2

    def test_descriptive_aspect_added(
        self, template_path, mark2_result_path, tmp_path,
        descriptive_config, descriptive_scores
    ):
        """記述問題で新しい観点が追加される"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(
            template_path, mark2_result_path, out,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
        )
        wb = load_workbook(out)
        ws = wb["観点別統計"]
        rows = 0
        for r in range(4, 20):
            if ws.cell(row=r, column=2).value is None:  # 満点列が空ならデータ終了
                break
            rows += 1
        # 観点1, 2(マーク) + 観点3(記述のみ) = 3
        assert rows == 3


# ============================================================
# 記述問題統合テスト
# ============================================================

class TestDescriptiveIntegration:
    """記述問題込みの統合検証"""

    def test_max_score_includes_descriptive(
        self, template_path, mark2_result_path, tmp_path,
        descriptive_config, descriptive_scores
    ):
        """満点にマーク+記述が含まれる"""
        out = str(tmp_path / "out.xlsx")
        stats = generate_exam_summary(
            template_path, mark2_result_path, out,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
        )
        # マーク: 2+3+5=10, 記述: 5+10=15 → 合計25
        assert stats['満点'] == 25

    def test_overview_shows_descriptive_count(
        self, template_path, mark2_result_path, tmp_path,
        descriptive_config, descriptive_scores
    ):
        """試験概要に記述式設問数が含まれる"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(
            template_path, mark2_result_path, out,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
        )
        wb = load_workbook(out)
        ws = wb["試験概要"]
        all_text = []
        for r in range(1, 25):
            for c in range(1, 5):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    all_text.append(str(v))
        assert any('記述式' in t for t in all_text)

    def test_no_descriptive_no_crash(self, template_path, mark2_result_path, tmp_path):
        """記述問題なしでもクラッシュしない"""
        out = str(tmp_path / "out.xlsx")
        stats = generate_exam_summary(template_path, mark2_result_path, out)
        assert stats['受験者数'] == 4
        assert Path(out).exists()


# ============================================================
# シート4: 相関係数・散布図行列テスト
# ============================================================

class TestSheet4CorrelationAndScatter:
    """観点別統計シートの相関係数テーブル・散布図行列"""

    def test_correlation_section_exists(self, template_path, mark2_result_path, tmp_path):
        """相関係数セクションが存在する"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["観点別統計"]
        all_text = [str(ws.cell(row=r, column=1).value) for r in range(1, 30)]
        assert any('相関係数' in t for t in all_text)

    def test_correlation_diagonal_is_one(self, template_path, mark2_result_path, tmp_path):
        """相関行列の対角要素が1.0"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["観点別統計"]
        # 相関行列のデータ開始行を探す
        corr_data_row = None
        for r in range(1, 30):
            v = str(ws.cell(row=r, column=1).value)
            if '相関係数' in v:
                corr_data_row = r + 2  # セクションヘッダ → 列ヘッダ → データ開始
                break
        assert corr_data_row is not None
        # 2観点: 対角はcol=2のrow[0]とcol=3のrow[1]
        assert ws.cell(row=corr_data_row, column=2).value == 1.0
        assert ws.cell(row=corr_data_row + 1, column=3).value == 1.0

    def test_scatter_matrix_image(
        self, template_path, mark2_result_path, tmp_path
    ):
        """散布図行列の画像が挿入されている"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(template_path, mark2_result_path, out)
        wb = load_workbook(out)
        ws = wb["観点別統計"]
        assert len(ws._images) >= 1

    def test_scatter_matrix_with_3_aspects(
        self, template_path, mark2_result_path, tmp_path,
        descriptive_config, descriptive_scores
    ):
        """3観点で散布図行列が生成される"""
        out = str(tmp_path / "out.xlsx")
        generate_exam_summary(
            template_path, mark2_result_path, out,
            descriptive_config=descriptive_config,
            descriptive_scores=descriptive_scores,
        )
        wb = load_workbook(out)
        ws = wb["観点別統計"]
        assert len(ws._images) >= 1
        # 相関行列が3x3
        corr_data_row = None
        for r in range(1, 40):
            v = str(ws.cell(row=r, column=1).value)
            if '相関係数' in v:
                corr_data_row = r + 2
                break
        assert corr_data_row is not None
        # 3行分のデータ
        for i in range(3):
            assert ws.cell(row=corr_data_row + i, column=1).value is not None
