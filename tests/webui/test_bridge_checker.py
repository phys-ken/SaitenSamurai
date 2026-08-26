"""webui マークチェックAPI（L1: 実 mark_checker・合成答案で検証）。"""
import sys
import time
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "webui"))
sys.path.insert(0, str(PROJECT_ROOT / "main_src"))
sys.path.insert(0, str(PROJECT_ROOT / "tests"))

from api.bridge import Bridge  # noqa: E402
from test_bridge_jobs import RecordingAdapter, _wait_job_done  # noqa: E402
from test_multidigit_image_e2e import (  # noqa: E402
    make_coord_xlsx, make_sheet, sym_positions,
)


@pytest.fixture
def recognized_bridge(tmp_path):
    """認識実行まで済ませた bridge（複数桁・答案2枚: 完答と 無マーク+ダブル）"""
    import cv2
    adapter = RecordingAdapter()
    bridge = Bridge(window_adapter=adapter)
    bridge.set_mode("mark_only", "multi_digit")

    coord = make_coord_xlsx(tmp_path / "coord.xlsx", 3)
    scans = tmp_path / "scans"
    scans.mkdir()
    s1 = {q: p for q, p in zip([1, 2, 3], sym_positions('-24'))}
    # s2: 行1 無マーク / 行2 ダブルマーク / 行3 正常
    s2 = {2: [sym_positions('2')[0], sym_positions('8')[0]],
          3: sym_positions('4')[0]}
    for name, filled in [("s1.png", s1), ("s2.png", s2)]:
        cv2.imwrite(str(scans / name), make_sheet(filled, with_markers=True))

    adapter.folder_returns = [str(scans)]
    adapter.file_returns = [str(coord)]
    bridge.select_image_folder()
    bridge.set_skip_questions(0)   # フォルダ選択は完全クリアするので後から設定
    bridge.select_coord_file()
    bridge.run_recognition()
    _wait_job_done(bridge)
    return bridge


class TestOpenAndFilter:
    def test_open_summarizes_categories(self, recognized_bridge):
        b = recognized_bridge
        res = b.open_mark_checker()
        assert res["ok"]
        c = res["state"]["checker"]
        assert c["total"] == 6            # 3行 × 2枚
        assert c["error_count"] == 2      # 無マーク1 + ダブル1
        names = {cat["name"]: cat["count"] for cat in c["categories"]}
        assert names["ノーマーク"] == 1 and names["複数マーク"] == 1

    def test_filter_errors_only(self, recognized_bridge):
        b = recognized_bridge
        b.open_mark_checker()
        res = b.get_checker_entries(category="__errors__")
        assert res["total"] == 2
        cats = {i["category"] for i in res["items"]}
        assert cats == {"ノーマーク", "複数マーク"}

    def test_entry_image_data_url(self, recognized_bridge):
        b = recognized_bridge
        b.open_mark_checker()
        first = b.get_checker_entries()["items"][0]
        res = b.get_entry_image(first["id"])
        assert res["ok"] and res["data_url"].startswith("data:image/png;base64,")
        assert len(res["data_url"]) > 200


class TestCorrections:
    def test_symbol_validation(self, recognized_bridge):
        b = recognized_bridge
        b.open_mark_checker()
        item = b.get_checker_entries(category="ノーマーク")["items"][0]
        assert not b.set_correction(item["id"], "x")["ok"]
        res = b.set_correction(item["id"], "A")   # 大文字→小文字受理
        assert res["ok"] and res["entry"]["after"] == "a"
        assert res["state"]["checker"]["corrected"] == 1

    def test_apply_writes_xlsx_and_reloads(self, recognized_bridge):
        import openpyxl
        b = recognized_bridge
        b.open_mark_checker()
        nomark = b.get_checker_entries(category="ノーマーク")["items"][0]
        double = b.get_checker_entries(category="複数マーク")["items"][0]
        b.set_correction(nomark["id"], "-")
        b.set_correction(double["id"], "2")
        res = b.apply_corrections()
        assert res["ok"] and res["applied"] == 2 and Path(res["backup"]).exists()

        # xlsx 実体が更新されている
        ws = openpyxl.load_workbook(b.state["omr_result"])["Sheet1"]
        rows = {ws.cell(r, 2).value: [str(ws.cell(r, c).value) for c in (3, 4, 5)]
                for r in (3, 4)}
        # 注: 数字の訂正はセルに数値型で入る（訂正CSV経由の tk 版と同じ挙動。
        # 採点側は str() 正規化するため実害なし）→ str 比較で固定
        assert rows["s2.png"] == ['-', '2', '4']
        # 反映後の再読込ではエラー0件
        assert b.state["checker"]["error_count"] == 0

    def test_corrections_persist_across_reopen(self, recognized_bridge):
        """訂正CSVは tk 互換 — 閉じて開き直しても訂正が残る"""
        b = recognized_bridge
        b.open_mark_checker()
        item = b.get_checker_entries(category="ノーマーク")["items"][0]
        b.set_correction(item["id"], "9")
        b.close_mark_checker()
        b.open_mark_checker()
        again = [i for i in b.get_checker_entries(category="ノーマーク")["items"]]
        assert again[0]["after"] == "9"
