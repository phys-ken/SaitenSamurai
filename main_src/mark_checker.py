"""
mark_checker.py - エラー検出・修正補助モジュール

マークシート読み取り結果のエラー検出（未マーク・ダブルマーク・不正値）と、
エラー修正CSV⇔Excel間のデータ連携、画像トリミング表示機能を提供する。

v3.9 高速化:
- CorrectedImageCache: 補正済み画像のメモリキャッシュ（同一画像の重複処理を排除）
- crop_from_corrected_image: キャッシュ済み補正画像からの高速クロップ
- 遅延CSV保存 / バックグラウンド先読みは gui_components.py 側で実装
"""

import logging
import re
import shutil
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

import cv2
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from PIL import Image, ImageTk

from constants import (
    DEFAULT_BACKUP_FOLDER,
    DEFAULT_SCALE_FACTOR,
    DEFAULT_EXPAND_FACTOR,
    DEFAULT_EXPAND_FACTOR_Y,
    MAX_DISPLAY_WIDTH,
    MAX_DISPLAY_HEIGHT,
    MARK2_BASE_WIDTH,
    MARK2_BASE_HEIGHT,
    ERROR_TYPE_NO_MARK,
    ERROR_TYPE_DOUBLE_MARK,
    ERROR_TYPE_INVALID,
)
from omr_engine import detect_corner_markers


# ========================================
# 補正済み画像キャッシュ（v3.9 高速化）
# ========================================

class CorrectedImageCache:
    """同一画像に対するディスクI/O・マーカー検出・射影変換を1回に削減するキャッシュ。
    
    マークチェック時、同じ画像ファイルに複数のエラーがある場合、
    補正済み画像をメモリに保持して再利用する。
    max_size=2 で現在の画像+先読み画像を保持。
    """
    
    def __init__(self, max_size=2):
        self._cache = {}  # {filename: corrected_img (numpy)}
        self._order = []  # LRU 順序管理
        self._max_size = max_size
    
    def get(self, filename):
        """キャッシュから補正済み画像を取得。なければNone。"""
        if filename in self._cache:
            # LRU更新
            self._order.remove(filename)
            self._order.append(filename)
            return self._cache[filename]
        return None
    
    def put(self, filename, corrected_img):
        """補正済み画像をキャッシュに格納。"""
        if filename in self._cache:
            self._order.remove(filename)
        elif len(self._cache) >= self._max_size:
            # 最も古いエントリを削除
            oldest = self._order.pop(0)
            del self._cache[oldest]
        self._cache[filename] = corrected_img
        self._order.append(filename)
    
    def has(self, filename):
        """キャッシュに指定ファイルがあるか"""
        return filename in self._cache
    
    def clear(self):
        """キャッシュをクリア"""
        self._cache.clear()
        self._order.clear()
    
    @property
    def size(self):
        return len(self._cache)


def _load_and_correct_image(image_path):
    """画像をディスクから読み込み、射影補正を適用して返す。
    
    Returns:
        corrected_img: 補正済みnumpy配列 (BGR)
    """
    img_array = np.fromfile(str(image_path), dtype=np.uint8)
    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    
    if img is None:
        raise FileNotFoundError(f"画像が読み込めません: {image_path}")
    
    img_height, img_width = img.shape[:2]
    res_scale_x = img_width / MARK2_BASE_WIDTH
    res_scale_y = img_height / MARK2_BASE_HEIGHT
    
    try:
        markers = detect_corner_markers(img, debug=False)
        
        xp1 = MARK2_BASE_WIDTH * (0.14 + 0.015)
        yp1 = MARK2_BASE_HEIGHT * (0.03 + 0.01)
        xp2 = MARK2_BASE_WIDTH * (0.83 + 0.015)
        yp2 = MARK2_BASE_HEIGHT * (0.03 + 0.01)
        xp3 = MARK2_BASE_WIDTH * (0.83 + 0.015)
        yp3 = MARK2_BASE_HEIGHT * (0.95 + 0.01)
        xp4 = MARK2_BASE_WIDTH * (0.14 + 0.015)
        yp4 = MARK2_BASE_HEIGHT * (0.95 + 0.01)
        
        src_points = np.float32(markers)
        dst_points = np.float32([
            [xp1 * res_scale_x, yp1 * res_scale_y],
            [xp2 * res_scale_x, yp2 * res_scale_y],
            [xp3 * res_scale_x, yp3 * res_scale_y],
            [xp4 * res_scale_x, yp4 * res_scale_y]
        ])
        
        transform_matrix = cv2.getPerspectiveTransform(src_points, dst_points)
        corrected_img = cv2.warpPerspective(img, transform_matrix, (img_width, img_height))
    except Exception:
        corrected_img = img
    
    return corrected_img


def crop_from_corrected_image(corrected_img, bbox, scale_factor=1.25,
                               expand_factor=1.3, expand_factor_y=1.0):
    """補正済み画像から指定領域をクロップしてPIL画像を返す（高速パス）。
    
    _load_and_correct_image() またはキャッシュから取得した corrected_img を受け取り、
    ベース座標系のbboxに基づいてクロップ・拡大・PIL変換を行う。
    ディスクI/O・マーカー検出・射影変換は含まない。
    
    Returns:
        (pil_img, crop_info)
        crop_info: {
            'crop_x': int, 'crop_y': int, 
            'scale_x': float, 'scale_y': float,
            'res_scale_x': float, 'res_scale_y': float
        }
    """
    img_height, img_width = corrected_img.shape[:2]
    res_scale_x = img_width / MARK2_BASE_WIDTH
    res_scale_y = img_height / MARK2_BASE_HEIGHT
    
    x, y, w, h = bbox
    x = int(x * res_scale_x)
    y = int(y * res_scale_y)
    w = int(w * res_scale_x)
    h = int(h * res_scale_y)
    
    center_x = x + w / 2
    center_y = y + h / 2
    expanded_w = w * expand_factor
    expanded_h = h * expand_factor * expand_factor_y
    expanded_x = center_x - expanded_w / 2
    expanded_y = center_y - expanded_h / 2
    
    x = int(expanded_x)
    y = int(expanded_y)
    w = int(expanded_w)
    h = int(expanded_h)
    
    x = max(0, min(x, img_width - 1))
    y = max(0, min(y, img_height - 1))
    w = min(w, img_width - x)
    h = min(h, img_height - y)
    
    cropped = corrected_img[y:y+h, x:x+w]
    
    new_width = int(w * scale_factor)
    new_height = int(h * scale_factor)
    if new_width < 1 or new_height < 1:
        new_width = max(new_width, 1)
        new_height = max(new_height, 1)
    scaled = cv2.resize(cropped, (new_width, new_height), interpolation=cv2.INTER_LINEAR)
    
    rgb_img = cv2.cvtColor(scaled, cv2.COLOR_BGR2RGB)
    rgb_img = cv2.cvtColor(scaled, cv2.COLOR_BGR2RGB)
    
    crop_info = {
        'crop_x': x,
        'crop_y': y,
        'scale_x': new_width / w if w > 0 else 1.0,
        'scale_y': new_height / h if h > 0 else 1.0,
        'res_scale_x': res_scale_x,
        'res_scale_y': res_scale_y,
    }
    
    return Image.fromarray(rgb_img), crop_info


def create_backup_checker(xlsx_path, backup_folder=DEFAULT_BACKUP_FOLDER):
    """xlsxファイルのバックアップを作成"""
    xlsx_path = Path(xlsx_path)
    backup_dir = xlsx_path.parent / backup_folder
    backup_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"{timestamp}_{xlsx_path.name}"
    backup_path = backup_dir / backup_filename
    
    shutil.copy2(xlsx_path, backup_path)
    logger.info("バックアップ作成: %s", backup_path)
    return backup_path


def update_xlsx_from_csv_checker(xlsx_path, error_csv_path):
    """エラー修正CSVの内容でxlsxファイルを更新"""
    error_df = pd.read_csv(error_csv_path, encoding='utf-8-sig')
    
    updates = error_df[
        error_df['after'].notna() & 
        (error_df['after'] != '') & 
        (error_df['after'] != 'skip')
    ]
    
    if len(updates) == 0:
        logger.info("更新対象がありません")
        return 0
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['Sheet1']
    
    second_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    try:
        file_col_idx = first_row.index('File') + 1
    except ValueError:
        logger.error("'File'列が見つかりません")
        wb.close()
        return 0
    
    question_col_map = {}
    for col_idx, value in enumerate(second_row, start=1):
        q_num = None
        try:
            if isinstance(value, (int, float)):
                if value == int(value):
                    q_num = int(value)
            elif isinstance(value, str):
                # 文字列の場合、"14" や "14.0" を考慮
                try:
                    f_val = float(value)
                    if f_val == int(f_val):
                        q_num = int(f_val)
                except ValueError:
                    pass
        except Exception:
            pass
            
        if q_num is not None and q_num >= 1:
            question_col_map[q_num] = col_idx
    
    file_to_row = {}
    for row_idx in range(3, ws.max_row + 1):
        filename = ws.cell(row=row_idx, column=file_col_idx).value
        if filename:
            file_to_row[filename] = row_idx
    
    update_count = 0
    correction_font = Font(bold=True, color="4B0082")  # 太字・濃い紫
    for _, update_row in updates.iterrows():
        filename = update_row['filename']
        question_no = int(update_row['question_no'])
        after_value = update_row['after']
        
        if filename not in file_to_row:
            logger.warning("ファイル名が見つかりません: %s", filename)
            continue
        
        row_idx = file_to_row[filename]
        
        if question_no not in question_col_map:
            logger.warning("問題番号%sの列が見つかりません", question_no)
            continue
        
        col_idx = question_col_map[question_no]
        cell = ws.cell(row=row_idx, column=col_idx)
        
        if after_value == '-1':
            cell.value = -1
        else:
            try:
                cell.value = int(after_value)
            except ValueError:
                cell.value = after_value
        
        cell.font = correction_font  # 修正セル: 太字+濃い紫
        
        update_count += 1
    
    wb.save(xlsx_path)
    wb.close()
    
    logger.info("xlsx更新完了: %d件", update_count)
    return update_count


def apply_corrections_checker(xlsx_path, error_csv_path, backup_folder=DEFAULT_BACKUP_FOLDER):
    """バックアップ作成→xlsx更新の一連処理"""
    backup_path = create_backup_checker(xlsx_path, backup_folder)
    update_count = update_xlsx_from_csv_checker(xlsx_path, error_csv_path)
    return backup_path, update_count


def detect_errors_checker(xlsx_path, output_csv_path, registered_questions=None):
    """Excelファイルから問題の回答エラーを検出
    
    Args:
        xlsx_path: Mark2結果Excelのパス
        output_csv_path: エラーCSV出力先
        registered_questions: 採点対象の問題番号セット。
            Noneの場合は全問題をチェック。
            指定された場合はその問題番号のみチェックする。
    """
    xlsx_path = Path(xlsx_path)
    
    if not xlsx_path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {xlsx_path}")
    
    logger.info("Excelファイルから読み込み: %s", xlsx_path.name)
    
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['Sheet1']
    
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    labels = [ws.cell(2, i).value for i in range(1, ws.max_column + 1)]
    
    column_names = []
    for header, label in zip(headers, labels):
        if label is None:
            column_names.append(header)
        else:
            column_names.append(label)
    
    data_rows = []
    for row_idx in range(3, ws.max_row + 1):
        row_data = [ws.cell(row_idx, i).value for i in range(1, ws.max_column + 1)]
        data_rows.append(row_data)
    
    wb.close()
    
    df = pd.DataFrame(data_rows, columns=column_names)
    
    logger.info("データ確認:")
    logger.info("  総行数: %d行（画像枚数）", len(df))
    
    available_questions = []
    for col_name in df.columns:
        # カラム名が数値（int/float）の場合もあるため、文字列に変換してチェック
        col_str = str(col_name).strip()
        # "1.0" のようなケースに対応するため、一度floatにしてからint判定
        try:
            f_val = float(col_str)
            if f_val == int(f_val):
                q_num = int(f_val)
                if q_num >= 1:
                    available_questions.append(col_name)
        except ValueError:
            pass
    
    available_questions.sort(key=lambda x: int(x))
    
    # 登録済み問題でフィルタリング
    if registered_questions is not None:
        reg_set = set(int(q) for q in registered_questions)
        before_count = len(available_questions)
        available_questions = [q for q in available_questions if int(q) in reg_set]
        skipped = before_count - len(available_questions)
        if skipped > 0:
            logger.info("  正答未登録の%d問をスキップ", skipped)
    
    logger.info("  チェック対象: %d問", len(available_questions))
    
    if available_questions:
        logger.info("  問題範囲: %s - %s", available_questions[0], available_questions[-1])
    
    errors = []
    
    for idx, row in df.iterrows():
        filename = row['File']
        
        for q_num in available_questions:
            value = row[q_num]
            error_type = None
            before_value = ""
            
            if pd.isna(value):  # type: ignore[arg-type]
                error_type = ERROR_TYPE_NO_MARK
                before_value = ''
            else:
                value_str = str(value).strip()
                before_value = value_str
                
                if value_str == '-1' or value_str == '-1.0':
                    continue
                
                if ';' in value_str or '|' in value_str:
                    error_type = ERROR_TYPE_DOUBLE_MARK
                else:
                    if '.' in value_str:
                        try:
                            num_value = float(value_str)
                            if num_value == int(num_value):
                                value_str = str(int(num_value))
                        except ValueError:
                            error_type = ERROR_TYPE_INVALID
                    
                    if error_type is None:
                        if re.match(r'^\d$', value_str):
                            continue
                        else:
                            error_type = ERROR_TYPE_INVALID
            
            if error_type:
                errors.append({
                    'filename': filename,
                    'question_no': int(q_num),
                    'before': before_value,
                    'after': '',
                    'error_type': error_type
                })
    
    if len(errors) > 0:
        error_df = pd.DataFrame(errors)
    else:
        error_df = pd.DataFrame(columns=['filename', 'question_no', 'before', 'after', 'error_type'])
    
    error_df.to_csv(output_csv_path, index=False, encoding='utf-8-sig')
    
    logger.info("エラー検出完了: %d件", len(errors))
    logger.info("保存先: %s", output_csv_path)
    
    if len(errors) > 0:
        logger.info("【エラータイプ別集計】")
        logger.info("%s", error_df['error_type'].value_counts())
    
    return len(errors)


def load_errors_checker(error_csv_path):
    """エラーCSVを読み込み"""
    error_csv_path = Path(error_csv_path)
    if error_csv_path.exists():
        try:
            df = pd.read_csv(error_csv_path, encoding='utf-8-sig')
            if len(df) == 0:
                return pd.DataFrame(columns=['filename', 'question_no', 'before', 'after', 'error_type'])
            # after 列は文字列 'skip' や空文字が混在するため object 型を保証
            if 'after' in df.columns and df['after'].dtype != object:
                df['after'] = df['after'].astype('object')
            return df
        except pd.errors.EmptyDataError:
            return pd.DataFrame(columns=['filename', 'question_no', 'before', 'after', 'error_type'])
    return pd.DataFrame(columns=['filename', 'question_no', 'before', 'after', 'error_type'])


def save_errors_checker(error_df, error_csv_path):
    """エラーCSVを保存"""
    error_df.to_csv(error_csv_path, index=False, encoding='utf-8-sig')


def load_coordinates_csv_checker(csv_path):
    """coordinates.csvを読み込み（選択肢ごとに展開）"""
    df = pd.read_csv(csv_path, encoding='utf-8')
    
    # 既存のcolumnsに choice, x, y, width, height が無ければ展開処理を行う
    if 'choice' not in df.columns and 'mark_coords' in df.columns:
        expanded_rows = []
        for _, row in df.iterrows():
            mark_coords_str = row['mark_coords']
            if pd.isna(mark_coords_str):
                continue
            
            # coords format: "x;y;w;h|x;y;w;h|..."
            choices = str(mark_coords_str).split('|')
            for idx, c_str in enumerate(choices):
                try:
                    parts = c_str.split(';')
                    if len(parts) == 4:
                        x, y, w, h = map(int, parts)
                        new_row = row.to_dict()
                        new_row['choice'] = idx
                        new_row['x'] = x
                        new_row['y'] = y
                        new_row['width'] = w
                        new_row['height'] = h
                        expanded_rows.append(new_row)
                except ValueError:
                    pass
        
        if expanded_rows:
            return pd.DataFrame(expanded_rows)
        else:
            # 展開できなかった場合は元のDFを返す（エラー回避）
            return df
    
    return df


def get_bbox_for_question_checker(coords_df, image_filename, question_no):
    """指定された画像と問題番号のバウンディングボックスを取得"""
    row = coords_df[(coords_df['image_path'] == image_filename) & 
                     (coords_df['question_no'] == question_no)]
    
    if row.empty:
        return None
    
    bbox_str = row.iloc[0]['choices_bbox']
    parts = bbox_str.split(';')
    
    if len(parts) != 4:
        return None
    
    x, y, w, h = map(int, parts)
    return (x, y, w, h)


def crop_and_scale_image_checker(image_path, bbox, scale_factor=3.0, expand_factor=1.3, expand_factor_y=1.0):
    """画像を元解像度で透視補正してから指定座標で切り出し、拡大
    
    座標CSVは補正済み画像(595x842)のベース座標系で記録されているため、
    元画像に透視補正を元解像度のまま適用し、ベース座標をスケーリングして切り出す。
    これにより高解像度を維持しつつ正確な位置合わせを実現する。
    
    Note: キャッシュ利用時は get_display_image_checker() + CorrectedImageCache を使用。
          この関数は後方互換のために維持。
    
    Args:
        image_path: 元画像のパス
        bbox: (x, y, w, h) ベース座標系(595x842)での切り出し範囲
        scale_factor: 最終表示用の拡大率
        expand_factor: 幅・高さ共通の拡張率
        expand_factor_y: 高さ方向の追加拡張率（上下マージン確保用）
    """
    corrected_img = _load_and_correct_image(image_path)
    return crop_from_corrected_image(corrected_img, bbox, scale_factor, expand_factor, expand_factor_y)


def get_display_image_checker(coords_df, image_folder, image_filename, question_no, 
                      scale_factor=DEFAULT_SCALE_FACTOR, expand_factor=DEFAULT_EXPAND_FACTOR,
                      expand_factor_y=DEFAULT_EXPAND_FACTOR_Y, cache=None):
    """表示用の画像を取得
    
    Returns:
        (pil_img, crop_info) or (None, None)
    
    Args:
        cache: CorrectedImageCache インスタンス。指定時はキャッシュを利用して
               ディスクI/O・マーカー検出・射影変換をスキップする。
    """
    bbox = get_bbox_for_question_checker(coords_df, image_filename, question_no)
    
    if bbox is None:
        logger.warning("座標が見つかりません: %s, Q%s", image_filename, question_no)
        return None, None
    
    image_path = Path(image_folder) / image_filename
    
    if not image_path.exists():
        logger.warning("画像ファイルが見つかりません: %s", image_path)
        return None, None
    
    try:
        if cache is not None:
            # キャッシュ高速パス
            corrected_img = cache.get(image_filename)
            if corrected_img is None:
                corrected_img = _load_and_correct_image(image_path)
                cache.put(image_filename, corrected_img)
            pil_img, crop_info = crop_from_corrected_image(
                corrected_img, bbox, scale_factor, expand_factor, expand_factor_y
            )
        else:
            # 従来パス（後方互換）
            pil_img, crop_info = crop_and_scale_image_checker(
                image_path, bbox, scale_factor, expand_factor, expand_factor_y
            )
        return pil_img, crop_info
    except Exception as e:
        logger.error("画像処理エラー: %s", e)
        return None, None


def fit_image_to_display(pil_img, max_width=MAX_DISPLAY_WIDTH, max_height=MAX_DISPLAY_HEIGHT):
    """PIL画像を最大表示サイズに収まるようリサイズ（アスペクト比維持）
    
    画像が max_width / max_height を超える場合のみ縮小する。
    拡大は行わない。
    
    Returns:
        リサイズ後の PIL.Image（超過していなければ元の画像をそのまま返す）
    """
    if pil_img.width <= max_width and pil_img.height <= max_height:
        return pil_img
    
    ratio_w = max_width / pil_img.width
    ratio_h = max_height / pil_img.height
    ratio = min(ratio_w, ratio_h)
    
    new_width = max(int(pil_img.width * ratio), 1)
    new_height = max(int(pil_img.height * ratio), 1)
    
    return pil_img.resize((new_width, new_height), Image.LANCZOS)


def pil_to_imagetk_checker(pil_img):
    """PIL.ImageをImageTk.PhotoImageに変換"""
    return ImageTk.PhotoImage(pil_img)
