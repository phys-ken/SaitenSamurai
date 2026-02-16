#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
test_name_trimmer.py — name_trimmer モジュール + generate_student_summary 氏名画像拡張テスト

テスト内容:
    Part 1: name_trimmer モジュール単体テスト
        1. モジュールインポート・クラスインスタンス化
        2. get_image_files() のファイル取得
        3. trim_images() の一括トリミング（ダミー画像）
        4. trim_images() のリサイズ動作
        5. trim_images() のクランプ動作（座標が画像外にはみ出す場合）
        6. trim_images() の空フォルダ対応
        7. NameTrimmer の cleanup 動作

    Part 2: generate_student_summary の name_images 拡張テスト
        8. name_images=None で従来通りの動作（回帰テスト）
        9. name_images 付きで氏名欄列が追加されること
        10. 氏名欄列に画像が埋め込まれていること
        11. freeze_panes の変化（C3 → D3）
        12. 列構成の正確性
"""

import os
import sys
import shutil
import tempfile
import unittest
from pathlib import Path

import numpy as np
import pandas as pd

# main_src をパスに追加
sys.path.insert(0, str(Path(__file__).parent.parent / "main_src"))

from PIL import Image


class TestNameTrimmerImport(unittest.TestCase):
    """モジュールのインポートとクラスインスタンス化テスト"""

    def test_01_import_module(self):
        """name_trimmer モジュールがインポートできること"""
        import name_trimmer
        self.assertTrue(hasattr(name_trimmer, 'select_region_on_image'))
        self.assertTrue(hasattr(name_trimmer, 'trim_images'))
        self.assertTrue(hasattr(name_trimmer, 'NameTrimmer'))
        self.assertTrue(hasattr(name_trimmer, 'get_image_files'))

    def test_02_instantiate_name_trimmer(self):
        """NameTrimmer クラスがインスタンス化できること"""
        from name_trimmer import NameTrimmer
        trimmer = NameTrimmer()
        self.assertIsNone(trimmer.last_trim_rect)
        self.assertIsNone(trimmer._temp_dir)

    def test_03_constants_defined(self):
        """定数が正しく定義されていること"""
        from name_trimmer import (
            IMAGE_EXTENSIONS, DEFAULT_MAX_HEIGHT,
            MAX_DISPLAY_WIDTH, MAX_DISPLAY_HEIGHT
        )
        self.assertIn('.jpg', IMAGE_EXTENSIONS)
        self.assertIn('.png', IMAGE_EXTENSIONS)
        self.assertEqual(DEFAULT_MAX_HEIGHT, 50)
        self.assertEqual(MAX_DISPLAY_WIDTH, 700)
        self.assertEqual(MAX_DISPLAY_HEIGHT, 700)


class TestGetImageFiles(unittest.TestCase):
    """get_image_files() のテスト"""

    def setUp(self):
        self.test_dir = tempfile.mkdtemp(prefix="test_img_files_")
        # ダミー画像ファイルを作成
        for name in ["a.jpg", "b.png", "c.txt", "d.jpeg", "e.bmp"]:
            filepath = Path(self.test_dir) / name
            if name.endswith('.txt'):
                filepath.write_text("not an image")
            else:
                img = Image.new('RGB', (10, 10), color='red')
                img.save(str(filepath))

    def tearDown(self):
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def test_01_returns_only_images(self):
        """画像ファイルのみを返すこと"""
        from name_trimmer import get_image_files
        files = get_image_files(self.test_dir)
        filenames = [Path(f).name for f in files]
        self.assertIn("a.jpg", filenames)
        self.assertIn("b.png", filenames)
        self.assertIn("d.jpeg", filenames)
        self.assertIn("e.bmp", filenames)
        self.assertNotIn("c.txt", filenames)

    def test_02_returns_sorted(self):
        """ソート済みで返すこと"""
        from name_trimmer import get_image_files
        files = get_image_files(self.test_dir)
        filenames = [Path(f).name for f in files]
        self.assertEqual(filenames, sorted(filenames))

    def test_03_empty_folder(self):
        """空フォルダの場合、空リストを返すこと"""
        from name_trimmer import get_image_files
        empty_dir = tempfile.mkdtemp(prefix="test_empty_")
        try:
            files = get_image_files(empty_dir)
            self.assertEqual(files, [])
        finally:
            shutil.rmtree(empty_dir, ignore_errors=True)

    def test_04_nonexistent_folder(self):
        """存在しないフォルダの場合、空リストを返すこと"""
        from name_trimmer import get_image_files
        files = get_image_files("/nonexistent/path")
        self.assertEqual(files, [])


class TestTrimImages(unittest.TestCase):
    """trim_images() のテスト"""

    def setUp(self):
        self.input_dir = tempfile.mkdtemp(prefix="test_trim_input_")
        self.output_dir = tempfile.mkdtemp(prefix="test_trim_output_")
        # 100x200 のダミー画像を3枚作成
        for i in range(3):
            img = Image.new('RGB', (100, 200), color=(i * 80, 100, 50))
            img.save(str(Path(self.input_dir) / f"test_{i:03d}.jpg"))

    def tearDown(self):
        shutil.rmtree(self.input_dir, ignore_errors=True)
        shutil.rmtree(self.output_dir, ignore_errors=True)

    def test_01_basic_trim(self):
        """基本的なトリミングが正しく動作すること"""
        from name_trimmer import trim_images
        # 画像の左上 (10,20) から (80,60) の領域をトリミング
        saved = trim_images(self.input_dir, (10, 20, 80, 60), self.output_dir, max_height=200)
        self.assertEqual(len(saved), 3)
        # トリミング後の画像サイズを確認
        for f in saved:
            img = Image.open(f)
            self.assertEqual(img.size, (70, 40))  # 80-10=70, 60-20=40
            img.close()

    def test_02_trim_with_resize(self):
        """max_heightを超える場合にリサイズされること"""
        from name_trimmer import trim_images
        # 画像の (0,0)-(100,200) をトリミング → 高さ200px。max_height=50でリサイズ
        saved = trim_images(self.input_dir, (0, 0, 100, 200), self.output_dir, max_height=50)
        self.assertEqual(len(saved), 3)
        for f in saved:
            img = Image.open(f)
            self.assertEqual(img.height, 50)  # リサイズ後は50px
            # アスペクト比保持: 100/200 * 50 = 25
            self.assertEqual(img.width, 25)
            img.close()

    def test_03_trim_no_resize_when_within_limit(self):
        """max_height以下の場合はリサイズされないこと"""
        from name_trimmer import trim_images
        # (10,20)-(80,50) → 高さ30px。max_height=50なのでリサイズ不要
        saved = trim_images(self.input_dir, (10, 20, 80, 50), self.output_dir, max_height=50)
        self.assertEqual(len(saved), 3)
        for f in saved:
            img = Image.open(f)
            self.assertEqual(img.size, (70, 30))
            img.close()

    def test_04_clamp_coordinates(self):
        """座標が画像範囲外にはみ出す場合にクランプされること"""
        from name_trimmer import trim_images
        # 画像は100x200。座標(-10, -20, 150, 250) → (0,0,100,200)にクランプ
        saved = trim_images(self.input_dir, (-10, -20, 150, 250), self.output_dir, max_height=200)
        self.assertEqual(len(saved), 3)
        for f in saved:
            img = Image.open(f)
            self.assertEqual(img.size, (100, 200))
            img.close()

    def test_05_empty_input_folder(self):
        """空フォルダの場合は空リストを返すこと"""
        from name_trimmer import trim_images
        empty_dir = tempfile.mkdtemp(prefix="test_empty_input_")
        try:
            saved = trim_images(empty_dir, (0, 0, 50, 50), self.output_dir)
            self.assertEqual(saved, [])
        finally:
            shutil.rmtree(empty_dir, ignore_errors=True)

    def test_06_output_folder_recreated(self):
        """出力フォルダが既存の場合、クリアして再作成されること"""
        from name_trimmer import trim_images
        # 先に出力先に適当なファイルを置く
        dummy_file = Path(self.output_dir) / "dummy.txt"
        dummy_file.write_text("should be removed")
        self.assertTrue(dummy_file.exists())

        saved = trim_images(self.input_dir, (10, 20, 80, 60), self.output_dir)
        self.assertEqual(len(saved), 3)
        # dummy.txt は消えているべき
        self.assertFalse(dummy_file.exists())

    def test_07_filenames_preserved(self):
        """トリミング後のファイル名が元画像と同じであること"""
        from name_trimmer import trim_images
        saved = trim_images(self.input_dir, (10, 20, 80, 60), self.output_dir)
        saved_names = sorted([Path(f).name for f in saved])
        expected_names = sorted(["test_000.jpg", "test_001.jpg", "test_002.jpg"])
        self.assertEqual(saved_names, expected_names)


class TestNameTrimmerCleanup(unittest.TestCase):
    """NameTrimmer.cleanup() のテスト"""

    def test_01_cleanup_removes_temp_dir(self):
        """cleanup() で一時ディレクトリが削除されること"""
        from name_trimmer import NameTrimmer
        trimmer = NameTrimmer()
        temp_dir = tempfile.mkdtemp(prefix="test_cleanup_")
        trimmer._temp_dir = temp_dir
        self.assertTrue(Path(temp_dir).exists())

        trimmer.cleanup()
        self.assertFalse(Path(temp_dir).exists())
        self.assertIsNone(trimmer._temp_dir)

    def test_02_cleanup_when_no_temp(self):
        """一時ディレクトリ未設定時に cleanup() がエラーにならないこと"""
        from name_trimmer import NameTrimmer
        trimmer = NameTrimmer()
        # 例外が発生しないことを確認
        trimmer.cleanup()
        self.assertIsNone(trimmer._temp_dir)


class TestTrimImagesHighresFallback(unittest.TestCase):
    """高解像度モードでマーカー検出に失敗した場合のフォールバックテスト"""

    def setUp(self):
        self.input_dir = tempfile.mkdtemp(prefix="test_trim_input_")
        self.orig_dir = tempfile.mkdtemp(prefix="test_trim_orig_")
        self.output_dir = tempfile.mkdtemp(prefix="test_trim_output_")
        # 100x200 のダミー画像を3枚作成 (input = 00_Processing相当)
        for i in range(3):
            img = Image.new('RGB', (100, 200), color=(i * 80, 100, 50))
            img.save(str(Path(self.input_dir) / f"test_{i:03d}.jpg"))
        # original_image_folder にも同名ファイルを配置 (マーカーなし画像)
        for i in range(3):
            img = Image.new('RGB', (400, 800), color=(i * 60, 80, 30))
            img.save(str(Path(self.orig_dir) / f"test_{i:03d}.jpg"))

    def tearDown(self):
        shutil.rmtree(self.input_dir, ignore_errors=True)
        shutil.rmtree(self.orig_dir, ignore_errors=True)
        shutil.rmtree(self.output_dir, ignore_errors=True)

    def test_fallback_when_marker_detection_fails(self):
        """マーカー検出失敗時にフォールバック(直接crop)で画像が生成されること"""
        from name_trimmer import trim_images
        # original_image_folder を指定 → 高解像度パスが使われるが、
        # ダミー画像にはマーカーがないため detect_corner_markers が失敗する。
        # フォールバックとして 00_Processing (input_dir) から直接cropされるべき。
        saved = trim_images(
            self.input_dir, (10, 20, 80, 60), self.output_dir,
            max_height=200, original_image_folder=self.orig_dir,
        )
        self.assertEqual(len(saved), 3, "マーカー検出失敗時もフォールバックで全画像処理されるべき")
        for f in saved:
            img = Image.open(f)
            # フォールバックは input_dir (100x200) からのcrop: 70x40
            self.assertEqual(img.size, (70, 40))
            img.close()

    def test_no_original_uses_direct_crop(self):
        """original_image_folder=None の場合は直接cropが使われること"""
        from name_trimmer import trim_images
        saved = trim_images(
            self.input_dir, (10, 20, 80, 60), self.output_dir,
            max_height=200, original_image_folder=None,
        )
        self.assertEqual(len(saved), 3)
        for f in saved:
            img = Image.open(f)
            self.assertEqual(img.size, (70, 40))
            img.close()


# ============================================================
# Part 2: generate_student_summary の name_images 拡張テスト
# ============================================================

def _make_dummy_mark2_files(tmpdir, n_students=5, n_questions=3, skip_questions=2, seed=42):
    """
    generate_student_summary テスト用のダミー Mark2形式ファイルを生成。
    test_ctt_integration.py の make_dummy_mark2_files() を参考に簡素化。
    """
    rng = np.random.default_rng(seed)

    # テンプレート (answer_key.xlsx)
    template_rows = []
    for i in range(1, n_questions + 1):
        template_rows.append({
            "問題番号": i,
            "正答": rng.integers(1, 6),
            "配点": rng.choice([2, 3]),
            "観点": rng.choice([1, 2]),
        })
    template_df = pd.DataFrame(template_rows)
    template_path = os.path.join(tmpdir, "answer_key.xlsx")
    template_df.to_excel(template_path, index=False)

    # Mark2結果 Excel
    total_cols = skip_questions + n_questions
    header_row = ["No", "File"] + [str(i) for i in range(1, total_cols + 1)]
    name_row = [np.nan, np.nan] + ["学年", "クラス"][:skip_questions] + [str(i) for i in range(1, n_questions + 1)]

    data_rows = []
    for s in range(n_students):
        row = [s + 1, f"page_{s+1:03d}.jpg"]
        row += [1, rng.integers(1, 4)][:skip_questions]
        for q in range(n_questions):
            row.append(int(rng.integers(1, 6)))
        data_rows.append(row)

    all_rows = [header_row, name_row] + data_rows
    result_df = pd.DataFrame(all_rows)
    result_path = os.path.join(tmpdir, "Mark2-Result.xlsx")
    result_df.to_excel(result_path, index=False, header=False)

    return template_path, result_path


def _make_dummy_name_images(tmpdir, n_students=5):
    """テスト用のダミー氏名欄トリミング画像を生成し、辞書を返す。"""
    name_images = {}
    img_dir = os.path.join(tmpdir, "name_trim_tmp")
    os.makedirs(img_dir, exist_ok=True)
    for i in range(n_students):
        filename = f"page_{i+1:03d}.jpg"
        img = Image.new('RGB', (120, 30), color=(200, 220, 255))
        img_path = os.path.join(img_dir, filename)
        img.save(img_path)
        name_images[filename] = img_path
    return name_images


class TestGenerateStudentSummaryWithNameImages(unittest.TestCase):
    """generate_student_summary の name_images 拡張テスト"""

    @classmethod
    def setUpClass(cls):
        """テストデータを一度だけ生成"""
        cls.test_dir = tempfile.mkdtemp(prefix="test_summary_name_")
        cls.template_path, cls.result_path = _make_dummy_mark2_files(
            cls.test_dir, n_students=5, n_questions=3, skip_questions=2
        )
        cls.name_images = _make_dummy_name_images(cls.test_dir, n_students=5)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.test_dir, ignore_errors=True)

    def test_08_without_name_images_regression(self):
        """name_images=None で従来通りの列構成になること（回帰テスト）"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_no_name.xlsx")
        df = generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=None
        )
        self.assertIsNotNone(df)
        self.assertTrue(os.path.exists(output_path))

        # Excel を開いてヘッダーを確認
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        # Row 2 のヘッダーを取得
        headers = [cell.value for cell in ws[2]]
        self.assertEqual(headers[0], 'No')
        self.assertEqual(headers[1], 'File')
        # 3番目は '学籍番号1'（氏名欄ではない）
        self.assertEqual(headers[2], '学籍番号1')
        # freeze_panes は C3
        self.assertEqual(ws.freeze_panes, 'C3')
        wb.close()

    def test_09_with_name_images_column_added(self):
        """name_images 付きで氏名欄列が追加されること"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_with_name.xlsx")
        df = generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=self.name_images
        )
        self.assertIsNotNone(df)
        self.assertTrue(os.path.exists(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[2]]
        self.assertEqual(headers[0], 'No')
        self.assertEqual(headers[1], 'File')
        self.assertEqual(headers[2], '氏名欄')  # 新規追加列
        self.assertEqual(headers[3], '学籍番号1')  # 1つ右にずれる
        wb.close()

    def test_10_images_embedded_in_excel(self):
        """氏名欄列に画像が埋め込まれていること"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_images.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=self.name_images
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        # openpyxl の _images リストに画像が含まれていること
        self.assertGreater(len(ws._images), 0)
        self.assertEqual(len(ws._images), 5)  # 5人分の画像
        wb.close()

    def test_11_freeze_panes_adjusted(self):
        """name_images がある場合、freeze_panes が D3 になること"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_freeze.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=self.name_images
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        self.assertEqual(ws.freeze_panes, 'D3')
        wb.close()

    def test_12_column_count_difference(self):
        """name_images ありの場合、無しと比べて1列多いこと"""
        from saitensamurai import generate_student_summary
        import openpyxl

        out_no = os.path.join(self.test_dir, "summary_cnt_no.xlsx")
        out_yes = os.path.join(self.test_dir, "summary_cnt_yes.xlsx")

        generate_student_summary(
            self.template_path, self.result_path, out_no,
            skip_questions=2, name_images=None
        )
        generate_student_summary(
            self.template_path, self.result_path, out_yes,
            skip_questions=2, name_images=self.name_images
        )

        wb_no = openpyxl.load_workbook(out_no)
        wb_yes = openpyxl.load_workbook(out_yes)

        max_col_no = wb_no.active.max_column
        max_col_yes = wb_yes.active.max_column

        self.assertEqual(max_col_yes, max_col_no + 1)
        wb_no.close()
        wb_yes.close()

    def test_13_empty_name_images_dict(self):
        """空の name_images 辞書の場合、従来通りの動作になること"""
        from saitensamurai import generate_student_summary
        import openpyxl

        output_path = os.path.join(self.test_dir, "summary_empty_dict.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images={}
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[2]]
        # 空辞書 → has_name_images=False → 氏名欄列なし
        self.assertEqual(headers[2], '学籍番号1')
        self.assertEqual(ws.freeze_panes, 'C3')
        wb.close()

    def test_14_partial_name_images(self):
        """一部の生徒だけ name_images がある場合でもエラーにならないこと"""
        from saitensamurai import generate_student_summary
        import openpyxl

        # 5人中3人分だけの画像
        partial_images = {
            k: v for i, (k, v) in enumerate(self.name_images.items()) if i < 3
        }
        output_path = os.path.join(self.test_dir, "summary_partial.xlsx")
        generate_student_summary(
            self.template_path, self.result_path, output_path,
            skip_questions=2, name_images=partial_images
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        # 氏名欄列は存在する
        headers = [cell.value for cell in ws[2]]
        self.assertEqual(headers[2], '氏名欄')
        # 画像は3枚だけ埋め込まれている
        self.assertEqual(len(ws._images), 3)
        wb.close()


if __name__ == '__main__':
    unittest.main()
