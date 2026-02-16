"""
OMR認識エンジン (omr_engine.py)

Mark2のOMR (Optical Mark Recognition) 処理を担うモジュール。
画像からのコーナーマーカー検出、射影変換補正、マーク認識、
座標管理、認識結果のExcel出力など、OMRパイプライン全体を提供する。

主な機能:
- imread_unicode: 日本語パス対応の画像読み込み
- parse_excel_coordinates: 座標定義Excelのパース
- detect_corner_markers: 四隅マーカー検出
- apply_perspective_transform: 射影変換による画像補正
- recognize_marks: マーク認識 (Mark2OSSロジック準拠)
- save_recognition_results: 認識結果のExcel出力
- process_box_drawer: フォルダ一括処理 (枠描画 + OMR認識)
"""

import csv
import json
import logging
import os
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime

import cv2
import pandas as pd
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from constants import (
    MARK2_WIDTH,
    MARK2_HEIGHT,
    OUTPUT_SCALE_MAX,
    RESULTS_FOLDER,
    BOXED_FOLDER,
    RESULTS_DATA_FOLDER,
    ANSWER_KEY_FILE,
    READING_RESULTS_FOLDER_NAME,
    MARKER_CACHE_FILE,
)

logger = logging.getLogger(__name__)


def imread_unicode(filepath):
    """日本語パスに対応した画像読み込み（np.fromfile + cv2.imdecode）"""
    try:
        img_array = np.fromfile(str(filepath), dtype=np.uint8)
        img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
        return img
    except Exception as e:
        logger.error("画像読み込みエラー (%s): %s", filepath, e)
        return None


def parse_excel_coordinates(excel_path, skip_questions=0):
    """
    Mark2の座標定義Excelファイルをパースして座標リストを取得
    
    Args:
        excel_path: 座標定義Excelファイルのパス
        skip_questions: スキップする問題数（出席番号エリアなど）
                        ※注意: この関数ではスキップ処理を行いません。
                        OMR出力の列ズレ防止のため全座標を読み込みます。
                        実際のスキップ判定はscore_answers() / load_mark2_results()で行います。
    
    Returns:
        coordinates: マーク領域のリスト
        question_groups: 設問ごとのグループ情報
    """
    df = pd.read_excel(excel_path, header=None)
    
    coordinates = []
    question_groups = {}  # 設問番号 -> 選択肢群の範囲
    renumber_offset = 0  # 再採番用のオフセット
    
    for row_idx in range(3, len(df)):
        row = df.iloc[row_idx]
        original_question_no = row[0]
        
        # 以前はここでskip_questionsに基づいてスキップしていましたが、
        # OMR出力の列ズレを防ぐため、すべての座標を読み込みます。
        # スキップ判定は採点時（score_answers / load_mark2_results）に行います。
        
        question_no = original_question_no
        question_name = row[1] if pd.notna(row[1]) else f"Q{question_no}"
        
        # この設問の選択肢座標を収集
        question_coords = []
        
        # この設問の選択肢座標を一時リストに収集
        temp_coords = []
        
        # 最大20選択肢まで確認（列がある限り）
        for raw_choice_idx in range(20):
            base_col = 4 + (raw_choice_idx * 4)
            
            if base_col + 3 < len(row):
                pos_x = row[base_col]
                pos_y = row[base_col + 1]
                size_x = row[base_col + 2]
                size_y = row[base_col + 3]
                
                if pd.notna(pos_x) and pd.notna(pos_y) and pd.notna(size_x) and pd.notna(size_y):
                    try:
                        coord = {
                            'question_no': question_no,
                            'question': question_name,
                            # choiceは後でX座標順に割り振るため、ここでは仮の値
                            'raw_choice': raw_choice_idx,
                            'x': int(pos_x),
                            'y': int(pos_y),
                            'width': int(size_x),
                            'height': int(size_y)
                        }
                        temp_coords.append(coord)
                    except (ValueError, TypeError):
                        continue
            else:
                break
        
        # X座標でソート（左から右へ）
        temp_coords.sort(key=lambda c: c['x'])
        
        # ソート順にchoice番号（0, 1, 2...）を割り当てて正式なリストに追加
        for i, coord in enumerate(temp_coords):
            coord['choice'] = i
            # raw_choiceは不要なら削除、デバッグ用に残しても良い
            coordinates.append(coord)
            question_coords.append(coord)
        
        # 設問ごとの選択肢群の範囲を計算
        if question_coords:
            x_list = [c['x'] for c in question_coords] + [c['x'] + c['width'] for c in question_coords]
            y_list = [c['y'] for c in question_coords] + [c['y'] + c['height'] for c in question_coords]
            
            min_x = min(x_list)
            max_x = max(x_list)
            min_y = min(y_list)
            max_y = max(y_list)
            
            question_groups[question_no] = {
                'question_name': question_name,
                'choices_bbox': {'x': min_x, 'y': min_y, 'width': max_x - min_x, 'height': max_y - min_y},
                'min_x': min_x,
                'min_y': min_y,
                'height': max_y - min_y
            }
    
    return coordinates, question_groups


def save_template_coordinates_debug(coordinates, output_path):
    """座標リストをCSVファイルに保存（デバッグ用・静的）"""
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # ヘッダー
            writer.writerow(['question_no', 'question_name', 'choice', 'x', 'y', 'width', 'height'])
            # データ
            for coord in coordinates:
                writer.writerow([
                    coord['question_no'],
                    coord['question'],
                    coord['choice'],
                    coord['x'],
                    coord['y'],
                    coord['width'],
                    coord['height']
                ])
        logger.info("テンプレート座標データを保存しました: %s", output_path)
    except Exception as e:
        logger.error("テンプレート座標データの保存に失敗しました: %s", e)


def load_coordinates_from_csv(csv_path):
    """CSVファイルから座標リストを読み込む"""
    coordinates = []
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                coord = {
                    'question_no': int(row['question_no']),
                    'question': row['question_name'],
                    'choice': int(row['choice']),
                    'x': int(row['x']),
                    'y': int(row['y']),
                    'width': int(row['width']),
                    'height': int(row['height'])
                }
                coordinates.append(coord)
        return coordinates
    except Exception as e:
        logger.error("座標データの読み込みに失敗しました: %s", e)
        return []


def detect_corner_markers(image, debug=False):
    """
    画像の四隅近くにある黒い正方形マーカーを検出
    
    Returns:
        markers: [(x1, y1), (x2, y2), (x3, y3), (x4, y4)] 左上、右上、右下、左下の順
    """
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    h, w = gray.shape
    
    # Mark2のアルゴリズム: 四隅から1%マージン、30%×8%のサーチエリア
    margin_x = int(w * 0.01)
    margin_y = int(h * 0.01)
    search_w = int(w * 0.3)
    search_h = int(w * 0.08)
    
    # 4つのサーチ領域を定義
    search_regions = [
        {'name': '左上', 'x': margin_x, 'y': margin_y, 'w': search_w, 'h': search_h},
        {'name': '右上', 'x': w - margin_x - search_w, 'y': margin_y, 'w': search_w, 'h': search_h},
        {'name': '右下', 'x': w - margin_x - search_w, 'y': h - margin_y - search_h, 'w': search_w, 'h': search_h},
        {'name': '左下', 'x': margin_x, 'y': h - margin_y - search_h, 'w': search_w, 'h': search_h},
    ]
    
    markers = []
    debug_img = image.copy() if debug else None
    
    for region in search_regions:
        x, y, rw, rh = region['x'], region['y'], region['w'], region['h']
        
        # サーチ領域を切り出し
        roi = gray[y:y+rh, x:x+rw]
        
        # 二値化（Otsu's法）
        _, binary = cv2.threshold(roi, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        
        # 連結成分解析
        num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary, connectivity=8)
        
        # 最大面積の成分を検出（背景を除く）
        max_area = 0
        max_label = -1
        
        for i in range(1, num_labels):  # 0はバックグラウンド
            area = stats[i, cv2.CC_STAT_AREA]
            if area > max_area:
                max_area = area
                max_label = i
        
        if max_label >= 0:
            # マーカーの中心座標（画像全体座標系）
            center_x = int(centroids[max_label][0]) + x
            center_y = int(centroids[max_label][1]) + y
            markers.append((center_x, center_y))
            
            if debug:
                # サーチ領域を描画
                cv2.rectangle(debug_img, (x, y), (x + rw, y + rh), (255, 0, 0), 2)
                # マーカー中心を描画
                cv2.circle(debug_img, (center_x, center_y), 10, (0, 0, 255), -1)
    
    if len(markers) != 4:
        raise ValueError(f"4個のマーカーが必要ですが、{len(markers)}個しか検出されませんでした")
    
    if debug:
        return markers, debug_img
    else:
        return markers


def apply_perspective_transform(image, markers, output_scale=1.0):
    """
    射影変換で画像を補正（Mark2アルゴリズムに準拠）
    
    Args:
        image: 入力画像
        markers: コーナーマーカー座標
        output_scale: 出力スケール倍率（1.0 = 595x842, 大きいほど高解像度）
    
    Returns:
        corrected_image: 補正後の画像（595*scale x 842*scale）
        transform_matrix: 変換行列
    """
    # Mark2の基準サイズ
    w = int(595 * output_scale)
    h = int(842 * output_scale)
    
    # Mark2のマーカー位置（基準座標系、スケール適用後）
    xp1 = w * (0.14 + 0.015)  # 左上 X
    yp1 = h * (0.03 + 0.01)   # 左上 Y
    xp2 = w * (0.83 + 0.015)  # 右上 X
    yp2 = h * (0.03 + 0.01)   # 右上 Y
    xp3 = w * (0.83 + 0.015)  # 右下 X
    yp3 = h * (0.95 + 0.01)   # 右下 Y
    xp4 = w * (0.14 + 0.015)  # 左下 X
    yp4 = h * (0.95 + 0.01)   # 左下 Y
    
    src_points = np.float32(markers)
    
    # 画像のマーカー位置を、Mark2の基準座標系のマーカー位置にマッピング
    dst_points = np.float32([
        [xp1, yp1],  # 左上
        [xp2, yp2],  # 右上
        [xp3, yp3],  # 右下
        [xp4, yp4]   # 左下
    ])
    
    # 変換行列を計算
    transform_matrix = cv2.getPerspectiveTransform(src_points, dst_points)
    
    # 射影変換を適用（595*scale x 842*scale のサイズで出力）
    corrected_image = cv2.warpPerspective(image, transform_matrix, (w, h))
    
    return corrected_image, transform_matrix


def compute_output_scale(image):
    """元画像の解像度からoutput_scaleを計算する
    
    元画像の解像度を極力活かすスケールを返す。OUTPUT_SCALE_MAXで上限を制限。
    
    Args:
        image: OpenCV画像 (BGR)
    Returns:
        float: output_scale値
    """
    img_h, img_w = image.shape[:2]
    scale_w = img_w / MARK2_WIDTH
    scale_h = img_h / MARK2_HEIGHT
    scale = min(scale_w, scale_h, OUTPUT_SCALE_MAX)
    return max(scale, 1.0)


def draw_all_areas(image, coordinates, question_groups):
    """
    全てのエリアを描画
    
    1. マーク領域（選択肢）- 緑の枠
    2. 設問単位の選択肢群 - 赤の枠
    """
    result_image = image.copy()
    
    # 1. マーク領域を描画（緑の枠）
    mark_count = 0
    for coord in coordinates:
        x = coord['x']
        y = coord['y']
        w = coord['width']
        h = coord['height']
        
        cv2.rectangle(result_image, (x, y), (x + w, y + h), (0, 255, 0), 2)
        mark_count += 1
    
    # 2. 設問ごとの選択肢群を描画（赤の枠）
    group_count = 0
    for question_no, group_data in question_groups.items():
        choices_bbox = group_data['choices_bbox']
        cv2.rectangle(result_image,
                     (choices_bbox['x'], choices_bbox['y']),
                     (choices_bbox['x'] + choices_bbox['width'], choices_bbox['y'] + choices_bbox['height']),
                     (0, 0, 255), 2)
        group_count += 1
        
    return result_image, mark_count, group_count


def generate_template(coord_excel_path, output_folder, skip_questions=0):
    """
    採点用正答データExcelを生成
    
    Args:
        coord_excel_path: 座標定義ファイルのパス
        output_folder: 出力先フォルダ
        skip_questions: スキップする問題数（出席番号エリアなど）
    
    Returns:
        template_path: 生成されたテンプレートファイルのパス
    """
    # 座標ファイルを読み込み
    df_coord = pd.read_excel(coord_excel_path, header=None)
    
    template_data = []
    
    # 3行目から問題データを読み取り
    for row_idx in range(3, len(df_coord)):
        row = df_coord.iloc[row_idx]
        original_question_no = row[0]
        
        if pd.isna(original_question_no):
            continue
        
        # スキップする問題は除外
        if original_question_no <= skip_questions:
            continue
        
        # 再採番された問題番号
        question_no = original_question_no - skip_questions
        
        template_data.append({
            '問題番号': question_no,
            '正答': '',  # 空欄（ユーザーが入力）
            '配点': '',  # 空欄（ユーザーが入力）
            '観点': ''   # 空欄（ユーザーが入力）
        })
    
    # DataFrameに変換
    df_template = pd.DataFrame(template_data)
    
    # Excelファイルとして出力
    output_folder = Path(output_folder)
    output_folder.mkdir(exist_ok=True)
    template_path = output_folder / ANSWER_KEY_FILE
    
    # ⚠ 既存のanswer_key.xlsxがある場合は上書きしない（ユーザー入力済みの正答・配点を保護）
    if template_path.exists():
        logger.info("テンプレートが既に存在します。上書きしません: %s", template_path.name)
        return template_path
    
    df_template.to_excel(template_path, index=False)
    
    return template_path


def save_coordinates_to_csv(csv_path, all_data):
    """
    座標データをCSV形式で保存（画像ごとの補正後座標）
    
    CSV形式:
    image_path,question_no,choices_bbox,mark_coords
    
    choices_bbox: x;y;w;h
    mark_coords: choice0_x;y;w;h|choice1_x;y;w;h|...
    
    Raises:
        PermissionError: ファイルが他のアプリで開かれている場合
    """
    try:
        _save_coordinates_to_csv_impl(csv_path, all_data)
    except PermissionError:
        logger.warning("%s への保存に失敗しました。ファイルが別のアプリで開かれている可能性があります。", csv_path)
        raise


def _save_coordinates_to_csv_impl(csv_path, all_data):
    """save_coordinates_to_csvの実装部"""
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # ヘッダー
        writer.writerow([
            'image_path',
            'question_no',
            'choices_bbox',
            'mark_coords'
        ])
        
        # データ行
        for data in all_data:
            image_path = data['image_path']
            question_no = data['question_no']
            
            # choices_bbox形式: x;y;w;h
            choices_bbox = data['choices_bbox']
            choices_str = f"{choices_bbox['x']};{choices_bbox['y']};{choices_bbox['width']};{choices_bbox['height']}"
            
            # マーク座標: choice0_x;y;w;h|choice1_x;y;w;h|...
            mark_list = []
            for mark in data['mark_coords']:
                mark_list.append(f"{mark['x']};{mark['y']};{mark['width']};{mark['height']}")
            mark_str = '|'.join(mark_list)
            
            writer.writerow([
                image_path,
                question_no,
                choices_str,
                mark_str
            ])


def recognize_marks(image, coordinates, color_threshold=0.1, area_threshold=0.4):
    """
    マーク認識を行う (Mark2OSSロジック準拠)
    
    Args:
        image: 補正後の画像 (Gray or BGR)
        coordinates: マーク領域のリスト
        color_threshold: 画素値の閾値 (0.0-1.0). (1 - color_threshold) * 255 より暗い画素をマークとみなす.
                         Default: 0.1 (255 * 0.9 = 229.5未満をマークとする)
        area_threshold: 面積閾値 (0.0-1.0). マーク画素の割合がこれを超えたらマークとみなす.
                        Default: 0.4 (40%以上)
    
    Returns:
        results: {question_no: [choice_idx, ...]}
    """
    if len(image.shape) == 3:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image
        
    results = {}
    
    # 閾値計算
    pixel_threshold = int((1.0 - color_threshold) * 255)
    
    # 二値化
    # 画素値 < pixel_threshold (暗い) -> 255 (白/カウント対象)
    _, binary = cv2.threshold(gray, pixel_threshold, 255, cv2.THRESH_BINARY_INV)
    
    for coord in coordinates:
        q_no = coord['question_no']
        choice_idx = coord['choice']
        x, y, w, h = coord['x'], coord['y'], coord['width'], coord['height']
        
        # ROI抽出
        roi = binary[y:y+h, x:x+w]
        
        # マーク画素数をカウント
        marked_pixels = cv2.countNonZero(roi)
        total_pixels = w * h
        
        if total_pixels == 0:
            continue
            
        ratio = marked_pixels / total_pixels
        
        if ratio > area_threshold:
            if q_no not in results:
                results[q_no] = []
            results[q_no].append(choice_idx)
            
    return results


def save_recognition_results(output_path, recognition_results, all_questions, question_names=None, choice_counts=None, coordinates=None):
    """
    認識結果をExcelファイルに保存 (Mark2OSS Survey.cs準拠)
    スタイリング: ヘッダー装飾, 罫線, NoMark背景色(オレンジ), DoubleMark背景色(薄い赤), ウィンドウ枠固定
    
    Args:
        output_path: 出力Excelファイルパス
        recognition_results: 認識結果リスト
        all_questions: 全設問番号リスト
        question_names: 設問名辞書 (optional)
        choice_counts: 設問番号 -> 選択肢数 の辞書 (optional, 未指定時は10)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # --- スタイル定義 ---
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    # raw_choiceルックアップ: {q_no: {sorted_choice_idx: raw_choice_value}}
    # Excelの列ヘッダ値（raw_choice）を表示値として使用する。
    # 従来の (c+1)%num_choices は横並び10列専用だったが、
    # raw_choice は縦並び・異なる選択肢数のテンプレートでも正しい値を返す。
    choice_to_display = {}
    if coordinates:
        for c in coordinates:
            q = c['question_no']
            if q not in choice_to_display:
                choice_to_display[q] = {}
            choice_to_display[q][c['choice']] = c['raw_choice']

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    label_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    label_font = Font(bold=True, size=9)
    no_mark_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    double_mark_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    # --- Row 1: ヘッダー (No, File, 設問番号) ---
    header_values = ['No', 'File'] + [str(q) for q in all_questions]
    ws.append(header_values)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- Row 2: 設問名ラベル ---
    label_values = ['', '']
    if question_names:
        for q in all_questions:
            label_values.append(question_names.get(q, ''))
    else:
        label_values.extend([''] * len(all_questions))
    ws.append(label_values)
    for cell in ws[2]:
        cell.font = label_font
        cell.fill = label_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- データ行 ---
    for idx, res in enumerate(recognition_results):
        row_values = [idx + 1, res['image']]
        marks = res['marks']

        for q_no in all_questions:
            if q_no in marks:
                choices = marks[q_no]
                num_choices = choice_counts.get(q_no, 10) if choice_counts else 10
                val_strs = []
                for c in sorted(choices):
                    # raw_choice（Excelの列ヘッダ値）を表示値として使用
                    # 座標パース時に保存された raw_choice は、テンプレートの
                    # レイアウト（横並び・縦並び）に依存しない正しい値を持つ。
                    if q_no in choice_to_display and c in choice_to_display[q_no]:
                        val = choice_to_display[q_no][c]
                    else:
                        # フォールバック（coordinates未指定時の後方互換）
                        val = (c + 1) % num_choices
                    val_strs.append(str(val))
                row_values.append(';'.join(val_strs))
            else:
                row_values.append('')  # No mark

        ws.append(row_values)
        row_idx = ws.max_row

        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = center_align
            elif col_idx >= 3:
                cell.alignment = center_align
                cell_value = cell.value
                if cell_value is None or (isinstance(cell_value, str) and cell_value == ''):
                    cell.fill = no_mark_fill
                elif isinstance(cell_value, str) and ';' in cell_value:
                    cell.fill = double_mark_fill

    # --- 列幅 ---
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    for col_idx in range(3, len(all_questions) + 3):
        col_letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[col_letter].width = 8

    # --- ウィンドウ枠固定 (ヘッダー2行 + No/File列) ---
    ws.freeze_panes = 'C3'

    try:
        wb.save(output_path)
    except PermissionError:
        logger.warning("%s への保存に失敗しました。ファイルが別のアプリで開かれている可能性があります。", output_path)
        raise


def _process_single_image(args: tuple) -> dict:
    """
    1枚の画像を処理するワーカー関数（ProcessPoolExecutor 用）。

    モジュールレベルに定義することで pickle 可能。
    各ワーカープロセスで独立して実行される。

    Args:
        args: (image_path_str, boxed_folder_str, coordinates,
               question_groups, color_threshold, area_threshold)

    Returns:
        dict with keys: filename, marks, marker_data, csv_data, success
    """
    (image_path_str, boxed_folder_str, coordinates,
     question_groups, color_threshold, area_threshold) = args

    image_path = Path(image_path_str)
    boxed_folder = Path(boxed_folder_str)

    with open(str(image_path), 'rb') as f:
        image_data_bytes = f.read()
    image = cv2.imdecode(np.frombuffer(image_data_bytes, np.uint8), cv2.IMREAD_COLOR)

    if image is None:
        raise ValueError(f"画像を読み込めません: {image_path.name}")

    markers = detect_corner_markers(image, debug=False)
    corrected_image, _ = apply_perspective_transform(image, markers)
    marker_data = [
        pt.tolist() if hasattr(pt, 'tolist') else list(pt) for pt in markers
    ]

    # OMR認識
    marks = recognize_marks(
        corrected_image, coordinates,
        color_threshold=color_threshold, area_threshold=area_threshold,
    )

    # 枠描画
    result_image, _mark_count, _group_count = draw_all_areas(
        corrected_image, coordinates, question_groups,
    )

    # 認識結果をオーバーレイ描画
    for q_no, choices in marks.items():
        for c_idx in choices:
            target = next(
                (c for c in coordinates
                 if c['question_no'] == q_no and c['choice'] == c_idx),
                None,
            )
            if target:
                cv2.rectangle(
                    result_image,
                    (target['x'], target['y']),
                    (target['x'] + target['width'], target['y'] + target['height']),
                    (255, 0, 0), 2,
                )

    # boxed画像を保存
    output_path = boxed_folder / image_path.name
    _, encoded = cv2.imencode('.jpg', result_image)
    with open(str(output_path), 'wb') as f:
        f.write(encoded)

    # CSVデータ構築
    csv_data = []
    for question_no, group_data in question_groups.items():
        question_marks = [c for c in coordinates if c['question_no'] == question_no]
        csv_data.append({
            'image_path': str(image_path.name),
            'question_no': question_no,
            'choices_bbox': group_data['choices_bbox'],
            'mark_coords': question_marks,
        })

    return {
        'filename': image_path.name,
        'marks': marks,
        'marker_data': marker_data,
        'csv_data': csv_data,
        'success': True,
    }


def process_box_drawer(image_folder, coord_excel_path, skip_questions=0, output_base_folder=None, debug=False, color_threshold=0.1, area_threshold=0.4, progress_callback=None, cancel_event=None):
    """
    フォルダ内の画像を一括処理（枠描画 + OMR認識）

    Args:
        progress_callback: 進捗コールバック(current, total)（オプション、GUIプログレスバー用）
        cancel_event: threading.Event — set()されると処理を中断
    """
    start_time = time.time()
    
    image_folder = Path(image_folder)
    coord_excel_path = Path(coord_excel_path)
    
    if output_base_folder is None:
        output_base_folder = image_folder
    else:
        output_base_folder = Path(output_base_folder)
    
    results_folder = output_base_folder / RESULTS_FOLDER
    results_folder.mkdir(exist_ok=True)
    
    boxed_folder = results_folder / BOXED_FOLDER
    boxed_folder.mkdir(exist_ok=True)

    results_data_folder = results_folder / RESULTS_DATA_FOLDER
    results_data_folder.mkdir(exist_ok=True)

    reading_results_folder = results_data_folder / READING_RESULTS_FOLDER_NAME
    reading_results_folder.mkdir(exist_ok=True)
    
    logger.info("出力フォルダ: %s", results_folder)
    logger.info("枠描画結果: %s/", boxed_folder.name)
    logger.info("読取結果: %s/%s/", results_data_folder.name, reading_results_folder.name)
    logger.info("座標ファイル: %s", coord_excel_path.name)
    logger.info("スキップする問題数: %s問", skip_questions)
    
    coordinates, question_groups = parse_excel_coordinates(coord_excel_path, skip_questions)
    logger.info("座標データ: %d個のマークエリア, %d個の設問", len(coordinates), len(question_groups))
    
    # 座標データをCSVとして保存（検証用）→ 01_Results/ に配置
    try:
        csv_output_path = results_data_folder / "template_coordinates.csv"
        save_template_coordinates_debug(coordinates, csv_output_path)
    except Exception:
        pass
    
    all_questions = sorted(list(question_groups.keys()))
    question_names = {q: g['question_name'] for q, g in question_groups.items()}
    
    try:
        template_path = generate_template(coord_excel_path, results_data_folder, skip_questions)
        logger.info("テンプレート生成: %s", template_path.name)
    except Exception as e:
        logger.warning("テンプレート生成エラー: %s", e)
    
    image_files = sorted(image_folder.glob('*.jpg')) + sorted(image_folder.glob('*.png'))
    image_files = [f for f in image_files if not str(f.parent).endswith(RESULTS_FOLDER) 
                   and RESULTS_FOLDER not in f.parts]
    
    if not image_files:
        logger.error("%s に画像ファイルが見つかりません", image_folder)
        return {'success_count': 0, 'error_count': 0, 'total_count': 0, 'elapsed_time': 0}
    
    logger.info("=" * 60)
    logger.info("処理対象: %d個の画像", len(image_files))
    logger.info("=" * 60)
    
    success_count = 0
    error_count = 0
    all_csv_data = []
    recognition_results_list = []
    marker_cache = {}  # マーカー座標キャッシュ（Step2高速化用）

    # --- 並列処理 (ProcessPoolExecutor) ---
    max_workers = max(1, (os.cpu_count() or 1) - 1)
    total = len(image_files)
    logger.info("並列ワーカー数: %d", max_workers)

    worker_args = [
        (str(img), str(boxed_folder), coordinates, question_groups,
         color_threshold, area_threshold)
        for img in image_files
    ]

    completed = 0
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        future_to_name = {
            executor.submit(_process_single_image, args): Path(args[0]).name
            for args in worker_args
        }

        for future in as_completed(future_to_name):
            # 中断チェック（新規結果の取得を停止）
            if cancel_event and cancel_event.is_set():
                # 未完了のfutureをキャンセル
                for f in future_to_name:
                    f.cancel()
                logger.info("中断されました (%d/%d件処理済み)", completed, total)
                break

            completed += 1
            fname = future_to_name[future]
            logger.info("[%d/%d] 完了: %s", completed, total, fname)

            if progress_callback:
                try:
                    progress_callback(completed, total)
                except Exception:
                    pass

            try:
                result = future.result()
                recognition_results_list.append({
                    'image': result['filename'],
                    'marks': result['marks'],
                })
                marker_cache[result['filename']] = result['marker_data']
                all_csv_data.extend(result['csv_data'])
                success_count += 1
            except Exception as e:
                logger.error("処理エラー (%s): %s", fname, e)
                error_count += 1
    
    csv_path = results_data_folder / 'coordinates.csv'
    save_coordinates_to_csv(csv_path, all_csv_data)
    logger.info("座標データCSV保存: %s", csv_path.name)

    # マーカーキャッシュをJSON保存（Step2での射影変換高速化用）
    try:
        marker_cache_path = results_data_folder / MARKER_CACHE_FILE
        with open(str(marker_cache_path), 'w', encoding='utf-8') as f:
            json.dump(marker_cache, f)
        logger.info("マーカーキャッシュ保存: %d件", len(marker_cache))
    except Exception:
        pass  # キャッシュ保存失敗は許容（Step2は再検出にフォールバック）
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    area_str = f"{int(area_threshold * 100):03d}"
    color_str = f"{int(color_threshold * 100):03d}"
    omr_result_path = reading_results_folder / f"Mark2-Result-A{area_str}-C{color_str}-{timestamp}.xlsx"
    
    # 設問ごとの選択肢数を構築
    choice_counts = {}
    for q_no in all_questions:
        q_coords = [c for c in coordinates if c['question_no'] == q_no]
        choice_counts[q_no] = len(q_coords)
    
    save_recognition_results(omr_result_path, recognition_results_list, all_questions, question_names, choice_counts, coordinates)
    logger.info("OMR認識結果保存: %s", omr_result_path.name)
    
    elapsed_time = time.time() - start_time
    
    logger.info("=" * 60)
    logger.info("処理完了: 成功 %d件 / エラー %d件", success_count, error_count)
    logger.info("実行時間: %.2f秒", elapsed_time)
    logger.info("=" * 60)
    
    return {
        'success_count': success_count,
        'error_count': error_count,
        'total_count': len(image_files),
        'elapsed_time': elapsed_time
    }


# process_folderエイリアス（後方互換性のため）
process_folder = process_box_drawer
