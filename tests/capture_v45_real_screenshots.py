"""
capture_v45_real_screenshots.py
————————————————————————————————————
main_src の 実クラス をそのまま起動してスクリーンショットを撮る。
モック不使用。実際の SaitenSamuraiGUI / MarkCheckerGUI を使用。

出力:
  docs/images/02a_main_mark_only.png   — v4.5 認識方式コンボボックス付き
  docs/images/18_mark_checker.png      — v4.5 タブ(薄い/濃い)＋グリッドビュー

使い方:
  python tests/capture_v45_real_screenshots.py
"""
from __future__ import annotations

import ctypes
import ctypes.wintypes as wintypes
import os
import shutil
import sys
import tempfile
import time
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd
from PIL import Image, ImageGrab

# — パス設定 —
ROOT = Path(__file__).resolve().parent.parent
MAIN_SRC = ROOT / "main_src"
SAMPLE_DIR = ROOT / "sample_basefile"
DOCS_IMAGES = ROOT / "docs" / "images"

sys.path.insert(0, str(MAIN_SRC))

import tkinter as tk
from tkinter import ttk

from constants import (
    MODE_MARK_ONLY,
    ERROR_TYPE_NO_MARK, ERROR_TYPE_DOUBLE_MARK,
)


# ============================================================
# Win32 PrintWindow キャプチャ（最高精度）
# ============================================================

class BITMAPINFOHEADER(ctypes.Structure):
    _fields_ = [
        ("biSize",          wintypes.DWORD),
        ("biWidth",         ctypes.c_long),
        ("biHeight",        ctypes.c_long),
        ("biPlanes",        wintypes.WORD),
        ("biBitCount",      wintypes.WORD),
        ("biCompression",   wintypes.DWORD),
        ("biSizeImage",     wintypes.DWORD),
        ("biXPelsPerMeter", ctypes.c_long),
        ("biYPelsPerMeter", ctypes.c_long),
        ("biClrUsed",       wintypes.DWORD),
        ("biClrImportant",  wintypes.DWORD),
    ]


def _get_hwnd(widget):
    try:
        frame_id = widget.wm_frame()
        if frame_id and frame_id != "0x0":
            return int(frame_id, 16)
    except Exception:
        pass
    return widget.winfo_id()


def _printwindow_capture(widget, path: Path) -> bool:
    """Win32 PrintWindow API でウィンドウ全体 (タイトルバー含む) をキャプチャ"""
    try:
        hwnd = _get_hwnd(widget)
        if not hwnd:
            return False
        rect = wintypes.RECT()
        ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(rect))
        width = rect.right - rect.left
        height = rect.bottom - rect.top
        if width < 10 or height < 10:
            return False
        wDC = ctypes.windll.user32.GetWindowDC(hwnd)
        if not wDC:
            return False
        try:
            dcObj = ctypes.windll.gdi32.CreateCompatibleDC(wDC)
            bmp = ctypes.windll.gdi32.CreateCompatibleBitmap(wDC, width, height)
            old = ctypes.windll.gdi32.SelectObject(dcObj, bmp)
            ok = ctypes.windll.user32.PrintWindow(hwnd, dcObj, 2)  # PW_RENDERFULLCONTENT
            if ok:
                bmi = BITMAPINFOHEADER()
                bmi.biSize = ctypes.sizeof(BITMAPINFOHEADER)
                bmi.biWidth = width
                bmi.biHeight = -height   # top-down
                bmi.biPlanes = 1
                bmi.biBitCount = 32
                bmi.biCompression = 0    # BI_RGB
                buf = ctypes.create_string_buffer(width * height * 4)
                ctypes.windll.gdi32.GetDIBits(dcObj, bmp, 0, height, buf, ctypes.byref(bmi), 0)
                img = Image.frombuffer("RGBA", (width, height), buf, "raw", "BGRA", 0, 1).convert("RGB")
                img.save(str(path))
            ctypes.windll.gdi32.SelectObject(dcObj, old)
            ctypes.windll.gdi32.DeleteObject(bmp)
            ctypes.windll.gdi32.DeleteDC(dcObj)
        finally:
            ctypes.windll.user32.ReleaseDC(hwnd, wDC)
        return bool(ok)
    except Exception as e:
        print(f"  [PrintWindow失敗] {e}")
        return False


def _imagegrab_capture(widget, path: Path) -> bool:
    """フォールバック: winfo_x/y (外枠座標) で ImageGrab"""
    try:
        outer_x = widget.winfo_x()
        outer_y = widget.winfo_y()
        client_x = widget.winfo_rootx()
        client_y = widget.winfo_rooty()
        client_w = widget.winfo_width()
        client_h = widget.winfo_height()

        border = client_x - outer_x
        # title_h = client_y - outer_y  (参考値)

        x1, y1 = outer_x, outer_y
        x2 = client_x + client_w + border
        y2 = client_y + client_h + border

        if (x2 - x1) < 10 or (y2 - y1) < 10:
            return False

        img = ImageGrab.grab(bbox=(x1, y1, x2, y2))
        img.save(str(path))
        return True
    except Exception as e:
        print(f"  [ImageGrab失敗] {e}")
        return False


def capture_window(widget, filename: str, delay_s: float = 0.9) -> Optional[Path]:
    widget.update_idletasks()
    widget.update()
    widget.lift()
    widget.focus_force()
    widget.update_idletasks()
    widget.update()
    time.sleep(delay_s)
    widget.update()

    path = DOCS_IMAGES / f"{filename}.png"
    if _printwindow_capture(widget, path):
        print(f"  ✔ {filename}.png  (PrintWindow)")
        return path
    if _imagegrab_capture(widget, path):
        print(f"  ✔ {filename}.png  (ImageGrab)")
        return path
    print(f"  ✗ {filename}.png  キャプチャ失敗")
    return None


# ============================================================
# ダミーデータセット構築
# ============================================================

_N_STUDENTS = 10
_N_QUESTIONS = 15   # skip=4 → 問題番号 1〜15
_SKIP = 4

# Mark2-003 ベース座標系(595×842)の典型的な選択肢領域
# 各問題は縦に約23px間隔で並ぶ。最初の問題を y=200 付近とする
def _q_bbox(q_no: int) -> str:
    """問題番号(1始まり)のchoices_bbox文字列を返す (x;y;w;h)"""
    x = 42
    y = 190 + (q_no - 1) * 23
    w = 380
    h = 21
    return f"{x};{y};{w};{h}"


def build_dummy_dataset(base_dir: Path) -> dict:
    """ダミー採点データを base_dir に生成し、パス群を返す"""
    img_dir = base_dir / "images"
    results_dir = base_dir / "results_data"
    img_dir.mkdir(parents=True, exist_ok=True)
    results_dir.mkdir(parents=True, exist_ok=True)

    # ── 答案画像: sample_marksheet.jpg を N 枚コピー ──
    src_img = SAMPLE_DIR / "sample_marksheet.jpg"
    if not src_img.exists():
        # フォールバック: グレー画像を生成
        grey = Image.new("RGB", (595, 842), (230, 230, 230))
        src_img = base_dir / "_grey.jpg"
        grey.save(str(src_img))

    filenames = []
    for i in range(1, _N_STUDENTS + 1):
        fn = f"student_{i:03d}.jpg"
        shutil.copy(str(src_img), str(img_dir / fn))
        filenames.append(fn)

    # ── Mark2-Result.xlsx ──
    # Row1: headers  Row2: labels(None or int)  Row3+: data
    xlsx_path = results_dir / "Mark2-Result.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ヘッダー行
    ws.cell(1, 1, "File")
    for q in range(1, _N_QUESTIONS + 1):
        ws.cell(1, q + 1, f"Q{q}")

    # ラベル行 (File列はNone → "File"がそのまま使われる)
    ws.cell(2, 1, None)
    for q in range(1, _N_QUESTIONS + 1):
        ws.cell(2, q + 1, q)

    # データ行: 故意にエラーを埋め込む
    # student_002 Q3 → DoubleMark (1;3)
    # student_004 Q5 → NoMark (empty)
    # student_006 Q3 → DoubleMark (2;4)
    # student_007 Q5 → NoMark (empty)
    # student_009 Q3 → DoubleMark (1;5)
    errors_map = {
        (2, 3): "1;3",
        (4, 5): None,   # NoMark
        (6, 3): "2;4",
        (7, 5): None,
        (9, 3): "1;5",
        (10, 8): None,
    }

    for i, fn in enumerate(filenames, start=1):
        ws.cell(i + 2, 1, fn)
        for q in range(1, _N_QUESTIONS + 1):
            key = (i, q)
            if key in errors_map:
                val = errors_map[key]  # None or string
                ws.cell(i + 2, q + 1, val)
            else:
                ws.cell(i + 2, q + 1, q % 4 + 1)  # normal answer

    wb.save(str(xlsx_path))

    # ── coordinates.csv ──
    # MarkCheckerGUI が使う: image_path, question_no, choices_bbox
    # 注: question_no は skip込みの元番号 (skip=4なので問題1=番号5)
    rows = []
    for fn in filenames:
        for q in range(1, _N_QUESTIONS + 1):
            original_q = q + _SKIP  # skip=4 なので 5,6,7...
            rows.append({
                "image_path": fn,
                "question_no": original_q,
                "choices_bbox": _q_bbox(q),
            })
    coords_csv = results_dir / "coordinates.csv"
    pd.DataFrame(rows).to_csv(str(coords_csv), index=False, encoding="utf-8")

    # ── エラー CSV を detect_errors_checker で生成 ──
    error_csv = results_dir / "tmp_checking_dm_nm.csv"
    try:
        from mark_checker import detect_errors_checker
        detect_errors_checker(str(xlsx_path), str(error_csv))
        print(f"  detect_errors_checker: エラーCSV生成 ({error_csv.name})")
    except Exception as e:
        # フォールバック: 手動生成
        print(f"  detect_errors_checker失敗 ({e}), フォールバック生成")
        err_rows = [
            {"filename": "student_002.jpg", "question_no": 3, "before": "1;3",  "after": "", "error_type": ERROR_TYPE_DOUBLE_MARK},
            {"filename": "student_004.jpg", "question_no": 5, "before": "",     "after": "", "error_type": ERROR_TYPE_NO_MARK},
            {"filename": "student_006.jpg", "question_no": 3, "before": "2;4",  "after": "", "error_type": ERROR_TYPE_DOUBLE_MARK},
            {"filename": "student_007.jpg", "question_no": 5, "before": "",     "after": "", "error_type": ERROR_TYPE_NO_MARK},
            {"filename": "student_009.jpg", "question_no": 3, "before": "1;5",  "after": "", "error_type": ERROR_TYPE_DOUBLE_MARK},
            {"filename": "student_010.jpg", "question_no": 8, "before": "",     "after": "", "error_type": ERROR_TYPE_NO_MARK},
        ]
        pd.DataFrame(err_rows).to_csv(str(error_csv), index=False, encoding="utf-8-sig")

    return {
        "img_dir": img_dir,
        "xlsx_path": xlsx_path,
        "coords_csv": coords_csv,
        "error_csv": error_csv,
        "filenames": filenames,
    }


# ============================================================
# 02a — 実 SaitenSamuraiGUI（マーク採点モード）
# ============================================================

def capture_02a(root: tk.Tk, data: dict):
    """02a_main_mark_only.png — 本物の SaitenSamuraiGUI を起動"""
    from main_gui import SaitenSamuraiGUI

    win = tk.Toplevel(root)
    win.geometry("1100x600+80+80")
    win.update_idletasks()

    try:
        gui = SaitenSamuraiGUI(win, mode=MODE_MARK_ONLY)

        # サンプルパスを事前入力（リアルな表示）
        gui.image_folder_path.set(str(data["img_dir"]))
        gui.coord_excel_path.set(str(SAMPLE_DIR / "M2-03-002_座標ファイル.xlsx"))
        gui.template_path.set(str(data["xlsx_path"]))

        win.deiconify()
        win.lift()
        win.update_idletasks()
        win.update()
        time.sleep(0.3)
        win.update()

        capture_window(win, "02a_main_mark_only")
    finally:
        try:
            win.destroy()
        except Exception:
            pass


# ============================================================
# 18 — 実 MarkCheckerGUI（グリッドビュー＋タブ表示）
# ============================================================

def capture_18(root: tk.Tk, data: dict):
    """18_mark_checker.png — 本物の MarkCheckerGUI グリッド + 薄い/濃いタブ"""
    from gui_components import MarkCheckerGUI

    # MarkCheckerGUI は __init__ 内で grab_set() するためルートが必要
    # root を非表示のまま parent にして Toplevel として起動
    checker: Optional[MarkCheckerGUI] = None

    def _after_load():
        """load_data 完了後に呼ばれるコールバック"""
        nonlocal checker
        if checker is None:
            return
        # グリッドビューに切り替え
        try:
            checker._switch_to_grid()
        except Exception:
            pass

        # サイドパネルの"選択肢 N"カテゴリを選ぶ → 薄い/濃いタブが出現
        # _get_selected_category と同じロジックでテキストを正規化して判定
        listbox = checker._category_listbox
        selected = False
        for i in range(listbox.size()):
            raw = listbox.get(i)
            text = raw.lstrip("\u2500 ")       # '─' (U+2500) と空白を除去
            if "(" in text:
                text = text[: text.rfind("(")].strip()  # "(N)" を除去
            if text.startswith("選択肢 "):
                listbox.selection_clear(0, tk.END)
                listbox.selection_set(i)
                listbox.event_generate("<<ListboxSelect>>")
                selected = True
                break

        if not selected:
            # 選択肢カテゴリが存在しない場合は先頭を選択
            listbox.selection_clear(0, tk.END)
            listbox.selection_set(0)
            listbox.event_generate("<<ListboxSelect>>")

        checker.window.update_idletasks()
        checker.window.update()
        time.sleep(0.8)
        checker.window.update()

        # キャプチャ
        capture_window(checker.window, "18_mark_checker", delay_s=0.3)

        # 後始末（grab を解放してから destroy）
        try:
            checker.window.grab_release()
            checker.window.destroy()
        except Exception:
            pass

    # MarkCheckerGUI を組み込みの load_data より前に after を仕掛けて制御
    # grab_set のせいで root のイベントループが止まらないよう
    # 少し遅延させてから _after_load を呼ぶ
    win_holder = tk.Toplevel(root)
    win_holder.geometry("1x1+-10000+-10000")  # 画面外に隔離（withdrawすると transient 子も隠れるため）

    checker = MarkCheckerGUI(
        parent_window=win_holder,
        image_folder=str(data["img_dir"]),
        coords_csv_path=str(data["coords_csv"]),
        xlsx_path=str(data["xlsx_path"]),
        skip_questions=_SKIP,
        template_path=None,
    )
    # load_data は window.after(100, load_data) でスケジュール済み
    # 画像ロード＋ホワイトネス構築（最大数秒）を考慮し 8000ms 待機してから _after_load を実行
    # transient 関係を解除して独立ウィンドウにする（親の状態に左右されないよう）
    checker.window.transient('')
    checker.window.after(8000, _after_load)
    checker.window.geometry("1200x750+50+50")
    checker.window.deiconify()

    # MarkCheckerGUI が grab_set しているのでそちらのウィンドウでブロッキング
    try:
        root.wait_window(checker.window)
    except Exception:
        pass

    try:
        win_holder.destroy()
    except Exception:
        pass


# ============================================================
# エントリーポイント
# ============================================================

def main():
    print("=" * 60)
    print("  v4.5 実アプリ スクリーンショット生成")
    print("=" * 60)

    root = tk.Tk()
    root.geometry("1x1+-10000+-10000")  # 画面外に配置（withdraw 不可：transient に影響するため）

    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td)
        print("\n[1/3] ダミーデータセットを構築中...")
        data = build_dummy_dataset(td_path)
        print(f"  images: {len(data['filenames'])} 枚")
        print(f"  xlsx:   {data['xlsx_path'].name}")
        print(f"  coords: {data['coords_csv'].name}")
        print(f"  errors: {data['error_csv'].name}")

        print("\n[2/3] スクリーンショットをキャプチャ中...")

        print("  → 02a_main_mark_only  (実 SaitenSamuraiGUI)")
        capture_02a(root, data)

        print("  → 18_mark_checker  (実 MarkCheckerGUI + グリッド + タブ)")
        capture_18(root, data)

    print("\n[3/3] 結果確認...")
    for name in ("02a_main_mark_only", "18_mark_checker"):
        p = DOCS_IMAGES / f"{name}.png"
        if p.exists():
            print(f"  ✔ {name}.png  ({p.stat().st_size // 1024} KB)")
        else:
            print(f"  ✗ {name}.png  — 失敗")

    root.destroy()
    print("\n完了！")


if __name__ == "__main__":
    main()
