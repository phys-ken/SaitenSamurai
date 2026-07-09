"""
test_answerkey_skip.py — ⑥ 正答未登録問題スキップのテスト

座標ファイルで問題エリアが定義されているが、模範解答テンプレートに
正答が登録されていない問題を採点対象外とする機能のテスト。
"""
import sys
import os
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "main_src"))

from scoring_engine import load_template, score_answers, normalize_value
from mark_checker import detect_errors_checker, load_errors_checker


# ── テンプレート生成ヘルパー ──────────────────────────


def _create_template_xlsx(path, questions):
    """テスト用 answer_key.xlsx を生成する。
    
    Args:
        path: 保存先パス
        questions: [{問題番号, 正答, 配点, 観点}, ...] のリスト。
                   正答が空文字 or None の行は「未登録」を意味する。
    """
    rows = []
    for q in questions:
        rows.append({
            '問題番号': q['問題番号'],
            '正答': q.get('正答', ''),
            '配点': q.get('配点', ''),
            '観点': q.get('観点', ''),
        })
    df = pd.DataFrame(rows)
    df.to_excel(path, index=False)


def _create_mark2_result_xlsx(path, n_students, question_numbers, skip_questions=0):
    """テスト用 Mark2結果xlsx を生成する。
    
    エラーチェックに使用。ヘッダー構造を模倣。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Row 1: No, File, 1, 2, 3, ...（元のインデックス）
    total_cols = skip_questions + len(question_numbers)
    headers = ['No', 'File'] + [str(i + 1) for i in range(total_cols)]
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)

    # Row 2: None, None, ID名..., 問題番号...
    id_labels = ['学年', 'クラス', '出席番号（十の位）', '出席番号（一の位）']
    labels = [None, None] + id_labels[:skip_questions] + [str(q) for q in question_numbers]
    for ci, l in enumerate(labels, 1):
        ws.cell(2, ci, l)

    # Row 3+: データ
    rng = np.random.default_rng(42)
    for s in range(n_students):
        row_idx = 3 + s
        ws.cell(row_idx, 1, s + 1)
        ws.cell(row_idx, 2, f'page_{s+1:03d}.png')
        # ID部
        for i in range(skip_questions):
            ws.cell(row_idx, 3 + i, 1)
        # 問題回答（1-5のランダム、一部に空欄やダブルマーク）
        for qi, q in enumerate(question_numbers):
            col = 3 + skip_questions + qi
            r = rng.random()
            if r < 0.05:
                ws.cell(row_idx, col, None)  # 空欄
            elif r < 0.1:
                ws.cell(row_idx, col, '2;3')  # ダブルマーク
            else:
                ws.cell(row_idx, col, int(rng.integers(1, 6)))
    
    wb.save(path)
    wb.close()


# ── テスト ──────────────────────────────────────────


class TestLoadTemplateSkip:
    """load_template() が正答未登録行をスキップするテスト"""

    def test_all_registered(self, tmp_path):
        """全問登録済み → 全問がtemplate_dictに含まれる"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '1', '配点': 3, '観点': 2},
            {'問題番号': 3, '正答': '5', '配点': 2, '観点': 1},
        ])
        td = load_template(path)
        assert set(td.keys()) == {1, 2, 3}

    def test_partial_registered(self, tmp_path):
        """一部未登録 → 未登録問題はスキップされる"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '',  '配点': '',  '観点': ''},   # 未登録
            {'問題番号': 3, '正答': '5', '配点': 2, '観点': 1},
            {'問題番号': 4, '正答': None, '配点': None, '観点': None}, # 未登録(NaN)
            {'問題番号': 5, '正答': '2', '配点': 3, '観点': 2},
        ])
        td = load_template(path)
        assert set(td.keys()) == {1, 3, 5}
        assert 2 not in td
        assert 4 not in td

    def test_all_unregistered(self, tmp_path):
        """全問未登録 → 空辞書"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '', '配点': '', '観点': ''},
            {'問題番号': 2, '正答': '', '配点': '', '観点': ''},
        ])
        td = load_template(path)
        assert td == {}

    def test_answer_registered_but_no_points(self, tmp_path):
        """正答はあるが配点なし → スキップ"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': '', '観点': 1},
            {'問題番号': 2, '正答': '1', '配点': 2, '観点': 1},
        ])
        td = load_template(path)
        assert set(td.keys()) == {2}

    def test_values_correct(self, tmp_path):
        """登録済み問題の値が正しく読まれる"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '',  '配点': '',  '観点': ''},
            {'問題番号': 3, '正答': '5', '配点': 3, '観点': 2},
        ])
        td = load_template(path)
        assert td[1] == {'正答': '3', '配点': 2, '観点': 1, '問題概要': '', '特例': ''}
        assert td[3] == {'正答': '5', '配点': 3, '観点': 2, '問題概要': '', '特例': ''}

    def test_no_crash_on_nan_points(self, tmp_path):
        """NaN配点でクラッシュしない（以前はint(NaN)でValueError）"""
        path = tmp_path / "key.xlsx"
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '1', '配点': None, '観点': None},
        ])
        # 以前のコードではここでValueError
        td = load_template(path)
        assert 1 in td
        assert 2 not in td

    def test_trailing_empty_rows(self, tmp_path):
        """Excel末尾の空行でクラッシュしない（問題番号がNaN）"""
        path = tmp_path / "key.xlsx"
        # 通常の行 + 末尾に空行（問題番号もNaN）
        _create_template_xlsx(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '1', '配点': 3, '観点': 2},
            {'問題番号': None, '正答': None, '配点': None, '観点': None},  # 空行
            {'問題番号': '', '正答': '', '配点': '', '観点': ''},           # 空行
        ])
        td = load_template(path)
        assert set(td.keys()) == {1, 2}

    def test_float_string_points(self, tmp_path):
        """配点が文字列 '2.0' でもクラッシュしない"""
        import openpyxl
        path = tmp_path / "key.xlsx"
        # openpyxlで直接テキスト形式のセルを作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['問題番号', '正答', '配点', '観点'])
        ws.append([1, '3', '2.0', '1.0'])  # 文字列として配点・観点
        ws.append([2, '1', 3, 2])           # 通常の数値
        wb.save(path)
        wb.close()
        td = load_template(path)
        assert set(td.keys()) == {1, 2}
        assert td[1]['配点'] == 2
        assert td[1]['観点'] == 1


class TestDetectErrorsWithFilter:
    """detect_errors_checker() の registered_questions フィルタテスト"""

    def test_no_filter_checks_all(self, tmp_path):
        """フィルタなし → 全問チェック"""
        xlsx = tmp_path / "result.xlsx"
        csv_out = tmp_path / "errors.csv"
        _create_mark2_result_xlsx(xlsx, n_students=5, question_numbers=list(range(1, 11)))
        detect_errors_checker(xlsx, csv_out, registered_questions=None)
        errors_df = load_errors_checker(csv_out)
        # 全10問がチェック対象
        if len(errors_df) > 0:
            assert errors_df['question_no'].max() <= 10

    def test_filter_limits_questions(self, tmp_path):
        """フィルタあり → 指定問題のみチェック"""
        xlsx = tmp_path / "result.xlsx"
        csv_out = tmp_path / "errors.csv"
        _create_mark2_result_xlsx(xlsx, n_students=10, question_numbers=list(range(1, 11)))
        detect_errors_checker(xlsx, csv_out, registered_questions={1, 3, 5})
        errors_df = load_errors_checker(csv_out)
        if len(errors_df) > 0:
            found_questions = set(errors_df['question_no'].unique())
            assert found_questions.issubset({1, 3, 5}), \
                f"未登録問題のエラーが含まれている: {found_questions - {1, 3, 5}}"

    def test_empty_filter_no_errors(self, tmp_path):
        """空セットでフィルタ → エラー0件"""
        xlsx = tmp_path / "result.xlsx"
        csv_out = tmp_path / "errors.csv"
        _create_mark2_result_xlsx(xlsx, n_students=5, question_numbers=list(range(1, 11)))
        detect_errors_checker(xlsx, csv_out, registered_questions=set())
        errors_df = load_errors_checker(csv_out)
        assert len(errors_df) == 0


class TestScoringWithPartialTemplate:
    """正答未登録問題がある場合の採点テスト"""

    def test_score_only_registered(self, tmp_path):
        """未登録問題は採点結果に含まれない"""
        key_path = tmp_path / "key.xlsx"
        _create_template_xlsx(key_path, [
            {'問題番号': 1, '正答': '2', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '',  '配点': '',  '観点': ''},   # 未登録
            {'問題番号': 3, '正答': '4', '配点': 3, '観点': 2},
        ])
        td = load_template(key_path)
        
        # 解答データ: 全3問に回答あり
        answers = {1: '2', 2: '3', 3: '4'}
        result = score_answers(answers, td)
        
        # 問題1,3のみ採点される（results は辞書 {問題番号: {...}}）
        scored_questions = set(result['results'].keys())
        assert 1 in scored_questions
        assert 3 in scored_questions
        assert 2 not in scored_questions, "未登録問題2が採点結果に含まれている"
        
        # 得点は問1(正答: 2点) + 問3(正答: 3点) = 5点
        assert result['total_score'] == 5

    def test_max_score_excludes_unregistered(self, tmp_path):
        """満点計算に未登録問題は含まれない"""
        key_path = tmp_path / "key.xlsx"
        _create_template_xlsx(key_path, [
            {'問題番号': 1, '正答': '1', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '',  '配点': '',  '観点': ''},
            {'問題番号': 3, '正答': '3', '配点': 3, '観点': 1},
        ])
        td = load_template(key_path)
        answers = {1: '1', 2: '1', 3: '3'}
        result = score_answers(answers, td)
        assert result['max_score'] == 5  # 2 + 3 = 5
