"""
test_multi_digit_mode.py — 複数桁設問モード（共通テスト数学式・完答採点）のテスト

トップ画面の「数学マーク採点」系モードで有効になる mark_format=multi_digit の挙動を検証する。

- constants: 記号対応表の整合
- omr_engine.parse_excel_coordinates: raw_choice が row0 の値ヘッダになること（標準テンプレは従来値と不変）
- choice_to_position_index: multi_digit の位置解決（-,0..9,a-d / 0⇔10非適用）
- load_template: 範囲表記「1-3」・バリデーション・標準モード非影響
- score_answers: グループ完答判定・無マーク/ダブルマーク無効・特例全員正解
"""
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "main_src"))

from constants import (
    MARK_FORMAT_STANDARD,
    MARK_FORMAT_MULTI_DIGIT,
    MULTI_DIGIT_VALUE_TO_SYMBOL,
    MULTI_DIGIT_SYMBOL_TO_VALUE,
)
from scoring_engine import (
    load_template,
    score_answers,
    choice_to_position_index,
    SPECIAL_ALL_CORRECT,
)
from omr_engine import parse_excel_coordinates, save_recognition_results


# ── ヘルパー ──────────────────────────────────────────


def _create_answer_key(path, questions, include_special_col=True):
    """テスト用 answer_key.xlsx を生成する（test_special_all_correct.py と同形式）"""
    rows = []
    for q in questions:
        row = {
            '問題番号': q['問題番号'],
            '正答': q.get('正答', ''),
            '配点': q.get('配点', ''),
            '観点': q.get('観点', ''),
        }
        if include_special_col:
            row['特例'] = q.get('特例', '')
        rows.append(row)
    df = pd.DataFrame(rows)
    df.to_excel(path, index=False)


def _create_coord_xlsx(path, header_values, num_questions=3):
    """テスト用座標ファイルを生成する。

    row0 に値ヘッダ（base_col=4,8,...）、row3以降に設問行（左から昇順のx座標）。
    """
    n = len(header_values)
    width = 4 + n * 4
    rows = []
    # row0: 値ヘッダ
    r0 = [''] * width
    r0[0], r0[2], r0[3] = 'no', 'page', 'type'
    for i, hv in enumerate(header_values):
        r0[4 + i * 4] = hv
    rows.append(r0)
    # row1, row2: position/size, x/y のダミー多段ヘッダ
    rows.append([''] * width)
    rows.append([''] * width)
    # row3以降: 設問行
    for q in range(1, num_questions + 1):
        r = [''] * width
        r[0] = q
        r[1] = f'設問{q}'
        for i in range(n):
            base = 4 + i * 4
            r[base] = 100 + i * 20      # x（ヘッダ順=物理順）
            r[base + 1] = 50 * q        # y
            r[base + 2] = 10            # width
            r[base + 3] = 10            # height
        rows.append(r)
    pd.DataFrame(rows).to_excel(path, index=False, header=False)


MULTI_DIGIT_HEADERS = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
STANDARD_HEADERS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 0]


# ── constants ──────────────────────────────────────────


class TestSymbolTables:
    def test_roundtrip(self):
        for v, s in MULTI_DIGIT_VALUE_TO_SYMBOL.items():
            assert MULTI_DIGIT_SYMBOL_TO_VALUE[s] == v

    def test_expected_symbols(self):
        assert MULTI_DIGIT_VALUE_TO_SYMBOL[-1] == '-'
        assert MULTI_DIGIT_VALUE_TO_SYMBOL[0] == '0'
        assert MULTI_DIGIT_VALUE_TO_SYMBOL[9] == '9'
        assert MULTI_DIGIT_VALUE_TO_SYMBOL[10] == 'a'
        assert MULTI_DIGIT_VALUE_TO_SYMBOL[13] == 'd'
        assert len(MULTI_DIGIT_VALUE_TO_SYMBOL) == 15


# ── parse_excel_coordinates: raw_choice = 値ヘッダ ──────


class TestParseCoordinatesRawChoice:
    def test_multi_digit_headers(self, tmp_path):
        """15列テンプレート: raw_choice が -1..13 のヘッダ値になる"""
        path = tmp_path / "coord.xlsx"
        _create_coord_xlsx(path, MULTI_DIGIT_HEADERS)
        coords, _ = parse_excel_coordinates(str(path))
        q1 = sorted([c for c in coords if c['question_no'] == 1], key=lambda c: c['choice'])
        assert [c['raw_choice'] for c in q1] == MULTI_DIGIT_HEADERS
        # 物理位置: choice = ヘッダ値 + 1
        for c in q1:
            assert c['choice'] == c['raw_choice'] + 1

    def test_standard_headers_regression(self, tmp_path):
        """標準10列テンプレート: raw_choice はヘッダ値(=従来の出現順と同値)のまま"""
        path = tmp_path / "coord.xlsx"
        _create_coord_xlsx(path, STANDARD_HEADERS)
        coords, _ = parse_excel_coordinates(str(path))
        q1 = sorted([c for c in coords if c['question_no'] == 1], key=lambda c: c['choice'])
        assert [c['raw_choice'] for c in q1] == STANDARD_HEADERS

    def test_missing_headers_fallback(self, tmp_path):
        """row0ヘッダが欠損している場合は従来どおり出現順インデックス"""
        path = tmp_path / "coord.xlsx"
        _create_coord_xlsx(path, ['', '', ''])
        coords, _ = parse_excel_coordinates(str(path))
        q1 = sorted([c for c in coords if c['question_no'] == 1], key=lambda c: c['choice'])
        assert [c['raw_choice'] for c in q1] == [0, 1, 2]


# ── choice_to_position_index ───────────────────────────


class TestChoiceToPositionMultiDigit:
    def test_symbols(self):
        md = MARK_FORMAT_MULTI_DIGIT
        assert choice_to_position_index('-', 15, md) == 0
        assert choice_to_position_index('0', 15, md) == 1
        assert choice_to_position_index('9', 15, md) == 10
        assert choice_to_position_index('a', 15, md) == 11
        assert choice_to_position_index('d', 15, md) == 14
        assert choice_to_position_index('A', 15, md) == 11  # 大文字も受理

    def test_numeric_values(self):
        """座標ヘッダ由来の数値表記(-1, 10〜13)も受理する"""
        md = MARK_FORMAT_MULTI_DIGIT
        assert choice_to_position_index('-1', 15, md) == 0
        assert choice_to_position_index(-1, 15, md) == 0
        assert choice_to_position_index('10', 15, md) == 11
        assert choice_to_position_index(13, 15, md) == 14

    def test_no_zero_ten_equivalence(self):
        """multi_digit では '10' は値10(=a)であり '0' と等価ではない"""
        md = MARK_FORMAT_MULTI_DIGIT
        assert choice_to_position_index('10', 15, md) != choice_to_position_index('0', 15, md)

    def test_out_of_range(self):
        md = MARK_FORMAT_MULTI_DIGIT
        assert choice_to_position_index('e', 15, md) is None
        assert choice_to_position_index('14', 15, md) is None
        assert choice_to_position_index('d', 10, md) is None  # num_choices超過

    def test_standard_unchanged(self):
        """standard は従来ロジック（1..9,0並び・0⇔10等価）のまま"""
        assert choice_to_position_index('1', 10) == 0
        assert choice_to_position_index('0', 10) == 9
        assert choice_to_position_index('10', 10) == 9
        assert choice_to_position_index('-', 10) is None


# ── load_template: 範囲表記 ────────────────────────────


class TestLoadTemplateMultiDigit:
    def test_range_and_single(self, tmp_path):
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-3', '正答': '-24', '配点': 3, '観点': 1},
            {'問題番号': 4, '正答': 'a', '配点': 2, '観点': 1},
            {'問題番号': '5-6', '正答': '15', '配点': 2, '観点': 2},
        ])
        td = load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert set(td.keys()) == {1, 4, 5}
        assert td[1] == {**td[1], '正答': '-24', 'span': 3, 'group_label': '1-3'}
        assert td[4]['span'] == 1 and td[4]['group_label'] == '4'
        assert td[5]['span'] == 2 and td[5]['正答'] == '15'

    def test_fullwidth_hyphen_and_uppercase(self, tmp_path):
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1－2', '正答': 'AB', '配点': 2, '観点': 1},
        ])
        td = load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert td[1]['span'] == 2
        assert td[1]['正答'] == 'ab'

    def test_numeric_answer_cell(self, tmp_path):
        """Excelが正答を数値(-24)として保持していても文字列化される"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-3', '正答': -24, '配点': 3, '観点': 1},
        ])
        td = load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert td[1]['正答'] == '-24'

    def test_length_mismatch_raises(self, tmp_path):
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-3', '正答': '-2', '配点': 3, '観点': 1},
        ])
        with pytest.raises(ValueError, match='文字列書式'):
            load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_invalid_symbol_raises(self, tmp_path):
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-3', '正答': '-2x', '配点': 3, '観点': 1},
        ])
        with pytest.raises(ValueError, match='使用できない文字'):
            load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_overlap_raises(self, tmp_path):
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-3', '正答': '-24', '配点': 3, '観点': 1},
            {'問題番号': 3, '正答': '5', '配点': 2, '観点': 1},
        ])
        with pytest.raises(ValueError, match='重複'):
            load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_reversed_range_raises(self, tmp_path):
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '3-1', '正答': '-24', '配点': 3, '観点': 1},
        ])
        with pytest.raises(ValueError, match='昇順'):
            load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_multiple_answers_rejected(self, tmp_path):
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-2', '正答': '12;34', '配点': 2, '観点': 1},
        ])
        with pytest.raises(ValueError, match='複数正答'):
            load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_all_correct_with_empty_answer(self, tmp_path):
        """特例=全員正解はグループでも正答空欄を許容"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-3', '正答': '', '配点': 3, '観点': 1, '特例': '全員正解'},
        ])
        td = load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert td[1]['特例'] == SPECIAL_ALL_CORRECT
        assert td[1]['span'] == 3

    def test_auto_span_from_answer_length(self, tmp_path):
        """単独表記+複数文字正答 → 正答文字数ぶん自動割付される"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1},
            {'問題番号': 4, '正答': 'a', '配点': 2, '観点': 1},
            {'問題番号': 5, '正答': '15', '配点': 2, '観点': 2},
        ])
        td = load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert td[1]['span'] == 3
        assert td[1]['group_label'] == '1-3'
        assert td[4]['span'] == 1
        assert td[5]['span'] == 2
        assert td[5]['group_label'] == '5-6'

    def test_auto_span_overlap_detected(self, tmp_path):
        """自動割付の消費先に別の登録があれば重複エラー"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1},  # 1〜3を自動消費
            {'問題番号': 3, '正答': '5', '配点': 2, '観点': 1},    # 3と衝突
        ])
        with pytest.raises(ValueError, match='重複'):
            load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_auto_span_and_explicit_range_mixed(self, tmp_path):
        """自動割付と明示範囲の併用が可能"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1},
            {'問題番号': '4-5', '正答': '15', '配点': 2, '観点': 1},
        ])
        td = load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert td[1]['span'] == 3
        assert td[4]['span'] == 2
        assert td[4]['group_label'] == '4-5'

    def test_explicit_range_still_validates_length(self, tmp_path):
        """明示範囲は従来どおり範囲長≠正答文字数でエラー（自動修正しない）"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-3', '正答': '-2', '配点': 3, '観点': 1},
        ])
        with pytest.raises(ValueError, match='文字列書式'):
            load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_standard_mode_unaffected(self, tmp_path):
        """標準モードは従来どおり（範囲表記なしのanswer_keyが同一結果）"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': 1, '正答': '3', '配点': 2, '観点': 1},
            {'問題番号': 2, '正答': '10', '配点': 2, '観点': 1},
        ])
        td = load_template(path)
        assert td[1]['正答'] == '3'
        assert 'span' not in td[1]
        assert 'group_label' not in td[1]


# ── score_answers: グループ完答判定 ────────────────────


def _md_template(tmp_path, questions):
    path = tmp_path / "key.xlsx"
    _create_answer_key(path, questions)
    return load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)


class TestScoreAnswersMultiDigit:
    def test_full_correct(self, tmp_path):
        td = _md_template(tmp_path, [
            {'問題番号': '1-3', '正答': '-24', '配点': 3, '観点': 1},
            {'問題番号': 4, '正答': 'a', '配点': 2, '観点': 2},
        ])
        res = score_answers({1: '-', 2: '2', 3: '4', 4: 'a'}, td,
                            mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert res['total_score'] == 5
        assert res['max_score'] == 5
        assert res['results'][1]['correct'] is True
        assert res['results'][1]['student_answer'] == '-24'
        assert res['results'][1]['span'] == 3
        assert res['results'][1]['group_label'] == '1-3'
        assert res['aspect_scores'] == {1: 3, 2: 2}

    def test_partial_is_zero(self, tmp_path):
        """1行でも違えばグループ全体0点（完答のみ）"""
        td = _md_template(tmp_path, [
            {'問題番号': '1-3', '正答': '-24', '配点': 3, '観点': 1},
        ])
        res = score_answers({1: '-', 2: '2', 3: '5'}, td,
                            mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert res['total_score'] == 0
        assert res['results'][1]['correct'] is False

    def test_blank_row_invalidates_group(self, tmp_path):
        """グループ内の無マーク行 → 不正解"""
        td = _md_template(tmp_path, [
            {'問題番号': '1-3', '正答': '-24', '配点': 3, '観点': 1},
        ])
        res = score_answers({1: '-', 3: '4'}, td, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert res['results'][1]['correct'] is False
        assert res['results'][1]['student_answer'] == '-,,4'

    def test_double_mark_invalidates_group(self, tmp_path):
        """グループ内のダブルマーク行(;連結) → 不正解"""
        td = _md_template(tmp_path, [
            {'問題番号': '1-3', '正答': '-24', '配点': 3, '観点': 1},
        ])
        res = score_answers({1: '-', 2: '2;3', 3: '4'}, td,
                            mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert res['results'][1]['correct'] is False

    def test_no_zero_ten_in_multi_digit(self, tmp_path):
        """'1'+'0' の連結 "10" が "0" に正規化されない"""
        td = _md_template(tmp_path, [
            {'問題番号': '1-2', '正答': '10', '配点': 2, '観点': 1},
        ])
        res = score_answers({1: '1', 2: '0'}, td, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert res['results'][1]['correct'] is True
        td2 = _md_template(tmp_path, [
            {'問題番号': 1, '正答': '0', '配点': 2, '観点': 1},
        ])
        res2 = score_answers({1: '10'}, td2, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert res2['results'][1]['correct'] is False

    def test_all_correct_special_group(self, tmp_path):
        """特例=全員正解はグループ単位で満点（無マークでも）"""
        td = _md_template(tmp_path, [
            {'問題番号': '1-3', '正答': '', '配点': 3, '観点': 1, '特例': '全員正解'},
        ])
        res = score_answers({}, td, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert res['results'][1]['correct'] is True
        assert res['total_score'] == 3

    def test_uppercase_student_answer(self, tmp_path):
        """生徒解答側の大文字も小文字化して比較"""
        td = _md_template(tmp_path, [
            {'問題番号': '1-2', '正答': 'ab', '配点': 2, '観点': 1},
        ])
        res = score_answers({1: 'A', 2: 'B'}, td, mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert res['results'][1]['correct'] is True


# ── CTT/R連携: グループ=1項目・ExactMatch ──────────────


class TestCttMultiDigit:
    def _convert(self, tmp_path):
        from ctt_analyzer import convert_mark2_to_ctt_data
        key_path = tmp_path / "key.xlsx"
        _create_answer_key(key_path, [
            {'問題番号': '1-3', '正答': '-24', '配点': 3, '観点': 1},
            {'問題番号': 4, '正答': 'a', '配点': 2, '観点': 1},
            {'問題番号': '5-6', '正答': '10', '配点': 2, '観点': 1},
        ])
        mark2_results = [
            {'image': 's1.jpg', 'answers': {1: '-', 2: '2', 3: '4', 4: 'a', 5: '1', 6: '0'}},
            {'image': 's2.jpg', 'answers': {1: '-', 2: '2', 3: '5', 4: 'b', 5: '1', 6: '0'}},
            {'image': 's3.jpg', 'answers': {1: '-', 3: '4', 4: 'a', 5: '0', 6: '1'}},
        ]
        return convert_mark2_to_ctt_data(
            str(key_path), None, mark2_results=mark2_results,
            mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_key_df_structure(self, tmp_path):
        _, key_df = self._convert(tmp_path)
        assert key_df['QuestionID'].tolist() == ['1-3', '4', '5-6']
        assert key_df['Key'].tolist() == ['-24', 'a', '10']  # 0⇔10正規化されない
        assert key_df['ExactMatch'].tolist() == [True, True, True]

    def test_ans_df_group_concat(self, tmp_path):
        ans_df, _ = self._convert(tmp_path)
        assert ans_df['1-3'].tolist() == ['-24', '-25', '無効回答']  # s3は無マーク行あり
        assert ans_df['5-6'].tolist() == ['10', '10', '01']

    def test_score_matrix_exact_match(self, tmp_path):
        from ctt_analyzer import CTTAnalyzer
        ans_df, key_df = self._convert(tmp_path)
        analyzer = CTTAnalyzer(ans_df, key_df)
        assert analyzer.score_matrix['1-3'].tolist() == [1, 0, 0]
        assert analyzer.score_matrix['4'].tolist() == [1, 0, 1]
        # '01' ≠ '10'、かつ '10' が '0' に正規化されないこと
        assert analyzer.score_matrix['5-6'].tolist() == [1, 1, 0]

    def test_distractor_analysis_no_crash(self, tmp_path):
        """選択肢分析が非数値の解答文字列('-24'等)でも動作する"""
        from ctt_analyzer import CTTAnalyzer
        ans_df, key_df = self._convert(tmp_path)
        analyzer = CTTAnalyzer(ans_df, key_df)
        df = analyzer.calculate_distractor_analysis()
        g = df[df['QuestionID'] == '1-3']
        choices = g['Choice'].tolist()
        assert '-24' in choices and '-25' in choices and '無効回答' in choices
        # 正答行のフラグ
        assert g[g['Choice'] == '-24']['IsKey'].iloc[0]
        # s3の無効回答が全体カウントに入る
        assert g[g['Choice'] == '無効回答']['Count_全体'].iloc[0] == 1

    def test_standard_conversion_unchanged(self, tmp_path):
        """標準モードのkey_df/正規化ロジックは従来どおり(ExactMatch=False)"""
        from ctt_analyzer import convert_mark2_to_ctt_data
        key_path = tmp_path / "key.xlsx"
        _create_answer_key(key_path, [
            {'問題番号': 1, '正答': '10', '配点': 2, '観点': 1},
        ])
        mark2_results = [{'image': 's1.jpg', 'answers': {1: '0'}}]
        ans_df, key_df = convert_mark2_to_ctt_data(
            str(key_path), None, mark2_results=mark2_results)
        assert key_df['Key'].tolist() == ['0']  # 0⇔10正規化される
        assert key_df['ExactMatch'].tolist() == [False]


# ── image_renderer: グループ描画 ───────────────────────


class TestRendererMultiDigit:
    """複数桁グループの描画（先頭行に○×1つ・各行に正答赤表示）"""

    ROW_Y = {5: 100, 6: 200, 7: 300}  # 座標側の設問番号(skip=4) -> y座標

    def _coords(self):
        coords = []
        for q_no, y in self.ROW_Y.items():
            for i in range(15):
                coords.append({
                    'question_no': q_no,
                    'choice': i,
                    'x': 30 + i * 30,
                    'y': y,
                    'width': 20,
                    'height': 20,
                })
        return coords

    def _white(self):
        return np.zeros((400, 520, 3), dtype=np.uint8) + 255

    def _box_has_ink(self, image, choice_idx, y):
        x = 30 + choice_idx * 30
        crop = image[y:y + 20, x:x + 20]
        return bool((crop != 255).any())

    def _group_result(self):
        return {
            'results': {
                1: {
                    'correct': False,
                    'correct_answer': '-24',
                    'student_answer': '-25',
                    'points': 0,
                    'max_points': 3,
                    'aspect': 1,
                    'special': '',
                    'span': 3,
                    'group_label': '1-3',
                }
            },
            'total_score': 0, 'max_score': 3,
            'aspect_scores': {1: 0}, 'aspect_max_scores': {1: 3},
        }

    def test_correct_answer_drawn_on_each_row(self):
        """誤答時、正答記号がグループ各行の正しいマーク位置に描かれる"""
        from image_renderer import draw_scoring_results
        settings = {'show_correct_answer': True, 'show_ox_mark': False,
                    'show_score': False, 'show_aspect': False}
        result = draw_scoring_results(
            self._white(), self._coords(), self._group_result(),
            skip_questions=4, rendering_settings=settings,
            mark_format=MARK_FORMAT_MULTI_DIGIT,
        )
        # '-' → 位置0(行1)、'2' → 位置3(行2)、'4' → 位置5(行3)
        assert self._box_has_ink(result, 0, 100)
        assert self._box_has_ink(result, 3, 200)
        assert self._box_has_ink(result, 5, 300)
        # 行2の位置0（誤った位置）には描かれない
        assert not self._box_has_ink(result, 0, 200)

    def test_ox_mark_only_on_first_row(self):
        """○×はグループ先頭行の既定位置(後ろから2番目)にのみ描かれる"""
        from image_renderer import draw_scoring_results
        settings = {'show_correct_answer': False, 'show_ox_mark': True,
                    'show_score': False, 'show_aspect': False}
        result = draw_scoring_results(
            self._white(), self._coords(), self._group_result(),
            skip_questions=4, rendering_settings=settings,
            mark_format=MARK_FORMAT_MULTI_DIGIT,
        )
        assert self._box_has_ink(result, 13, 100)       # 先頭行 num_choices-2
        assert not self._box_has_ink(result, 13, 200)   # 2行目には無し
        assert not self._box_has_ink(result, 13, 300)   # 3行目には無し

    def test_all_correct_star_on_first_symbol(self):
        """特例★はグループ先頭行の正答1文字目位置に描かれる"""
        from image_renderer import draw_scoring_results
        res = self._group_result()
        res['results'][1]['correct'] = True
        res['results'][1]['special'] = '全員正解'
        settings = {'show_correct_answer': False, 'show_ox_mark': False,
                    'show_score': False, 'show_aspect': False,
                    'show_all_correct_star': True}
        result = draw_scoring_results(
            self._white(), self._coords(), res,
            skip_questions=4, rendering_settings=settings,
            mark_format=MARK_FORMAT_MULTI_DIGIT,
        )
        assert self._box_has_ink(result, 0, 100)  # '-' = 位置0

    def test_bg_white_no_crash(self):
        """白塗りON(2パス描画)でもグループ描画がクラッシュしない"""
        from image_renderer import draw_scoring_results
        settings = {'mark_result_bg_white': True}
        result = draw_scoring_results(
            self._white(), self._coords(), self._group_result(),
            skip_questions=4, rendering_settings=settings,
            mark_format=MARK_FORMAT_MULTI_DIGIT,
        )
        assert result is not None


# ── load_template: Excel日付変換・モード違いの検出 ─────


class TestLoadTemplateInputGuards:
    def _key_with_date_cell(self, tmp_path):
        """Excelが「1-3」を日付(1月3日)に自動変換した状態を再現"""
        import datetime
        from openpyxl import Workbook
        path = tmp_path / "key.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(['問題番号', '正答', '配点', '観点'])
        ws.cell(row=2, column=1, value=datetime.datetime(2026, 1, 3))
        ws.cell(row=2, column=2, value='-24')
        ws.cell(row=2, column=3, value=3)
        ws.cell(row=2, column=4, value=1)
        wb.save(path)
        return path

    def test_date_cell_hint_multi_digit(self, tmp_path):
        path = self._key_with_date_cell(tmp_path)
        with pytest.raises(ValueError, match='日付に自動変換'):
            load_template(path, mark_format=MARK_FORMAT_MULTI_DIGIT)

    def test_date_cell_hint_standard(self, tmp_path):
        """標準モードでもTypeError生クラッシュではなくヒント付きValueError"""
        path = self._key_with_date_cell(tmp_path)
        with pytest.raises(ValueError, match='日付に自動変換'):
            load_template(path)

    def test_range_notation_in_standard_mode_hint(self, tmp_path):
        """標準モードで範囲表記を読むと数学モードへの誘導メッセージ"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': '1-3', '正答': '-24', '配点': 3, '観点': 1},
        ])
        with pytest.raises(ValueError, match='数学マーク採点'):
            load_template(path)

    def test_garbage_question_no_standard(self, tmp_path):
        """解釈不能な問題番号は日本語メッセージのValueError"""
        path = tmp_path / "key.xlsx"
        _create_answer_key(path, [
            {'問題番号': 'あ', '正答': '3', '配点': 2, '観点': 1},
        ])
        with pytest.raises(ValueError, match='数値として解釈できません'):
            load_template(path)


# ── マークチェック訂正: 複数桁記号の対応 ───────────────


class TestCheckerCorrectionSymbols:
    def test_update_xlsx_writes_minus_without_escape(self, tmp_path):
        """訂正値'-'がエスケープされず素の文字列でセルに書かれる"""
        from openpyxl import Workbook, load_workbook
        from mark_checker import update_xlsx_from_csv_checker
        xlsx = tmp_path / "omr.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws.append(['No', 'File', '5', '6'])
        ws.append(['', '', 1, 2])
        ws.append([1, 'img1.jpg', '', '3'])
        wb.save(xlsx)
        csv = tmp_path / "err.csv"
        pd.DataFrame([
            {'filename': 'img1.jpg', 'question_no': 1, 'before': '', 'after': '-', 'error_type': 'NoMark'},
            {'filename': 'img1.jpg', 'question_no': 2, 'before': '3', 'after': 'a', 'error_type': ''},
        ]).to_csv(csv, index=False, encoding='utf-8-sig')
        n = update_xlsx_from_csv_checker(str(xlsx), str(csv))
        assert n == 2
        ws2 = load_workbook(xlsx)['Sheet1']
        assert ws2.cell(row=3, column=3).value == '-'   # "'-" ではない
        assert ws2.cell(row=3, column=4).value == 'a'

    def _stub_checker(self, mark_format):
        import tkinter as tk
        from unittest.mock import MagicMock
        from conftest import get_shared_tk_root
        from gui_components import MarkCheckerGUI
        root = get_shared_tk_root()
        chk = object.__new__(MarkCheckerGUI)
        chk.window = root
        chk.mark_format = mark_format
        chk.current_index = 0
        chk._all_entries_df = pd.DataFrame([
            {'filename': 'img1.jpg', 'question_no': 1, 'before': '', 'after': '', 'error_type': 'NoMark'},
        ])
        chk.correction_entry = tk.Entry(root)
        chk._csv_dirty = False
        chk._unsaved_count = 0
        chk._save_interval = 100
        return chk

    def test_ui_accepts_symbols_in_multi_digit(self):
        chk = self._stub_checker(MARK_FORMAT_MULTI_DIGIT)
        chk.correction_entry.insert(0, '-')
        assert chk.save_current_correction() is True
        assert chk._all_entries_df.at[0, 'after'] == '-'
        chk.correction_entry.delete(0, 'end')
        chk.correction_entry.insert(0, 'A')  # 大文字→小文字化
        assert chk.save_current_correction() is True
        assert chk._all_entries_df.at[0, 'after'] == 'a'

    def test_ui_rejects_invalid_symbol_in_multi_digit(self):
        from unittest.mock import patch
        chk = self._stub_checker(MARK_FORMAT_MULTI_DIGIT)
        chk.correction_entry.insert(0, 'x')
        with patch('gui_components.messagebox') as mb:
            assert chk.save_current_correction() is False
            assert mb.showwarning.called

    def test_ui_standard_still_integer_only(self):
        from unittest.mock import patch
        chk = self._stub_checker(MARK_FORMAT_STANDARD)
        chk.correction_entry.insert(0, 'a')
        with patch('gui_components.messagebox') as mb:
            assert chk.save_current_correction() is False

    def test_count_pending_corrections(self):
        chk = self._stub_checker(MARK_FORMAT_STANDARD)
        chk._all_entries_df = pd.DataFrame([
            {'after': '3'}, {'after': ''}, {'after': 'skip'},
            {'after': None}, {'after': '-1'},
        ])
        assert chk._count_pending_corrections() == 2  # '3' と '-1'


# ── GUI: モード選択とセッションガード ──────────────────


class TestGuiMultiDigit:
    def test_startup_dialog_select_sets_mark_format(self):
        """_selectでapp_modeとmark_formatの両方が設定される"""
        from unittest.mock import MagicMock
        from gui_components import StartupModeDialog
        from constants import MODE_MARK_ONLY
        d = object.__new__(StartupModeDialog)
        d.dialog = MagicMock()
        d._select(MODE_MARK_ONLY, MARK_FORMAT_MULTI_DIGIT)
        assert d.result == MODE_MARK_ONLY
        assert d.mark_format == MARK_FORMAT_MULTI_DIGIT

    def test_startup_dialog_default_is_standard(self):
        """既存3ボタン相当の_select呼び出しはstandardのまま"""
        from unittest.mock import MagicMock
        from gui_components import StartupModeDialog
        from constants import MODE_DESCRIPTIVE_ONLY
        d = object.__new__(StartupModeDialog)
        d.dialog = MagicMock()
        d._select(MODE_DESCRIPTIVE_ONLY)
        assert d.mark_format == MARK_FORMAT_STANDARD

    def test_gui_holds_mark_format_and_title(self):
        """SaitenSamuraiGUIがmark_formatを保持し、タイトルに数学表記が入る"""
        import tkinter as tk
        from conftest import get_shared_tk_root
        from main_gui import SaitenSamuraiGUI
        from constants import MODE_MARK_ONLY
        root = get_shared_tk_root()
        top = tk.Toplevel(root)
        top.withdraw()
        try:
            app = SaitenSamuraiGUI(top, mode=MODE_MARK_ONLY,
                                   mark_format=MARK_FORMAT_MULTI_DIGIT)
            assert app.mark_format == MARK_FORMAT_MULTI_DIGIT
        finally:
            top.destroy()

    def test_coord_mismatch_warning(self, tmp_path):
        """数学モード×10列座標／標準モード×15列座標で警告が出る"""
        import tkinter as tk
        from unittest.mock import patch
        from conftest import get_shared_tk_root
        from main_gui import SaitenSamuraiGUI
        from constants import MODE_MARK_ONLY
        std_coord = tmp_path / "std.xlsx"
        _create_coord_xlsx(std_coord, STANDARD_HEADERS, num_questions=8)
        md_coord = tmp_path / "md.xlsx"
        _create_coord_xlsx(md_coord, MULTI_DIGIT_HEADERS, num_questions=8)

        root = get_shared_tk_root()
        top = tk.Toplevel(root)
        top.withdraw()
        try:
            app = SaitenSamuraiGUI(top, mode=MODE_MARK_ONLY,
                                   mark_format=MARK_FORMAT_MULTI_DIGIT)
            # 数学モード×15列 → 警告なしでTrue
            with patch('main_gui.messagebox') as mb:
                assert app._confirm_mark_format_coord_match(str(md_coord), 4) is True
                assert not mb.askyesno.called
            # 数学モード×10列(標準座標) → 警告ダイアログ
            with patch('main_gui.messagebox') as mb:
                mb.askyesno.return_value = False
                assert app._confirm_mark_format_coord_match(str(std_coord), 4) is False
                assert mb.askyesno.called
        finally:
            top.destroy()
        top2 = tk.Toplevel(root)
        top2.withdraw()
        try:
            app2 = SaitenSamuraiGUI(top2, mode=MODE_MARK_ONLY,
                                    mark_format=MARK_FORMAT_STANDARD)
            # 標準モード×15列 → 警告ダイアログ
            with patch('main_gui.messagebox') as mb:
                mb.askyesno.return_value = True
                assert app2._confirm_mark_format_coord_match(str(md_coord), 4) is True
                assert mb.askyesno.called
        finally:
            top2.destroy()

    def test_session_restore_rejects_format_mismatch(self):
        """標準セッションを数学モードで復元しようとするとエラーで中止"""
        import tkinter as tk
        from unittest.mock import patch
        from conftest import get_shared_tk_root
        from main_gui import SaitenSamuraiGUI
        from constants import MODE_MARK_ONLY
        root = get_shared_tk_root()
        top = tk.Toplevel(root)
        top.withdraw()
        try:
            app = SaitenSamuraiGUI(top, mode=MODE_MARK_ONLY,
                                   mark_format=MARK_FORMAT_MULTI_DIGIT)
            state = {"version": 1, "image_folder": "C:/nonexistent",
                     "mark_format": MARK_FORMAT_STANDARD}
            with patch('main_gui.messagebox') as mb:
                assert app._apply_session_state(state) is False
                assert mb.showerror.called
        finally:
            top.destroy()


# ── save_recognition_results: 記号変換 ─────────────────


class TestSaveRecognitionResultsSymbols:
    def _run(self, tmp_path, mark_format):
        coord_path = tmp_path / "coord.xlsx"
        _create_coord_xlsx(coord_path, MULTI_DIGIT_HEADERS, num_questions=3)
        coords, _ = parse_excel_coordinates(str(coord_path))
        out = tmp_path / "omr.xlsx"
        # 設問1: 「-」(choice0)、設問2: 「a」(choice11)、設問3: ダブルマーク(choice1=0, choice2=1)
        results = [{'image': 'img1.jpg', 'marks': {1: [0], 2: [11], 3: [1, 2]}}]
        save_recognition_results(str(out), results, [1, 2, 3],
                                 coordinates=coords, mark_format=mark_format)
        from openpyxl import load_workbook
        ws = load_workbook(out).active
        return [ws.cell(row=3, column=col).value for col in range(1, 6)]  # Row3 = 最初のデータ行

    def test_multi_digit_symbols(self, tmp_path):
        row = self._run(tmp_path, MARK_FORMAT_MULTI_DIGIT)
        assert row[2] == '-'
        assert row[3] == 'a'
        assert row[4] == '0;1'  # ダブルマークは記号を;連結

    def test_standard_keeps_raw_values(self, tmp_path):
        row = self._run(tmp_path, MARK_FORMAT_STANDARD)
        assert row[2] == '-1'
        assert row[3] == '10'
