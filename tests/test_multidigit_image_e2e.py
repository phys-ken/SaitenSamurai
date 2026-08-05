"""
test_multidigit_image_e2e.py — 複数桁設問モードの画像一気通貫テスト

既存テストは部品（score_answers 等）を厚く覆う一方、部品を組んで回す駆動部
（process_box_drawer / process_scoring）と「画像から始まる複数桁の検証」が
実行されていなかった（fable-review/test-plan.md 優先1）。

合成画像（白地の numpy 配列に矩形を黒塗り）で以下を固定する:
  A. 画像 → recognize_marks → 記号Excel → load_mark2_results → score_answers
     の完答方式・記号表記・skip・全員正解・混在
  B. process_box_drawer 駆動部（コーナーマーカー付き完全版、並列ワーカー実パス）
  C. process_scoring 駆動部（採点結果の描画が実際に画像へ載ること）

紙面レイアウトは Mark2 基準座標系 595x842。マークは 15 個/行
（-, 0〜9, a〜d）。コーナーマーカーのサーチ領域（四隅、幅30%×高さ8%相当）を
避けるため、マーク行は y=150 以降に置く。
"""
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "main_src"))
sys.path.insert(0, str(PROJECT_ROOT / "tests"))

from constants import (
    MARK_FORMAT_MULTI_DIGIT,
    MARKER_X_FRAC_LEFT, MARKER_X_FRAC_RIGHT,
    MARKER_Y_FRAC_TOP, MARKER_Y_FRAC_BOTTOM,
    MULTI_DIGIT_SYMBOL_TO_VALUE,
)
from omr_engine import (
    parse_excel_coordinates,
    recognize_marks,
    save_recognition_results,
)
from scoring_engine import load_template, load_mark2_results, score_answers
from test_multi_digit_mode import _create_answer_key

# ── 紙面レイアウト（Mark2 基準座標系 595x842） ────────────────

SHEET_W, SHEET_H = 595, 842
N_CHOICES = 15
MARK_X0, MARK_DX = 100, 24     # 15個目: x=436..448 (<595)
MARK_Y0, MARK_DY = 150, 44     # コーナーマーカーのサーチ領域(y<56, y>786)を回避
MARK_SIZE = 12
MULTI_DIGIT_HEADERS = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]


def _mark_rect(q_no, position):
    """マーク行 q_no（1始まり）・位置 position（0始まり）の矩形 (x, y, w, h)"""
    return (MARK_X0 + position * MARK_DX,
            MARK_Y0 + (q_no - 1) * MARK_DY,
            MARK_SIZE, MARK_SIZE)


def make_coord_xlsx(path, num_rows):
    """複数桁テンプレート相当の座標Excelを生成（row0=値ヘッダ -1..13）"""
    width = 4 + N_CHOICES * 4
    rows = []
    r0 = [''] * width
    r0[0], r0[2], r0[3] = 'no', 'page', 'type'
    for i, hv in enumerate(MULTI_DIGIT_HEADERS):
        r0[4 + i * 4] = hv
    rows.append(r0)
    rows.append([''] * width)
    rows.append([''] * width)
    for q in range(1, num_rows + 1):
        r = [''] * width
        r[0] = q
        r[1] = f'設問{q}'
        for i in range(N_CHOICES):
            x, y, w, h = _mark_rect(q, i)
            base = 4 + i * 4
            r[base], r[base + 1], r[base + 2], r[base + 3] = x, y, w, h
        rows.append(r)
    pd.DataFrame(rows).to_excel(path, index=False, header=False)
    return path


def make_sheet(filled, with_markers=False):
    """白地の答案画像を生成する。

    Args:
        filled: {行番号: 位置 or [位置,...]} 黒塗りするマーク
        with_markers: 四隅コーナーマーカーを描く（駆動部テスト用）。
            マーカー中心は Mark2 基準比率の位置なので、検出→射影変換が
            ほぼ恒等変換になり、座標Excelの値がそのまま通用する。
    Returns:
        グレースケール ndarray (SHEET_H, SHEET_W)
    """
    img = np.full((SHEET_H, SHEET_W), 255, dtype=np.uint8)
    if with_markers:
        m = 12  # マーカー半サイズ（24px角）
        for fx, fy in [(MARKER_X_FRAC_LEFT, MARKER_Y_FRAC_TOP),
                       (MARKER_X_FRAC_RIGHT, MARKER_Y_FRAC_TOP),
                       (MARKER_X_FRAC_RIGHT, MARKER_Y_FRAC_BOTTOM),
                       (MARKER_X_FRAC_LEFT, MARKER_Y_FRAC_BOTTOM)]:
            cx, cy = int(SHEET_W * fx), int(SHEET_H * fy)
            img[cy - m:cy + m, cx - m:cx + m] = 0
    for q_no, positions in filled.items():
        if not isinstance(positions, (list, tuple)):
            positions = [positions]
        for pos in positions:
            x, y, w, h = _mark_rect(q_no, pos)
            img[y:y + h, x:x + w] = 0
    return img


def sym_positions(answer):
    """正答文字列 '-24' → 各行の物理位置 [0, 3, 5]（位置 = ヘッダ値+1）"""
    return [MULTI_DIGIT_SYMBOL_TO_VALUE[ch] + 1 for ch in answer]


# ── A. 画像 → 認識 → 記号Excel → 採点 ────────────────────────


class TestImageToScore:
    """recognize_marks から採点までの一気通貫（駆動部は使わない高速系）"""

    def _run(self, tmp_path, filled, key_rows, num_rows=4, skip=0):
        """合成画像1枚を認識→保存→読込→採点して (Excelセル値リスト, 採点結果) を返す"""
        coord = make_coord_xlsx(tmp_path / "coord.xlsx", num_rows)
        coords, _groups = parse_excel_coordinates(str(coord))
        img = make_sheet(filled)
        marks = recognize_marks(img, coords)

        xlsx = tmp_path / "omr.xlsx"
        all_q = sorted({c['question_no'] for c in coords})
        save_recognition_results(
            str(xlsx), [{'image': 's1.jpg', 'marks': marks}], all_q,
            question_names={q: str(q) for q in all_q},
            choice_counts={q: N_CHOICES for q in all_q},
            coordinates=coords, mark_format=MARK_FORMAT_MULTI_DIGIT)

        import openpyxl
        ws = openpyxl.load_workbook(xlsx)['Sheet1']
        cells = [ws.cell(3, 2 + i).value for i in range(1, num_rows + 1)]

        key = tmp_path / "key.xlsx"
        _create_answer_key(key, key_rows)
        td = load_template(key, mark_format=MARK_FORMAT_MULTI_DIGIT)
        students = load_mark2_results(str(xlsx), skip_questions=skip)
        assert len(students) == 1
        result = score_answers(students[0]['answers'], td,
                               mark_format=MARK_FORMAT_MULTI_DIGIT)
        return cells, result

    def test_full_correct_with_minus(self, tmp_path):
        """完答（'-24'）: 記号がExcelに載り満点になる"""
        filled = {q: p for q, p in zip([1, 2, 3], sym_positions('-24'))}
        filled[4] = sym_positions('7')[0]
        cells, res = self._run(
            tmp_path, filled,
            [{'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1},
             {'問題番号': 4, '正答': '7', '配点': 2, '観点': 1}])
        assert cells == ['-', '2', '4', '7']
        assert res['total_score'] == 5

    def test_one_wrong_row_zero(self, tmp_path):
        """グループ内1行だけ誤マーク → 完答方式で0点"""
        filled = {1: sym_positions('-')[0], 2: sym_positions('9')[0],  # '2'のはずが'9'
                  3: sym_positions('4')[0]}
        _, res = self._run(
            tmp_path, filled,
            [{'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1}])
        assert res['total_score'] == 0

    def test_no_mark_row_zero(self, tmp_path):
        """グループ内1行無マーク → 0点（Excelセルは空欄）"""
        filled = {1: sym_positions('-')[0], 3: sym_positions('4')[0]}  # 行2を塗らない
        cells, res = self._run(
            tmp_path, filled,
            [{'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1}])
        assert cells[1] in (None, '')
        assert res['total_score'] == 0

    def test_double_mark_row_zero(self, tmp_path):
        """グループ内1行ダブルマーク → 0点（Excelセルは 記号;記号）"""
        filled = {1: sym_positions('-')[0],
                  2: [sym_positions('2')[0], sym_positions('8')[0]],
                  3: sym_positions('4')[0]}
        cells, res = self._run(
            tmp_path, filled,
            [{'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1}])
        assert cells[1] == '2;8'
        assert res['total_score'] == 0

    def test_letters_a_to_d(self, tmp_path):
        """a〜d を含む正答（15マーク目までの物理位置解決）"""
        filled = {q: p for q, p in zip([1, 2], sym_positions('ad'))}
        cells, res = self._run(
            tmp_path, filled,
            [{'問題番号': '1-2', '正答': 'ad', '配点': 2, '観点': 1}],
            num_rows=2)
        assert cells[:2] == ['a', 'd']
        assert res['total_score'] == 2

    def test_skip_questions_offset(self, tmp_path):
        """skip=2（ID行2行）でも行番号がズレない"""
        # 座標行1-2がID行、3-5が解答行（'-24'）
        filled = {1: 5, 2: 8,  # ID行のマーク（値は採点に使われない）
                  3: sym_positions('-')[0], 4: sym_positions('2')[0],
                  5: sym_positions('4')[0]}
        cells, res = self._run(
            tmp_path, filled,
            [{'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1}],
            num_rows=5, skip=2)
        assert res['total_score'] == 3

    def test_all_correct_special_group(self, tmp_path):
        """特例「全員正解」グループ（正答空欄・範囲表記）は無マークでも満点"""
        _, res = self._run(
            tmp_path, {},  # 何も塗らない
            [{'問題番号': '1-2', '正答': '', '配点': 4, '観点': 1,
              '特例': '全員正解'}],
            num_rows=2)
        assert res['total_score'] == 4

    def test_mixed_single_and_group(self, tmp_path):
        """択一設問（span=1）とグループの混在"""
        filled = {q: p for q, p in zip([1, 2, 3], sym_positions('-24'))}
        filled[4] = sym_positions('a')[0]
        _, res = self._run(
            tmp_path, filled,
            [{'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1},
             {'問題番号': 4, '正答': 'a', '配点': 2, '観点': 2}])
        assert res['total_score'] == 5
        assert res['aspect_scores'] == {1: 3, 2: 2}


def _find_omr_xlsx(img_folder):
    """process_box_drawer が書き出したOMR結果Excelを探す（GUIと同じくglob方式）"""
    from constants import RESULTS_FOLDER, RESULTS_DATA_FOLDER, READING_RESULTS_FOLDER_NAME
    folder = img_folder / RESULTS_FOLDER / RESULTS_DATA_FOLDER / READING_RESULTS_FOLDER_NAME
    candidates = sorted(folder.glob('Mark2-Result-*.xlsx'))
    assert candidates, f"OMR結果Excelが見つからない: {folder}"
    return candidates[-1]


# ── B. process_box_drawer 駆動部 ─────────────────────────────


class TestProcessBoxDrawerSmoke:
    """フォルダ一括処理の駆動部（コーナーマーカー検出→射影変換→並列認識→Excel出力）"""

    @pytest.fixture
    def workspace(self, tmp_path):
        import cv2
        coord = make_coord_xlsx(tmp_path / "coord.xlsx", 4)
        img_folder = tmp_path / "scans"
        img_folder.mkdir()
        # 生徒1: '-24' 完答 + 択一 '7'
        s1 = {q: p for q, p in zip([1, 2, 3], sym_positions('-24'))}
        s1[4] = sym_positions('7')[0]
        # 生徒2: 行2が無マーク
        s2 = {1: sym_positions('-')[0], 3: sym_positions('4')[0],
              4: sym_positions('7')[0]}
        for name, filled in [('s1.png', s1), ('s2.png', s2)]:
            cv2.imwrite(str(img_folder / name),
                        make_sheet(filled, with_markers=True))
        return img_folder, coord

    def test_driver_produces_symbol_xlsx(self, workspace, tmp_path):
        from omr_engine import process_box_drawer
        from constants import RESULTS_FOLDER, BOXED_FOLDER, CLEAN_FOLDER
        img_folder, coord = workspace

        progress = []
        result = process_box_drawer(
            str(img_folder), str(coord),
            progress_callback=lambda cur, total: progress.append((cur, total)),
            mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert result['success_count'] == 2 and result['error_count'] == 0

        results_folder = img_folder / RESULTS_FOLDER
        assert (results_folder / BOXED_FOLDER).is_dir()
        assert (results_folder / CLEAN_FOLDER).is_dir()
        assert len(list((results_folder / BOXED_FOLDER).glob('*.png'))) == 2
        assert progress and progress[-1] == (2, 2)

        # 出力Excelのセルが記号表記であること（駆動部経由の本丸）
        # 戻り値にxlsxパスは含まれない（GUIはglobで探す方式）ため同様に探す
        import openpyxl
        omr_xlsx = _find_omr_xlsx(img_folder)
        ws = openpyxl.load_workbook(omr_xlsx)['Sheet1']
        rows = {ws.cell(r, 2).value: [ws.cell(r, c).value for c in range(3, 7)]
                for r in (3, 4)}
        assert rows['s1.png'] == ['-', '2', '4', '7']
        assert rows['s2.png'][0] == '-'
        assert rows['s2.png'][1] in (None, '')   # 無マーク行
        assert rows['s2.png'][2] == '4'


# ── C. process_scoring 駆動部（描画） ─────────────────────────


class TestProcessScoringSmoke:
    """採点結果が実際に答案画像へ描かれること（駆動部経由）"""

    def test_scored_images_written_with_red_marks(self, tmp_path):
        import cv2
        from omr_engine import process_box_drawer
        from image_renderer import process_scoring
        from constants import RESULTS_FOLDER, SCORED_FOLDER

        coord = make_coord_xlsx(tmp_path / "coord.xlsx", 3)
        img_folder = tmp_path / "scans"
        img_folder.mkdir()
        correct = {q: p for q, p in zip([1, 2, 3], sym_positions('-24'))}
        wrong = {1: sym_positions('9')[0], 2: sym_positions('9')[0],
                 3: sym_positions('9')[0]}
        cv2.imwrite(str(img_folder / 'ok.png'), make_sheet(correct, with_markers=True))
        cv2.imwrite(str(img_folder / 'ng.png'), make_sheet(wrong, with_markers=True))

        box_result = process_box_drawer(
            str(img_folder), str(coord), mark_format=MARK_FORMAT_MULTI_DIGIT)
        assert box_result['success_count'] == 2

        key = tmp_path / "key.xlsx"
        _create_answer_key(key, [{'問題番号': 1, '正答': '-24', '配点': 3, '観点': 1}])

        process_scoring(
            str(img_folder), str(coord), str(key),
            str(_find_omr_xlsx(img_folder)),
            mark_format=MARK_FORMAT_MULTI_DIGIT)

        scored = img_folder / RESULTS_FOLDER / SCORED_FOLDER
        outs = {p.name: cv2.imread(str(p)) for p in scored.glob('*.png')}
        assert set(outs) == {'ok.png', 'ng.png'}

        def red_count(bgr):
            b, g, r = bgr[:, :, 0].astype(int), bgr[:, :, 1].astype(int), bgr[:, :, 2].astype(int)
            return int(((r > 150) & (g < 110) & (b < 110)).sum())

        # 両者とも何かしら描かれている（○× と得点）
        assert red_count(outs['ok.png']) > 0
        # 誤答者は各行に正答の赤記号が追加で載るため、赤画素が明確に多い
        assert red_count(outs['ng.png']) > red_count(outs['ok.png'])


# ── 優先3: 15マーク行での認識アルゴリズム ─────────────────────


class TestRecognitionAlgorithms15Marks:
    """K-means とキャリブレーション駆動部が 15マーク/行の紙面で動くこと。

    既存の test_kmeans_omr.py は最大10択の合成のみ。複数桁紙面は
    1行15マーク・行あたり塗りは高々1個（塗り率 1/15 ≈ 6.7%）という
    偏ったクラス比になるため、その条件で固定する。
    """

    def test_kmeans_on_15_mark_rows(self, tmp_path):
        from omr_engine import recognize_marks_kmeans
        coord = make_coord_xlsx(tmp_path / "coord.xlsx", 4)
        coords, _ = parse_excel_coordinates(str(coord))
        assert len(coords) == 60  # 4行×15マーク ≥ KMEANS_MIN_SAMPLES(50)

        filled = {q: p for q, p in zip([1, 2, 3], sym_positions('-24'))}
        filled[4] = sym_positions('7')[0]
        img = make_sheet(filled)

        results, info = recognize_marks_kmeans(img, coords, min_samples=50)
        assert info is not None, "フォールバックせず実K-means経路で走ること"
        expected = {q: [p] for q, p in filled.items()}
        assert results == expected
        assert info['n_marked'] == 4 and info['n_empty'] == 56

    def test_calibration_driver_on_multidigit_sheets(self, tmp_path):
        import cv2
        from threshold_calibrator import (
            run_threshold_calibration,
            collect_mark_fill_ratios,
            reclassify_with_threshold,
        )
        coord = make_coord_xlsx(tmp_path / "coord.xlsx", 4)
        img_folder = tmp_path / "scans"
        img_folder.mkdir()
        s1 = {q: p for q, p in zip([1, 2, 3], sym_positions('-24'))}
        s1[4] = sym_positions('7')[0]
        s2 = {q: p for q, p in zip([1, 2, 3], sym_positions('a05'))}
        s2[4] = sym_positions('0')[0]
        for name, filled in [('s1.png', s1), ('s2.png', s2)]:
            cv2.imwrite(str(img_folder / name),
                        make_sheet(filled, with_markers=True))

        result = run_threshold_calibration(str(img_folder), str(coord))
        assert result['image_count'] == 2
        assert result['error_images'] == []
        assert 0.0 < result['recommended_color_threshold'] < 1.0
        assert 0.0 < result['recommended_area_threshold'] < 1.0

        # スライダー再分類（GUIのリアルタイム更新経路）: 塗った8個だけがマーク側
        all_ratios = []
        for name, gray in result['corrected_images']:
            all_ratios.extend(collect_mark_fill_ratios(
                gray, result['coordinates'],
                result['recommended_color_threshold']))
        re = reclassify_with_threshold(all_ratios, 0.5)
        assert re['total_count'] == 120  # 60マーク×2枚
        assert re['marked_count'] == 8   # 4マーク×2枚
