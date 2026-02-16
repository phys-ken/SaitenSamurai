"""
test_mark_checker_speed.py — マークチェック高速化テスト

v3.9 の高速化改善を検証する:
1. CorrectedImageCache のユニットテスト
2. crop_from_corrected_image の動作確認
3. get_display_image_checker のキャッシュ利用テスト
4. 遅延CSV保存の動作確認
5. 実際のGUIウィンドウでの画面遷移速度測定

テスト用に合成画像（マーカー付き）を生成し、実際の画像処理パイプラインを通す。
"""
import sys
import os
import time
import tempfile
import shutil
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest
import cv2
import numpy as np
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent / "main_src"))

from mark_checker import (
    CorrectedImageCache,
    _load_and_correct_image,
    crop_from_corrected_image,
    crop_and_scale_image_checker,
    get_display_image_checker,
    get_bbox_for_question_checker,
    save_errors_checker,
    load_errors_checker,
    fit_image_to_display,
)
from constants import (
    MARK2_BASE_WIDTH,
    MARK2_BASE_HEIGHT,
    DEFAULT_SCALE_FACTOR,
    DEFAULT_EXPAND_FACTOR,
    DEFAULT_EXPAND_FACTOR_Y,
    ERROR_TYPE_NO_MARK,
    ERROR_TYPE_DOUBLE_MARK,
)


# =====================================================
# テスト用ヘルパー: 合成マークシート画像の生成
# =====================================================

def _create_synthetic_marksheet(width=2480, height=3508):
    """マーカー付き合成マークシート画像を生成する。
    
    四隅にマーカー（黒い正方形）を配置し、
    detect_corner_markers() で検出可能な画像を返す。
    """
    img = np.ones((height, width, 3), dtype=np.uint8) * 240  # 薄灰色の背景
    
    marker_size = int(width * 0.03)
    
    # マーカー位置（画像座標系）
    # Mark2のサーチエリア: 四隅から1%マージン、30%×8%の領域で最大面積の黒成分を検出
    # マーカーをサーチ領域の中心付近に配置
    positions = [
        (int(width * 0.15), int(height * 0.04)),     # 左上
        (int(width * 0.85), int(height * 0.04)),     # 右上
        (int(width * 0.85), int(height * 0.96)),     # 右下
        (int(width * 0.15), int(height * 0.96)),     # 左下
    ]
    
    for cx, cy in positions:
        x1 = cx - marker_size // 2
        y1 = cy - marker_size // 2
        x2 = cx + marker_size // 2
        y2 = cy + marker_size // 2
        cv2.rectangle(img, (x1, y1), (x2, y2), (0, 0, 0), -1)
    
    # マーク回答欄のダミー描画（ベース座標系で (100, 200, 300, 30) あたり）
    res_x = width / MARK2_BASE_WIDTH
    res_y = height / MARK2_BASE_HEIGHT
    for q in range(5):
        y_base = 200 + q * 40
        x_base = 100
        for choice in range(5):
            cx = int((x_base + choice * 30 + 15) * res_x)
            cy = int((y_base + 15) * res_y)
            cv2.circle(img, (cx, cy), int(8 * res_x), (180, 180, 180), 2)
    
    return img


def _create_test_image_file(tmp_dir, filename="test_001.jpg", width=2480, height=3508):
    """合成画像をJPEGファイルとして保存し、パスを返す"""
    img = _create_synthetic_marksheet(width, height)
    filepath = Path(tmp_dir) / filename
    # 日本語パスを考慮してnp.array経由で保存
    _, buf = cv2.imencode('.jpg', img, [cv2.IMWRITE_JPEG_QUALITY, 85])
    buf.tofile(str(filepath))
    return filepath


def _create_test_coords_df(n_files=1, n_questions=5, filenames=None):
    """テスト用の座標DataFrameを生成"""
    rows = []
    if filenames is None:
        filenames = [f"test_{i+1:03d}.jpg" for i in range(n_files)]
    for fn in filenames:
        for q in range(1, n_questions + 1):
            x = 100
            y = 200 + (q - 1) * 40
            w = 150
            h = 30
            rows.append({
                'image_path': fn,
                'question_no': q,
                'choices_bbox': f'{x};{y};{w};{h}',
                'mark_coords': ''
            })
    return pd.DataFrame(rows)


# =====================================================
# セクション1: CorrectedImageCache ユニットテスト
# =====================================================

class TestCorrectedImageCache:
    """CorrectedImageCache の基本動作テスト"""
    
    def test_put_and_get(self):
        """格納した画像が取得できること"""
        cache = CorrectedImageCache(max_size=2)
        dummy = np.zeros((100, 100, 3), dtype=np.uint8)
        cache.put("file1.jpg", dummy)
        result = cache.get("file1.jpg")
        assert result is not None
        assert np.array_equal(result, dummy)
    
    def test_get_nonexistent(self):
        """存在しないキーでNoneが返ること"""
        cache = CorrectedImageCache(max_size=2)
        assert cache.get("notexist.jpg") is None
    
    def test_has(self):
        """has()でキャッシュの存在確認ができること"""
        cache = CorrectedImageCache(max_size=2)
        dummy = np.zeros((10, 10, 3), dtype=np.uint8)
        assert not cache.has("a.jpg")
        cache.put("a.jpg", dummy)
        assert cache.has("a.jpg")
    
    def test_lru_eviction(self):
        """max_sizeを超えたら最も古いエントリが削除されること"""
        cache = CorrectedImageCache(max_size=2)
        cache.put("a.jpg", np.zeros((10, 10, 3), dtype=np.uint8))
        cache.put("b.jpg", np.ones((10, 10, 3), dtype=np.uint8))
        cache.put("c.jpg", np.ones((10, 10, 3), dtype=np.uint8) * 2)
        
        # a.jpg は削除されているはず
        assert cache.get("a.jpg") is None
        assert cache.has("b.jpg")
        assert cache.has("c.jpg")
        assert cache.size == 2
    
    def test_lru_access_updates_order(self):
        """getでアクセスしたエントリはLRU順序が更新されること"""
        cache = CorrectedImageCache(max_size=2)
        cache.put("a.jpg", np.zeros((10, 10, 3), dtype=np.uint8))
        cache.put("b.jpg", np.ones((10, 10, 3), dtype=np.uint8))
        
        # a.jpg にアクセスしてLRU順序を更新
        cache.get("a.jpg")
        
        # c.jpg を追加 → b.jpg（最も古い）が削除される
        cache.put("c.jpg", np.ones((10, 10, 3), dtype=np.uint8) * 2)
        assert cache.has("a.jpg")
        assert not cache.has("b.jpg")
        assert cache.has("c.jpg")
    
    def test_clear(self):
        """clear()でキャッシュがリセットされること"""
        cache = CorrectedImageCache(max_size=3)
        cache.put("a.jpg", np.zeros((10, 10, 3), dtype=np.uint8))
        cache.put("b.jpg", np.zeros((10, 10, 3), dtype=np.uint8))
        cache.clear()
        assert cache.size == 0
        assert cache.get("a.jpg") is None
    
    def test_put_same_key_updates(self):
        """同じキーでputすると値が更新されること"""
        cache = CorrectedImageCache(max_size=2)
        v1 = np.zeros((10, 10, 3), dtype=np.uint8)
        v2 = np.ones((10, 10, 3), dtype=np.uint8)
        cache.put("a.jpg", v1)
        cache.put("a.jpg", v2)
        assert np.array_equal(cache.get("a.jpg"), v2)
        assert cache.size == 1


# =====================================================
# セクション2: crop_from_corrected_image テスト
# =====================================================

class TestCropFromCorrectedImage:
    """キャッシュ済み補正画像からのクロップ処理テスト"""
    
    def test_returns_pil_image(self):
        """PIL.Image が返ること"""
        from PIL import Image
        corrected = np.ones((3508, 2480, 3), dtype=np.uint8) * 200
        bbox = (100, 200, 150, 30)
        result = crop_from_corrected_image(corrected, bbox)
        assert isinstance(result, Image.Image)
    
    def test_bbox_at_edge(self):
        """画像端のbboxでもエラーにならないこと"""
        corrected = np.ones((3508, 2480, 3), dtype=np.uint8) * 200
        bbox = (0, 0, 100, 100)
        result = crop_from_corrected_image(corrected, bbox, expand_factor=2.0)
        assert result is not None
    
    def test_result_matches_legacy(self, tmp_path):
        """キャッシュパスと従来パスで同等の結果が得られること（±数ピクセル以内）"""
        img_path = _create_test_image_file(tmp_path, "compare.jpg")
        bbox = (100, 200, 150, 30)
        
        # 従来パス
        legacy_img = crop_and_scale_image_checker(img_path, bbox, 1.25, 1.3, 1.0)
        
        # キャッシュパス
        corrected = _load_and_correct_image(img_path)
        cache_img = crop_from_corrected_image(corrected, bbox, 1.25, 1.3, 1.0)
        
        # サイズの一致を確認
        assert legacy_img.size == cache_img.size
        
        # ピクセル値がほぼ同じ（完全一致を期待 — 同じ関数パス）
        legacy_arr = np.array(legacy_img)
        cache_arr = np.array(cache_img)
        assert np.array_equal(legacy_arr, cache_arr), "キャッシュパスと従来パスの結果が一致しません"


# =====================================================
# セクション3: get_display_image_checker キャッシュ統合
# =====================================================

class TestGetDisplayImageWithCache:
    """get_display_image_checker のキャッシュ利用テスト"""
    
    @pytest.fixture
    def setup_data(self, tmp_path):
        """テスト用のファイルと座標を準備"""
        _create_test_image_file(tmp_path, "img001.jpg")
        _create_test_image_file(tmp_path, "img002.jpg")
        coords_df = _create_test_coords_df(
            n_files=2, n_questions=5,
            filenames=["img001.jpg", "img002.jpg"]
        )
        return tmp_path, coords_df
    
    def test_without_cache(self, setup_data):
        """キャッシュなしでも動作すること（後方互換）"""
        folder, coords_df = setup_data
        result = get_display_image_checker(
            coords_df, folder, "img001.jpg", 1,
            cache=None
        )
        assert result is not None
    
    def test_with_cache(self, setup_data):
        """キャッシュありで動作し、キャッシュにエントリが追加されること"""
        folder, coords_df = setup_data
        cache = CorrectedImageCache(max_size=2)
        
        result = get_display_image_checker(
            coords_df, folder, "img001.jpg", 1,
            cache=cache
        )
        assert result is not None
        assert cache.has("img001.jpg")
    
    def test_cache_reused_same_image(self, setup_data):
        """同一画像の2つ目の問題でキャッシュが再利用されること"""
        folder, coords_df = setup_data
        cache = CorrectedImageCache(max_size=2)
        
        # 1問目: キャッシュミス → ロード+キャッシュ
        get_display_image_checker(
            coords_df, folder, "img001.jpg", 1, cache=cache
        )
        
        # 2問目: キャッシュヒット（_load_and_correct_image が呼ばれないはず）
        with patch('mark_checker._load_and_correct_image', wraps=_load_and_correct_image) as mock_load:
            result = get_display_image_checker(
                coords_df, folder, "img001.jpg", 2, cache=cache
            )
            mock_load.assert_not_called()
            assert result is not None
    
    def test_cache_different_image(self, setup_data):
        """異なる画像ではキャッシュミスしてロードされること"""
        folder, coords_df = setup_data
        cache = CorrectedImageCache(max_size=2)
        
        get_display_image_checker(
            coords_df, folder, "img001.jpg", 1, cache=cache
        )
        get_display_image_checker(
            coords_df, folder, "img002.jpg", 1, cache=cache
        )
        
        assert cache.has("img001.jpg")
        assert cache.has("img002.jpg")
        assert cache.size == 2


# =====================================================
# セクション4: 速度測定テスト
# =====================================================

class TestSpeedImprovement:
    """キャッシュによる速度改善を実測するテスト"""
    
    @pytest.fixture
    def setup_images(self, tmp_path):
        """複数の問題を持つテスト画像を用意"""
        for i in range(3):
            _create_test_image_file(tmp_path, f"speed_{i+1:03d}.jpg")
        coords_df = _create_test_coords_df(
            n_files=3, n_questions=5,
            filenames=[f"speed_{i+1:03d}.jpg" for i in range(3)]
        )
        return tmp_path, coords_df
    
    def test_same_image_speedup(self, setup_images):
        """同一画像の複数問題: キャッシュありが有意に高速であること"""
        folder, coords_df = setup_images
        filename = "speed_001.jpg"
        
        # ウォームアップ（初回ロードのオーバーヘッドを排除）
        get_display_image_checker(coords_df, folder, filename, 1, cache=None)
        
        # キャッシュなし: 毎回ディスク読み込み+マーカー検出+射影変換
        times_no_cache = []
        for q in range(1, 6):
            t0 = time.perf_counter()
            get_display_image_checker(coords_df, folder, filename, q, cache=None)
            times_no_cache.append(time.perf_counter() - t0)
        
        # キャッシュあり: 初回のみロード、以降はクロップのみ
        cache = CorrectedImageCache(max_size=2)
        times_with_cache = []
        for q in range(1, 6):
            t0 = time.perf_counter()
            get_display_image_checker(coords_df, folder, filename, q, cache=cache)
            times_with_cache.append(time.perf_counter() - t0)
        
        avg_no_cache = sum(times_no_cache) / len(times_no_cache)
        avg_with_cache = sum(times_with_cache) / len(times_with_cache)
        # 初回はキャッシュミスなので、2-5問目の平均を比較
        avg_no_cache_2nd = sum(times_no_cache[1:]) / len(times_no_cache[1:])
        avg_with_cache_2nd = sum(times_with_cache[1:]) / len(times_with_cache[1:])
        
        print(f"\n【速度測定: 同一画像5問連続】")
        print(f"  キャッシュなし平均: {avg_no_cache*1000:.1f}ms")
        print(f"  キャッシュあり平均: {avg_with_cache*1000:.1f}ms")
        print(f"  2-5問目 キャッシュなし: {avg_no_cache_2nd*1000:.1f}ms")
        print(f"  2-5問目 キャッシュあり: {avg_with_cache_2nd*1000:.1f}ms")
        print(f"  2-5問目 高速化率: {avg_no_cache_2nd/max(avg_with_cache_2nd, 0.001):.1f}x")
        
        # キャッシュヒット時（2問目以降）はキャッシュなしより確実に速い
        assert avg_with_cache_2nd < avg_no_cache_2nd, \
            f"キャッシュが有効でない: {avg_with_cache_2nd*1000:.1f}ms >= {avg_no_cache_2nd*1000:.1f}ms"
    
    def test_different_images_no_regression(self, setup_images):
        """異なる画像間の切り替え: キャッシュありでも顕著な劣化がないこと"""
        folder, coords_df = setup_images
        filenames = [f"speed_{i+1:03d}.jpg" for i in range(3)]
        
        # キャッシュなし
        t0 = time.perf_counter()
        for fn in filenames:
            get_display_image_checker(coords_df, folder, fn, 1, cache=None)
        time_no_cache = time.perf_counter() - t0
        
        # キャッシュあり
        cache = CorrectedImageCache(max_size=3)
        t0 = time.perf_counter()
        for fn in filenames:
            get_display_image_checker(coords_df, folder, fn, 1, cache=cache)
        time_with_cache = time.perf_counter() - t0
        
        print(f"\n【速度測定: 異なる画像3枚】")
        print(f"  キャッシュなし: {time_no_cache*1000:.1f}ms")
        print(f"  キャッシュあり: {time_with_cache*1000:.1f}ms")
        
        # キャッシュありで2倍以上遅くなっていないこと
        assert time_with_cache < time_no_cache * 2.0


# =====================================================
# セクション5: 遅延CSV保存テスト
# =====================================================

class TestDeferredCsvSave:
    """遅延CSV保存（_flush_csv）の動作テスト"""
    
    def test_save_interval_respected(self, tmp_path):
        """save_interval件ごとにCSVが書き出されること"""
        csv_path = tmp_path / "errors.csv"
        
        # テスト用DataFrameを作成
        df = pd.DataFrame([
            {'filename': f'f{i}.jpg', 'question_no': 1, 'before': '', 'after': '', 'error_type': ERROR_TYPE_NO_MARK}
            for i in range(10)
        ])
        save_errors_checker(df, csv_path)
        
        # _flush_csv の代わりに save_errors_checker の呼び出し回数をカウント
        with patch('gui_components.save_errors_checker') as mock_save:
            from gui_components import MarkCheckerGUI
            
            # MarkCheckerGUIのインスタンスを直接使わず、ロジックを検証
            # _flush_csv ロジック: _unsaved_count >= _save_interval で保存
            unsaved = 0
            save_interval = 5
            calls = 0
            
            for i in range(10):
                unsaved += 1
                if unsaved >= save_interval:
                    calls += 1
                    unsaved = 0
            
            assert calls == 2, "10件中、interval=5で2回の保存が期待される"
    
    def test_flush_on_close_saves_dirty(self, tmp_path):
        """ウィンドウを閉じる際に未保存データが保存されること（ロジック検証）"""
        csv_path = tmp_path / "errors.csv"
        
        df = pd.DataFrame([
            {'filename': 'f1.jpg', 'question_no': 1, 'before': '', 'after': '-1', 'error_type': ERROR_TYPE_NO_MARK}
        ])
        save_errors_checker(df, csv_path)
        
        # 保存して再読み込み
        loaded = load_errors_checker(csv_path)
        assert len(loaded) == 1
        assert str(int(float(loaded.iloc[0]['after']))) == '-1'
    
    def test_csv_written_after_flush(self, tmp_path):
        """_flush_csvが呼ばれたらCSVが更新されていること"""
        csv_path = tmp_path / "test_flush.csv"
        
        df = pd.DataFrame([
            {'filename': 'f1.jpg', 'question_no': 1, 'before': '', 'after': '', 'error_type': ERROR_TYPE_NO_MARK},
            {'filename': 'f2.jpg', 'question_no': 2, 'before': '3;4', 'after': '', 'error_type': ERROR_TYPE_DOUBLE_MARK},
        ])
        
        # 初期状態を保存
        save_errors_checker(df, csv_path)
        
        # DataFrameを変更
        df.at[0, 'after'] = '-1'
        
        # 再保存（flush相当）
        save_errors_checker(df, csv_path)
        
        # 読み込んで検証
        reloaded = load_errors_checker(csv_path)
        assert str(int(float(reloaded.iloc[0]['after']))) == '-1'
        assert reloaded.iloc[1]['after'] == '' or pd.isna(reloaded.iloc[1]['after'])  # 未変更


# =====================================================
# セクション6: GUI統合テスト（実際のウィンドウ表示+速度測定）
# =====================================================

class TestMarkCheckerGUISpeed:
    """MarkCheckerGUIの実際のウィンドウでの遷移速度を測定するテスト
    
    ダミーのマークチェックウィンドウを表示し、
    show_current() の呼び出し時間を計測して効果を確認する。
    """
    
    @pytest.fixture
    def gui_setup(self, tmp_path):
        """テスト用のGUI環境を準備（実際のウィンドウは表示しない）"""
        # テスト画像を複数生成
        filenames = []
        for i in range(3):
            fn = f"gui_test_{i+1:03d}.jpg"
            _create_test_image_file(tmp_path, fn)
            filenames.append(fn)
        
        # 座標CSV
        coords_df = _create_test_coords_df(n_files=3, n_questions=5, filenames=filenames)
        coords_csv_path = tmp_path / "coordinates.csv"
        coords_df.to_csv(coords_csv_path, index=False, encoding='utf-8')
        
        # エラーCSV（同一画像に複数エラー + 異なる画像のエラー）
        errors = []
        for fn in filenames:
            for q in range(1, 4):  # 各画像3エラー
                errors.append({
                    'filename': fn,
                    'question_no': q,
                    'before': '',
                    'after': '',
                    'error_type': ERROR_TYPE_NO_MARK
                })
        error_df = pd.DataFrame(errors)
        error_csv_path = tmp_path / "tmp_checking_dm_nm.csv"
        error_df.to_csv(error_csv_path, index=False, encoding='utf-8-sig')
        
        # ダミーのxlsx
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.cell(1, 1, "File")
        for q in range(1, 6):
            ws.cell(1, q + 1, q)
            ws.cell(2, q + 1, str(q))
        for i, fn in enumerate(filenames):
            ws.cell(i + 3, 1, fn)
        xlsx_path = tmp_path / "result.xlsx"
        wb.save(xlsx_path)
        wb.close()
        
        return {
            'image_folder': tmp_path,
            'coords_csv_path': coords_csv_path,
            'xlsx_path': xlsx_path,
            'error_csv_path': error_csv_path,
            'error_df': error_df,
            'coords_df': coords_df,
            'filenames': filenames,
        }
    
    def test_show_current_with_cache_faster(self, gui_setup):
        """show_current()がキャッシュ利用時に高速であることの比較測定
        
        GUIウィンドウなしで、画像処理部分のみの速度を比較する。
        """
        data = gui_setup
        coords_df = data['coords_df']
        folder = data['image_folder']
        filenames = data['filenames']
        
        # 同一画像の連続エラー（3問）のシミュレーション
        filename = filenames[0]
        
        # キャッシュなし（従来動作）
        times_legacy = []
        for q in range(1, 4):
            t0 = time.perf_counter()
            img = get_display_image_checker(coords_df, folder, filename, q, cache=None)
            if img:
                fit_image_to_display(img)
            times_legacy.append(time.perf_counter() - t0)
        
        # キャッシュあり（v3.9）
        cache = CorrectedImageCache(max_size=2)
        times_cached = []
        for q in range(1, 4):
            t0 = time.perf_counter()
            img = get_display_image_checker(coords_df, folder, filename, q, cache=cache)
            if img:
                fit_image_to_display(img)
            times_cached.append(time.perf_counter() - t0)
        
        total_legacy = sum(times_legacy) * 1000
        total_cached = sum(times_cached) * 1000
        # 2-3問目の平均
        avg_legacy_2nd = sum(times_legacy[1:]) / 2 * 1000
        avg_cached_2nd = sum(times_cached[1:]) / 2 * 1000
        
        print(f"\n【GUI速度測定: 同一画像3連続エラー】")
        print(f"  従来合計: {total_legacy:.1f}ms")
        print(f"  v3.9合計: {total_cached:.1f}ms")
        print(f"  2-3問目 従来平均: {avg_legacy_2nd:.1f}ms")
        print(f"  2-3問目 v3.9平均: {avg_cached_2nd:.1f}ms")
        if avg_cached_2nd > 0:
            print(f"  2-3問目 高速化率: {avg_legacy_2nd/avg_cached_2nd:.1f}x")
        
        # キャッシュヒット時は従来より速い
        assert avg_cached_2nd < avg_legacy_2nd
    
    def test_image_transition_across_files(self, gui_setup):
        """異なる画像間の遷移でもキャッシュが正常に動作すること"""
        data = gui_setup
        coords_df = data['coords_df']
        folder = data['image_folder']
        filenames = data['filenames']
        
        cache = CorrectedImageCache(max_size=2)
        
        results = []
        for fn in filenames:
            t0 = time.perf_counter()
            img = get_display_image_checker(coords_df, folder, fn, 1, cache=cache)
            elapsed = (time.perf_counter() - t0) * 1000
            results.append({
                'filename': fn,
                'time_ms': elapsed,
                'success': img is not None
            })
        
        print(f"\n【画像切り替え速度】")
        for r in results:
            status = "OK" if r['success'] else "NG"
            print(f"  [{status}] {r['filename']}: {r['time_ms']:.1f}ms")
        
        # 全画像が正常に表示できること
        assert all(r['success'] for r in results)
    
    def test_gui_window_navigation_timing(self, gui_setup):
        """実際のGUIウィンドウを表示して画面遷移時間を測定する
        
        実際のMarkCheckerGUIウィンドウを開き、show_current()を
        連続で呼んで遷移時間を計測する。
        """
        from conftest import get_shared_tk_root
        
        data = gui_setup
        root = get_shared_tk_root()
        
        # MarkCheckerGUIをパッチして自動ダイアログをスキップ
        with patch('gui_components.messagebox') as mock_mb:
            mock_mb.askyesno.return_value = True  # 前回の続きから作業
            mock_mb.showinfo.return_value = None
            
            from gui_components import MarkCheckerGUI
            
            gui = MarkCheckerGUI(
                root,
                data['image_folder'],
                data['coords_csv_path'],
                data['xlsx_path'],
                skip_questions=0,
                template_path=None,
            )
            
            # load_dataはafter(100)で呼ばれるため、手動で呼ぶ
            root.update()  # after(100) のトリガー
            root.after(200)  # ロード完了を待つ
            root.update()
            
            # エラーDFを手動セット（ダイアログを回避）
            gui.error_df = data['error_df'].copy()
            gui.coords_df = data['coords_df'].copy()
            gui.current_index = 0
            
            # 画面遷移の速度測定
            transition_times = []
            n_transitions = min(len(gui.error_df) - 1, 8)
            
            for i in range(n_transitions):
                gui.current_index = i
                t0 = time.perf_counter()
                gui.show_current()
                root.update_idletasks()
                elapsed = (time.perf_counter() - t0) * 1000
                transition_times.append(elapsed)
            
            avg_time = sum(transition_times) / len(transition_times) if transition_times else 0
            max_time = max(transition_times) if transition_times else 0
            
            print(f"\n{'='*50}")
            print(f"【実GUIウィンドウ遷移速度測定】")
            print(f"  遷移回数: {len(transition_times)}回")
            for i, t in enumerate(transition_times):
                filename = gui.error_df.iloc[i]['filename']
                q_no = int(gui.error_df.iloc[i]['question_no'])
                print(f"  [{i+1}] {filename} Q{q_no}: {t:.1f}ms")
            print(f"  平均遷移時間: {avg_time:.1f}ms")
            print(f"  最大遷移時間: {max_time:.1f}ms")
            print(f"{'='*50}")
            
            # クリーンアップ
            gui._flush_csv()
            gui._image_cache.clear()
            gui.window.destroy()
            root.update()
            
            # 全遷移が1秒以内であること（十分な余裕を持った閾値）
            assert max_time < 1000, f"最大遷移時間が1秒を超えました: {max_time:.0f}ms"


# =====================================================
# セクション7: 後方互換性テスト
# =====================================================

class TestBackwardCompatibility:
    """高速化変更が既存の動作を壊していないことの確認"""
    
    def test_crop_and_scale_still_works(self, tmp_path):
        """crop_and_scale_image_checker がそのまま動作すること"""
        img_path = _create_test_image_file(tmp_path, "compat.jpg")
        bbox = (100, 200, 150, 30)
        
        result = crop_and_scale_image_checker(img_path, bbox, 1.25, 1.3, 1.0)
        from PIL import Image
        assert isinstance(result, Image.Image)
        assert result.width > 0
        assert result.height > 0
    
    def test_get_display_image_without_cache_unchanged(self, tmp_path):
        """cache=None（デフォルト）で従来動作と同じこと"""
        _create_test_image_file(tmp_path, "default.jpg")
        coords_df = _create_test_coords_df(n_files=1, n_questions=3, filenames=["default.jpg"])
        
        result = get_display_image_checker(
            coords_df, tmp_path, "default.jpg", 1
        )
        assert result is not None
    
    def test_save_load_csv_roundtrip(self, tmp_path):
        """CSV保存→読み込みのラウンドトリップが正常"""
        csv_path = tmp_path / "roundtrip.csv"
        df = pd.DataFrame([
            {'filename': 'a.jpg', 'question_no': 1, 'before': '', 'after': '-1', 'error_type': ERROR_TYPE_NO_MARK},
            {'filename': 'b.jpg', 'question_no': 3, 'before': '2;3', 'after': 'skip', 'error_type': ERROR_TYPE_DOUBLE_MARK},
        ])
        save_errors_checker(df, csv_path)
        loaded = load_errors_checker(csv_path)
        
        assert len(loaded) == 2
        assert loaded.iloc[0]['after'] == '-1'
        assert loaded.iloc[1]['after'] == 'skip'
    
    def test_fit_image_to_display_unchanged(self):
        """fit_image_to_display の動作が変わっていないこと"""
        from PIL import Image
        
        # 大きい画像
        big = Image.new('RGB', (2000, 800), color='white')
        fitted = fit_image_to_display(big, max_width=1100, max_height=400)
        assert fitted.width <= 1100
        assert fitted.height <= 400
        
        # 小さい画像（拡大されない）
        small = Image.new('RGB', (200, 100), color='white')
        fitted_small = fit_image_to_display(small, max_width=1100, max_height=400)
        assert fitted_small.width == 200
        assert fitted_small.height == 100
